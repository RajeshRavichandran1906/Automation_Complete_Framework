from common.lib.base import BasePage
from common.lib.utillity import UtillityMethods as utillobject
from common.pages.active_miscelaneous import Active_Miscelaneous as miscelaneousobject
from common.pages.ia_miscelaneous import IA_Miscelaneous as ia_mis_object
from common.pages.ia_run import IA_Run as iarunobject
from common.pages.ia_ribbon import IA_Ribbon as iaribbonobject
from common.pages.active_filter_selection import Active_Filter_Selection as activefilter_selectionobject
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from common.pages.active_pivot_comment import Active_Pivot_Comment as active_pivot_commentobj

class Active_Report(BasePage):
    """ Inherit attributes of the parent class = Baseclass """

    def __init__(self, driver):
        super(Active_Report, self).__init__(driver)
        
    
    """****************************************** This is for Common Section. ************************************"""
    
    def wait_for_number_of_element(self, element_css, expected_number=None, time_out=60):
        miscelaneousobject.wait_for_object(self, element_css, 'number', expected_number=expected_number, time_out=time_out)
    
    def wait_for_visible_text(self, element_css, visble_element_text=None, time_out=60):
        miscelaneousobject.wait_for_object(self, element_css, 'text', visble_element_text=visble_element_text, time_out=time_out)
    
    def invoke_report_tool_using_api(self, master_file_path, mrid=None, mrpass=None, repository_path=None):
        '''
        Desc:- This function will invoke report tool using api call.
        '''
        ia_mis_object.invoke_ia_tool_using_api_(self, tool='report', master=master_file_path, mrid=mrid, mrpass=mrpass, folder_path=repository_path)
        
    def invoke_report_in_edit_mode_using_api(self, fex, mrid=None, mrpass=None, repository_path=None):
        '''
        Desc:- This function will invoke report tool in edit mode using api call.
        '''
        ia_mis_object.invoke_ia_tool_in_edit_mode_using_api(self, fex, tool='chart', mrid=mrid, mrpass=mrpass, folder_path=repository_path)
        
    def run_active_report_using_api(self, fex_name, column_css="#ITableData0 tr:nth-child(3) td:nth-child(1)", synchronize_visible_element_text=None, repository_path=None):
        '''
        Desc:- This function is used to switch to a new window.
        '''
        miscelaneousobject.run_active_reports_using_api(self, fex_name, mrid='mrid', mrpass='mrpass', folder_path=repository_path)
        miscelaneousobject.wait_for_object(self, column_css, option='text', visble_element_text=synchronize_visible_element_text, time_out=290)
        
    def create_active_report_dataset(self, file_name, table_css="#ITableData0", desired_no_of_rows=None, starting_rownum=None):
        '''
        Desc :-This function is used to create active report dataset in the form of xlsx format
        reportobj.create_active_report_dataset('C123456_Ds01.xlsx', desired_no_of_rows=5, starting_rownum=5)
        '''
        Active_Report.create_reporttable_dataset(self, file_name, table_css=table_css, desired_no_of_rows=desired_no_of_rows, starting_rownum=starting_rownum)
           
    def verify_active_report_dataset(self, file_name, msg, table_css=None, desired_no_of_rows=None, starting_rownum=None):
        '''
        Desc : Verify the compared table data using start and end rowcolumn
        :Usage: verify_table_data_using_start_end_rowcolumn
        '''
        x=Active_Report.compare_table_dataset(self, file_name, table_css, desired_no_of_rows=desired_no_of_rows, starting_rownum=starting_rownum)
        utillobject.asequal(self, len(x), 1, msg+ ' Row,Column -'+ str(x))
        
    def verify_no_of_autolinks_in_activereport(self, table_css, colnum, expected_autolink_rows, msg, color_name ='cerulean_blue_2'):
        '''
        Desc:- This function is used to verify the number of autolinks are present in the active report
        report_obj.verify_no_of_autolinks_in_activereport(self, table_css, colnum, expected_autolink_rows, msg) 
        '''
        iarunobject.verify_activereport_autolink(self, table_css, colnum, expected_autolink_rows, msg, color_name)
        
    def create_data_set_using_table_rowid(self,table_id, tr_id, file_name, color='default'):
        '''
        This function will create data in the form of xlsx format and can be used to create highlight data also.
        '''
        utillobject.create_data_set(self,table_id, tr_id, file_name, color=color)
        
    def verify_data_set_using_table_rowid(self, table_id, tr_id, file_name, msg, color='default'):
        '''
        Desc:- This function is used to create active report data at runtime by passing tablerowid: Visualization active rpeort can be done using this function.
        To verify highlight data also can be done using this function.
        report_obj.verify_data_set_using_table_rowid("ITableData0", "I0r", "example.xlsx")
        '''
        utillobject.verify_data_set(self, table_id, tr_id, file_name, msg, color=color)
        
    def verify_visualize_bar_added_in_activereport(self, table_css, field_col_no, visualize_bar_color, message, expected_visual_bars=None):
        '''
        Desc:- This function is used to verify thevisualization present in the report
        report_obj.verify_active_report_visualization("#ITableData0", 1, 'green', "Step X: Verify visualization")
        '''
        miscelaneousobject.verify_visualize_bar_added_in_activereport(self, table_css, field_col_no, visualize_bar_color, message, expected_visual_bars)
        
    def verify_column_heading(self, table_id, expected_list, msg):
        '''
        This function will verrify active report column heading
        '''
        miscelaneousobject.verify_column_heading(self, table_id, expected_list, msg)
        
    def verify_page_summary(self, page_num, expected_title, msg):
        '''
        Desc :-This function is used to verifypage summary in active report.
        report_obj.verify_page_summary('0','9of107records,Page1of1', 'Step 05: Verify Page summary')
        '''
        miscelaneousobject.verify_page_summary(self, page_num, expected_title, msg)
        
    def move_active_popup(self,index, x, y,**kwargs):
        '''
        This functiion will move active popup from one location to desired location.
        '''
        miscelaneousobject.move_active_popup(self, index, x, y,**kwargs)
        
    def create_filter(self, rownum, condition_name, value_list_type='small', popup_id='wall1', large_list_pgdn=0, table_id='ITableData0', **kwargs):
        '''
        Desc :- This function is used to create filter
        '''
        activefilter_selectionobject.create_filter(self, rownum, condition_name, value_list_type, popup_id, large_list_pgdn, table_id, **kwargs)
        
    def close_filter_dialog(self, popup_id='wall1'):
        '''
        Desc:-This function is used to close the filter dialog ok button
        '''
        activefilter_selectionobject.close_filter_dialog(self, popup_id)
        
    def select_filter_values(self, index1, list_type, css, val, index, large_list_pgdn,**kwargs):
        '''
        Desc :This function is used to create filter
        '''
        activefilter_selectionobject.select_filter_values(self, index1, list_type, css, val, index, large_list_pgdn,**kwargs)
    
    def verify_filter_values(self, rownum, expected_menu_list, msg):
        '''
        Desc :This function is used to verify filter values
        '''
        activefilter_selectionobject.verify_filter_values_menu_list(self, rownum, expected_menu_list, msg)
        
    def filter_button_click(self, btn_name, popup_id='wall1'):
        '''
        Desc:- This function is used to click filter button in the filter dialog
        '''
        activefilter_selectionobject.filter_button_click(self, btn_name, popup_id)
        
    def select_menu_items(self, table_id, column_no, column_name, *args, **kwargs):
        '''
        Desc:- This fucntion is used to select column menu items in active report column starts with 0
        report_obj.select_menu_items("ITableData0", 4, 'Revenue')
        '''
        miscelaneousobject.select_menu_items(self, table_id, column_no, column_name, *args, **kwargs)
    
    def verify_menu_items(self, table_id, colnum, navigate_pattern, expected_menu_list, msg, all_items='yes', **kwargs):
        '''
        Desc:- This fucntion is used to verify column menu items in active report column starts with 0
        report_obj.select_menu_items("ITableData0", 4, None, ['Filter'], 'Step 9: verify')
        '''
        miscelaneousobject.verify_menu_items(self, table_id, colnum, navigate_pattern, expected_menu_list, msg, all_items=all_items, **kwargs)
        
    def create_new_item(self, popup_id, popup_instance, item_list, **kwargs):
        '''
        Desc :- This function is used to select menu items from the first dropdown options in the active chart toolbar
        '''
        active_pivot_commentobj.create_new_item(self, popup_id, popup_instance, item_list, **kwargs)
        
    def verify_filter_selection_dialog(self,verify, msg,*args,popup_id='wall1'):
        '''
        Desc:-This function will verify filter selection dialog added fields condition.
        '''
        activefilter_selectionobject.verify_filter_selection_dialog(self,verify, msg,*args,popup_id=popup_id)
    
    def select_activereport_field_menu_items (self, table_id, rownum, colnum, field_val,**kwargs):#please don't use this function
        '''
        Desc:- This function is used to select active report field
        report_obj.select_activereport_filed_menu_items('ITableData0', 1, 1, 'New York')
        '''
        miscelaneousobject.select_field_menu_items(self, table_id, rownum, colnum, field_val, **kwargs)
        
    def select_option_from_field_menu(self, option_name, table_id='ITableData0', rownum=0, colnum=0):
        '''
        Desc :- This function is used to select field menu items from report field
        '''
        miscelaneousobject.select_field_menu_items(self, table_id, rownum, colnum, option_name)
        
    def verify_filter_buttons(self, expected_btn_level_text, msg):
        '''
        Desc :- This function will verify filter buttons.
        '''
        activefilter_selectionobject.verify_filter_buttons(self, expected_btn_level_text, msg)
    
    def verify_activereport_field_menu_items(self, table_id, rownum, colnum, expected_value, msg, **kwargs):#please don't use this function
        '''
        Desc:-This function is used to verify field menu items
        report_obj.verify_activereport_field_menu_items(self, 'ITableData0', 1, 2, ['Comments', 'Filter Cell'], "Step X: Verify menu")
        '''
        miscelaneousobject.verify_field_menu_items(self, table_id, rownum, colnum, expected_value, msg, **kwargs)
    
    def verify_field_menu_options(self, expected_value, msg, table_id='ITableData0', rownum=0, colnum=0):
        '''
        Desc :- This function is used to verify field menu items from report field
        '''
        miscelaneousobject.verify_field_menu_items(self, table_id, rownum, colnum, expected_value, msg)
        
    def add_global_condition_field(self,field_name, parent_menu_css=None):
        '''
        Desc :- This function will add fields in the global filter
        '''
        activefilter_selectionobject.add_global_condition_field(self, field_name, parent_menu_css=parent_menu_css)
    
    def verify_comment_tooltip(self, expected_tooltip_list, msg, table_id='ITableData0', rownum=0, colnum=0):
        '''
        Desc :- This function is used to verify tool-tip from report field
        '''
        miscelaneousobject.verify_comment_tooltip(self, table_id, rownum, colnum, expected_tooltip_list, msg)
    
    def verify_comment_window(self, msg):
        '''
        Desc :- This function is used to verify comment window title
        '''
        utillobject.synchronize_until_element_is_visible(self, '#wall1 #wtop1 .arWindowBarTitle', 30)
        title_text = utillobject.validate_and_get_webdriver_object(self, '#wall1 #wtop1 .arWindowBarTitle', 'Comment window title').text.strip()
        utillobject.asequal(self, 'Add Comment', title_text, msg)
    
    def enter_text_on_comment_dialog(self, comment_text_string, popup_id='wall1'):
        '''
        Desc :- This function is used to enter value in comment box
        '''
        active_pivot_commentobj.create_comment(self, popup_id, comment_text_string)
    
    def close_comment_dialog(self):
        '''
        Desc :- This function is used to close comment window
        '''
        active_pivot_commentobj.close_comment_dialog(self)
        
    def verify_comment_field(self, msg, table_id='ITableData0', rownum=0, colnum=0, expected_field_val='[*]'):
        '''
        Desc :- This function is used to verify comment cell from report field
        '''
        miscelaneousobject.verify_comment_field(self, table_id, rownum, colnum, expected_field_val, msg)
    
    def verify_expandable_comment_tooltip(self, expected_tooltip_list, msg, table_id='ITableData0', rownum=0, colnum=0):
        '''
        Desc :- This function is used to verify expand comment cell from report field
        '''
        miscelaneousobject.verify_expandable_comment_tooltip(self, table_id, rownum, colnum, expected_tooltip_list, msg)
    
    def verify_horizontal_page_scroll(self, msg, table='IWindowBody0', scroll='yes'):
        '''
        Desc :- This function is used to verify horizontal page scroll
        '''
        miscelaneousobject.verify_horizontal_page_scroll(self, table, scroll, msg)
    
    def verify_freeze_column(self, msg, scroll='yes', scroll_window_id='IScrollWindowBody0'):
        '''
        Desc :- This function is used to verify freeze column
        '''
        miscelaneousobject.verify_freeze_column(self, scroll, scroll_window_id, msg)
    
    def navigate_page(self, navigate_type, *args):
        '''
        Desc :-This function is used to move next page from page summary in active report.
        Usage : navigate_page('next_page')
        '''
        miscelaneousobject.navigate_page(self, navigate_type, *args)
        
    def active_report_options(self, tab_name, **kwargs):
        '''
        Desc :- This function is used to verify active report options.
        '''
        iaribbonobject.active_report_options(self, tab_name,**kwargs)
        
    def create_reporttable_dataset(self, file_name, table_css=None, desired_no_of_rows=None, starting_rownum=None):
        '''
        Desc :-This function is used to create report dataset in the form of xlsx format
        reportobj.create_reporttable_data_set('C123456_Ds01.xlsx', table_css='#ITableData0', desired_no_of_rows=10, starting_rownum=2)
        '''
        rows_css = table_css + " > tbody > tr"
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        no_of_rows=desired_no_of_rows if desired_no_of_rows!=None else len(table_rows) 
        start_rownum=starting_rownum if starting_rownum!=None else 0        
        for r in range(start_rownum,no_of_rows):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text
                s.cell(row=r + 1, column=c + 1).value = str(value)
        wb.save(os.path.join(os.path.join(os.getcwd(), "data"), file_name))
            
    def compare_table_dataset(self, file_name, table_css=None, desired_no_of_rows=None, starting_rownum=None):
        wb1 = load_workbook(os.path.join(os.path.join(os.getcwd(), "data"), file_name))
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        rows_css = table_css + " > tbody > tr"
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        no_of_rows=desired_no_of_rows if desired_no_of_rows!=None else len(table_rows)
        start_rownum=starting_rownum if starting_rownum!=None else 0
        for r in range(start_rownum, no_of_rows):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(0, len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text
                if s1.cell(row=r + 1, column=c + 1).value != None and str(value) != '':
                    if s1.cell(row=r + 1, column=c + 1).value.replace(' ','').lower() == str(value).replace(' ','').lower():
                        status=[0]
                        continue
                    else:
                        status=[r+1,c, 'expected',s1.cell(row=r + 1, column=c + 1).value.replace(' ','').lower(), 'actual',str(value).replace(' ','').lower()]
                        return (status)
        return (status)