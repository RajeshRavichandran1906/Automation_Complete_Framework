from common.lib.base import BasePage
from common.lib.utillity import UtillityMethods as utillobject
from common.lib.core_utility import CoreUtillityMethods as coreutillobject
from common.pages.visualization_metadata import Visualization_Metadata as visual_metaobj
from common.pages.core_metadata import CoreMetaData as core_metadata_obj
from common.pages.visualization_ribbon import Visualization_Ribbon as visual_ribbonobj
from common.pages.ia_miscelaneous import IA_Miscelaneous as miscelaneousobject
from common.pages.ia_ribbon import IA_Ribbon as ribbonobject
from common.pages.ia_resultarea import IA_Resultarea as ia_resultobject
from common.pages.ia_group_dialog import GroupDialog
from common.pages.visualization_resultarea import Visualization_Resultarea as visual_resultareaobj
from selenium.webdriver.support.color import Color
from openpyxl import Workbook
from openpyxl import load_workbook
import os
from common.pages.ia_run import IA_Run as iarunobj
from common.lib.javascript import JavaScript as js_obj
from common.pages.filter_metadata import FilterMetaData as filter_metadata_obj
from common.pages.ia_metadata import IA_Metadata as metaobject
from common.pages.ia_ribbon import IA_Ribbon as ia_ribbon_obj
from common.locators.ia_ribbon_locators import IaRibbonLocators
from common.pages.ia_join import IAJoin
from common.pages.define_compute import DefineCompute

class Report(BasePage):
    """ Inherit attributes of the parent class = Baseclass """

    def __init__(self, driver):
        super(Report, self).__init__(driver)
        

    """****************************************** This is for Common Section. ************************************"""
    
    def wait_for_number_of_element(self, element_css, expected_number=None, time_out=60):
        miscelaneousobject.wait_for_object(self, element_css, 'number', expected_number=expected_number, time_out=time_out)
    
    def wait_for_visible_text(self, element_css, visble_element_text=None, time_out=60):
        miscelaneousobject.wait_for_object(self, element_css, 'text', visble_element_text=visble_element_text, time_out=time_out)
    
    def take_preview_snapshot(self, file_name, step_num):
        '''
        Desc:- This function will take screenshot only for preview area
        :param file_name:- 'C123435'
        :param step_num:- '05'
        :usage take_preview_snapshot('C123435' ,'05')
        '''
        element=self.driver.find_element_by_id('resultArea')
        file_name=file_name + '_Actual_Step_' + str(step_num)
        utillobject.take_snapshot(self, element, file_name)
        
    def take_run_window_snapshot(self, file_name, step_num):
        '''
        Desc:- This function will take screenshot only for run window
        :param file_name:- 'C123435'
        :param step_num:- '05'
        :usage take_preview_snapshot('C123435' ,'05')
        '''
        file_name=file_name + '_Actual_Step_' + str(step_num)
        utillobject.take_browser_snapshot(self,file_name)
        
    def switch_to_new_window(self):
        '''
        Desc:- This function is used to switch to a new window.
        '''
        coreutillobject.switch_to_new_window(self)
        
    def switch_to_previous_window(self):
        '''
        Desc:- This function is used to switch to a previous window.
        '''
        coreutillobject.switch_to_previous_window(self)
        
    def switch_to_frame(self, frame_css="[id^='ReportIframe']"):
        '''
        Desc:- This function is used to switch to a Frame. Here you need to provide the css of the frame to be switched to.
        '''
        coreutillobject.switch_to_frame(self, frame_css=frame_css, timeout=60)
    
    def switch_to_default_content(self):
        '''
        Desc:- This function is used to switch back to default content from a frame.
        '''
        coreutillobject.switch_to_default_content(self)
        
    def api_logout(self):
        """
        Desc: This function uses new function to pass the api logout url
        """
        utillobject.wf_logout(self)
        
    def set_browser_window_size(self, x='998', y='768'):
        '''
        Desc:-This function is used to resize the browser based on the given width and height
        '''
        utillobject.set_browser_window_size(self, x, y)
        
    def maximize_browser(self):
        '''
        Desc:-This function is used to maximize the browser window
        '''
        self.driver.maximize_window()
    
    """*****************************************************************RUN AND EDIT FEX*********************************************************************************"""
            
    def run_fex_using_api_url(self, folder_name, fex_name=None, mrid=None, mrpass=None, home_page='New', run_table_css="table[summary='Summary'] > tbody > tr", no_of_element=1, wait_time=0):
        '''
        Desc:- This function will run report/chart using api link and sign in as user defined in config file
        Usage: report_obj.run_fex_using_api(subfolder_name='Auto_Link', fex_name='Report_Store_Product_Metrics', mrid='mrid', mrpass='mrpass')
        '''
        wait_time= self.report_long_timesleep if wait_time==0 else wait_time
        miscelaneousobject.run_fex_using_api(self, folder_name, fex_name=fex_name, mrid=mrid, mrpass=mrpass, home_page=home_page)
        utillobject.synchronize_until_element_is_visible(self, run_table_css, wait_time)
        
    def edit_fex_using_api_url(self, fex_name, folder_name=None, mrid="mrid", mrpass="mrpass"):
        '''
        Desc:- This function will edit retail samples report using api link and sign in as user defined in config file
        Usage: report_obj.edit_retailsamples_using_api(subfolder_name='Auto_Link',fex_name='Report_Store_Product_Metrics', mrid='mrid', mrpass='mrpass')
        '''
        miscelaneousobject.edit_fex_using_api(self, folder_name, tool="report", fex_name=fex_name, mrid=mrid, mrpass=mrpass)
        
    def invoke_ia_tool_using_api_login(self, tool='report', master='baseapp', mrid=None, mrpass=None, report_css="#resultArea", no_of_element=1, wait_time=0):
        '''Desc:-This function is used to invoke ia tool
        Usage: report_obj.invoke_ia_tool_using_api_login(master='ibisamp/car')
        '''
        wait_time= self.report_medium_timesleep if wait_time==0 else wait_time
        miscelaneousobject.invoke_ia_tool_using_api(self, tool, master, mrid, mrpass)
        Report.wait_for_number_of_element(self, report_css, no_of_element, wait_time)
    
    def invoke_report_without_master_file_by_api(self, mrid='mrid', mrpass='mrpass', folder=None):
        """
        Description : Invoke IA report tool without master file by using api url. 
        Usage : invoke_report_without_master_file_by_api()
        """
        utillobject.invoke_ia_tool_without_master_file_using_api(self, 'report', mrid, mrpass, folder)
        utillobject.synchronize_with_visble_text(self, "#IbfsOpenFileDialog7_btnOK", "Open", self.report_long_timesleep)
    
    def invoke_ia_tool_using_new_api_login(self, tool='report', master='baseapp', mrid=None, mrpass=None, report_css="#resultArea", no_of_element=1, wait_time=0):
        '''Desc:-This function is used to invoke ia tool usng the new folder structure
        Usage: report_obj.invoke_ia_tool_using_new_api_login(master='ibisamp/car')
        '''
        wait_time= self.report_medium_timesleep if wait_time==0 else wait_time
        miscelaneousobject.invoke_ia_tool_using_api_(self, tool, master, mrid, mrpass)
        Report.wait_for_number_of_element(self, report_css, no_of_element, wait_time)
        
    """**************************************************DATA PANE***************************************************************************************************"""
    def double_click_on_datetree_item(self, field_name, field_position=1):
        '''
        Desc:- This function is used to double click and select the field from data tree.
        :usage double_click_on_datetree_item('Revenue', 1)
        '''
        core_metadata_obj.double_click_on_data_filed(self, field_name, field_position)
    
    def right_click_on_datetree_item(self, field_name, field_position, context_menu_path):
        '''
        Desc:- This function is used to right click on the field in data tree and select the item from the context menu.
        :usage right_click_on_datetree_item("Product,Category",1,'Add To Query->Horizontal Axis')
        '''
        core_metadata_obj.right_click_on_data_filed(self, field_name, 'right', field_position, context_menu_path)
    
    def select_data_field_context_menu(self, data_field_path, context_menu_path, field_position=1):
        """
        This method used to expand and right click on specific data field and select context menu
        :usage : select_data_field_context_menu("CAR", "Sort")
        """
        core_metadata_obj.select_data_field_context_menu(self, data_field_path, context_menu_path, field_position)
        
    def select_datatree_field_context_menu(self, data_field_path, context_menu_path):
        '''
        Desc:-This function is used to right click on datafield and select the menu from the bipopup menu
        '''
        core_metadata_obj.select_data_field_context_menu(self, data_field_path, context_menu_path)
        
    def collapse_datatree_field_section(self, data_field_section_path, find_from_top=True):
        '''
        Desc:-This function is used to collapse the data field 
        '''
        core_metadata_obj.collapse_data_field_section(self, data_field_section_path, find_from_top)
           
    def expand_datatree_field_section(self, data_field_section_path, find_from_top=True):
        '''
        Desc:-This function is used to expand data field
        '''
        core_metadata_obj.expand_data_field_section(self, data_field_section_path, find_from_top)
        
    def verify_datapane_toggle_button_tooltip_in_enablemode(self):
        '''
        Desc:-This function will hover on datapane toggle button to verify tooltip
        '''
        visual_metaobj.verify_datapane_toggle_button_tooltip(self, "Enable Path Enforcement")
        
    def verify_datapane_toggle_button_tooltip_in_disablemode(self):
        '''
        Desc:-This function will hover on datapane toggle button to verify tooltip
        '''
        visual_metaobj.verify_datapane_toggle_button_tooltip(self, "Disable Path Enforcement")
        
    def verify_grayedout_field_in_data_pane(self, field_type, field_name, position, color_to_verify='Trolley_Grey', msg="Step X"):
        '''
        Desc:-This function will verify the given field is grayed out
        '''
        visual_metaobj.verify_field_in_data_pane(self,field_type, field_name, position, color_to_verify=color_to_verify, msg=msg)
        
    def verify_grayedout_field_in_data_pane_list_view(self,field_name_value,field_title_value,Stepno,color_to_verify='Trolley_Grey'):
        '''
        Desc:-This function will verify the given field is grayed out in datapane list view
        '''
        field_name = field_name_value+' '+field_title_value
        visual_metaobj.verify_field_in_data_pane_list_view(self, field_name=field_name,msg=Stepno,color_to_verify=color_to_verify)

    def click_datapane_toggle_button(self):
        '''
        Desc:- This function will click on datapane toggle button
        '''
        visual_metaobj.click_datapane_toggle_button(self)
        
    def right_click_on_field_under_query_tree(self, field_name, field_position, context_menu_path):
        '''
        Desc:- This function is used to right click on the field in query tree and select the item from the context menu
        '''
        metaobject.select_querytree_field(self, field_name, 'right', field_position, context_menu_path)
        
    def right_click_on_field_in_filterbox(self, field_name, field_position=1, context_menu_path=None):
        '''
        Desc:- This function is used to right click on the field in filter box and select the item from the context menu
        '''
        visual_metaobj.select_filterbox_field(self, field_name, field_position, 'right', context_menu_path)
        
    """*******************************************************DRAG AND DROP******************************************************************************************"""
        
    def drag_field_from_data_tree_to_query_pane(self, field_name, field_position, bucket_name, bucket_position=1,ele_loc='bottom_middle'):
        '''
        Desc:- This function is used to drag and drop data field to the specified bucket in query tree.
        :param - field_position: 1,2 .. (first match is 1)
        :param - bucket_position: 1,2 .. (first match is 1)
        :param - To replace field in query pan ele_loc="middle"
        :usage drag_field_from_data_tree_to_query_pane("PRICE_DOLLARS_BIN_1",1,"Model",1)
        '''
        visual_metaobj.drag_and_drop_from_data_tree_to_query_tree(self, field_name, field_position, bucket_name,bucket_position,ele_loc=ele_loc)
    
    def drag_and_drop_from_data_tree_to_filter(self, field_name, field_position=1):
        '''
        Desc:- This function is used to drag from data tree and drop in filter.
        :usage drag_and_drop_from_data_tree_to_filter('Discount','Color')
        '''
        visual_metaobj.drag_and_drop_from_data_tree_to_filter(self, field_name, field_position)
        
    def drag_field_within_query_pane(self, source_item_name, target_item_name):
        '''
        Desc:- This function is used to drag and drop data field within query tree.
        :usage drag_field_within_query_pane('Discount','Color')
        '''
        visual_metaobj.drag_and_drop_within_query_tree(self, source_item_name, target_item_name)
    
    def drag_query_filed_to_another_query_field(self, source_field, target_field, source_field_position=1, target_field_position=1, source_field_loc='middle', target_field_loc='bottom_middle'):
        """
        Description : Drag the query field to another query filed.
        :Usage - drag_query_filed_to_another_query_field("SALES", "DEALER_COST", target_field_loc='top_middle')
        """
        core_metadata_obj.drag_query_filed_to_another_query_field(self, source_field, target_field, source_field_position, target_field_position, source_field_loc, target_field_loc)
        
    """***************************************************QUERY PANE**************************************************************************************************"""
    
    def verify_all_fields_in_query_pane(self, expected_fields, msg):
        """
        Desc: This function will verify all fields listed in querypane
        """
        visual_metaobj.verify_query_panel_all_field(self, expected_fields,msg)
        
    def verify_field_in_querypane(self, bucket_type,field_name, position, msg='Step X', color_to_verify=None, font_to_verify=None):
        '''
        Desc :- This function is used to verify field which is available in querytree, field colour and field font style also we can verify.
        report_obj.verify_field_in_querypane('By','Gross Profit',1, msg="Step x: Verify query tree field name and colur", color_to_verify="light_gray", font_to_verify="gray")
        '''
        visual_metaobj.verify_field_in_query_pane(self,bucket_type,field_name, position, msg=msg, color_to_verify=color_to_verify, font_to_verify=font_to_verify)
    
    """*******************************************************************FILTER BOX*********************************************************************************"""
    
    def verify_filter_pane_field(self, field_name, position, msg, **kwargs):
        """
        Desc: This function will verify fields in filter pane
        """
        visual_metaobj.verify_filter_pane_field(self, field_name, position, msg, **kwargs)
        
    def save_as_from_application_menu_item(self, file_name, target_table_path=None, application_menu_item_name='save_as', file_type=None):
        """
        Desc: This function will Click IA application button and click save as under SmokeTest MyContent folder
        """
        visual_ribbonobj.select_visualization_application_menu_item(self, application_menu_item_name)
        if target_table_path != None :
            utillobject.expand_domain_folders_in_open_dialog(self, target_table_path=target_table_path)
        utillobject.ibfs_save_as(self, file_name, file_type=file_type)
    
    def verify_grayedout_fields_in_query_pane(self, expected_fields_list, step_num, comparison_type='asequal'):
        """
        Description : Verify grayed out fields in query pane
        :Usage : verify_grayedout_fields_in_query_pane(['SALES'], '02.01')
        """
        core_metadata_obj.verify_grayedout_fields_in_query_pane(self, expected_fields_list, step_num, comparison_type)
    
    def verify_descending_fields_in_query_pane(self, expected_fields_list, step_num, comparison_type='asequal'):
        """
        Description : Verify descending fields in query pane
        :Usage : verify_grayedout_fields_in_query_pane(['SALES'], '02.01')
        """
        core_metadata_obj.verify_descending_fields_in_query_pane(self, expected_fields_list, step_num, comparison_type)
    
    """*******************************************************************AUTO LINK***************************************************************************************"""
        
    def select_report_autolink_tooltip_runtime(self, parent_table_css, row, col, tooltip_path, verify_tooltip=None, msg=None, **kwargs):
        '''
        Desc:-This function is used to select hyperlink value at runtime
        kwargs: verify_type='asin'
        report_obj.select_report_autolink_tooltip('table[summary='Summary']', 5, 6, "Auto Links->AUTOLINK_TARGET01")
        '''
        iarunobj.select_report_autolink_tooltip(self, parent_table_css, row, col, tooltip_path, verify_tooltip=verify_tooltip, msg=msg, **kwargs)
        
    def verify_autolink_in_preview(self, table_css, field_value, expected_no_of_hlinks, msg, **kwargs):
        '''
        Desc :-This function is used to verify number of autolinks are present in the report
        report_obj.verify_autolink_in_reportdata('TableChart_1', 5, 6, Auto Links->AUTOLINK_TARGET01, verify_tooltip=['Drill up to Model'], msg="Step x:Verify tooltip")
        '''
        ia_resultobject.verify_autolink(self, table_css, field_value, expected_no_of_hlinks, msg, **kwargs)
        
    def verify_autolink_in_runtime(self, table_css, field_value, rownum, colnum, expected_no_of_hlinks, msg, color_name='cerulean_blue_2'):
        '''
        Desc:- This function is used to verify auto links in runtime
        report_obj.verify_autolink_in_runtime(self, table_css, field_value, rownum, colnum, expected_no_of_hlinks, msg, color_name='cerulean_blue_2')
        '''
        iarunobj.verify_autolink(self, table_css, field_value, rownum, colnum, expected_no_of_hlinks, msg, color_name=color_name)
    
    def select_report_autolink_tooltip_using_actionchains(self, row, col, tooltip_path, parent_table_css="table[summary='Summary']"):
        """
        Description : Click on specified table cell and select drill down menu using action chains.
        :usage - select_report_autolink_tooltip_using_actionchain(5, 1, "Drill up to Store Business Sub Region")
        """
        iarunobj.select_report_autolink_tooltip_using_actionchain(self, row, col, tooltip_path, parent_table_css)
        
    """********************************************************************RIBBON BAR*************************************************************************************"""
    
    def verify_checked_class_property_for_selected_object(self, web_element, msg, check_property=True):
        '''
        Desc :-Verify selected class property for the given webelement
        report_obj.verify_checked_class_property_for_selected_object(self, web_element, "Step x:Verify the selected object value")
        '''
        utillobject.verify_checked_class_property(self, web_element, msg, check_property=check_property)
        
    def select_visualization_application_menu_item(self, application_menu_item_name):
        '''
        Desc: This function is specific to click on top toolbar buttons -.
        param: item_name: 'save_as' OR 'run' OR 'exit' OR 'new'.
        '''
        visual_ribbonobj.select_visualization_application_menu_item(self, application_menu_item_name)
        
    def switch_ia_ribbon_tab(self, tab_name):
        '''
        Desc:-switch the ribbon tab like Home Format Data
        report_obj.switch_ia_ribbon_tab('Foramt')
        '''
        visual_ribbonobj.switch_ribbon_tab(self, tab_name)
        
    def select_ia_ribbon_item(self, tab_name, ribbon_button_name_path):
        ''''
        Desc :- Switching ia ribbon item : Format->AutoLink
        report_obj.switch_ia_ribbon_item(self, tab_name,ribbon_button_name_path)
        '''
        visual_ribbonobj.select_visualization_ribbon_item(self, tab_name, ribbon_button_name_path)
        
    def run_report_from_toptoolbar(self):
        '''
        Desc: This is used to run visualization from the top toolbar.
        '''
        ribbonobject.select_ia_top_toolbar_item(self, 'run')
        
    def save_report_from_toptoolbar(self):
        '''
        Desc:-This function will select save button from top toolbar.
        '''
        ribbonobject.select_ia_top_toolbar_item(self, 'save')
        
    def select_new_from_toptoolbar(self):
        '''
        Desc:- This function will select new from top toolbar.
        '''
        ribbonobject.select_ia_top_toolbar_item(self, 'new')
        
    def save_file_in_save_dialog(self, file_name, file_type=None,**kwargs):
        '''
        Desc:-This function is used to save file in the save dialog
        '''
        utillobject.ibfs_save_as(self, file_name, file_type,**kwargs)
        
    def select_ia_toolbar_item(self, item_name, **kwargs):#need to remove
        '''
        Desc:- select toolbar item run, save
        visual_ribbonobj.select_ia_toolbar_item('Run')
        '''
        visual_ribbonobj.select_top_toolbar_item(self, item_name, **kwargs)
    
    def verify_ribbon_item_is_disabled(self, ribbon_button_name, step_num):
        """
        Description : This method will verify whether ribbon is disabled 
        :usage : verify_ribbon_item_is_disabled("format_auto_drill", "01.02")
        """
        ia_ribbon_obj.verify_ribbon_item_is_disabled_or_enabled(self, ribbon_button_name, step_num, enabled=False)
    
    def verify_ribbon_item_is_enabled(self, ribbon_button_name, step_num):
        """
        Description : This method will verify whether ribbon is disabled 
        :usage : verify_ribbon_item_is_enabled("format_auto_drill", "01.02")
        """
        ia_ribbon_obj.verify_ribbon_item_is_disabled_or_enabled(self, ribbon_button_name, step_num)
        
    def ia_exit_save(self, btn_name, parent_object=None):
        '''
        Desc:-This function is used to click on button like Yes, No and Cancel
        '''
        ia_resultobject.ia_exit_save(self, btn_name, parent_object)
        
    def click_any_bibutton_in_dialog(self, dialog_css, btn_name):
        '''
        Desc:-This function is used to click on button in the dialog
        '''
        utillobject.click_any_bibutton_in_dialog(self, dialog_css, btn_name)
    
    def set_record_limit_in_home_tab(self, record_limits):
        '''
        Desc: This is used to set the record limit from home tab.
        '''
        ribbonobject.set_record_limit_homepage(self, record_limits)
        
    def change_output_format_type(self, output_format_type, location='Home'):
        '''
        Desc:- This function is used to change the output format
        '''
        visual_ribbonobj.change_output_format_type(self, output_format_type, location)
    
    def verify_fexcode_syntax(self, expected_syntax_list, msg):
        '''
        Desc :-This function is used to verify syntax in view code.
        Usage : verify_fexcode_syntax('ARGRAPHENGINE=JSCHART', 'Step 9: Verify')
        '''
        ia_resultobject.verify_fexcode_syntax(self, expected_syntax_list, msg)
    
    def verify_syntax_not_in_fexcode(self, expected_syntax_list, msg):
        '''
        Desc :-This function is used to verify syntax not in view code.
        Usage : verify_syntax_not_in_fexcode('ARGRAPHENGINE=JSCHART', 'Step 9: Verify')
        '''
        ia_resultobject.verify_syntax_not_in_fexcode(self, expected_syntax_list, msg)
            
    def close_ia_without_save(self):
        """
        Description : This method click IA button and click on exit to close IA then click on No button.
        """
        ia_resultobject.close_ia_without_save(self)
    
    def select_output_format_from_ribbon(self, outpu_fromat_path):
        """
        Description : Click on format button in ribbon bar and select output format options
        :uage : select_output_format_from_ribbon("Excel->Excel (xlsx)")
        """
        ia_ribbon_obj.select_output_format(self, outpu_fromat_path)
        
    def select_output_format_from_status_bar(self, outpu_fromat_path):
        """
        Description : Click on format button in status bar and select output format options
        :uage : select_output_format_from_status_bar("Excel->Excel (xlsx)")
        """
        ia_ribbon_obj.select_output_format(self, outpu_fromat_path,  select_from="status_bar")
        
    def verify_ribbon_item_selected(self, ribbon_button_name, step_num):
        """
        Description : Verify ribbon bar item is selected by using background color
        :usage : verify_ribbon_item_selected("format_auto_drill", "02.01")
        """
        ia_ribbon_obj.verify_ribbon_item_selected_or_not(self, ribbon_button_name, step_num)
        
    def verify_ribbon_item_not_selected(self, ribbon_button_name, step_num):
        """
        Description : Verify ribbon bar item is not selected by using background color
        :usage : verify_ribbon_item_not_selected("format_auto_drill", "02.01")
        """
        ia_ribbon_obj.verify_ribbon_item_selected_or_not(self, ribbon_button_name, step_num, selected=False)
    
    def select_ia_application_menu(self, menu_name):
        """
        Description : Click on application button and select menu
        """
        visual_ribbonobj.select_visualization_application_menu_item(self, menu_name)
    
    """**************************************************************AUTOPROMPT REPORT********************************************************************************"""
    def run_auto_prompt_report(self):
        '''
        This function is to select run button in the auto prompt
        '''
        iarunobj.select_amper_menu(self, 'Run')
        
    def select_field_filter_values_dropdown_in_auto_prompt(self, field_name):
        '''
        This function is used to select field dropdown
        select_field_filter_values_dropdown_in_auto_prompt('MODEL')
        '''
        iarunobj.select_field_filter_values_dropdown_in_auto_prompt(self, field_name)
        
    def select_single_field_filter_value_in_auto_prompt(self, value_list):
        '''
        This function will select a single value in field autoprompt
        select_single_field_filter_value_in_auto_prompt(['BMW'])
        '''
        iarunobj.select_field_filter_values_in_auto_prompt(self, value_list, 'list')
    
    def select_input_single_field_filter_value_in_auto_prompt(self, value_list):
        '''
        This function will select a single value in field autoprompt
        select_single_field_filter_value_in_auto_prompt(['BMW'])
        '''
        iarunobj.select_field_filter_values_in_auto_prompt_(self, value_list, 'input')
        
    def select_single_field_filter_value_in_listbox_auto_prompt(self, value_list):
        '''
        This function will select a single value from list box auto prompt.(scroll bar will be in list of values)
        select_single_field_filter_value_in_auto_prompt(['BMW'])
        '''
        iarunobj.select_field_filter_values_in_auto_prompt(self, value_list, 'listbox')
        
    def select_multiple_filter_values_from_field_auto_prompt(self,value_list, ):
        '''
        This function will select multiple values in field autoprompt
        select_multiple_filter_values_from_field_auto_prompt(['AUDI','BMW'])
        '''
        iarunobj.select_field_filter_values_in_auto_prompt(self, value_list, 'check')
        
    def verify_input_type_field_filter_values_in_auto_prompt(self, expected_value_list, msg):
        '''
        This function will verify input type filter values in auto prompt
        verify_input_type_field_filter_values_in_auto_prompt(['AUDI','BMW'], "Step X: Verify CAR field filter values in auto prompt)
        '''
        iarunobj.verify_field_filter_values_in_auto_prompt(self, expected_value_list, msg, 'input')
        
    def verify_field_filter_values_checked_property_in_auto_prompt(self, expected_value_list, msg, verify_type):
        '''
        This function will verify checked or unchecked property of the autoprompt list of values
        verify_field_filter_values_checked_property_in_auto_prompt(['AUDI','BMW'], "Step X: Verify CAR field filter values in auto prompt, 'checked')
        '''
        iarunobj.verify_field_filter_values_checked_property_in_auto_prompt(self, expected_value_list, msg, verify_type)
        
    def verify_option_type_field_filter_values_in_auto_prompt(self, expected_value_list, msg):
        '''
        This function will verify input type filter values in auto prompt
        verify_option_type_field_filter_values_in_auto_prompt(['AUDI','BMW'], "Step X: Verify CAR field filter values in auto prompt)
        '''
        iarunobj.verify_field_filter_values_in_auto_prompt(self, expected_value_list, msg, 'option')
        
    def get_autoprompt_field_object(self, field_name):
        '''
        This function will get field object in auto prompt
        get_autoprompt_field_object('CAR')
        '''
        iarunobj.get_autoprompt_field_object(self, field_name)
        
    def verify_selected_field_dropdown_value_in_autoprompt(self, field_name, expected_selected_value, msg):
        '''
        This function will verify selected field dropdown value in auto prompt
        verify_selected_field_dropdown_value_in_autoprompt('CAR', 'BMW', "Step x: Verify selected drop down value")
        '''
        iarunobj.verify_selected_field_dropdown_value_in_autoprompt(self, field_name, expected_selected_value, msg)
        
    def select_auto_prompt_value_back_button(self):
        '''
        This function will select back button in auto prompt
        select_auto_prompt_value_back_button()
        '''
        iarunobj.select_auto_prompt_value_back_button(self)
        
    def verify_selected_field_dropdown_value_count_in_autoprompt(self, field_name, expected_selected_value, msg):
        '''
        This fucntion will verify the selected dropdown values count in auto prompt
        verify_selected_field_dropdown_value_count_in_autoprompt('MODEL','3 DOOR',"Step x: Verify selected field dropdown value count in auto prompt
        '''
        iarunobj.verify_selected_field_dropdown_value_count_in_autoprompt(self, field_name, expected_selected_value, msg)
        
    def verify_autoprompt_field_labels_using_asin(self, expected_field_label_list, msg):
        '''
        This function will verify autoprompt field labels
        verify_autoprompt_field_labels(['CAR','COUNTRY','MODEL'], "Step x:Verify autoprompt field labels", 'asin') 
        '''
        iarunobj.verify_autoprompt_field_labels(self, expected_field_label_list, msg, 'asin')
        
    def verify_autoprompt_field_labels_using_asequal(self, expected_field_label_list, msg):
        '''
        This function will verify autoprompt field labels
        verify_autoprompt_field_labels(['CAR','COUNTRY','MODEL'], "Step x:Verify autoprompt field labels", 'asin') 
        '''
        iarunobj.verify_autoprompt_field_labels(self, expected_field_label_list, msg, 'asequal')
    
    def verify_field_textbox_value_in_autoprompt(self, field_name, expected_selected_value, msg):
        '''
        This function will verify field text box value in auto prompt
        '''
        iarunobj.verify_field_textbox_value_in_autoprompt(self, field_name, expected_selected_value, msg)    
        
    def select_radio_button_in_auto_prompt_values(self, radio_button_name):
        '''
        This function will select radio button in auto prompt field filter values
        '''
        iarunobj.select_radio_button_in_auto_prompt_values(self, radio_button_name)
    
    def select_value_button_in_auto_prompt(self, button_name):
        '''
        This function will select all or none value in auto prompt
        '''
        iarunobj.select_value_button_in_auto_prompt(self, button_name)
            
    def select_close_button_in_field_filter_values_in_auto_prompt(self):
        '''
        This function will click close button which is in field filter values
        '''
        iarunobj.select_close_button_in_field_filter_values_in_auto_prompt(self)
        
    def enter_value_field_textbox_in_auto_prompt(self, field_name, input_value):
        '''
        This function will enter value field in textbox in auto prompt
        '''
        iarunobj.enter_value_field_textbox_in_auto_prompt(self, field_name, input_value)
    
    def enter_value_search_textbox_popup_in_auto_prompt(self, input_value):
        '''
        This function will enter value in search box in popup inside auto prompt
        '''
        iarunobj.enter_value_search_textbox_popup_in_auto_prompt(self, input_value)
        
    def select_month_in_calendardatepicker_dialog_in_run_window(self, month_to_select):
        '''
        Desc:-This function is used to select month_in_calendardatepicker_dialog.
        '''
        iarunobj.select_month_in_calendardatepicker_dialog(self, month_to_select)
        
    def select_year_in_calendardatepicker_dialog_in_run_window(self, year_to_select):
        '''
        Desc:-This function is used to select year_in_calendardatepicker_dialog.
        '''
        iarunobj.select_year_in_calendardatepicker_dialog(self, year_to_select)
    
    def select_date_in_calendardatepicker_dialog_in_run_window(self, date_to_select):
        '''
        Desc:-This function is used to select date_in_calendardatepicker_dialog.
        '''
        iarunobj.select_date_in_calendardatepicker_dialog(self, date_to_select)
    
    def verify_filter_chained_group_icon_is_visible_in_autoprompt(self, msg):
        '''
        Desc:-This function is used to verify filter chained group icon is displayed in auto prompt window
        '''
        iarunobj.verify_filter_chained_group_icon_is_visible_in_autoprompt(self, msg)
    
    def verify_all_values_button_selected_autoprompt(self, step_num):
        """
        Description : Verify All Values button is selected in auto prompt filter filed popup.
        :uage : verify_all_values_button_selected_autoprompt("01.02")
        """
        iarunobj.verify_all_or_select_values_button_selected_in_auto_prompt(self, ['All Values'], step_num)
        
    """****************************************************************REPORT PREVIEW**********************************************************************************"""
    def group_dialog(self):
        """
        Description : This method will return the GroupDialog class object. Using this object we can perform all group dialog related action.
        """
        return GroupDialog(self.driver)
    
    def select_preview_report_context_menu(self, cell_num, context_menu_path, parent_css="#TableChart_1"):
        """
        Description : This method will right click on preview report cell and select context menu
        :Note - cell_num should start from 1
        :Usage - select_preview_report_context_menu(2, "Sort->Rank->On")
        """
        ia_resultobject.select_preview_report_context_menu(self, cell_num, context_menu_path, parent_css)
        
    def verify_column_title_on_preview(self,colnum,no_of_cells,table_id,expected_list, msg):
        '''
        Desc :- Verify column titles on preview in ia tool
        resultobj.verify_report_titles_on_preview(2, 4, "TableChart_1", coln_list, "Step 02.1: Verify column titles")
        '''  
        visual_resultareaobj.verify_report_titles_on_preview(self,colnum, no_of_cells, table_id,expected_list, msg)
        
    def verify_report_column_titles_on_preview(self, no_of_colnum, no_of_column_cells, expected_list, table_css='#TableChart_1', msg='Step X'):
        '''
        Desc :- Verify column titles on preview in ia tool
        report_obj.verify_report_column_titles_on_preview(4, 8, expected_list, msg='Step 09:03: Verify report preview titles')
        '''
        ia_resultobject.verify_report_column_titles_on_preview(self, no_of_colnum, no_of_column_cells, expected_list, table_css, msg)
        
    def verify_report_header_footer_title_in_preview(self, expected_title_list, report_title_css="#TableChart_1 div span[class^='x']", msg="Step X"):
        '''
        Desc : This function is used to verify report header footer titles. 
        Please note : We have already created fn in report styling where we need to pass index value which will verify whether the given title is present in the report or not.
        If anything is added additionally then we won know.
        report_obj.verify_report_header_footer_title(['Monthly YTD Stats', 'For All PRODUCT_CATEGORY'], msg="Step x: Verify report")
        '''
        report_title_elems=self.driver.find_elements_by_css_selector(report_title_css)
        actual_title_list=[i.text.strip() for i in report_title_elems if i.text.strip()!='']
        utillobject.asequal(self, expected_title_list, actual_title_list, msg)    
    
    def verify_visualize_bar_added_in_previewreport(self, visualize_bar_color, exp_visualize_bars, message, table_css='#TableChart_1'):
        '''
        Desc:-This function is used to verify no of visualization bar is present in the preview report and check the color of the visual bar
        report_obj.verify_visualize_bar_added_in_previewreport('light_gray', 3, "Step X: Verify visualization bar is added in preview report")
        '''
        ia_resultobject.verify_visualize_bar_added_in_previewreport(self,visualize_bar_color, exp_visualize_bars, message, table_css=table_css)
        
    def verify_report_preview_cell_background(self, cell_css, expected_color, msg):
        '''
        Desc: This function will verify report preview background color
        Usage:
        '''
        bg_elem =self.driver.find_element_by_css_selector(cell_css)
        expected_background_color=utillobject.color_picker(self, expected_color, 'rgba')
        actual_background_color=Color.from_string(bg_elem.value_of_css_property("background-color")).rgba
        utillobject.asin(self, actual_background_color, expected_background_color , msg)
    
    def verify_report_cell_property(self, parent_id, cell_no, **kwargs):
        '''
        Desc:-This function is used to verify report data styling in preview
        report_obj.verify_report_cell_property("TableChart_1", 22, font_color='green', text_value='40.17%', msg="Step 14:03: Verify font color shows in green and the value returns 40.17")
        '''
        ia_resultobject.verify_report_cell_property(self, parent_id, cell_no, **kwargs)
        
    def verify_table_cell_property(self,rownum, colnum, table_css=None, **kwargs):
        '''
        Desc:-This function is used to verify font style, font color, bg_color and hyper link.(a tag hyper link verification)
        report_obj.verify_table_cell_property(5, 2, table_css="table[summary='Summary'] > tbody > tr:nth-child(5) td:nth-child(2)", bg_color='chelsea_cucumber', font_color='white', text_value='Model', msg="Step X: Verify background color for Rank")
        '''
        iarunobj.verify_table_cell_property_(self, rownum, colnum, table_css=table_css, **kwargs)
        
    def verify_column_total(self, parent_id, expected_column_total_value_list, msg):
        '''
        Desc:This function is used to verify column total in preview reoprt
        report_obj.verify_column_total('TableChart_1', ['TOTAL', '', '143,794', '173,204'], msg="Sstep X:"
        '''
        ia_resultobject.verify_column_total(self, parent_id, expected_column_total_value_list, msg)
    
    def verify_row_total_report_titles_on_preview(self,colnum,no_of_cells,table_id,expected_list, msg):
        '''
        Desc :- This function is used to verify row_total_titles on preeview
        report_obj.verify_row_total_reeport_titles_on_preview(1, 5, 'TableChart_1', ['Revenue', Total'], "Step :X")
        '''
        ia_resultobject.verify_row_total_report_titles_on_preview(self, colnum, no_of_cells, table_id, expected_list, msg)
        
    def verify_visualize_bar_added_in_htmlreport(self, column_name, visualize_bar_color, expected_total_visual_bars, msg, table_css="table[summary='Summary']", column_position=1):
        '''
        Desc:- This fucntion is used to verify visualization bar added in the htmlreport.
        report_obj.verify_visualize_bar_added_in_htmlreport('AVEMargin', 'light_gray', 7, "Step 03:03: Verify visualization bar are added in the report")[AVE Margin is wrapped into two row]
        '''
        iarunobj.verify_visualize_bar_added_in_htmlreport(self, column_name, visualize_bar_color, expected_total_visual_bars, msg, table_css, column_position)
        
    def select_field_on_canvas(self, table_id, click_expected_col, **kwargs):
        '''
        Desc:-This function will select report field on canvas
        '''
        ia_resultobject.select_field_on_canvas(self, table_id, click_expected_col, **kwargs)
        
        
    def click_preview_canvas(self, canvas_css,element_location='middle', xoffset=0, yoffset=0, action_chain_click=False, mouse_move_duration=0.5):
        '''
        Desc:-This function will click on preview
        '''
        ia_resultobject.click_preview_canvas(self,canvas_css,element_location=element_location, xoffset=xoffset, yoffset=yoffset, action_chain_click=action_chain_click, mouse_move_duration=mouse_move_duration)
        
    """ *************************************************************PREVIEW STATUS BAR ************************************************************************************"""
    def switch_to_report_panel(self, panel_name,**kwargs):
        '''
        Desc:-Switch to report, chart, document from status bar
        '''
        visual_ribbonobj.switch_to_report_panel(self, panel_name, **kwargs)
        
    def create_hold_type(self, hold_type, **kwargs):
        '''
        Decs:-Select create hold type from one component to an another component
        '''
        ia_resultobject.create_hold_type(self, hold_type,**kwargs)
        
    """**************************************************************REPORT PREVIEW DATA VERIFICATIONS*********************************************************************"""
    
    def create_acrossreport_data_set_in_preview(self, parent_id, header_rows, header_cols, data_rows, data_cols, file_name, **kwargs):
        '''
        Desc :- Create_acrossreport_data_set_in_preview by giving header rows(example the header row is two then need to give 2
        report_obj.create_acrossreport_data_set_in_preview("TableChart_1", 2, 2, 10, 2, "Test.xlsx")
        '''
        ia_resultobject.create_across_report_data_set(self, parent_id, header_rows, header_cols, data_rows, data_cols, file_name, **kwargs)
        
    def verify_across_report_data_set_in_preview(self, parent_id, header_rows, header_cols, data_rows, data_cols, file_name, msg, **kwargs):
        '''
        Desc :- verify_acrossreport_data_set_in_preview by giving header rows(example the header row is two then need to give 2
        report_obj.create_acrossreport_data_set_in_preview("TableChart_1", 2, 2, 10, 2, "Test.xlsx", "Step x:Verify across report data in preview)
        '''
        ia_resultobject.verify_across_report_data_set(self, parent_id, header_rows, header_cols, data_rows, data_cols, file_name, msg, **kwargs)
    
    def create_report_dataset_using_start_end_rowcolumn(self, file_name, table_css="table[summary='Summary']", start_rownum=None, end_rownum=None, start_colnum=None, end_colnum=None):
        '''
        Desc :-This function is used to create dataset at runtime of html report in the form xlsx format
        reportobj.create_report_dataset_using_start_end_rowcolumn('C121646_Ds01.xlsx', table_css="table[summary='Summary']", start_rownum=3, end_rownum=10, start_colnum=1, end_colnum=3)
        '''
        utillobject.create_table_data_start_end_rowcolumn(self, file_name, table_css=table_css, start_rownum=start_rownum, end_rownum=end_rownum, start_colnum=start_colnum, end_colnum=end_colnum)
    
    def verify_report_dataset_using_start_end_rowcolumn(self, file_name, table_css, msg, start_rownum=None, end_rownum=None, start_colnum=None, end_colnum=None):
        '''
        Desc :-This function is used to create dataset at runtime of html report in the form xlsx format
        reportobj.create_report_dataset_using_start_end_rowcolumn('C121646_Ds01.xlsx', table_css="table[summary='Summary']", start_rownum=3, end_rownum=10, start_colnum=1, end_colnum=3)
        '''
        utillobject.verify_table_data_using_start_end_rowcolumn(self, file_name, table_css, msg, start_rownum=start_rownum, end_rownum=end_rownum, start_colnum=start_colnum, end_colnum=end_colnum)
    
    def create_report_data_set_in_preview(self, parent_id, rows, cols, file_name, **kwargs):
        '''
        Desc :- This function is used to create reportdataset for the preview data
        report_obj.create_report_data_set('TableChart_1', 10, 6, 'Test_1.xlsx', **kwargs)
        '''
        ia_resultobject.create_report_data_set(self, parent_id, rows, cols, file_name, **kwargs)
        
    def verify_report_data_set_in_preview(self, parent_id, rows, cols, file_name, msg, **kwargs):
        '''
        Desc :- This function is used to create reportdataset for the preview data
        report_obj.create_report_data_set('TableChart_1', 10, 6, 'Test_1.xlsx', **kwargs)
        '''
        ia_resultobject.verify_report_data_set(self, parent_id, rows, cols, file_name, msg, **kwargs)
        
    """************************************************************RUNTIME VERIFICATIONS************************************************************************************"""
    def select_accordion_row(self, row_css):
        '''
        Desc :- This function will expand/collapse accordion row
        '''
        row_element=self.driver.find_element_by_css_selector(row_css)
        row_element.click()
    
    def verify_accordion_symbol(self, row_css , msg, attribute='content'): 
        '''
        Desc: This function will verify accordion symbol '+' in run window
        ''' 
        obj=self.driver.find_element_by_css_selector(row_css)
        actual_accordion_symbol = js_obj.get_element_before_style_properties(self, obj, attribute).replace('"','')
        utillobject.asequal(self, '+', actual_accordion_symbol, msg)  
        
    def verify_message_in_html_body(self, message="/bipgqashare/qaauto_lnx_apps/smoke_retailsamples_alphaformat/wf_retail_lite.mas", msg="Step X"):
        '''
        Desc: This function is used to verify message from html>body 
        :Param :msg="/bipgqashare/qaauto_lnx_apps/smoketest/wf_retail_lite.mas"
        @Usage :verify_message_in_html_body() 
        '''
        miscelaneousobject.verify_message_in_html_body(self, message, msg)
        
    def create_html_report_dataset(self, file_name, table_css="table[summary='Summary']"):
        '''
        Desc :-This function is used to create dataset at runtime of html report in the form xlsx format
        reportobj.create_html_report_dataset('C121646_Ds01.xlsx', desired_no_of_rows=10, starting_rownum=3)
        '''
        Report.create_table_data_set(self, table_css, file_name)
        
    def verify_html_report_dataset(self, file_name, msg, table_css="table[summary='Summary']"):
        '''
        Desc :- This function is used to verify html report data which is already taken using create_html_report_dataset
        reportobj.verify_html_report_dataset("C123456_Ds01.xslx',table_css="table[summary='Summary']", desired_no_of_rows=10, starting_rownum=2) 
        '''
        Report.verify_table_data_set(self, table_css, file_name, msg)
        
    def verify_table_data_set(self, table_css, file_name, msg, **kwargs):
        """
        Usage: utillobj.verify_table_data_set("table[summary='Summary']", "test1.xlsx","Step 10: fail data set", desired_no_of_rows=1)
        """
        if table_css==None:
            table_css="table[summary='Summary']"
        else:
            table_css=table_css
        iarunobj.verify_table_data_set(self, table_css, file_name, msg, **kwargs)
    
    def create_table_data_set(self, table_css, file_name, **kwargs):
        """    
        @param: table_css= ".arPivot tr:nth-child(1) table" or "table[summary='Summary']"   =  Need to provide the full parent path till table
        @param: file_name: "test1.xlsx" 
        Usage: create_table_data_set("table[summary='Summary']", "test1.xlsx" )
        """
        if table_css==None:
            table_css="table[summary='Summary']"
        else:
            table_css=table_css
        iarunobj.create_table_data_set(self, table_css, file_name, **kwargs)
        
    def expand_colapse(self, node_path, expand_colapse_toggle):
        """
        iarub_obj.expand_colapse("South America->Camcorder", "expand") or iarunobj.expand_colapse("North America->Camcorder", "colapse")
        """
        iarunobj.expand_colapse(self, node_path, expand_colapse_toggle)
        
    def verify_column_heading(self, table_css, expected_title, msg):
        """
        iarunobj.verify_column_heading("#divTreeTable #treetable thead", ['GrossProfitDiscountMSRPCOGSQty'], "Step x:Verify report heading")
        """
        iarunobj.verify_column_heading(self, table_css, expected_title, msg)
        
    def verify_report_titles_on_preview(self,colnum,no_of_cells,table_id,expected_list, msg):
        '''
        Desc:-This function is used to verify report titles in the preview.
        '''
        visual_resultareaobj.verify_report_titles_on_preview(self,colnum,no_of_cells,table_id,expected_list, msg)
        
    """******************************************************************************COMMON FUNCTIONS*****************************************************************"""
    
    def verify_item_in_open_dialog(self, target_table_path, item_name, item_exist, msg):
        '''
        Desc:-This function is used to verify items in the open dialog
        '''
        utillobject.verify_item_in_open_dialog(self, target_table_path, item_name, item_exist, msg)
        
    def select_masterfile_in_open_dialog(self, target_table_path, master_file_name, file_type=None):
        '''
        Desc:-This function is used to select files in the open dialog.Example master file or any file can be selected using this function.
        '''
        utillobject.select_masterfile_in_open_dialog(self, target_table_path, master_file_name, file_type)
        
    def verify_element_visibility(self, element=None, element_css=None, visible=True, msg='Step X'):
        '''
        Desc:-This function is used to verify the element visibility
        '''
        utillobject.verify_element_visiblty(self, element=element, element_css=element_css, visible=visible, msg=msg)
        
    
    
    """************************************************************************Filter dialog functions******************************************************************"""
    
    def open_filter_where_value_dialog(self, **kwargs):
        '''
        Desc:- This function is used to open the where value in filter condition dialog box
        '''
        ribbonobject.open_where_value_popup(self, **kwargs)
    
    def open_where_field_popup_in_filter_dialog(self, where_row):
        """
        Description : This method will open where field popup by double click.
        :arg - where_row : Row number of where condition. where_row should start from 1
        :Usage : open_where_field_popup_in_filter_dialog(1)
        """
        ribbonobject.open_where_popup_in_filter_dialog(self, where_row, "field")
    
    def open_where_operator_popup_in_filter_dialog(self, where_row):
        """
        Description : This method will open where operator popup by double click.
        :arg - where_row : Row number of where condition. where_row should start from 1
        :Usage : open_where_operator_popup_in_filter_dialog(1)
        """
        ribbonobject.open_where_popup_in_filter_dialog(self, where_row, "operator")
    
    def open_where_value_popup_in_filter_dialog(self, where_row):
        """
        Description : This method will open where where value popup by double click.
        :arg - where_row : Row number of where condition. where_row should start from 1
        :Usage : open_where_value_popup_in_filter_dialog(1)
        """
        ribbonobject.open_where_popup_in_filter_dialog(self, where_row, "value")
        
    def select_filter_type(self, filter_type):
        '''
        Desc:- This function is used to select the filter type - Constant/Parameter/Field
        usage: - 
        select_filter_type(Constant/Parameter/Field) 
        '''
        utillobject.select_combobox_item(self, 'id_wv_combo_type', filter_type)       
    
    def select_filter_parameter_type(self, parameter_type):
        '''
        Desc:- This function is used to select the radio button - Simple/Static/Dynamic
        usage:-
        select_filter_parameter_type(Simple/Static/Dynamic)
        '''
        ribbonobject.select_parameter_type(self, parameter_type)
    
    def select_filter_parameter_checkbox(self, ParamOptional=False, ParamMultiple=False):
        '''
        Desc:- This function is used to check the parameter checbox - Optional, Select multiple values at runtime
        '''
        ribbonobject.check_parameter_checkbox(self, ParamOptional=ParamOptional, ParamMultiple=ParamMultiple)
        
    def select_field_in_filter_tree(self, filter_field_path, field_position):
        '''
        Desc:- This function is used to expans the field tree and select the field name.
        usage:- select_field_in_filter_tree("Dimensions->Product->Product->Product,Subcategory", 1)
        '''
        filter_metadata_obj.click_on_filter_field(self, filter_field_path, field_position)
    
    def double_click_field_in_filter_tree(self, data_field_path, field_position=1):
        '''
        This method used to expand and right click on specific data field
        data_field_path = 'Dimensions->COUNTRY->COUNTRY->COUNTRY'
        field_position = 1 , 2, 3
        Example Usage : double_click_on_filter_field('Dimensions->COUNTRY->COUNTRY->COUNTRY', field_position=3)
        '''
        filter_metadata_obj.double_click_on_filter_field(self, data_field_path, field_position)
    
    def close_filter_where_value_popup_dialog(self):
        '''
        Desc:- This function closes the filter_where value popup dialog
        '''
        ribbonobject.close_where_value_popup(self)
    
    def close_filter_dialog(self):
        '''
        Desc:- this function closed the filter dialog
        '''
        ribbonobject.close_filter_dialog(self)
    
    def enter_where_value_field_textbox_in_filter_dialog(self, field_name):
        '''
        Desc:- This function enter the field name in where value filter dialog box
        '''
        ribbonobject.enter_where_value_text_field(self, field_name)
        
    def click_equal_to_condition_in_filter_dialog(self, filter_condition_type):
        '''
        This function will click equal to condition in filter dialog.
        click_equal_to_condition_in_filter_dialog('Great than')
        '''
        ribbonobject.click_equal_to_condition_in_filter_dialog(self, filter_condition_type)
    
    def select_static_fields_in_filter_dialog(self, get_values_option, fields_list):
        """
        Description : This method will select get values option and double click on fields when static option is checked in filter dialog
        select_static_field_in_filter_dialog('All', ['BMW', 'AUDI'])
        """
        ribbonobject.select_filter_values(self, "static", get_values_option, fields_list)
        
    """*****************************************************************************************************************************************************************"""   
    def create_reporttable_dataset(self, file_name, table_css=None, desired_no_of_rows=None, starting_rownum=None):
        '''
        Desc :-This function is used to create report dataset in the form of xlsx format
        reportobj.create_reporttable_data_set('C123456_Ds01.xlsx', table_css='#ITableData0', desired_no_of_rows=10, starting_rownum=2)
        '''
        rows_css = table_css + " > tbody > tr"
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        no_of_rows=desired_no_of_rows if desired_no_of_rows!=None else len(table_rows) 
        start_rownum=starting_rownum if starting_rownum!=None else 0        
        for r in range(start_rownum,no_of_rows):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text
                s.cell(row=r + 1, column=c + 1).value = str(value)
        wb.save(os.path.join(os.path.join(os.getcwd(), "data"), file_name))
        
    def compare_table_dataset(self, file_name, table_css=None, desired_no_of_rows=None, starting_rownum=None):
        '''
        Desc :- This function is used to verify html report data which is already taken using create_reporttable_dataset.
        reportobj.compare_table_dataset('C123456_Ds01.xlsx', table_css="table[summary='Summary']", desired_no_of_rows=10, starting_rownum=2)
        ''' 
        wb1 = load_workbook(os.path.join(os.path.join(os.getcwd(), "data"), file_name))
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        rows_css = table_css + " > tbody > tr"
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        no_of_rows=desired_no_of_rows if desired_no_of_rows!=None else len(table_rows)
        start_rownum=starting_rownum if starting_rownum!=None else 0
        for r in range(start_rownum, no_of_rows):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(0, len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text
                if s1.cell(row=r + 1, column=c + 1).value != None and str(value) != '':
                    if s1.cell(row=r + 1, column=c + 1).value.replace(' ','').lower() == str(value).replace(' ','').lower():
                        status=[0]
                        continue
                    else:
                        status=[r+1,c, 'expected',s1.cell(row=r + 1, column=c + 1).value.replace(' ','').lower(), 'actual',str(value).replace(' ','').lower()]
                        return (status)
        return (status)
    
    '''*************************************************************Active_report_options*********************************************************'''
   
    def verify_general_tab_dropdown_label(self, text_label, step_no=None):
        '''
        Desc:- This function is used to verify dropdown under general tab
        text_label : "general_window" or "general_freeze_columns" or "general_location" or "general_records_per_page" 
        usage: document_obj.verify_general_tab_dropdown_label("general_window_label", step_no="1")
        '''
        ia_ribbon_obj.dropdown_row_verification(self, text_label, IaRibbonLocators.general_tab_parent_css, label= text_label, step=step_no)

    def verify_general_tab_default_value(self, dropdown_name, value, step_no=None):
        '''
        Desc:- This function is used to verify dropdown value under general tab
        dropdown_name : "general_window" or "general_freeze_columns" or "general_location" or "general_records_per_page"
        usage: document_obj.verify_general_tab_default_combobox_values("general_window_label","Cascade", step_no='1')
        '''
        ia_ribbon_obj.dropdown_row_verification(self, dropdown_name, IaRibbonLocators.general_tab_parent_css, default_value=value, input_type='combo_box', step=step_no)
    
    def verify_general_tab_default_inputbox_values(self, dropdown_name, value, step_no=None):
        '''
        Desc:- This function is used to verify inputbox value under general tab
        dropdown_name : "general_window" or "general_freeze_columns" or "general_location" or "general_records_per_page"
        usage: document_obj.verify_general_tab_default_inputbox_values("general_records_per_page_default_value","57", step_no='1')
        '''
        ia_ribbon_obj.dropdown_row_verification(self, dropdown_name, IaRibbonLocators.general_tab_parent_css, default_value=value, input_type='input_box', step=step_no)

    def verify_general_tab_dropdown_values(self, dropdown_name, dropdown_values_list, step_no=None):
        '''
        Desc:- This function is used to verify dropdown values under general tab
        dropdown_name : "general_window" or "general_freeze_columns" or "general_location" or "general_records_per_page"
        usage: document_obj.verify_general_tab_dropdown_values("general_window_dropdown",["Cascade", "Tabs"], step_no='1')
        '''
        ia_ribbon_obj.dropdown_row_verification(self, dropdown_name, IaRibbonLocators.general_tab_parent_css, dropdown_list=dropdown_values_list, step=step_no)

    def verify_general_tab_propery_enabled(self, dropdown_name, enable_status=True, step_no=None):
        '''
        Desc:- This function is used to verify properties under general tab is enabled or not
        dropdown_name : "general_window" or "general_freeze_columns" or "general_location" or "general_records_per_page"
        usage: document_obj.verify_general_tab_propery_enabled("general_window_label",enable_status=True, step_no="1")
        '''
        ia_ribbon_obj.dropdown_row_verification(self, dropdown_name, IaRibbonLocators.general_tab_parent_css, enable=enable_status, step=step_no)
    
    def verify_general_tab_checkbox_verification(self, checkbox_name, enable_status=True, step_no=None):
        '''
        Desc:- This function is used to verify checkbox under general tab is enabled or not
        checkbox_name : "general_display_page_info"
        usage: document_obj.verify_general_tab_propery_enabled("general_display_page_info",enable_status=True, step_no="1")
        '''
        ia_ribbon_obj.checkbox_verification(self, checkbox_name, IaRibbonLocators.general_tab_parent_css, checked=enable_status, step=step_no)
    
    def join_dialog(self):
        """
        Description : Return the object of common.pages.ia_join class
        """
        return IAJoin(self.driver)
    
    def define_compute_dialog(self):
        """
        Description : This method returns DefineCompute class object. Using this object we can call all define and compute dialog dialog related methods 
        """
        return DefineCompute(self.driver)
    
    '''*************************************************************ENDS*********************************************************'''   