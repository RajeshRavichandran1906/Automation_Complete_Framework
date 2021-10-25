import unittest
import re
import os,io
from selenium.webdriver.support.ui import WebDriverWait
import time
from PIL import Image
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from common.locators.mobile_web_locators import MobileLoginPageLocators
import platform
from configparser import ConfigParser, NoOptionError

class UtillityMethods(object):

    asert_failure_count=0
    browser_height=0
    
    def __init__(self, driver):
        self.driver = driver
        self.shortwait = WebDriverWait(self.driver, 50)
        self.mediumwait = WebDriverWait(self.driver, 100)
        self.longwait = WebDriverWait(self.driver, 150)
        self.test = UtillityMethods.parseinitfile(self, 'test')
    
    def wait_for_element(self,element,wait_method):
        
        '''
        Params : element=element
        params : wait_method= 0 (short time wait )
        params : wait_method= 1 (medium time wait ) 
        params : wait_method= 2 (long time wait )
        usage : wait_for_element(element,0) 
        usage : wait_for_element(element,2) 
        '''
        if wait_method==0 :
            self.shortwait.until(EC.visibility_of_element_located(element))
        elif wait_method==1 :
            self.mediumwait.until(EC.visibility_of_element_located(element))
        else :
            self.longwait.until(EC.visibility_of_element_located(element))
    
    
    def parseinitfile(self, key):
        init_file = 'config.init'
        config_pair = {}
        try:
            fileObj = open(init_file, "r")
            line = fileObj.readline()
            while line:
                lineObjbj = re.match(r'(\S*)\s(.*)', line)
                keyName = lineObjbj.group(1)
                config_pair[keyName] = lineObjbj.group(2)
                line = fileObj.readline()
            fileObj.close()
        except IOError:
            exit()

        if key in config_pair:
            return (config_pair[key])
        else:
            return ('Key not found')
        
    def invoke_mobile_faves(self):
        
        node = self.parseinitfile('nodeid')
        port = self.parseinitfile('httpport')
        context = self.parseinitfile('wfcontext')
        setup_url = 'http://' + node + ':' + port +''+ context + '/mobile/signin'
        self.driver.get(setup_url)
        element=(By.CSS_SELECTOR,"a[role='button'][onclick*='javascript:submitSignonRequest']")
        self.wait_for_element(element,2)
    
    def login_mobile_faves_web(self,mrid,mrpass):
        '''
        usage : login_mobile_faves()
        Author : Prabhakaran
        '''
        self.invoke_mobile_faves()
        user_id = self.parseinitfile(mrid)
        user_password = self.parseinitfile(mrpass)
        username=self.driver.find_element(*MobileLoginPageLocators.username)
        passwd=self.driver.find_element(*MobileLoginPageLocators.password)
        time.sleep(1)
        UtillityMethods.browser_height=self.get_browser_height()
        username.send_keys(user_id)
        if mrpass!=None :
            passwd.send_keys(user_password)
        login=self.driver.find_element(*MobileLoginPageLocators.signin)
        if login.is_displayed() :
            login.click()
            time.sleep(2)
        try :
            a=self.driver.switch_to_alert()
            a.accept()
        except :
            pass
        element=(By.CSS_SELECTOR,"div[data-role='collapsible-set']")
        self.wait_for_element(element,0)
        time.sleep(1)
    
    def select_mobilefav_fex(self,folder,fex):
        '''
        usage : select_mobilefav_fex('Active','ahtml_001')
        Author : Prabhakaran
        '''
        folder_elements=self.driver.find_elements_by_css_selector("div[data-role='collapsible'] a[class^='ui-collapsible-heading-toggle']")
        folder_names=[name.text.replace('\nclick to expand contents','').strip() for name in folder_elements]
        select_folder=folder_elements[folder_names.index(folder)]
        self.driver.execute_script("return arguments[0].scrollIntoView();",select_folder)
        select_folder.click()
        time.sleep(2)
        parent_element=self.driver.find_elements_by_css_selector("div[data-role='collapsible']")
        fex_elements=parent_element[folder_names.index(folder)].find_elements_by_css_selector("ul[data-role='listview'] li h2")
        fex_names=[name.get_attribute('textContent').strip() for name in fex_elements]
        #print(fex_names)
        fex_file=fex_elements[fex_names.index(fex)]
        self.driver.execute_script("return arguments[0].scrollIntoView();",fex_file)
        fex_file.click()
        time.sleep(2)
    
    def mobilefav_login_and_select_fex(self,fex,mrid,mrpass):
        '''
        usage : mobilefav_login_and_select_fex('Active','ahtml_001')
        Author : Prabhakaran
        '''
        
        self.login_mobile_faves_web(mrid,mrpass)
        folder=self.parseinitfile('suite_folder')
        self.select_mobilefav_fex(folder,fex)
        
    def color_picker(self, color, rgb_type='rgb'):
        """
        Usage: color_picker(self,'comment', 'rgba')
        params:color='green'
        params:rgb_type='rgb' or 'rgba' or 'transparent'
        return:rgb(0, 0, 0) or rgba(0, 0, 0)
        """
        os_platform=platform.system()
        if os_platform=='Windows' :
            color_file=os.getcwd() + "\\common\\lib\\color.data"
        else :
            color_file=os.getcwd() + "/common/lib/color.data"
        section = 'transparent' if color == 'transparent' else 'DEFAULT'
        color_pair = {}
        parser = ConfigParser()
        parser.optionxform=str
        parser.read(color_file)
        try:
            color_pair[color] = parser[section][color]
            if rgb_type=='rgb':
                return (rgb_type + color_pair[color])
            elif rgb_type=='transparent':
                return(color_pair[color])
            else:
                return(rgb_type + re.sub('\)', ', 1)', color_pair[color]))
        except (KeyError, NoOptionError) as e:
            print("Specified Color is not available in color.data. " + str(e))
            return
        
    def as_List_equal(self,*params):
        try:
            testobj = unittest.TestCase()
            testobj.assertListEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            UtillityMethods.asert_failure_count += 1
    
    def asequal(self, *params):
        try:
            testobj = unittest.TestCase()
            testobj.assertEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            print(params[0])
            print(params[1])
            UtillityMethods.asert_failure_count += 1
            
    def as_not_equal(self, *params):
        try:
            testobj = unittest.TestCase()
            testobj.assertNotEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            UtillityMethods.asert_failure_count += 1
    
    def asin(self, *params):
        try:
            #testobj = unittest.TestCase()
            assert params[0] in params[1], params[2]
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            UtillityMethods.asert_failure_count += 1
            
    def as_notin(self, *params):
        try:
            #testobj = unittest.TestCase()
            assert params[0] not in params[1], params[2]
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            UtillityMethods.asert_failure_count += 1
    
    def as_GE(self, *params):
        try:
            testobj = unittest.TestCase()
            testobj.assertGreaterEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            UtillityMethods.asert_failure_count += 1

    def as_LE(self, *params):
        try:
            testobj = unittest.TestCase()
            testobj.assertLessEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            UtillityMethods.asert_failure_count += 1
            
    def compare_data_set(self, table_id, tr_id, file_name, color='default'):
        if tr_id == 'rgb':
            css="[id='{0}'] tr[style*='{1}']"
        elif tr_id == 'notrgb':
            css="[id='{0}'] tr:not([style*='rgb'])[id*='I0']"
        elif tr_id=='bgcolor':
            bgcolor=self.color_picker(color, 'rgb')
            css="[id='{0}'] tr[style*='background-color: " + bgcolor + "']"
        else:
            css="[id='{0}'] tr[id ^='{1}']"
        os_platform=platform.system()
        if os_platform=='Windows' :
            wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        else :
            wb1 = load_workbook(os.getcwd() + "/data/" + file_name)
        status=[]
        #sheet = wb.get_active_sheet
        #s = wb.get_sheet_by_name('Sheet')
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(css.format(table_id, tr_id))
        for r in range(0, len(table_rows)):
            columns = table_rows[r].find_elements_by_css_selector("td")
            for c in range(0, len(columns)):
                if s1.cell(row=r + 1, column=c + 1).value != None:
                    if s1.cell(row=r + 1, column=c + 1).value.strip() == str(columns[c].get_attribute('textContent').strip()):
                        status=[0]
                        continue
                    else:
                        status=[r+1,c]
                        return (status)
        return (status)

    def create_data_set(self,table_id,tr_id,file_name, color='default'):
        """
        Usage: utillobj.create_data_set('ITableData0','I0r','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','rgb','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','notrgb','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','bgcolor','C2140681.xlsx', color='green')
        """
        if tr_id == 'rgb':
            css="[id='{0}'] tr[style*='{1}']"
        elif tr_id == 'notrgb':
            css="[id='{0}'] tr:not([style*='rgb'])[id*='I0']"
        elif tr_id=='bgcolor':
            bgcolor=self.color_picker(color, 'rgb')
            css="[id='{0}'] tr[style*='background-color: " + bgcolor + "']"
        else:
            css="[id='{0}'] tr[id ^='{1}']"
        #from openpyxl import Workbook
        wb = Workbook()
        #sheet = wb.get_active_sheet
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(css.format(table_id, tr_id))
        for r in range(0, len(table_rows)):
            columns = table_rows[r].find_elements_by_css_selector("td")
            for c in range(0, len(columns)):
                s.cell(row=r + 1, column=c + 1).value = str(columns[c].get_attribute('textContent').strip())
        os_platform=platform.system()
        if os_platform=='Windows' :
            wb.save(os.getcwd() + "\data\\" + file_name)
        else :
            wb.save(os.getcwd() + "/data/" + file_name)
        
    def verify_data_set(self,table_id, tr_id, file_name,msg, color='default'):
        """
        Usage: utillobj.verify_data_set('ITableData0','I0r','C2140681.xlsx',"Step 10: fail data set")
        Usage: utillobj.verify_data_set('ITableData0','rgb','C2140681.xlsx',"Step 10: fail data set")
        Usage: utillobj.verify_data_set('ITableData0','notrgb','C2140681.xlsx',"Step 10: fail data set")
        Usage: utillobj.verify_data_set('ITableData0','bgcolor','C2140681.xlsx',"Step 10: fail data set", color='green')
        """
        x= self.compare_data_set(table_id,tr_id,file_name, color)
        self.asequal(len(x),1,msg+ ' Row,Column -'+ str(x))
        
    def compare_two_image(self,base_image,actual_image,msg):
        '''
        params : base_image='C2266374_Base_Step_01'
        params : actual_image='C2266374_Actual_Step_01'
        params : msg='Step X : Veriy image'
        usage  : verify_two_image("C2266374_Base_Step_03","C2266374_Actual_Step_03","Step 03 : Verify Live preview report image got updated")
        Author : Prabhakaran
        '''
        os_platform=platform.system()
        siklui_loc=None
        src_loc=None
        file_path=None
        if os_platform=='Windows' :
            siklui_loc=os.getcwd() +"\\common\\lib\\sikuli\\runsikulix.cmd"
            src_loc=os.getcwd() +"\\sikuli\\compare_img --arg "+os.getcwd()+"\\images\\"+base_image+" "+os.getcwd()+"\\actual_images\\"+actual_image
            file_path=os.getcwd()+"\\sikuli\\sikuli.txt"
        else :
            siklui_loc="echo 'Automate1' | sudo -S "+os.getcwd() +"/common/lib/sikuli/runsikulix.cmd"
            src_loc=os.getcwd() +"/sikuli/compare_img --arg "+os.getcwd()+"/images/"+base_image+" "+os.getcwd()+"/actual_images/"+actual_image 
            file_path=os.getcwd()+"/sikuli/sikuli.txt"
        os.system(siklui_loc+' -r '+ src_loc)
        f=open(file_path,'r')
        result_output=f.read()
        f.close()
        with open(file_path,'w') : pass
        self.asequal('Result->True',result_output.strip(),msg)
    
    def take_mobile_screenshot(self,element,file_name,image_type='actual',x=0,y=0,w=0,h=0):
        """
            params:elemnt=element1
            params:file_name='C2266365'
            params:image_type='actual' or image_type='base'
            Usage:stake_mobile_screenshot(eleemnt1,'C2266365',image_type='base')
            @author: Prabhakaran 
        """  
        location = element.location
        size = element.size
        IMG=self.driver.get_screenshot_as_png()
        im = Image.open(io.BytesIO(IMG)) # uses PIL library to open image in memory
        left = location['x']+x
        top = location['y']+y
        right = location['x'] + size['width']+w
        bottom = location['y'] + size['height']+h
        im = im.crop((left, top, right, bottom)) # defines crop points
        file_name=file_name+'_ios' if self.test=='ios_web' or self.test=='ios_app' else file_name+'_android'
        location_path='actual_images' if image_type=='actual' else 'images'
        os_platform=platform.system()
        if os_platform=='Windows' :
            file_path=os.getcwd() + "\\" + location_path + "\\" + file_name + ".png"
        else :
            file_path=os.getcwd() + "/" + location_path + "/" + file_name + ".png"
        im.save(file_path) 
        
    def take_mobile_monitor_screenshot(self,file_name,image_type='actual'):
        """
            params:file_name='C2266365'
            params:image_type='actual' or image_type='base'
            Usage:stake_mobile_screenshot(eleemnt1,'C2266365',image_type='base')
            @author: Prabhakaran 
        """  
        os_platform=platform.system()
        file_name=file_name+'_ios' if self.test=='ios_web' or self.test=='ios_app' else file_name+'_android'
        location_path='actual_images' if image_type=='actual' else 'images'
        if os_platform=='Windows' :
            file_path=os.getcwd() + "\\" + location_path + "\\" + file_name + ".png"
        else :
            file_path=os.getcwd() + "/" + location_path + "/" + file_name + ".png"
        self.driver.save_screenshot(file_path)
    
    def move_to_element(self,element,x=None,y=None):
        '''
        :param : x=1 , y=1 ( if want to move element with offset )
        :usage : move_to_element(self,element, x=1, y=1)
        :usage : move_to_element(self,element)
        '''
        action=ActionChains(self.driver)
        if x!=None and y!=None :
            action.move_to_element_with_offset(element,x,y).perform()
        else :
            action.move_to_element(element).perform()
        #time.sleep(2)
        del action 
    
    def create_pivot_data_set(self, table_id, file_name,**kwargs):
        """
        Syntax: utillobj.create_pivot_data_set('piv1', 'C2140681_Ds01.xlsx')
        """
        if 'fullscreen' in kwargs :
            rows_css = "#" + table_id + ">table>tbody>tr"
        else :
            rows_css = "#" + table_id + " > tbody > tr:nth-child(2) > td >table > tbody > tr"
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        for r in range(len(table_rows)):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").get_attribute('textContent').strip()
                s.cell(row=r + 1, column=c + 1).value = str(value)
        os_platform=platform.system()
        if os_platform=='Windows' :
            wb.save(os.getcwd() + "\data\\" + file_name)
        else :
            wb.save(os.getcwd() + "/data/" + file_name)
    
    def verify_pivot_data_set(self, table_id, file_name, msg,**kwargs):
        """
        Usage: utillobj.verify_pivot_data_set('piv1', 'C2140681_Ds01.xlsx,"Step 10: fail data set")
        """
        if 'fullscreen' in kwargs :
            x= self.compare_pivot_data_set(table_id, file_name,fullscreen=True)
        else :
            x= self.compare_pivot_data_set(table_id, file_name)
        self.asequal(len(x),1,msg+ ' Row,Column -'+ str(x))
    
    def compare_pivot_data_set(self, table_id, file_name,**kwargs):
        """
        Usage: utillobj.compare_pivot_data_set('piv1', C2140681_Ds01.xlsx)
        """
        os_platform=platform.system()
        if os_platform=='Windows' :
            wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        else :
            wb1 = load_workbook(os.getcwd() + "/data/" + file_name)
        if 'fullscreen' in kwargs :
            rows_css = "#" + table_id + ">table>tbody>tr"
        else :
            rows_css = "#" + table_id + " > tbody > tr:nth-child(2) > td >table > tbody > tr"
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        for r in range(len(table_rows)):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                if (s1.cell(row=r + 1, column=c + 1).value):
                    value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").get_attribute('textContent').strip()
                    if s1.cell(row=r + 1, column=c + 1).value.strip() == str(value):
                        status=[0]
                        continue
                    else:
                        status=[r+1,c]
                        return (status)
        return (status)
    
    def switch_to_window(self,window_index,time_interval=5):
        '''
        :param : window_index=0 (window index start from 0 )
        :param : time_interval=5 ( How long want to wait after switch to window )
        '''
        test=self.parseinitfile('test')
        time.sleep(4)
        if test=='ios_web' :
            contexts=self.driver.contexts
            contexts.remove('NATIVE_APP')
            self.driver.switch_to.context(contexts[window_index])
        else :
            windows=self.driver.window_handles
            self.driver.switch_to_window(windows[window_index])
        time.sleep(time_interval)
            
    def get_browser_height(self):
        '''
        :usage : browser_height=get_browser_height()
        '''
        availHeight = self.driver.execute_script("return screen.availHeight")
        innerHeight = self.driver.execute_script("return window.innerHeight|| document.documentElement.clientHeight|| document.body.clientHeight;")
        browser_height=availHeight - innerHeight
        return browser_height
    
    
    def get_element_screen_coordinate(self,element, coordinate_type='start', **kwargs):
        
        """
        element :- This is the object for which x,y coordinate to be returned.
        coordinate_type='start' OR 'top_middle' OR 'top_right' OR 'left' OR 'middle' OR 'right' OR 'bottom_left' OR 'bottom_middle' OR 'bottom_right' OR 'offset'
        The return type is a dictionary like = {'x': 524, 'y': 993}
        """
        dict_obj={}
        broswer_height=UtillityMethods.browser_height if 'browser_height' in kwargs else 0
        x_offset=kwargs['x_offset'] if 'x_offset' in kwargs else 0
        y_offset=kwargs['y_offset'] if 'y_offset' in kwargs else 0
        elem_x=element.location['x']
        elem_y=element.location['y'] + broswer_height
        elem_h=element.size['height']
        elem_w=element.size['width']
        if coordinate_type=='start':
            dict_obj['x'] = elem_x + x_offset
            dict_obj['y'] = elem_y+ y_offset
        if coordinate_type=='top_middle':
            dict_obj['x'] = elem_x + (elem_w/2) + x_offset
            dict_obj['y'] = elem_y + y_offset
        if coordinate_type=='top_right':
            dict_obj['x'] = elem_x + elem_w + x_offset
            dict_obj['y'] = elem_y + y_offset
        if coordinate_type=='left':
            dict_obj['x'] = elem_x + x_offset
            dict_obj['y'] = elem_y + (elem_h/2) + y_offset
        if coordinate_type=='middle':
            dict_obj['x'] = elem_x + (elem_w/2) + x_offset
            dict_obj['y'] = elem_y + (elem_h/2) + y_offset
        if coordinate_type=='right':
            dict_obj['x'] = elem_x + elem_w + x_offset
            dict_obj['y'] = elem_y + (elem_h/2) + y_offset
        if coordinate_type=='bottom_left':
            dict_obj['x'] = elem_x + x_offset
            dict_obj['y'] = elem_y + elem_h + y_offset
        if coordinate_type=='bottom_middle':
            dict_obj['x'] = elem_x + (elem_w/2) + x_offset
            dict_obj['y'] = elem_y + elem_h + y_offset
        if coordinate_type=='bottom_right':
            dict_obj['x'] = elem_x + elem_w + x_offset
            dict_obj['y'] = elem_y + elem_h + y_offset
        if coordinate_type=='offset':
            dict_obj['x'] =  x_offset
            dict_obj['y'] =  y_offset
        return(dict_obj)
    
    def webview_touch_actions(self,element,action='tap',element_postion='middle',**kwargs):
        
        if action=='tap' :
            element_coordinate=self.get_element_screen_coordinate(element,coordinate_type=element_postion,browser_height=True)
            self.driver.execute_script('mobile: tap',element_coordinate);
        if action=='long_press' :
            element_coordinate=self.get_element_screen_coordinate(element,coordinate_type=element_postion,browser_height=True)
            element_coordinate.update({'duration':2.0})
            self.driver.execute_script('mobile: touchAndHold',element_coordinate);
    
    def signout_wf(self):

        node = self.parseinitfile('nodeid')
        port = self.parseinitfile('httpport')
        context = self.parseinitfile('wfcontext')
        setup_url = 'http://' + node + ':' + port +''+ context + '/service/wf_security_logout.jsp'
        self.driver.get(setup_url)
        time.sleep(3)
    
    def rotate_screen(self,mode):
        '''
        :Param :  mode='P' ( P means PORTRAIT mode) mode='L' ( L mens LANDSCAPE)
        :Usage :  utiliobj.rotate_screen('P') or utiliobj.rotate_screen('L')
        '''
        if mode=='P' :
            self.driver.orientation='PORTRAIT'
            time.sleep(5)
            self.asequal('PORTRAIT',self.driver.orientation,'Step X : Verify screen rotated to PORTRAIT Mode')
        elif mode=='L' :
            self.driver.orientation='LANDSCAPE'
            time.sleep(5)
            self.asequal('LANDSCAPE',self.driver.orientation,'Step X : Verify screen rotated to LANDSCAPE Mode')
    
    def wait_for_webelements(self,element_css,expected_total_elements,timeout=120,**kwargs):
        '''
        :param : element_css="#MAINTABLE_wbody0_f rect[class='riser!']"
        :param : expected_total_elements= 7
        :param : timeout=60 ( Default value is 120 )
        :param : kwargs['element_text']=='State BY Product'
        :Usage : wait_for_webelements("#MAINTABLE_wbody0_f foreignObject div>span>a",1,element_text='Home')
        '''
        if 'element_text' in kwargs :
            end_time=time.time()+timeout
            while True :
                if time.time()>end_time :
                    print("Condition does not match")
                    break
                try :
                    if self.driver.find_element_by_css_selector(element_css).text.strip()==kwargs['element_text'] :
                        return False
                        break
                except :
                    pass
        else :
            try :
                WebDriverWait(self.driver,timeout).until(lambda driver:len(self.driver.find_elements_by_css_selector(element_css))==expected_total_elements)
            except :
                print ("Condition does not match")