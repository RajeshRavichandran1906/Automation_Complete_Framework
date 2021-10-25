import unittest
import re
import os
import pyautogui
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from common.locators.loginpage_locators import LoginPageLocators
import time
from selenium.common.exceptions import NoSuchElementException, NoAlertPresentException,StaleElementReferenceException
from PIL import Image, ImageGrab
import warnings
from openpyxl import Workbook
from openpyxl import load_workbook
import subprocess
from selenium.webdriver.common import alert
from selenium.webdriver import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common import keys
from selenium.webdriver.common.keys import Keys
from common.lib import utillity
import keyboard
import os
import subprocess
import uiautomation as automation
from common.pages import as_ribbon
from selenium.webdriver.support.ui import Select
from common.lib.webdriverfactory.WebDriverFactory import WebDriverFactory
from common.lib.core_utility import CoreUtillityMethods as coreutillobject
from selenium.webdriver.support.color import Color
import pywinauto
from common.lib.utillity import UtillityMethods
from asyncio.tasks import wait

class AS_Utillity_Methods(object):
    
    asert_failure_count=0
    windows=[]
    browser_x = 0
    browser_y = 0
    SHORTWAIT = 2
    WINDOW_WAIT=30
    
    def __init__(self, driver):
        self.driver = driver
        self.shortwait = WebDriverWait(self.driver, 50)
        self.mediumwait = WebDriverWait(self.driver, 100)
        self.longwait = WebDriverWait(self.driver, 150)
        
    def parseinitfile(self, key):
        init_file = 'config.init'
        config_pair = {}
        try:
            fileObj = open(init_file, "r")
            line = fileObj.readline()
            while line:
                lineObjbj = re.match(r'(\S*)\s(.*)', line)
                keyName = lineObjbj.group(1)
                config_pair[keyName] = lineObjbj.group(2)
                line = fileObj.readline()
            fileObj.close()
        except IOError:
            exit()

        if key in config_pair:
            return (config_pair[key])
        else:
            return ('Key not found')
        
    def asequal(self, *params):
        try:
            testobj = unittest.TestCase()
            testobj.assertEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            AS_Utillity_Methods.asert_failure_count += 1
            
    def as_not_equal(self, *params):
        try:
            testobj = unittest.TestCase()
            testobj.assertNotEqual(params[0], params[1], params[2])
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            AS_Utillity_Methods.asert_failure_count += 1
        
    def asin(self, *params):
        try:
            #testobj = unittest.TestCase()
            assert params[0] in params[1], params[2]
            print(params[2] + " - PASSED.")
        except AssertionError as e:
            m = re.match(r'.*(Step.*)...', str(e.args))
            print(m.group(1) + " - FAILED.")
            utillity.UtillityMethods.asert_failure_count += 1
        
    def close_se_driver(self,se_driver):
        
        se_driver.quit()
        time.sleep(1)    
            
    def login_wf(self, se_driver,website):
        loginid = self.parseinitfile('mrid')
        loginpwd = self.parseinitfile('mrpass')
  
        website =website+"signin" 
        se_driver.get(website)
        time.sleep(2)
        se_driver.find_element_by_id("SignonUserName").send_keys(loginid)
        se_driver.find_element_by_id("SignonPassName").send_keys(loginpwd)
        se_driver.find_element_by_id("SignonbtnLogin").click()
        time.sleep(2)

    def get_se_driver(self, file_name):        
        node = self.parseinitfile('clientid')
        port = self.parseinitfile('port')
        context = self.parseinitfile('context')
        browser = self.parseinitfile('browser')
        setup_url = 'http://' + node + ':' + port + context + '/'
        project = self.parseinitfile('project_id')
        suite = self.parseinitfile('suite_id')
        folder = project + '/' + suite
        api_url = setup_url + 'run.bip?BIP_REQUEST_TYPE=BIP_RUN&BIP_folder=IBFS%253A%252FWFC%252FRepository%252F' + "/" + folder + '&BIP_item=' + file_name        
        browser_driver = self.parseinitfile('browser_driver')
        if browser == "Chrome":
            options = webdriver.ChromeOptions()
            prefs = {"download.prompt_for_download": True}
            options.add_experimental_option("prefs", prefs)
            options.add_argument('--disable-print-preview')
            options.add_argument("--start-maximized")
            options.add_argument('--disable-infobars')
            options.binary_location = (r"c:\Users\bigscm\AppData\Local\Google\Chrome\Application\chrome.exe")
            se_driver = webdriver.Chrome(browser_driver,chrome_options=options)
        elif browser == 'Firefox':
            se_driver=webdriver.Firefox()
        #elif browser == "Ie":
            #se_driver = webdriver.Ie()
        else:
            #print("No browser selected")
            cap = DesiredCapabilities.INTERNETEXPLORER
            cap['requireWindowFocus']=True
            se_driver = webdriver.Ie(capabilities=cap)
        print(setup_url)
        self.login_wf(se_driver,setup_url)
        time.sleep(5)
        se_driver.get(api_url)
        time.sleep(10)
        return(se_driver)
    
    def click_html_webpage(self, title, name):
        '''
        title: get this from driver.title (Selenium browser driver)
        name: get this from driver.name (Selenium browser driver)
        The self.driver used in this function must be the Winium driver
        Example:
        se_driver = webdriver.Chrome()
        click_html_webpage(se_driver.title, se_driver.name)
        '''
        driver = self.driver
        browsers = {"chrome": "Google Chrome", "internet explorer": "Internet Explorer", "firefox": "Mozilla Firefox"}
        #print(title + " - " + browsers[name])
        #driver.find_element_by_name(title + " - " + browsers[name]).click()
        browser = AS_Utillity_Methods.parseinitfile(self,'browser')
        if browser=="Chrome":
            object=automation.PaneControl
        else:
            object=automation.WindowControl
            
        object(Name=title + " - " + browsers[name]).SetFocus()
    
    def select_tree_view_pane_item_fast_no_check(self,parent_obj, tree_path):                
        """
        Usage: select_item_from_tree_view("driver.find_element_by_name("Static"), "jawin7->Domains->P20")
        params:parent_obj="driver.find_element_by_name("Static" ) ;  "jawin7->Domains->P20"
        """        
        pyautogui.hotkey("home")
        time.sleep(2)                   
        path_list=tree_path.split("->")  
        for arg in path_list:
            print(arg)
            ac = ActionChains(self.driver)
            ac.double_click(self.driver.find_element_by_name(arg)).perform()
            time.sleep(2)
            del ac
            time.sleep(1)
    
    def select_tree_view_pane_item(self,tree_path,**kwargs):
        
        """
        Usage: select_tree_view_pane_item(jawin7->Domains->P20", Parent_Object = driver.find_element_by_name("Static"))
        params:parent_obj="driver.find_element_by_name("Static" );"jawin7->Domains->P20"
        """ 
        try:
            pyautogui.FAILSAFE=False
            driver = self.driver
            if "Parent_Object" in kwargs:
                driver.find_element(*kwargs["Parent_Object"]).click()
            else:
                tree_view = automation.TreeControl(ClassName="SysTreeView32")
                tree_view.Click(ratioX=10,ratioY=140)
                automation.SendKey(automation.Keys.VK_HOME,waitTime=3) 
            path_list=tree_path.split("->")
            for i in path_list:
                time.sleep(5)
                automation.TreeItemControl(Name=i).ScrollIntoView()
                automation.TreeItemControl(Name=i).Expand(waitTime=3)
                if (automation.TreeItemControl(Name=i).CurrentExpandCollapseState()==3):
                    automation.TreeItemControl(Name=i).DoubleClick(waitTime=3)
        except:
                print("Tree element does not exist")
                
    def ui_double_click_tree_view_item(self, tree_path,**kwargs):
        
        '''==========================================================================
        Description : To expand any folder or to double click any file.
        using click in right click menu for ENV Tree: as_utilo_obj.select_tree_view_item_using_ui(self,tree_path,**kwargs):
        @author: Robert
        ========================================================================='''
        
        try:
            pyautogui.FAILSAFE=False
            driver = self.driver
            if "Parent_Object" in kwargs:
                driver.find_element(*kwargs["Parent_Object"]).click()
            else:
                driver.find_element_by_class_name("SysTreeView32").click()
            pyautogui.hotkey("home")
            path_list=tree_path.split("->")
            for i in path_list:
                automation.TreeItemControl(Name=i).ScrollIntoView()
                automation.TreeItemControl(Name=i).DoubleClick()
        except:
                print("No Element Found")

    def select_list_view_pane_item(self,parent_obj, list_view_item, **kwargs): 
        """
        :param: parent_obj=driver.find_element_by_name("Save As")
        :param: list_view_item="C2022160.htm"
        :Syntax: select_items_from_list_view(driver.find_element_by_name("Save As"), "C2022160.htm")
        @author = Jesmin  
        """
        if parent_obj.is_displayed(): 
            ''' Click on 0th element '''
            parent_obj.find_element_by_id("1514").find_elements()[0].click()
            time.sleep(1)
            temp=0
            '''click on pagedown if item is not in view'''
            if 'int_off_set' in kwargs:
                while kwargs['int_off_set']>temp:
                    pyautogui.hotkey("pagedown")
                    temp=temp+1
                    time.sleep(1)
        parent_obj.find_element_by_id("1514").find_element_by_name(list_view_item).click()
        if 'more_items' in kwargs:
            pyautogui.keyDown("ctrl")
            parent_obj.find_element_by_id("1514").find_element_by_name(kwargs['more_items']).click()
            pyautogui.keyUp("ctrl")
        time.sleep(2)         
         
    def open_file_dialog(self,parent_obj,list_view_item,**kwargs): 
        """
        :param: parent_obj=driver.find_element_by_name("Open File") ; tree_path="jawin7->Domains->P20->S6811" ; list_view_item= "zip3_baselayer.fex"
        :Syntax: open_file_dialog(self,parent_obj,list_view_item,)
        @author = Jesmin  
        """
        if 'tree_path' in kwargs:
            self.select_tree_view_pane_item(parent_obj, kwargs['tree_path'])
             
        '''open an existing item'''
        self.select_list_view_pane_item(parent_obj, list_view_item)        
        #TODO:   Add function to verify selected item name   
        #selected_item=parent_obj.find_element_by_id("1516").get_attribute("value") 
        #print(selected_item)  
        #utillity.UtillityMethods.asequal(list_view_item,select_item,"Verify selected item appears in File Name edit box")
        if 'more_items' in kwargs:
            for x in kwargs['more_items']:
                self.select_list_view_pane_item(parent_obj, x) 
        '''Click on OK button '''
        parent_obj.find_element_by_name("OK").click() 
    
    def save_as_dialog(self,parent_obj,file_name,**kwargs): 
        """
        :param: parent_obj=driver.find_element_by_name("Open File") ;
        :Syntax: open_file_dialog(self,parent_obj,tree_path)
        @author = Jesmin  
        """
        if 'tree_path' in kwargs:
            self.select_tree_view_pane_item(parent_obj, kwargs['tree_path']) 
        
        '''delete default name & enter new file name'''
        parent_obj.find_element_by_id("1516").click()
        time.sleep(1)
        pyautogui.hotkey('ctrl','a')
        time.sleep(1)
        parent_obj.find_element_by_id("1516").send_keys(file_name)
        time.sleep(2)
        
        '''Click on OK button to save new or override '''
        parent_obj.find_element_by_name("OK").click()
        try:
            self.driver.find_element_by_name("Yes").click()
        except NoSuchElementException:
            pass
        #TODO:   Add function to verify selected item name   
        #selected_item=parent_obj.find_element_by_id("1516").get_attribute("value") 
        #print(selected_item)  
        #utillity.UtillityMethods.asequal(list_view_item,select_item,"Verify selected item appears in File Name edit box")     
        
    def menu_keyboard_navigation(self, path):
        '''
        Send in path as a tuple of individual characters such as: ('c', 'o', 'e', 'm') to get ESRI Maps. CO: Components, EM: ESRI Maps
        Hotkeys will navigate through the ribbon/toolbar/application menu
        
        '''
        import pyautogui
        pyautogui.PAUSE = 1
        driver = self.driver
        #app_studio = driver.find_element_by_id("59398")
        #app_studio.click()
        self.element_clicker(self.driver, (By.ID, "59398"), "left", "offset", 1000, 10)
        time.sleep(2)
        pyautogui.hotkey('alt')
        for keys in path:
            pyautogui.hotkey(keys)  
            time.sleep(1)
        time.sleep(4)

    def element_clicker(self, parent_obj, object, click_type, move_type, *args):
        '''
        Usage: element_clicker(driver, (By.NAME, "Environments"), "left", "no offset")
               element_clicker(driver.find_element_by_name("Settings"), (By.NAME, "Label"), "double", "offset", 100, 5)
        Param: parent_obj - can be driver or the full path starting from driver until the parent of the element to click on
                object - use the tuple system to input the find type and identifier
                click_type - options below 
                move_type - options below (can input "none" if you will not be moving for some reason)
                *args - currently just add two parameters at the end for x,y coordinates (offset) as shown in the second usage
        Description: Uses action chains to move to and click on any element or relative to any element. The getattr() functions employed takes your function
                sent in as a parameter and applies your clicks/moves accordingly
        Author: Lawrence Yu
        Date: 4/26/17 (for usage)
        '''
        driver = self.driver
        click_types = {'left': 'click', 'right': 'context_click', 'double': 'double_click', 'hold': 'click_and_hold'}
        move_types = {'no offset': 'move_to_element', 'offset': 'move_to_element_with_offset'}
        ac_menu = ActionChains(driver)
        if move_type in move_types:
            getattr(ac_menu, move_types[move_type])(parent_obj.find_element(*object), *args)
            getattr(ac_menu, click_types[click_type.lower()])().perform()
        else:
            getattr(ac_menu, click_types[click_type.lower()])(parent_obj.find_element(*object)).perform()
        del ac_menu
        time.sleep(1)
             
    def release(self, *args):
        '''
        ActionChains release method in case a click is held 
        '''
        ac_release = ActionChains(self.driver)
        ac_release.release(*args).perform()
        time.sleep(1)
    
    '''def select_item_from_dropdown_menu(self, parent_object, click_type, menu_item, move_type, *args):
        driver = self.driver
        click_types = {'left': 'click', 'right': 'context_click', 'double': 'double_click'}
        move_types = {'no offset': 'move_to_element', 'offset': 'move_to_element_with_offset'}
        ac_menu = ActionChains(driver)
        getattr(ac_menu, move_types[move_type])(driver.find_element(*parent_object), *args)
        getattr(ac_menu, click_types[click_type.lower()])()
        ac_menu.perform()
        time.sleep(1)
        del ac_menu
        driver.find_element_by_name(menu_item).click()  
        time.sleep(1)'''
        
    def select_item_from_dropdown_menu(self, parent_object, click_type, menu_path, move_type, *args):
        '''
        select_item_from_dropdown_menu((By.ID, "23456"), "Left", "City", "no offset")
        select_item_from_dropdown_menu((By.NAME, "Geographic Role"), "Left", "City", "offset", 100, 0)
        
        parent_object: takes in a tuple starting with find method and ending with the find value
        click_type: left, right, double - these are mouse clicks
        menu_item: select key from item_dict by item name
        move_type: offset, no offset - tell the function whether or not you will be including an offset when opening the menu
        *args: either input nothing more or two ints, x and y, to include an offset
        Extra usage note: in the case of multiple submenus, menu_path will still be a single string, but include => in between each item in the path as follows: "Security=>Rules..."
        '''
        driver = self.driver
        self.element_clicker(driver, parent_object, click_type, move_type, *args)
        path_list=menu_path.split("=>")  
        for x in path_list:
            self.element_clicker(driver, (By.NAME, x), "left", "no offset")
        time.sleep(1)     
        
    def context_switch_to_as(self, file_name):
        '''
        Switch active window to App Studio
        Currently only usable when an active HTML is open and saved 
        '''
        driver = self.driver
        as_name = "App Studio - " + file_name
        driver.find_elements_by_name(as_name)[0].click()
        #print(as_name)
        #print("Total tabs with requested name: " + str(len(driver.find_elements_by_name(as_name))))
        '''try:
            if driver.find_element_by_id("59648").is_displayed() == False:
                driver.find_elements_by_name(as_name)[0].click()
            else:
                print(driver.find_element_by_id("59648").is_enabled())
        except:
            pass'''
        time.sleep(1)

    def control_click(self, parent_obj, child_obj):
        driver = self.driver
        object = parent_obj.find_element(*child_obj)
        ac = ActionChains(driver)
        ac.key_down(Keys.CONTROL)
        ac.click(object)
        ac.key_up(Keys.CONTROL)
        ac.perform()
        
    def save_esri_map_georoles_in_xl(self):
        #TODO: make this function generic
        time.sleep(3)
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        ac_geo = ActionChains(self.driver)
        ac_geo.move_to_element_with_offset(self.driver.find_element_by_name('Geographic Role'), 100, 5).click().perform()   
        box = self.driver.find_element_by_name("Settings").find_elements_by_class_name("ComboLBox")[1]
        children = box.find_elements()
        time.sleep(1)
        r=1
        for x in children:
            s.cell(row=r, column=1).value = str(x.get_attribute("Name")).strip()
            r=r+1
        wb.save("D:\\test.xlsx")  
        
    def compare_menu_list_from_xl(self,file_name,sheet_name,obj,x_loc,y_loc,menu_box_obj,msg):
        """
        :param: file_name="C132553.xlsx"
                sheet_name="Georole"
                obj=driver.find_element_by_name('Geographic Role')
                x_loc=100
                y_loc=5
                menu_box_obj=driver.find_element_by_name("Settings").find_elements_by_class_name("ComboLBox")[1]
                msg="Step 01: " 
        :Syntax: compare_menu_list_from_excel(self,file_name,sheet_name,obj,x_loc,y_loc,menu_box_obj,msg):
        @author = Jesmin  
        """
        wb= load_workbook(os.getcwd() + "\data\\" + file_name)
        get_sheet= wb.get_sheet_by_name(sheet_name)
        ac_geo = ActionChains(self.driver)
        ac_geo.move_to_element_with_offset(obj,x_loc,y_loc).click().perform()   
        box = menu_box_obj
        children = box.find_elements()
        time.sleep(1)
        status=True
        for r in range(0,len(children)):
            print(get_sheet.cell(row=r+1, column=1).value)
            print(str(children[r].get_attribute("Name")))
            if get_sheet.cell(row=r+1, column=1).value == str(children[r].get_attribute("Name")):
                status= True
            else:
                status= False
                break
        print(status)        
        utillity.UtillityMethods.asequal(self,True, status, msg+"Verify list of geo roles")      
    
    def select_item_from_listbox_control(self,parent_obj,item):
        """
        :PARAM: parent_obj=driver.find_element_by_name("Layer Visualizations")
        :SYNTAX: select_item_from_listbox_control()
        :Author: Jesmin
        NEED TO TEST IT
        """
        self.driver.find_element_by_name(parent_obj).find_element_by_name(item).click()
        time.sleep(1)
        
    def set_grid_control_value(self,parent_obj,x,y,row_num,value,**kwargs):
        '''
        :PARAM: parent_obj=driver.find_element_by_name("Layer Visualizations")
                x=150 ; y=20
                row_num=7
                value="22222"
                optional param: 'selection_type'- selection_type="delete" ; 'set_value'-set_value="5"
        '''
        driver=self.driver
        time.sleep(3)
        ac_col = ActionChains(driver)
        ac_col.move_to_element_with_offset(parent_obj, x,y).click().perform()  
        time.sleep(1)  
        pyautogui.hotkey('home')  
        time.sleep(1)  
        tmp=1
        while tmp < row_num:
            pyautogui.press('down')
            time.sleep(1)
            tmp=tmp+1
        pyautogui.hotkey("tab")
        time.sleep(1)
        elem = driver.find_element_by_id("3")      
        elem_name = elem.get_attribute("Name")
        time.sleep(1)
        
        if 'selection_type' in kwargs: 
            self.delete_and_set_new_value(value,**kwargs)
            #print(self.driver.find_element_by_name("Settings").driver.find_element_by_name("Transparency").get_attribute("value")) 
        else:
            while elem_name != value:
                pyautogui.press('down')
                time.sleep(1)             
                elem_name = elem.get_attribute("Name")       
                time.sleep(1)
                print(elem_name)                 
                utillity.UtillityMethods.asequal(self,value,elem_name,"Verify new value was set") 
           
    def delete_and_set_new_value(self,set_value, **kwargs):
        if 'parent_obj' in kwargs:
            ac_col = ActionChains(self.driver)
            ac_col.move_to_element_with_offset(kwargs['parent_obj'], kwargs['x'],kwargs['y']).click().perform() 
        time.sleep(1)
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(1)
        pyautogui.typewrite(set_value)
        time.sleep(1) 
    
    def save_all(self, **kwargs):
        '''
        File_Name = "C151437"
        Response = "Yes"
        '''
        driver = self.driver
        driver.find_element_by_name("Save All").click()
        try:
            file = kwargs['File_Name']
            pyautogui.typewrite(kwargs['File_Name'])
            time.sleep(1)
            driver.find_element_by_name("OK").click()
            time.sleep(1)    
            try:
                driver.find_element_by_id("65535")
                driver.find_element_by_name("Yes").click()
                time.sleep(1)
            except: 
                print("Not prompted to overwrite existing HTML")
        except:
            pass
        
        try:
            response = kwargs['Response']
            responses = ["Yes", "No"]
            if response in responses:
                driver.find_element_by_class_name("#32770").find_element_by_name(response).click()
            else:
                print("Response not found")
        except:
            print("Did not handle close dialog")
        time.sleep(1)
        
    def is_positive(self, value):
        return (value > 0)
    
    def take_monitor_screenshot(self, file_name, image_type='actual', left=0, top=0, right=0, bottom=0):#Need to delete
        """
        :param file_name: file for saving
        :param image_type: where you want to save your image in directory
        :param left: how much you want to reduce the size from left in output of your image
        :param top: how much you want to reduce the size from top in output of your image
        :param right: how much you want to reduce the size from right in output of your image
        :param bottom: how much you want to reduce the size from bottom in output of your image
        :usage take_monitor_screenshot('full_monitor_screenshot', image_type='actual', left=10, top=25, right=10, bottom=25)
        """        
        if image_type=='actual' :
            location='actual_images'
        elif image_type=='fail' :
            location='failure_captures'
        else:
            location='images'
        file_path=os.getcwd() + "\\" + location + "\\" + file_name + ".png"
        im=ImageGrab.grab()
        im.save(file_path)
        resolution=pyautogui.size()
        left_rs = left
        top_rs = top
        right_rs=resolution[0]-right
        bottom_rs=resolution[1]-bottom
        bbox = (left_rs, top_rs, right_rs, bottom_rs)
        base_image = Image.open(file_path)
        cropped_image = base_image.crop(bbox)
        cropped_image.save(file_path)
    
    def select_data_source_dialog(self,list_view_item,button,**kwargs): 
        
        '''========================================================================================================================================
        :param: list_view_item= "zip3_baselayer.fex","Finish",tree_path="jawin7->Domains->P20->S6811"
        :Syntax: as_utilobj.select_data_source_dialog("car.mas","Finish",tree_path="ibisamp")
        ========================================================================================================================================'''
        
        list_item="SysListView32"
        key_1="home"
        
        if 'tree_path' in kwargs:
            self.select_tree_view_pane_item(kwargs['tree_path'])
             
        '''open an existing item''' 
             
        automation.ListControl(ClassName=list_item).Click()
        time.sleep(1)
        
        list_view = self.driver.find_element_by_class_name('SysListView32')
        if list_view.find_element_by_name(list_view_item).is_displayed():
            list_view.find_element_by_name(list_view_item).click()
        else:
            pyautogui.hotkey(key_1)
            for i in range(10):
                if list_view.find_element_by_name(list_view_item).is_displayed():
                    list_view.find_element_by_name(list_view_item).click()
                    time.sleep(2)
                    break
                else:
                    automation.Win32API.MouseWheelDown(5)
                    time.sleep(1)
                if i == 10:
                    print("The Specified file is not listed")
                    break     
                   
        '''Click on OK button'''
        self.driver.find_element_by_name(button).click()
        time.sleep(1)
    
    def double_click_element(self,element):
        
        '''=====================================================================
        @author: Adithyaa AK : Description : To double click on element
        element : driver.find_element_by_name()
        Usage : as_utility.double_click_element(driver.find_element_by_name('Home'))
        ====================================================================='''
        
        try:
            ac = ActionChains(self.driver)
            ac.double_click(element).perform()
            time.sleep(2)
            del ac
        except NoSuchElementException:
            print('Required Element not found')
            
    def locate_item_env_detail(self,element):
        
        '''=====================================================================
        @author: Adithyaa AK : Description : To locate environment detail files
        element :
        Usage : as_utility.locate_item_env_detail('carnist.fex')
        ========================================================'''
        try:
            ac = ActionChains(self.driver)
            ac.move_to_element(self.driver.find_element_by_class_name('SysListView32')).perform()
            self.driver.find_element_by_name(element).click()
            time.sleep(4)
            del ac
        except NoSuchElementException:
            print('Required Element not found')
            
    def move_to_tree_item(self,file_name):
    
        '''===================================================================================================
        @author: Adithyaa AK : Description : To move on tree element in environments detail or environments tree
        Usage : move_to_tree_item('file_name')
        ========================================================'''
        
        ac = ActionChains(self.driver)
        ac.move_to_element(self.driver.find_element_by_class_name('SysTreeView32').find_element_by_name(file_name)).click().perform()
        time.sleep(5)
        del ac
        
    def double_click_element_using_offset(self,x,y): 
        
        '''=====================================================================
        @author: Adithyaa AK : Description : To double click on element using offset
        element : double_click_element_using_offset(self,x,y)
        ========================================================'''
        
        ac = ActionChains(self.driver)
        ac.move_by_offset(x,y).double_click().perform()
        time.sleep(2)
    
    def click_and_drag(self,dialog,x_to,y_to):
        
        '''=====================================================================
        @author: Adithyaa AK : Description : To drag a element using offset value
        element : driver.find_element_by_name("Select Data Source")
        item: "Select Data Source" 
        Usage : as_utility.click_and_drag(self,driver.find_element_by_name("Select Data Source"),x_to,y_to)
        =================================================================================================='''
        
        element=self.driver.find_element_by_name(dialog)
        w1=element.size['width']
        h1=element.size['height']
        actions = ActionChains(self.driver)
        actions.move_to_element_with_offset(element, w1-10, h1-10).perform()
        pyautogui.mouseDown()
        time.sleep(2)
        pyautogui.moveTo(x_to,y_to)
        time.sleep(2)
        pyautogui.mouseUp()
        time.sleep(5)
    
    def select_component_by_right_click(self,**kwargs):
        
        '''==========================================================================
        @author: Adithyaa AK
        Description : To select any component by right clicking in environment tree or environment detail
        Usage: Use below function after "as_utilobj.select_tree_view_pane_item"
        
        using click in right click menu: as_utilobj.select_component_by_right_click(right_click_folder='S9100',click='Refresh Descendants')
                                                (((((or)))))
        using click in right click sub menu: as_utilobj.select_component_by_right_click("FWSubFolder",send_keys=['down','down','right'],click='Refresh Descendants')
        ========================================================================='''
        
        env_list_class_name="SysListView32"
        right_click='right'
         
        try:
                          
            if 'right_click_folder' in kwargs:
                
                tree_view = automation.TreeControl(ClassName="SysTreeView32")
                tree_view.Click(ratioX=10,ratioY=140)
                 
                if tree_view.TreeItemControl(Name=kwargs['right_click_folder']).Exists():
                    tree_view.TreeItemControl(Name=kwargs['right_click_folder']).ScrollIntoView()
                    tree_view.TreeItemControl(Name=kwargs['right_click_folder']).Click()
                    time.sleep(1)
                    automation.SendKey(automation.Keys.VK_ESCAPE,waitTime=3)   
                    time.sleep(3)
#                 else:
#                     print("Right_click_folder does not exist")
            
            if 'right_click_item' in kwargs:
                
                tree_view = automation.TreeControl(ClassName="SysTreeView32")
                tree_view.Click(ratioX=10,ratioY=140)
                
                verify_tree_item=tree_view.TreeItemControl(Name=kwargs['right_click_item']).Exists()
                
                if verify_tree_item != False:
                    tree_view.TreeItemControl(Name=kwargs['right_click_item']).ScrollIntoView()
                    tree_view.TreeItemControl(Name=kwargs['right_click_item']).Click()
                    time.sleep(1)
                    automation.SendKey(automation.Keys.VK_ESCAPE,waitTime=3)  
                    time.sleep(3) 
#                 else:
#                     print("Right_click_folder does not exist")
                
            if 'right_click_item_env_detail' in kwargs:
                
                list_view=automation.ListControl(ClassName=env_list_class_name)
                list_view.Click(ratioX=45,ratioY=60)
                
                verify_list_item=list_view.ListItemControl(Name=kwargs['right_click_item_env_detail']).Exists()
            
                if verify_list_item != False:
                    list_view.ListItemControl(Name=kwargs['right_click_item_env_detail']).ScrollIntoView()
                    list_view.ListItemControl(Name=kwargs['right_click_item_env_detail']).Click()
                    time.sleep(1)
                    automation.SendKey(automation.Keys.VK_ESCAPE,waitTime=3) 
                    time.sleep(3)
#                 else:
#                     print("Right_click_env_detail_item could not be found") 
                
            pyautogui.click(button=right_click)
                
            if 'send_keys' in kwargs:
                pyautogui.press(kwargs['send_keys'],interval=0.25)
                
            if 'click' in kwargs:
                automation.MenuItemControl(Name=kwargs['click']).Click()
                
            if 'click_sub_menu' in kwargs:
                automation.MenuItemControl(Name=kwargs['click_sub_menu']).Click()
                    
        except Exception or TimeoutError or LookupError:
            print("No avaliability of searching element")
                       
    def select_any_dialog(self,button,**kwargs): 
        
        """========================================================================================
        Description:
        * To select any dialog which invokes in App studio by selecting tree and file element. 
        * To select list item in in any list. 
        * To rename a file in a dialog and save it
        Usage: as_utilobj.select_any_dialog_("Cancel",rename_file="C2287689")
        @author:adithyaa
        ========================================================================================"""
        
        list_item_locator="SysListView32"
        edit_box="Edit"
        dialog_locator="#32770"
        
        automation.WindowControl(ClassName=dialog_locator).SetFocus()
            
        if 'tree_path' in kwargs:
            self.ui_double_click_tree_view_item(kwargs['tree_path'])
            time.sleep(2)
        
        if 'select_file' in kwargs:  
             
            list_view=automation.ListControl(AutomationId=list_item_locator)
            list_view.Click(ratioX=45,ratioY=60)
            
            verify_list_item=list_view.ListItemControl(Name=kwargs['select_file']).Exists()
        
            if verify_list_item != False:
                list_view.ListItemControl(Name=kwargs['select_file']).ScrollIntoView()
                list_view.ListItemControl(Name=kwargs['select_file']).Click() 
                
            else:
                print("File to be selected does not exist") 
            
        if 'rename_file' in kwargs:  
            automation.EditControl(ClassName=edit_box).DoubleClick()
            time.sleep(1)
            keyboard.write(kwargs['rename_file'])
            time.sleep(3) 
               
        '''Click on OK button'''
        automation.ButtonControl(Name=button).Click(waitTime=1)
        
    def select_file_in_dialogs(self,button,**kwargs): 
        
        '''========================================================================================================================================
        :param: parent_obj=driver.find_element_by_name("Open File"),"Finish",tree_path="jawin7->Domains->P20->S6811",list_item="car.mas"
        :Syntax: as_utilobj.select_file_in_list_view("OK",tree_path="ibisamp",select_file="ggsales.mas")
        @author: adithyaa
        ========================================================================================================================================'''
        
        dialog_locator="#32770" 
        list_item_locator="SysListView32"
        
        automation.WindowControl(ClassName=dialog_locator).SetFocus()
        
        if 'tree_path' in kwargs:
            self.select_tree_view_pane_item(kwargs['tree_path'])
            time.sleep(2)
        
        if 'select_file' in kwargs:  
             
            list_view=automation.ListControl(ClassName=list_item_locator)
            list_view.Click(ratioX=90,ratioY=90)
            
            verify_list_item=list_view.ListItemControl(Name=kwargs['select_file']).Exists()
        
            if verify_list_item != False:
                list_view.ListItemControl(Name=kwargs['select_file']).ScrollIntoView()
                list_view.ListItemControl(Name=kwargs['select_file']).Click()  
            else:
                print("File to be selected does not exist")
    
        '''Click on OK button'''
        automation.ButtonControl(Name=button).Click()
        time.sleep(2)
        
    def select_multiple_files(self,initial_file,files_to_select,**kwargs):
        
        '''===========================================================================================
        :param: initial_file="first_file_name_to_be_clicked", files_to_select=['CarTwoParms','Chart1']
        :Syntax: as_utilobj.select_multiple_files("CarParm",['CarTwoParms','Chart1'])
        @author: adithyaa
        =============================================================================================='''
        
        tree_view=self.driver.find_element_by_class_name("SysTreeView32")
        
        automation.TreeItemControl(Name=initial_file).Click()
        pyautogui.keyDown('ctrl')
        time.sleep(2)
        
        for items in files_to_select:
            if tree_view.find_element_by_name(items).is_displayed():
                automation.TreeItemControl(Name=items).Click()
                time.sleep(2)
                            
        if "send_keys" in kwargs:
            pyautogui.click(button='right')
            pyautogui.keyUp('ctrl')
            time.sleep(1)
            pyautogui.press(kwargs["send_keys"])
            
        if "click" in kwargs:
            pyautogui.click(button='right')
            pyautogui.keyUp('ctrl')
            time.sleep(1)
            automation.MenuItemControl(Name=kwargs["click"]).Click()
            
        pyautogui.keyUp('ctrl')
        time.sleep(2)
              
    def click_tree_item(self,**kwargs):
         
        '''==========================================================
        Description : To search and select tree_item
        Usage : as_utilobj.select_tree_item(self,parent_obj,**kwargs):
        @author: adithyaa
        ===========================================================''' 
        
        tree_locator="SysTreeView32"
        home_key="Home"
         
        if 'tree_path' in kwargs:
            self.select_tree_view_pane_item(kwargs['tree_path'])
              
        '''open an existing item'''
             
        if 'click_file' in kwargs:
            
            self.driver.find_element_by_class_name(tree_locator).click()
            time.sleep(1)
            tree_view = self.driver.find_element_by_class_name(tree_locator)
            
            if tree_view.find_element_by_name(kwargs['click_file']).is_displayed():
                    tree_view.find_element_by_name(kwargs['click_file']).click()
            else:
                pyautogui.hotkey(home_key)     
                for i in range(20):
                    if tree_view.find_element_by_name(kwargs['click_file']).is_displayed():
                        tree_view.find_element_by_name(kwargs['click_file']).click()
                        break
                    else:
                        automation.Win32API.MouseWheelDown(2)
                        time.sleep(1)
                    if i == 20:
                        print("The Specified file is not listed")
                        break
        
    def select_application_menu_options(self,**kwargs):
        
        '''==========================================================================================================================================
        Description : To select the options in application menu of app studio. Using shortcut keys.
        Usage : as_utilobj.select_application_menu_options("App Studio - 110045NonRIA*",save_as=['down','down','down','enter'])
        @author:adithyaa
        ==================================================================================================================================='''
        
        application_menu="Application Menu" 
        
        if 'open_application_menu' in kwargs:
            automation.ButtonControl(Name=application_menu).Click()
         
        if 'options' in kwargs:
            automation.SendKey(automation.Keys.VK_MENU, waitTime=3)
            time.sleep(2)
            automation.SendKey(automation.Keys.VK_F, waitTime=3)
            time.sleep(2)
            automation.SendKey(automation.Keys.VK_T, waitTime=3)
            time.sleep(2)
            
        if 'send_keys' in kwargs:
            automation.ButtonControl(Name=application_menu).Click()
            time.sleep(1)
            pyautogui.press(kwargs['send_keys'])
            automation.SendKey(automation.Keys.VK_RETURN,waitTime=2)
            time.sleep(2)
        
    def click_by_offset(self,xoffset,yoffset):
        
        '''@author: Adithyaa AK : Description : To click on any element using offset.
        ============================================================================
        Usage : as_ribbon_obj.click_by_offset(self,244,40)'''
             
        action = ActionChains(self.driver)
        action.move_by_offset(xoffset,yoffset).click().perform()
        time.sleep(2)
        del action
        
    def select_home_button(self,**kwargs):
        
        '''@author: Adithyaa AK : Description : To click at any point on the screen from common element "Home".
        ============================================================================
        Usage : select_home_button() or select_home_button(self,move_x=319,move_y=114):'''
        
        automation.TabItemControl(Name="Home").Click(waitTime=1)
        
        if 'move_x' in kwargs:
            AS_Utillity_Methods.click_by_offset(self,kwargs['move_x'],kwargs['move_y'])
            
    def logout_conf_env(self,**kwargs):
        
        '''============================================================
        Adithyaa : To logout from environment of Appstudio
        parameters: envpath->config.init
        Usage : as_utilobj.logout_environment()
        @author: adithyaa
        ===========================================================''' 
        
        if "webfocus_environment" in kwargs:
            
            tree_path=self.parseinitfile("webfocus_environment")
            self.select_tree_view_pane_item(tree_path)
            automation.TreeItemControl(Name=tree_path).RightClick()
            pyautogui.press(['down','enter'])
            time.sleep(4)
            
        if "envpath" in kwargs:
            
            tree_path=self.parseinitfile("envpath")
            self.select_tree_view_pane_item(tree_path)
            automation.TreeItemControl(Name=tree_path).RightClick()
            pyautogui.press(['down','enter'])
            time.sleep(4)
                    
        if "local_environment" in kwargs:
            
            tree_path=self.parseinitfile("local_environment")
            self.select_tree_view_pane_item(tree_path)
            automation.TreeItemControl(Name=tree_path).RightClick()
            pyautogui.press(['down','enter'])
            time.sleep(4)  
        
    def web_element_click(self,title,**kwargs):
        
        '''============================================================
        Description : To logout from environment of Appstudio
        parameters: envpath->config.init
        Usage : as_utilobj.web_element_click("App Studio - Page Designer",click_button="Create"):
        @author: adithyaa
        ===========================================================''' 
        
        browser_locator=title
        
        automation.WindowControl(ClassName=browser_locator).SetFocus()   
        web_element=automation.WindowControl(ClassName=browser_locator)  
        web_element
            
        if "button_click" in kwargs:
            web_element.ButtonControl(Name=kwargs['click_button']).Click(waitTime=3)
            
        if "text_click" in kwargs:
            web_element.TextControl(Name=kwargs['text_click']).Click(waitTime=3)
            
        if "group_click" in kwargs:
            web_element.GroupControl(Id=kwargs['group_click']).Click(waitTime=3)
            
        
    def check_show_project_area(self,**kwargs):
        
        '''==========================================================
        Description : To enable show project area
        parameters: enable='Show Projects area'
        Usage : as_utilobj.check_show_project_area(self,enable='Show Projects area'):
        @author: adithyaa
        ===========================================================''' 
        
        environment_list="Environments"
        ok_button='OK'
        
        self.select_application_menu_options(options=True)
        
        self.click_element_using_ui(list_item=environment_list)

        if 'enable' in kwargs:
            
            get_state=automation.CheckBoxControl(Name=kwargs['enable']).CurrentToggleState()
            if get_state==0:
                automation.CheckBoxControl(Name=kwargs['enable']).Click()
                automation.ButtonControl(Name=ok_button).Click()
            if get_state==1:
                print("Project area already enabled")
                automation.ButtonControl(Name=ok_button).Click()  
                    
        if 'disable' in kwargs:
            
            get_state=automation.CheckBoxControl(Name=kwargs['disable']).CurrentToggleState()
            if get_state==0:
                print("Project area already disabled")
                automation.ButtonControl(Name=ok_button).Click()
            if get_state==1:
                automation.CheckBoxControl(Name=kwargs['disable']).Click()
                automation.ButtonControl(Name=ok_button).Click()
                
    def click_element_using_ui(self,**kwargs): 
        
        '''==========================================================
        Description : To click the elements using ui automation functionality
        Usage : as_utility.click_element_using_ui(self,menu_item=True,name='Show Projects area')
        @author: adithyaa
        ===========================================================''' 
        
        try:
            
            if "window_click" in kwargs:
                automation.WindowControl(Name=kwargs['window_click']).Click()
                        
            if "window_click_with_offset" in kwargs:
                automation.WindowControl(Name=kwargs['window_click_with_offset']).Click(ratioX=kwargs["x"],ratioY=kwargs["y"])            
            
            if 'menu_item' in kwargs:
                
                if 'name' in kwargs:
                    automation.MenuItemControl(Name=kwargs['name']).Click(waitTime=4)
                elif 'class_name' in kwargs:
                    automation.MenuItemControl(ClassName=kwargs['class_name']).Click()
                    
            if 'edit_element' in kwargs:
                
                if "name" in kwargs:
                    automation.EditControl(Name=kwargs['name']).DoubleClick()
                    
                    if 'send_key' in kwargs:
                        pyautogui.hotkey(kwargs['send_key'])                                       
                        time.sleep(1)
                    if 'write' in kwargs: 
                        keyboard.write(kwargs['write'])
                        
                elif "id" in kwargs:
                    automation.EditControl(AutomationId=kwargs['id']).DoubleClick()
                    if 'send_key' in kwargs:
                        pyautogui.hotkey(kwargs['send_key'])                                       
                        time.sleep(1)
                    if 'write' in kwargs:
                        keyboard.write(kwargs['write'])                                     
                    
            if 'button_item' in kwargs:
                
                if 'name' in kwargs:
                    automation.ButtonControl(Name=kwargs['name']).Click(waitTime=2)
                elif 'id' in kwargs:
                    automation.ButtonControl(AutomationId=kwargs['id']).Click(waitTime=2)
                    
            if 'split_button' in kwargs:
                automation.SplitButtonControl(Name=kwargs['split_button']).Click(waitTime=1)
                if 'send_keys' in kwargs:
                    pyautogui.press(kwargs['send_keys'],interval=0.25)
                    
            if 'split_button_with_offset' in kwargs:
                automation.SplitButtonControl(Name=kwargs['split_button_with_offset']).Click(ratioX=kwargs["x"], ratioY=kwargs["y"])
                if 'send_keys' in kwargs:
                    pyautogui.press(kwargs['send_keys'])
                    automation.SendKey(automation.Keys.VK_RETURN,waitTime=2)       
            
            if "text_click" in kwargs:
                automation.TextControl(Name=kwargs['text_click']).Click(waitTime=2)
                
            if "text_double_click" in kwargs:
                automation.TextControl(Name=kwargs['text_double_click']).DoubleClick(waitTime=3)   
                      
            if 'pane_control' in kwargs:
                automation.PaneControl(AutomationId=kwargs['pane_control']).DoubleClick(ratioX=kwargs['x'],ratioY=kwargs['y'])
                
            if "group_click" in kwargs:
                automation.GroupControl(AutomationId=kwargs['group_click']).Click()
                
            if 'radio_button' in kwargs:
                automation.RadioButtonControl(AutomationId=kwargs['radio_button']).Click(waitTime=2) 
                
            if 'tree_item' in kwargs:
                automation.TreeItemControl(Name=kwargs['tree_item']).DoubleClick(waitTime=2)
                
            if 'tree_item_click' in kwargs:
                automation.TreeItemControl(Name=kwargs['tree_item_click']).Click()
            
            if 'tab_item' in kwargs:
                automation.TabItemControl(Name=kwargs['tab_item']).Click(waitTime=2)
                
            if 'tab_control' in kwargs:
                automation.TabControl(AutomationId=kwargs['tab_control']).Click(ratioX=kwargs["x"],ratioY=kwargs["y"]) 
                
            if 'combo_box' in kwargs:
                automation.ComboBoxControl(AutomationId=kwargs['combo_box']).Click(waitTime=2)
                
            if 'combo_box_with_offset' in kwargs:
                automation.ComboBoxControl(AutomationId=kwargs['combo_box_with_offset']).Click(ratioX=kwargs['x'],ratioY=kwargs['y'])
                if 'combo_box_list_item' in kwargs:
                    automation.ListItemControl(Name=kwargs['combo_box_list_item']).Click()
                
            if 'check_box' in kwargs:
                
                if 'name' in kwargs:
                    automation.CheckBoxControl(Name=kwargs['name']).Click()
                elif 'id' in kwargs:
                    automation.CheckBoxControl(AutomationId=kwargs['id']).Click(waitTime=3)
                    
            if "header_item" in kwargs:
                automation.HeaderItemControl(Name=kwargs['header_item']).Click()
                
            if "list_item" in kwargs:
                automation.ListItemControl(Name=kwargs['list_item']).Click()
                
            if "list_item_double_click" in kwargs:
                automation.ListItemControl(Name=kwargs["list_item_double_click"]).DoubleClick() 
                
            if "hyperlink_control" in kwargs:
                automation.HyperlinkControl(Name=kwargs["hyperlink_control"]).Click(waitTime=2)
                
            if "close_dialog_window" in kwargs:
                automation.WindowControl(Name=kwargs["close_dialog_window"]).Close(waitTime=2)
                       
        except LookupError or TimeoutError:
            print("UI Click element cannot be found")
            
    def select_file(self,**kwargs):
        
        '''==========================================================
        Description : To select any file if it is avaliable
        Usage : as_ut_obj.select_file(self,list_item="car.mas"):
        @author: adithyaa
        ===========================================================''' 
        
        if "list_item" in kwargs:
            automation.ListItemControl(Name=kwargs['list_item']).Select()
                
        if "tree_item" in kwargs:
            automation.TreeItemControl(Name=kwargs['tree_item']).Select()
                
                    
    def select_element_appstudio_options(self,**kwargs):
        
        '''==========================================================
        Description : To click the elements in app studio option dialog, Enable\Disable
        parameters : as required to click in app studio options
        Usage : as_utilobj.select_element_appstudio_options(list_item="HTML Page",edit_element="15002",write="0")
        @author: adithyaa
        ===========================================================''' 
        
        appstudio_options="App Studio Options"
        
        try:
            
            automation.WindowControl(Name=appstudio_options).SetFocus()
            
            if 'list_item' in kwargs:
                automation.ListItemControl(Name=kwargs['list_item']).Click()
                time.sleep(2)
            
            if 'edit_element' in kwargs:
                automation.EditControl(AutomationId=kwargs['edit_element']).DoubleClick()
                pyautogui.typewrite(kwargs['write'])                                       
                time.sleep(2)
                
            if 'button' in kwargs:
                automation.ButtonControl(Name=kwargs['button']).Click()
                
            if 'radio_button' in kwargs:
                automation.RadioButtonControl(AutomationId=kwargs['radio_button']).Click()
                 
            if 'check_box' in kwargs:
                automation.CheckBoxControl(AutomationId=kwargs['check_box']).Click()
                
            if 'combo_box' in kwargs:
                automation.ComboBoxControl(AutomationId=kwargs['combo_box']).Click()
                            
        except LookupError or TimeoutError:
            print("App Studio options not been displayed")
    
    def drag_drop_component(self,tab,component,**kwargs):
        
        '''==========================================================
        Description : To drag and drop components from ribbon tab
        Usage : as_utilobj.drag_drop_component(self,"Insert","Report",**kwargs):
        @author: adithyaa
        ===========================================================''' 
        
        automation.TabItemControl(Name=tab).Click(waitTime=1)
        automation.ButtonControl(Name=component).Click(waitTime=1)
        
        if 'source_x' in kwargs:
            automation.DragDrop(kwargs['source_x'],kwargs['source_y'],kwargs['target_x'],kwargs['target_y'])
            time.sleep(4)
        
    def drag_drop(self,source_x,source_y,target_x,target_y):
        
        '''==========================================================
        Description : To drag and drop
        Usage : as_utilobj.drag_drop_component(self,source_x,source_y,target_x,target_y):
        @author: adithyaa
        ===========================================================''' 
        
        automation.DragDrop(source_x,source_y,target_x,target_y)
        time.sleep(4)
                        
    def verify_google_browser(self,title,msg,**kwargs):
        
        '''==========================================================
        Description : To verify google browser and elements
        Usage : as_utilobj.verify_google_browser("WebFOCUS Report1","msg",browser_close=True)
        @author: adithyaa
        ===========================================================''' 
         
        try:
            
            maximize_window="Maximize" 
            automation.PaneControl(Name=title).CustomControl(Name="Google Chrome").CustomControl(Name="").ButtonControl(Name=maximize_window).Click(waitTime=1)
            
            if "verify_browser" in kwargs:
                verify_browser=automation.PaneControl(Name=kwargs["verify_browser"]).Exists()
                AS_Utillity_Methods.asequal(self,verify_browser,True,msg)
            
            if "verify_image" in kwargs:
                siklui_loc=os.getcwd() +"\\common\\lib\\sikuli\\runsikulix.cmd"
                kwargs['verify_image']=os.getcwd() +"\\data\\"+kwargs["verify_image"]
                src_loc=os.getcwd() +"\\common\\lib\\sikuli\\verify_picture --arg " + kwargs["verify_image"]
                p=subprocess.Popen(siklui_loc+' -r '+ src_loc, shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
                AS_Utillity_Methods.asin(self, 'Exit code: 123', str(p.communicate()[0]), msg)
        
            if "browser_close" in kwargs:  
                automation.PaneControl(Name=title).CustomControl(Name="Google Chrome").CustomControl(Name="").ButtonControl(Name="Close").Click(waitTime=1)
                
        except:
            print("google_browser is missing")
              
    def verify_dialog_properties(self,dialog,x_to,y_to,**kwargs):
        
        '''==========================================================
        Description : To verify height_width properties of a dialog
        Usage : as_utilobj.verify_dialog_properties("dialog","msg")
        @author: adithyaa
        ===========================================================''' 
        
        dialog_locator="#32770"
        header_locator="SysHeader32"
        
        automation.WindowControl(ClassName=dialog_locator).SetFocus()
        get_width=self.driver.find_element_by_class_name(header_locator).size['width']
          
        AS_Utillity_Methods.click_and_drag(self,dialog,x_to,y_to)
        time.sleep(2)
        
        automation.WindowControl(ClassName=dialog_locator).SetFocus()
        verify_width=self.driver.find_element_by_class_name(header_locator).size['width']

        if "resize_verification_msg" in kwargs:
            
            if verify_width>get_width:
                print(kwargs["resize_verification_msg"] +  " - PASSED")
            else:
                print(kwargs["resize_verification_msg"] +  " - FAILED")
                AS_Utillity_Methods.asert_failure_count += 1
                
        if "verification_msg" in kwargs:
            
            if verify_width<get_width:
                print( kwargs["verification_msg"] +  " - PASSED")
            else:
                print(kwargs["verification_msg"] +  " - FAILED")
                AS_Utillity_Methods.asert_failure_count += 1
                    
                    
    def verify_element_using_ui(self,msg,**kwargs): 
    
        '''==========================================================
        Description : To verify the elements using ui automation functionality
        Usage : as_utility.verify_element_using_ui(self,**kwargs)
        @author: adithyaa
        ===========================================================''' 
        
        try:
#             AS_Utillity_Methods.select_home_button(self)
#             time.sleep(3)    
#             get_appstudio=automation.WindowControl(RegexName="App Stud.*").GetWindowText()
#             automation.WindowControl(Name=get_appstudio).SetFocus() 
            
            if "window_control_item" in kwargs:
                window=automation.WindowControl(Name=kwargs["window_control_item"])   
                window.SetFocus()
                window_existing=window.Exists()
                AS_Utillity_Methods.asequal(self,window_existing,True,msg)
                
                if "window_close" in kwargs:
                    window.Close()
            
            if "radio_button_id" in kwargs:
                automation.RadioButtonControl(AutomationId=kwargs["radio_button_id"]).CurrentIsSelected()
                radio_button=automation.RadioButtonControl(AutomationId=kwargs["radio_button_id"]).CurrentIsSelected()
                AS_Utillity_Methods.asequal(self,radio_button,True,msg)
                
            if "check_box" in kwargs:
                automation.CheckBoxControl(Name=kwargs["check_box"])
                check_box=automation.CheckBoxControl(Name=kwargs["check_box"]).CurrentToggleState()
                if check_box==1:
                    AS_Utillity_Methods.asequal(self,check_box,1,msg)
                if check_box==0:
                    AS_Utillity_Methods.asequal(self,check_box,0,msg)
                    
            if "menu_item_enabled" in kwargs:
                
                if "enabled" in kwargs:
                    menu_item=automation.MenuItemControl(Name=kwargs["menu_item_enabled"]).IsEnabled
                    AS_Utillity_Methods.asequal(self,menu_item,True,msg)
                    
                elif "disabled" in kwargs:
                    menu_item=automation.MenuItemControl(Name=kwargs["menu_item_enabled"]).IsEnabled
                    AS_Utillity_Methods.asequal(self,menu_item,False,msg)
                
            if "pane_item" in kwargs:
                
                automation.PaneControl(AutomationId=kwargs["pane_item"])
                pane_item=automation.PaneControl(AutomationId=kwargs["pane_item"]).Exists()
                AS_Utillity_Methods.asequal(self,pane_item,True,msg)
                
            if "text_item" in kwargs:
                
                automation.TextControl(Name=kwargs["text_item"])
                text_item=automation.TextControl(Name=kwargs["text_item"]).Exists()
                AS_Utillity_Methods.asequal(self,text_item,True,msg)
                
            if "button_item" in kwargs:
                
                if "enabled" in kwargs:
                    button_item=automation.ButtonControl(Name=kwargs["button_item"]).IsEnabled
                    AS_Utillity_Methods.asequal(self,button_item,1,msg)
                    
                elif "disabled" in kwargs:
                    button_item=automation.ButtonControl(Name=kwargs["button_item"]).IsEnabled
                    AS_Utillity_Methods.asequal(self,button_item,0,msg)
                    
            if "edit_control_item" in kwargs:
                
                if "enabled" in kwargs:
                    edit_control_item=automation.EditControl(Name=kwargs["edit_control_item"]).IsEnabled
                    AS_Utillity_Methods.asequal(self,edit_control_item,1,msg)
                    
                elif "disabled" in kwargs:
                    edit_control_item=automation.EditControl(Name=kwargs["edit_control_item"]).IsEnabled
                    AS_Utillity_Methods.asequal(self,edit_control_item,0,msg)

            if "split_button_item" in kwargs:
                
                if "enabled" in kwargs:
                    split_button_item=automation.SplitButtonControl(Name=kwargs["split_button_item"]).IsEnabled
                    AS_Utillity_Methods.asequal(self,split_button_item,1,msg)
                    
                elif "disabled" in kwargs:
                    split_button_item=automation.SplitButtonControl(Name=kwargs["split_button_item"]).IsEnabled
                    AS_Utillity_Methods.asequal(self,split_button_item,0,msg)
                    
            if "tab_item" in kwargs:
                
                if "enabled" in kwargs:
                    tab_item=automation.TabItemControl(Name=kwargs["tab_item"]).Exists()
                    AS_Utillity_Methods.asequal(self,tab_item,True,msg)
                    
                elif "disabled" in kwargs:
                    tab_item=automation.TabItemControl(Name=kwargs["tab_item"]).Exists()
                    AS_Utillity_Methods.asequal(self,tab_item,False,msg)
                    
            if "tab_item_status" in kwargs:
                
                if "active" in kwargs:
                    tab_item=automation.TabItemControl(Name=kwargs["tab_item_status"]).CurrentIsSelected()
                    AS_Utillity_Methods.asequal(self,tab_item,True,msg)
                    
                elif "inactive" in kwargs:
                    tab_item=automation.TabItemControl(Name=kwargs["tab_item_status"]).CurrentIsSelected()
                    AS_Utillity_Methods.asequal(self,tab_item,False,msg)
                    
            if "tree_item" in kwargs:
                
                if "available" in kwargs:
                    tree_item=automation.TreeItemControl(Name=kwargs["tree_item"]).Exists()
                    AS_Utillity_Methods.asequal(self,tree_item,True,msg)
                    
                if "unavailable" in kwargs:
                    tree_item=automation.TreeItemControl(Name=kwargs["tree_item"]).Exists()
                    AS_Utillity_Methods.asequal(self,tree_item,False,msg)    
            
            if "list_item" in kwargs:
                
                if "selected" in kwargs:  
                    list_item=automation.ListItemControl(Name=kwargs["list_item"]).CurrentIsSelected()
                    AS_Utillity_Methods.asequal(self,list_item,True,msg)
                    
                elif "unselected" in kwargs:
                    list_item=automation.ListItemControl(Name=kwargs["list_item"]).CurrentIsSelected()
                    AS_Utillity_Methods.asequal(self,list_item,False,msg)
                    
            if "list_item_exist" in kwargs:
                
                if "available" in kwargs:  
                    list_item=automation.ListItemControl(Name=kwargs["list_item_exist"]).Exists()
                    AS_Utillity_Methods.asequal(self,list_item,True,msg)
                    
                elif "unavailable" in kwargs:
                    list_item=automation.ListItemControl(Name=kwargs["list_item_exist"]).Exists()
                    AS_Utillity_Methods.asequal(self,list_item,False,msg)
                    
            if 'multi_text_item' in kwargs:#pass text values in list
                for text in kwargs['multi_text_item']:
                    items=automation.TextControl(Name=text).Exists(0,0)
                    AS_Utillity_Methods.asequal(self,items,True,msg +" verify" + text + " displayed")  
                    
            if "pane_control" in kwargs:
                
                if "available" in kwargs:
                    tree_item=automation.PaneControl(Name=kwargs["pane_control"]).Exists()
                    AS_Utillity_Methods.asequal(self,tree_item,True,msg)
                    
                if "unavailable" in kwargs:
                    tree_item=automation.TreeItemControl(Name=kwargs["pane_control"]).Exists()
                    AS_Utillity_Methods.asequal(self,tree_item,False,msg)   

        except LookupError or TimeoutError:
            print("Verification Failed")
            AS_Utillity_Methods.asert_failure_count += 1
            
    def verify_edit_control_value(self,edit_control_id,verify_value,msg):
        
        '''==========================================================
        Description : To verify the elements using ui automation functionality
        Usage : as_utility.verify_edit_control_value(self,"34578","car.mas",msg)
        @author: adithyaa
        ===========================================================''' 
        
        get_value=automation.EditControl(AutomationId=edit_control_id).AccessibleCurrentValue()
        time.sleep(1)
        AS_Utillity_Methods.asequal(self,get_value,verify_value,msg)
            
    def verify_picture_using_sikuli(self,image_name, msg):
        """      
        Description:- In order to verify any unavoidable image comparison situation, then keep your image under data folder.
        :param image_name='image1.png'
        :param msg='your message'
        :Usage: verify_picture_using_sikuli('image1.png', 'your message') 
        """
        siklui_loc=os.getcwd() +"\\common\\lib\\sikuli\\runsikulix.cmd"
        image_name=os.getcwd() +"\\data\\"+image_name
        src_loc=os.getcwd() +"\\common\\lib\\sikuli\\verify_picture --arg " + image_name
        p=subprocess.Popen(siklui_loc+' -r '+ src_loc, shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        AS_Utillity_Methods.asin(self, 'Exit code: 123', str(p.communicate()[0]), msg)
    
    def verify_picture_with_tolerance_using_sikuli(self,image_name, image_tolerance,msg):
        """      
        Description:- In order to verify any unavoidable image comparison situation, then keep your image under data folder.
        :param image_name='image1.png'
        :param image_tolerance='0.5'
        :param msg='your message'
        :Usage: verify_picture_with_tolerance_using_sikuli('image1.png', '0.5','your message') 
        """
        siklui_loc=os.getcwd() +"\\common\\lib\\sikuli\\runsikulix.cmd"
        image_name=os.getcwd() +"\\data\\"+image_name
        src_loc=os.getcwd() +"\\common\\lib\\sikuli\\verify_picture_tolerance --arg " + image_name + image_tolerance
        p=subprocess.Popen(siklui_loc+' -r '+ src_loc, shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        AS_Utillity_Methods.asin(self, 'Exit code: 123', str(p.communicate()[0]), msg)
        
    def verify_exact_picture_using_sikuli(self,image_name, msg):
        """      
        Description:- In order to verify any unavoidable image comparison situation, then keep your image under data folder.
        :param image_name='image1.png'
        :param msg='your message'
        :Usage: verify_picture_using_sikuli('image1.png', 'your message') 
        """
        siklui_loc=os.getcwd() +"\\common\\lib\\sikuli\\runsikulix.cmd"
        image_name=os.getcwd() +"\\data\\"+image_name
        src_loc=os.getcwd() +"\\common\\lib\\sikuli\\verify_picture_exact --arg " + image_name
        p=subprocess.Popen(siklui_loc+' -r '+ src_loc, shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
        AS_Utillity_Methods.asin(self, 'Exit code: 123', str(p.communicate()[0]), msg)
        
    def verify_accesible_current_value(self,verify_value,msg,**kwargs):
        
        if "edit_control" in kwargs:
            automation.EditControl(AutomationId=kwargs["edit_control"]).DoubleClick() 
            edit_control_value=automation.EditControl(AutomationId=kwargs["edit_control"]).AccessibleCurrentValue()
            AS_Utillity_Methods.asequal(self,edit_control_value,verify_value,msg)
            
    def verify_text_appstudio_canvas(self,msg,**kwargs):
        
        '''==========================================================================================================================================
        @author: adithyaa
        Description : To verify text elements in browser
        Usage : as_util_obj.Verify_Browser_Content('WebFOCUS Report - Internet Explorer',"Step 01: Element avaliable",item_list=['CAR','ALFA ROMEO','AUDI'])
        ==========================================================='''        
        
        appstudio_title="App Studio*"
        AS_Utillity_Methods.select_home_button(self)
        time.sleep(3)    
        get_tool=automation.WindowControl(RegexName=appstudio_title).GetWindowText()
        automation.WindowControl(Name=get_tool)  
                    
        if "item_list" in kwargs:
            for text in kwargs['item_list']:
                items=automation.TextControl(Name=text).Exists(0,0)
                AS_Utillity_Methods.asequal(self,items,True,msg +" verified item " + text + " displayed")
                time.sleep(2)
        
    def verify_notepad_content(self,title,msg):
        
        '''==========================================================================================================================================
        Description : To verify text elements in notepad.
        Usage : as_util_obj.Verify_Browser_Content('WebFOCUS Report - Internet Explorer',"Step 01: Element avaliable",item_list=['CAR','ALFA ROMEO','AUDI'])
        @author: Adithyaa AK
        ==========================================================='''        
                
        notepad_window=automation.WindowControl(Name=title)   
        notepad_window.SetFocus()
        text_existing=notepad_window.Exists()
        time.sleep(2)
        AS_Utillity_Methods.asequal(self,text_existing,True,msg)
        notepad_window.Close()
        
        if text_existing==False:
            print(msg + "- FAILED.")
            AS_Utillity_Methods.asert_failure_count += 1
        
    def Verify_Tooltip(self,tab_name,component,msg,**kwargs):
        
        '''@author: Adithyaa AK : Description : To verify document canvas panel tooltips 
           ================================================================================
           Usage : as_panels_obj.Verifypanel_Tooltip('5','5','Auto Hide',"Verified - Tooltip is Auto Hide",move_x=-43,move_y=-49)'''
        
        tootltip_locator='tooltips_class32'
        get_tooltip_name="Name"
        
        self.switch_as_tab(tab_name) 
        time.sleep(1)
        as1 = self.driver.find_element_by_name(component)
        action = ActionChains(self.driver)
        action.move_to_element_with_offset(as1,10,5).perform()
        time.sleep(2)
        del action
            
        if 'move_x' in kwargs:
            action = ActionChains(self.driver)
            action.move_by_offset(kwargs['move_x'], kwargs['move_y']).perform()
            time.sleep(1)
            
        tooltip=self.driver.find_element_by_class_name(tootltip_locator).get_attribute(get_tooltip_name)
        AS_Utillity_Methods.asequal(self,tooltip,component,msg)
        del action  
        
    def Verify_Element(self,element,msg,**kwargs):
        
        '''=======================================================
        @author: Adithyaa AK 
        Description : To verify element
        Usage : Verify_Element(self,"Procedure","True",msg):
        ==========================================================='''
        
        if "available" in kwargs: 
            
            try:
                
                booln=self.driver.find_element_by_name(element).is_displayed()
                AS_Utillity_Methods.asequal(self,booln,True,msg)
                
            except NoSuchElementException:
                print(msg + " - FAILED")
                AS_Utillity_Methods.asert_failure_count += 1
        
        if "unavailable" in kwargs:    
            
            try:
                booln=self.driver.find_element_by_name(element).is_displayed()
                if booln==True:
                    print(msg + "- FAILED")
                    AS_Utillity_Methods.asert_failure_count += 1
                    
            except NoSuchElementException:
                print(msg + " - PASSED")
                
    def Verify_Current_Dialog_Opens(self,dialog,msg):
        
        '''=======================================================
        @author: Adithyaa AK
        Description : To verify current window title active in current window
        Usage : as_utilobj.Verify_Current_Dialog_Opens("#32770","Step 09 - SQL Report been displayed")
        ==========================================================='''
        
        try:
            automation.WindowControl(ClassName="#32770").Click(ratioX=0.10, ratioY=0.10,waitTime=2)
            get_dialog_name=self.driver.find_element_by_class_name('#32770').get_attribute("Name")
            AS_Utillity_Methods.asequal(self,get_dialog_name,dialog,msg)
            
        except:
            print(dialog + "not displayed")
                
    def Verify_Current_Dialog_Closes(self,dialog,msg):
        
        '''=======================================================
        @author: Adithyaa AK
        Description : To verify current window title active in current window
        Usage : as_utilobj.Verify_Current_Dialog_Closes("Select Data Source","Step 09 - SQL Report been displayed")
        ==========================================================='''
        
        try:
            booln=self.driver.find_element_by_name(dialog).is_displayed()
            if booln==True:
                print(msg + "- FAILED.")
                AS_Utillity_Methods.asert_failure_count += 1
                
        except NoSuchElementException:
            print(msg + " - PASSED")
                          
    def Verify_Browser_Content(self,title,msg,**kwargs):
        
        '''==========================================================================================================================================
        @author: Adithyaa AK
        Description : To verify text elements in browser
        Usage : as_utilobj.Verify_Browser_Content('IEFrame',"Step 01: Element avaliable",item_list=['CAR','ALFA ROMEO','AUDI'],**kwargs)
        ==========================================================='''        
        
        try:       
            automation.WindowControl(ClassName=title)   
            browser_window=automation.WindowControl(ClassName=title)
            browser_window.SetFocus()
            browser_window.Maximize()
            time.sleep(2)
            
            """To verify browser title"""
        
            if 'verify_browser' in kwargs:
                
                verify_browser=browser_window.Exists()
                AS_Utillity_Methods.asequal(self,verify_browser,True,msg)
                
            if 'verify_pane' in kwargs:
                
                automation.PaneControl(Name=kwargs['verify_pane']).Exists()
                verify_pane=automation.PaneControl(Name=kwargs['verify_pane']).Exists()
                AS_Utillity_Methods.asequal(self,verify_pane,True,msg)
                    
            """To verify any_list inside browser"""
            
            if 'item_list' in kwargs:
                
                for text in kwargs['item_list']:
                    items=automation.TextControl(Name=text).Exists(0,0)
                    AS_Utillity_Methods.asequal(self,items,True,msg + text)
                    
                if items==False:
                    print(msg + "- FAILED.")
                    AS_Utillity_Methods.asert_failure_count += 1
                
            """To verify specific item inside browser"""  
             
            if 'item' in kwargs:
                
                item=browser_window.GroupControl(Name=kwargs['item']).Exists(0,0)
                AS_Utillity_Methods.asequal(self,item,True,msg)
                time.sleep(2)
                
            if 'button_list' in kwargs:
                
                for button in kwargs['button_list']:
                    button_item=automation.ButtonControl(Name=button).Exists(0,0)
                    AS_Utillity_Methods.asequal(self,button_item,True,msg + button)
                                  
            if 'browser_close' in kwargs:
                browser_window.Close()  
                
        except:
            print(msg + "- FAILED.")
            AS_Utillity_Methods.asert_failure_count += 1
            
    def verify_active_tool(self,verify_tool,msg):
        
        '''=======================================================
        Description : To verify current window title active in current window
        Usage : as_utilobj.verify_current_active_tool("Step 3.1: Alert Assist Web tool opens")
        ==========================================================='''
        
        time.sleep(2)
        get_tool=automation.WindowControl(RegexName="App Studi.*").GetWindowText()
        AS_Utillity_Methods.asequal(self,get_tool,verify_tool,"Tool: " + get_tool + " : " +  msg)
        
    def reset_layout(self,app_studio_path): 
        
        '''=======================================================
        Description : To call reset_layout option in Appstudio
        Usage : as_utilobj.reset_layout("C:/ibi/AppStudio82/bin/FocShell.exe")
        ==========================================================='''
        
        automation.SendKeys('{Alt}(E)')
        keys.press(['down','down','down','enter'])
        time.sleep(2)          
        self.click_element_using_ui(radio_button="1688")
        self.select_any_dialog("OK")  
        self.select_application_menu_options(send_keys=['up','enter'])
        time.sleep(10)
        self.driver = webdriver.Remote(command_executor='http://localhost:9999',desired_capabilities={'debugConnectToRunningApp': 'false','app': app_studio_path}) 
        time.sleep(20)
        self.click_element_using_ui(button_item=True,name="Close")
        time.sleep(1)
        
    def close_canvas_item(self):
        
        '''=======================================================
        Description : To close any file in as canvas
        Usage : as_utilobj.close_canvas_item() or if prompt displays as yes or no use "as_utilobj.close_canvas_item(button="Yes")"
        ==========================================================='''
        
        try:
            env_tree=automation.PaneControl(AutomationId="319").Exists()
            if env_tree !=False:
                automation.TreeControl(AutomationId="2").Click(ratioX=10,ratioY=140)
                automation.TreeItemControl(Name="Domains").ScrollIntoView()
                automation.TreeItemControl(Name="Domains").RightClick(waitTime=2)
                automation.MenuItemControl(Name="Refresh Descendants").Click(waitTime=4)
        except:
            pass  
        
        try:
            env_details=automation.PaneControl(AutomationId="318").Exists()
            if env_details !=False:
                automation.TreeControl(AutomationId="59648").Click(ratioX=10,ratioY=140)
                automation.TreeItemControl(Name="Domains").ScrollIntoView()
                automation.TreeItemControl(Name="Domains").RightClick(waitTime=2)
                automation.MenuItemControl(Name="Refresh Descendants").Click(waitTime=4)
        except: 
            pass  
            
        automation.SendKey(automation.Keys.VK_MENU, waitTime=3)
        time.sleep(2)
        automation.SendKey(automation.Keys.VK_F, waitTime=3)
        time.sleep(2)
        automation.SendKey(automation.Keys.VK_C, waitTime=3)
        time.sleep(2)
              
    def run_item_from_tree_pane_context_menu(self, tree_path, item, **kwargs):
        
        """
        Usage: select_item_from_tree_view("driver.find_element_by_name("Static"), "jawin7->Domains->P20")
        params:parent_obj="driver.find_element_by_name("Static" );"jawin7->Domains->P20"
        """ 
        wait_time=2
        env=utillity.UtillityMethods.parseinitfile(self,"envpath")
        tree_path=tree_path.split("->")
        automation.TabItemControl(Name="Home").Click()
        automation.PaneControl(Name="Environments Tree").Click()
        automation.TreeItemControl(Name="Configured Environments").DoubleClick()
        time.sleep(wait_time)
        automation.TreeItemControl(Name=env).DoubleClick()
        time.sleep(wait_time)
        
        for x in tree_path:
            try:
                automation.TreeItemControl(Name=x).DoubleClick()
                time.sleep(wait_time)
            except LookupError:
                print(x+"  Item not found")
        automation.TreeItemControl(Name=item).RightClick()
        time.sleep(wait_time)
        automation.MenuItemControl(Name="Run").Click()
    def select_item_from_tree_pane_context_menu(self, tree_path, item, menu_item, **kwargs):
        
        """
        Usage: select_item_from_tree_view("driver.find_element_by_name("Static"), "jawin7->Domains->P20")
        params:parent_obj="driver.find_element_by_name("Static" );"jawin7->Domains->P20"
        """ 
        host= AS_Utillity_Methods.parseinitfile(self,'envpath')
        path1="Configured Environments->"+host
        menu_list=path1.split("->")
        for i in menu_list:
            if automation.ExpandCollapseState.Collapsed==automation.TreeItemControl(Name=i).CurrentExpandCollapseState():
                automation.TreeItemControl(Name=i).ScrollIntoView()
                time.sleep(3)
                automation.TreeItemControl(Name=i).Expand(waitTime=2)
                time.sleep(3)
                automation.TreeItemControl(Name=i).Click()
            if (automation.TreeItemControl(Name=i).CurrentExpandCollapseState()==3):
                time.sleep(3)
                automation.TreeItemControl(Name=i).DoubleClick()
                time.sleep(3)
        time.sleep(5)
        menu_list=tree_path.split("->")
        for i in menu_list:
            automation.TreeItemControl(Name=i).ScrollIntoView()
            time.sleep(3)
           
            if (automation.ExpandCollapseState.Collapsed==automation.TreeItemControl(Name=i).CurrentExpandCollapseState()):
                print (i)
                
                time.sleep(3)
                automation.TreeItemControl(Name=i).Expand(waitTime=2)
                time.sleep(3)
                automation.TreeItemControl(Name=i).Click()
            if (automation.TreeItemControl(Name=i).CurrentExpandCollapseState()==3):
                print (i)
                
                time.sleep(3)
                automation.TreeItemControl(Name=i).DoubleClick()
                time.sleep(3)
        try:
            automation.TreeItemControl(Name=item).ScrollIntoView()
            time.sleep(3)
            automation.TreeItemControl(Name=item).RightClick()
            menu_item_list=menu_item.split("->")
            for i in menu_item_list:
                menu_ctrl=automation.MenuItemControl(Name=i)
                automation.WaitForExist(menu_ctrl,10)
                time.sleep(5)
                automation.MenuItemControl(Name=i).DoubleClick()
                
        except:
            print(item+"  Item not found")
    def click_any_canvas_button(self,btn_name):
        '''
        Desc: This is used to click on any button whose ID start with 'button'.
        :param: dialog: "div[id^='BiDialog']"
        :param: btn_name: "OK"
        :Usage: utillobj.click_any_canvas_button("OK")
        '''        
        dialog_css = "input[id^='button']"
        dialog_btns=self.driver.find_elements_by_css_selector(dialog_css)
        btn_text_list=[el.text.strip() for el in dialog_btns]
        dialog_btns[btn_text_list.index(btn_name)].click()
        time.sleep(5)
        
    def select_or_verify_html_drop_down_option(self,dropdown_css,option,**kwargs):
        '''
        Desc: This function is to select or verify pure html drop down options.
        :param : dropdown_css="#IBILAYOUTDIV0TABS select[class='arDashboardMergeDropdown']"
        :param : option=United State
        :Usage : select_or_verify_html_drop_down_option("#IBILAYOUTDIV0TABS select[class='arDashboardMergeDropdown']",'North America and South America',verify_default_option='EMEA',verify_options=expected_options,msg='Step 18.1 : Verify BUSINESS_REGION_1 options')
        '''
        drop_down_items=Select(self.driver.find_element_by_css_selector(dropdown_css))
        if 'verify_default_option' in kwargs :
            default_item=drop_down_items.first_selected_option.text.strip()
            print(default_item)
            self.asequal(default_item,kwargs['verify_default_option'],'Step X : Verify '+kwargs['verify_default_option']+' is selected as default')
        if 'verify_options' in kwargs :
            actual_options=[value.text.strip() for value in drop_down_items.options]
            print(actual_options)
            self.asequal(actual_options,kwargs['verify_options'],kwargs['msg'])
        if option!=None :
            drop_down_items.select_by_visible_text(option)
    def login_wf_page(self, se_driver,website):
        loginid = AS_Utillity_Methods.parseinitfile(self,'mrid')
        loginpwd = AS_Utillity_Methods.parseinitfile(self,'mrpass')
  
        website =website+"signin" 
        se_driver.get(website)
        se_driver.set_window_size(400,400)
        time.sleep(2)
        browser = AS_Utillity_Methods.parseinitfile(self,'browser')
        if browser.lower()=="chrome":
            chrome_loc=automation.PaneControl(Name="Sign in to WebFOCUS - Google Chrome").ButtonControl(Name="Maximize")
            if (chrome_loc.Exists()):
                automation.SendKeys('{Alt}{Space}')
                automation.SendKeys('x')
            else:
                automation.PaneControl(Name="Sign in to WebFOCUS - Google Chrome").PaneControl(Name="Google Chrome").PaneControl(Name="").ButtonControl(Name="Maximize").SetFocus()
                automation.PaneControl(Name="Sign in to WebFOCUS - Google Chrome").PaneControl(Name="Google Chrome").PaneControl(Name="").ButtonControl(Name="Maximize").Click()
        elif browser.lower()=="ie":
            automation.WindowControl(Name="Sign in to WebFOCUS - Internet Explorer").Maximize(waitTime=5)
        elif browser.lower()=="firefox":
            automation.WindowControl(Name="Sign in to WebFOCUS - Mozilla Firefox").Maximize(waitTime=5)
        else:
            automation.WindowControl(RegexName="Sign in to WebFOCUS.*").Maximize(waitTime=5)
        
        time.sleep(8)
        corobj=coreutillobject(self.se_driver)
        corobj.update_current_working_area_browser_specification()
#         se_driver.find_element_by_id("SignonUserName").click()  
#         keyboard.write(loginid,delay=0.5)
#         time.sleep(2)
#         se_driver.find_element_by_id("SignonPassName").click()
#         keyboard.write(loginpwd,delay=0.5)
        se_driver.find_element_by_id("SignonUserName").click()
        time.sleep(1) 
        se_driver.find_element_by_id("SignonUserName").send_keys(loginid)
        time.sleep(2)
        se_driver.find_element_by_id("SignonPassName").click()
        time.sleep(1)
        se_driver.find_element_by_id("SignonPassName").send_keys(loginpwd)
        time.sleep(2)
        se_driver.find_element_by_id("SignonbtnLogin").click()
        AS_Utillity_Methods.windows=self.se_driver.window_handles
        corobj.windows_handles=self.se_driver.window_handles
        time.sleep(2)
    
    def run_item_in_browser(self, folderpath,file_name):
        
        browser = AS_Utillity_Methods.parseinitfile(self,'browser')
        self.se_driver = WebDriverFactory.getInstance(browser)
        self.se_driver.maximize_window()
                
        node = AS_Utillity_Methods.parseinitfile(self,'nodeid')
        port = AS_Utillity_Methods.parseinitfile(self,'httpport')
        context = AS_Utillity_Methods.parseinitfile(self,'wfcontext')
        
        setup_url = 'http://' + node + ':' + port + context + '/'
        
        folder = folderpath
        api_url = setup_url + 'run.bip?BIP_REQUEST_TYPE=BIP_RUN&BIP_folder=IBFS%253A%252FWFC%252FRepository%252F' + "/" + folder + '&BIP_item=' + file_name        
        
        AS_Utillity_Methods.login_wf_page(self,self.se_driver,setup_url)
        time.sleep(10)
        self.se_driver.get(api_url)
        self.se_driver.maximize_window()
        global url
        global session_id
        url = self.se_driver.command_executor._url
        session_id = self.se_driver.session_id
        
        time.sleep(10)
        return(self.se_driver)
    
    def wait_for_UI_object(self, ui_object, wait_time):
        '''
        Desc: This function is used to wait for a UI element
        :param : object="ButtonControl(AutomationID="button5")"
        :param : wait_time=30
        :Usage : wait_for_UI_object("ButtonControl(AutomationID="button5")", 30)
        '''
#         ui_object="automation." +ui_object
#         print (ui_object)
#                 
        automation.WaitForExist(ui_object, wait_time)
        
        
    def wait_for_UI_object_close(self, object,wait_time):
        '''
        Desc: This function is used to wait for a UI element
        :param : object="ButtonControl(AutomationID="button5")"
        :param : wait_time=30
        :Usage : wait_for_UI_object("ButtonControl(AutomationID="button5")", 30)
        '''
        object="automation."+object
        
        automation.WaitForDisappear(object, wait_time)
        
    def verify_object_visible(self, ui_object, visible, msg, **kwargs):
        """
        :params css="#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']"
        :params visible=True Or False
        :params msg='Step 5: Filter Button Visible'
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", True, 'Step 5: Filter Button Visible')
                OR
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", False, 'Step 5: Filter Button Removed')
        """
        try:
            
            status=ui_object.Exists(2,2)

        except:
            
            status = False
        utillity.UtillityMethods.asequal(self, status, visible, msg)    

    def verify_child_count(self, ui_object, count, msg, **kwargs):
        """
        :params css="#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']"
        :params visible=True Or False
        :params msg='Step 5: Filter Button Visible'
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", True, 'Step 5: Filter Button Visible')
                OR
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", False, 'Step 5: Filter Button Removed')
        """
        try:
             
            actual_count=len(ui_object.GetChildren())
        except:
            actual_count = 0
        print (actual_count)
        utillity.UtillityMethods.asequal(self, actual_count, count, msg)    
    
    def synchronize_with_number_of_element(self, element_css, expected_number, expire_time, pause_time=1):

        '''
        This function synchronize with expected number of element in current browser.
        :param parent_css:-element css need to pass
        :param expected_number:-0, 1, 2,........
        :param expire_time:-1, 5,........
        :param pause_time=0.2,1...
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                print(str(element_css) + " Parent Css not having " + str(expected_number) + " expected number of element.")
                run_=False
                break
            try:
                temp_obj = self.se_driver.find_elements_by_css_selector(element_css)
                check_obj_exist = temp_obj[0]
                del check_obj_exist
            except IndexError:
                time.sleep(pause_time)
                continue
            if len(temp_obj) == int(expected_number):
                run_=False
                break
            else:
                time.sleep(pause_time)
        time.sleep(pause_time)
        
        
    def select_dropdown(self, select_css, select_type, value):    
        '''
        select_css: #combobox_dsPROMPT_1
        type: index or value or visible_text 
        value: 'Jan, 1987
        Usage: select_dropdown('#combobox_dsPROMPT_1', 'visible_text', 'Jan, 1987')
        '''
        select=Select(self.se_driver.find_element_by_css_selector(select_css))
        if select_type=='index':
            select.select_by_index(value)
        elif select_type=='value':
            select.select_by_value(value)
        elif select_type=='visible_text':
            select.select_by_visible_text(value)     
    def close_current_window(self,tool_title):
        
        automation.WindowControl(RegexName=tool_title).Close()
    
    def close_web_window(self):
        
#         se_driver = webdriver.Remote(command_executor=url,desired_capabilities={})
#         se_driver.session_id = session_id
        try:
            self.se_driver.close()
        except:
            time.sleep(10)
            pyautogui.hotkey('alt','f4')
            
    
    def verify_web_object_visible(self, css, visible, msg, **kwargs):
        """
        :params css="#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']"
        :params visible=True Or False
        :params msg='Step 5: Filter Button Visible'
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", True, 'Step 5: Filter Button Visible')
                OR
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", False, 'Step 5: Filter Button Removed')
        """
        
        try:
            if 'elem_obj' in kwargs:
                status = kwargs['elem_obj'].is_displayed()
            else:
                status = self.se_driver.find_element_by_css_selector(css).is_displayed()
                print (status)
        except:
            status = False
        utillity.UtillityMethods.asequal(self, status, visible, msg)      
    
    def verify_web_object_count(self, css, count, msg, **kwargs):
        """
        :params css="#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']"
        :params visible=True Or False
        :params msg='Step 5: Filter Button Visible'
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", True, 'Step 5: Filter Button Visible')
                OR
        :Usage verify_object_visible(self, "#MAINTABLE_0 .arChartMenuBar div[title][onclick*='Filter']", False, 'Step 5: Filter Button Removed')
        """
        count=self.se_driver.find_elements_by_css_selector(css)
        try:
            acount=self.se_driver.find_elements_by_css_selector(css)
        except:
            acount = 0
        utillity.UtillityMethods.asequal(self, count, acount, msg)  
    
    
    def maximize_setfocus_ui_window(self,regex_name):    
        title=automation.WindowControl(RegexName=regex_name).GetWindowText()
        automation.WindowControl(Name=title).Maximize(5)
        automation.WindowControl(Name=title).SetFocus()

                  
    def beautifulsoup_object_creation(self):
        from bs4 import BeautifulSoup
        page_source_text=self.se_driver.page_source
        soup = BeautifulSoup(page_source_text, 'html.parser')
        return (soup)
     
    def create_table_data(self, table_css, file_name):
        '''
        This function is used to create run time data set like table structure.
        :param table_css="table[summary='Summary'] > tbody > tr"
        :param: file_name="test1.xlsx"
        :Usage: create_table_data("table[summary='Summary']", "test1.xlsx")
        :@author AAkhan
        '''
#         self.se_driver = WebDriverFactory.getInstance(browser)
        le=self.se_driver.find_elements_by_css_selector(table_css)
        len(le)
#         self.se_driver.session_id = session_id
#         print ('se driver session id attachment',self.se_driver.session_id)
#         print ("inside create table",self.se_driver)
        from bs4 import BeautifulSoup
        page_source_text=self.se_driver.page_source
        soup = BeautifulSoup(page_source_text, 'html.parser')
        #print ('soup', soup)
        #soup = AS_Utillity_Methods.beautifulsoup_object_creation(self)
        wb = Workbook()
        #s = wb.get_sheet_by_name('Sheet')
        s = wb['Sheet']
        table_rows = soup.select(table_css)
        for r in range(0, len(table_rows)):
            total_col = table_rows[r].find_all('td')
            for c in range(len(total_col)):
                s.cell(row=r + 1, column=c + 1).value = total_col[c].get_text(strip=True)
        wb.save(os.getcwd() + "\data\\" + file_name)
         
    def verify_table_data(self, table_css, file_name, msg):
        '''
        This function is used to verify run time data set like table structure.
        :param table_css="table[summary='Summary'] > tbody > tr"
        :param: file_name="test1.xlsx"
        :Usage: verify_table_data("table[summary='Summary']", "test1.xlsx")
        :@author AAkhan
        '''
        #soup = AS_Utillity_Methods.beautifulsoup_object_creation(self)
        from bs4 import BeautifulSoup
        page_source_text=self.se_driver.page_source
        soup = BeautifulSoup(page_source_text, 'html.parser')
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        status=[]
        s1 = wb1['Sheet']
        table_rows = soup.select(table_css)
        table_row_lenght = len(table_rows)
        for r in range(0, table_row_lenght):
            total_col = table_rows[r].find_all('td')
            for c in range(len(total_col)):
                browser_value=str(total_col[c].get_text(strip=True).replace(' ',''))
                if s1.cell(row=r + 1, column=c + 1).value != None and s1.cell(row=r + 1, column=c + 1).value != ' ':
                    excel_value=s1.cell(row=r + 1, column=c + 1).value.replace(' ','')
                     
                else:
                    excel_value=None
                     
                if (excel_value != None and browser_value != '') or (excel_value != None) or (browser_value != ''):
                    if excel_value == browser_value:
                         
                        status=[0]
                        continue
                    else:
                        status=[r+1,c+1]
                        return (status)
        utillity.UtillityMethods.asequal(self,status,[0], msg)
         
    def switch_to_frame(self, frame_css):
        """
        Switches focus to the specified frame
        :param pause=0,1,2...        (pass Interger/Float number)
        :kwargs frame_css="[id^='ReportIframe']"   (pass frame css)
        :kwargs frame_height_value=0        (Interger number)
        :Usage switch_to_frame(pause=2, frame_css='frame[src]')
            OR
        :Usage switch_to_frame(pause=2, frame_css='frame[src]', frame_height_value=0)
        """
        path_css=frame_css
        frame_element_obj = self.se_driver.find_element_by_css_selector(path_css)
        
        frame_actual_location = utillity.UtillityMethods.get_object_screen_coordinate(self, frame_element_obj, coordinate_type='start')
        WebDriverWait(self.se_driver, 100).until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, path_css)))
        time.sleep(5)
        AS_Utillity_Methods.browser_x=frame_actual_location['x']
        AS_Utillity_Methods.browser_y=frame_actual_location['y']
        time.sleep(5)
        
        
    def switch_to_window_handle_test(self, wndnum, **kwargs):
        old_windows=AS_Utillity_Methods.windows
        if 'custom_windows' in kwargs:
            old_windows=kwargs['custom_windows']
        new_windows=self.se_driver.window_handles
        new_set=set(new_windows)
        old_set=set(old_windows)
        diff_set=new_set-old_set
        last_window=list(diff_set)
        time.sleep(2)
        if wndnum>0:
            self.se_driver.switch_to.window(last_window[0])
        else:
            self.se_driver.switch_to.window(old_windows[-1])
        time.sleep(4)
        
    def switch_to_window(self,wndnum, pause=15, **kwargs):
        """
        Usage: switch_to_window(1)
        """
        
        status=0
        num = kwargs['win_num'] if 'win_num' in kwargs else 0
                 
        AS_Utillity_Methods.switch_to_window_handle_test(self, wndnum, **kwargs)
        time.sleep(pause)
        self.se_driver.maximize_window()
        window_widht = self.se_driver.execute_script("return window.innerWidth|| document.documentElement.clientWidth|| document.body.clientWidth || document.body.scrollWidth;")
#         if window_widht < self.se_driver.execute_script("return screen.availWidth;"):
#             self.se_driver.maximize_window()
        time.sleep(pause/2)
    
    def switch_to_main_window(self, **kwargs):
        """
        Usage : switch_to_main_window()
        Close all open window and switch to main window, this is useful when you want to return back to main window and close other opened window
        @author: AAKhan
        """
        time.sleep(9)
        total_window=len(self.se_driver.window_handles)
        i = total_window - 1
        time.sleep(1)
        if total_window > 1:
            timeout=0
            while i:
                self.se_driver.switch_to.window(self.se_driver.window_handles[i])
                time.sleep(5)
                self.se_driver.close()
                time.sleep(7)
                i-=1
                if i == 0:
                    break
                if timeout == 250:
                    print('No window found')
                    break
                timeout+=1
        self.se_driver.switch_to.window(self.se_driver.window_handles[i])
        
    def switch_to_default_content(self):
        """
        Switches focus to the default content
        :param pause=0,1,2...        (pass Interger/Float number)
        :Usage switch_to_default_content(pause=1)
        """
        self.se_driver.switch_to.default_content()
        time.sleep(4)
        
    def select_UI_tree_view_item(self,tree_item,list_item,button_name):
        object=automation.WindowControl(Name="Select Data Source")
        AS_Utillity_Methods.wait_for_UI_object(self, object, wait_time=30)
        
        automation.TreeItemControl(Name="Data Servers").ScrollIntoView()
        time.sleep(4)
        automation.TreeItemControl(Name="Data Servers").Expand()
        time.sleep(4)
        if automation.TreeItemControl(Name="EDASERVE").Exists():
            automation.TreeItemControl(Name="EDASERVE").ScrollIntoView()
            time.sleep(4)
            automation.TreeItemControl(Name="EDASERVE").Expand()
        time.sleep(4)
        if automation.TreeItemControl(Name="Applications").Exists():
            automation.TreeItemControl(Name="Applications").ScrollIntoView()
            time.sleep(4)
            automation.TreeItemControl(Name="Applications").Expand()
            
            
        automation.TreeItemControl(Name=tree_item).ScrollIntoView()
        time.sleep(4)
        automation.TreeItemControl(Name=tree_item).Click()
        time.sleep(6)
        automation.ListControl(AutomationId="59648").Click()
        for i in range(0,6):
            stat=automation.TextControl(Name=list_item).IsOffScreen
            time.sleep(2)
            if stat==True:
                automation.SendKeys('{PageDown}')
                time.sleep(2)
            else:
                automation.TextControl(Name=list_item).Click(40,10)
                break
        #automation.TextControl(Name=list_item).IsOffScreen
        time.sleep(4)
        automation.ButtonControl(Name=button_name).Click()
        
    def verify_chart_color(self, parent_id, riser_class, color, msg, **kwargs):
        """
        parent_id='MAINTABLE_0'
        riser_class='riser!s1!g8!mbar!'
        color='green' OR 'text-green'(text-green means it will verify only green not the rgb of green)
        color='green'
        kwargs = attribute_type='fill' or 'stroke' (pass attribute type to get css value of color)
        kwargs = attribute='yes'(if want to verify attribute value for color instead pf css property)
        Syntax:verify_chart_color('MAINTABLE_0', 'riser!s1!g8!mbar!', 'green', 'Step 10: Verify Color')
        Syntax:verify_chart_color('MAINTABLE_0', 'riser!s1!g8!mbar!', 'green', 'Step 10: Verify Color',attribute='yes')
        Syntax:verify_chart_color('MAINTABLE_0', 'riser!s1!g8!mbar!', 'green', 'Step 10: Verify Color',attribute_type='stroke')
        Syntax:verify_chart_color('MAINTABLE_0', 'riser!s1!g8!mbar!', 'green', 'Step 10: Verify Color',attribute_type='stroke',attribute='yes')
        @author : Niranjan
        """
        attribute_type=kwargs['attribute_type'] if 'attribute_type' in kwargs else "fill"
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else "[class*='" + riser_class + "']"
        raiser_css="#"+ parent_id + " " + custom_css
        if 'attribute' in kwargs:
            if self.browser == 'IE':
                temp_obj=((self.se_driver.find_element_by_css_selector(raiser_css).get_attribute(attribute_type))[:-9]+")")[4:]
            else:
                temp_obj=((self.se_driver.find_element_by_css_selector(raiser_css).get_attribute(attribute_type))[:-10]+")")[4:]
            actual_color = "rgb"+temp_obj
            expected_color=utillity.UtillityMethods.color_picker(self, color, 'rgb')
        else:
            actual_color = Color.from_string(self.se_driver.find_element_by_css_selector(raiser_css).value_of_css_property(attribute_type)).rgba
            expected_color=utillity.UtillityMethods.color_picker(self, color, 'rgba')
        utillity.UtillityMethods.asequal(self, actual_color, expected_color, msg)
    
    def verify_number_of_risers(self, parent_css_with_tagname, risers_per_segment, expected_number, msg):
        '''
        Desc: This function is used to verify the number of risers. As this covers wide range of risers with different tags, user must 
        pass the parent css upto tag. Also we need to provide the number of risers per segment. It means if there is a simple bar chart,
        then number of risers per segment is '1', for stacked bar it will differ. 
        '''
        riser_css=parent_css_with_tagname + " [class^='riser']"
        total_risers=len(self.se_driver.find_elements_by_css_selector(riser_css))
        actual_number=int(total_risers/risers_per_segment)
        utillity.UtillityMethods.asequal(self, expected_number, actual_number, msg)
    
    def verify_data_labels(self, parent_id, expected_datalabel, msg, **kwargs):
        """
        :param :parent_id = 'MAINTABLE_wbody1'
        :param :kwargs['custom_css'] = ".chartPanel text[class^='mdataLabels']"       (Other css value user pass)
        :param :expected_datalabel: ['1,546', '1,446', '1,646']
        :param :msg: 'Step 10: Verify number of pie chart segments displayed'
        :Usage :verify_data_labels('TableChart_1', ['1,546', '1,446', '1,646'], 'Step 4: Verify data labels', custom_css=".chartPanel .groupPanel text[class^='mdataLabels']")
                OR
        :Usage :verify_data_labels('MAINTABLE_0', ['1,546', '1,446', '1,646'], 'Step 4: Verify data labels')
        @author: AAkhan
        """
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else ".chartPanel text[class^='mdataLabels']"
        parent="#" + parent_id + " " + custom_css
        actual_datalabel=[]
        total_datalabel=self.se_driver.find_elements_by_css_selector(parent)
        if 'data_label_length' in kwargs:
            my_iter_x1=[i.text for i in total_datalabel]
            print (my_iter_x1)
            my_iter_x=(i.text for i in total_datalabel)
            for label_x in expected_datalabel:
                if label_x[:kwargs['data_label_length']] == next(my_iter_x)[:kwargs['data_label_length']]:
                    statex= True
                else:
                    statex=False
                    break
            del my_iter_x
            utillity.UtillityMethods.asequal(self, statex, True, msg)
        else:
            for i in range(len(total_datalabel)):
                actual_datalabel.append(total_datalabel[i].text.strip())
            utillity.UtillityMethods.as_List_equal(self, expected_datalabel, actual_datalabel, msg)
    
    def synchronize_with_visble_text(self, element_css, visble_element_text, expire_time, pause_time=1, text_option='dom_visible_text'):
        '''
        This function synchronize with expected visible text of element in current browser.
        :param parent_css:-element css need to pass
        :param visble_element_text:-visible text need to pass
        :param expire_time:-1, 5,........
        :param pause_time=0.2,1...
        :usage utillityobject.synchronize_with_visble_text(self, "[id*='RangeValuesBox'] [id*='From'] > input", '2014', 190, text_option='text_value')
        @author: Aftab_Alam_khan
        '''
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == int(expire_time)+1:
                print(str(element_css) + " Parent Css not having " + str(visble_element_text) + " visible text.")
                run_=False
                break
            try:
                temp_str_value_elem=self.se_driver.find_element_by_css_selector(element_css)
                temp_str_value=utillity.UtillityMethods.get_attribute_value(self, temp_str_value_elem, text_option)
            except NoSuchElementException:
                time.sleep(pause_time)
                continue
            except StaleElementReferenceException:
                time.sleep(pause_time)
                continue
            str_value = re.sub(' ','',temp_str_value[text_option]).replace('\n','')
            if str_value == str(visble_element_text.replace(' ','')):
                run_=False
                break
        time.sleep(pause_time)
        
           
    def verify_xy_axis_title(self,popup_id,expected_xaxis_title,expected_yaxis_title,msg,**kwargs):#Need to be deleted.
        """
        Params: table_id='MAINTABLE_wbody0_f'
        Params: expected_title='RCOUNTRY : CAR'
        Syntax: verify_xaxis_title('MAINTABLE_wbody0_f', 'COUNTRY : CAR', 'Step 03: Verify Title')
        @Author: Kiruthika 
        """
        x_css="text[class^='xaxis'][class$='title']"
        y_css="text[class^='yaxis'][class$='title']"
        x_sync_css="#"+popup_id+" " + x_css
        y_sync_css="#"+popup_id+" " + y_css
        AS_Utillity_Methods.synchronize_with_number_of_element(self, x_sync_css, 1, 90, 1)
        time.sleep(5)
        
        xtitle=self.se_driver.find_element_by_css_selector(x_css).text
        ytitle=self.se_driver.find_element_by_css_selector(y_css).text
        utillity.UtillityMethods.asequal(self,xtitle,expected_xaxis_title,msg+"for X-Axis")
        utillity.UtillityMethods.asequal(self,ytitle,expected_yaxis_title,msg+"for Y-Axis")
    
    def close_item_from_canvas(self):
        
        '''=======================================================
        Description : To close any file in as canvas
        Usage : as_utilobj.close_canvas_item() or if prompt displays as yes or no use "as_utilobj.close_canvas_item(button="Yes")"
        ==========================================================='''
        
#         AS_Utillity_Methods.selfect_component_by_right_click(self,right_click_item="Domains",click="Refresh Descendants")
#         time.sleep(4)
        application_menu="Application Menu"
        automation.ButtonControl(Name=application_menu).Click()

        automation.SendKey(automation.Keys.VK_MENU, waitTime=3)
        time.sleep(5)
        automation.SendKey(automation.Keys.VK_F, waitTime=3)
        time.sleep(5)
        automation.SendKey(automation.Keys.VK_C, waitTime=3)
        time.sleep(5)
        
    def save_as_UI_dialog(self, file_name):
        """      
        Description:- Save content from the Save AS dialog
        :param file_name= the file name to be typed
        
        :Usage: save_as_UI_dialog('onetwo.htm') 
        """
        win_control=automation.WindowControl(Name="Save As")
        AS_Utillity_Methods.wait_for_UI_object(self, win_control, 30)
        automation.EditControl(AutomationId="1516").Click()
        time.sleep(2)
        pyautogui.hotkey('ctrl','a')
        time.sleep(2)
        pyautogui.hotkey('del')
        time.sleep(2)
        automation.EditControl(AutomationId="1516").SetValue(file_name, 2)
        
        automation.ButtonControl(Name="OK").Click()
        time.sleep(6)
        stat=automation.ButtonControl(Name="Yes").Exists()
        if stat==True:
            automation.ButtonControl(Name="Yes").Click()
        
    def verify_web_object_text(self,web_object_css,text_verify,msg,**kwargs):#Need to be deleted.
        """      
        Description:- Verify any given web object's text by passing the css value
        :param web_object_css= css of the web object
        :param text_verify= the text to verify
        :param msg= msg to be displayed on the verification point result.
        :Usage: verify_web_object_text('#button1','My button','Step x. Verify the button') 
        """
        x_css=web_object_css
        
        AS_Utillity_Methods.synchronize_with_number_of_element(self, x_css, 1, 90, 1)
        time.sleep(5)
        
        itext=self.se_driver.find_element_by_css_selector(x_css).text.replace(' ','')
        utillity.UtillityMethods.asequal(self,itext,text_verify,msg)
        
    
    def select_new_request_datasource_menu(self, menu_path):
        """      
        Description:- From Tasks and Animations panel, select run request and the path.
        :param menu_path= 'Run request->Run'
        :Usage: select_run_request_actions_menu('Run request->run') 
        """
        menu_list=menu_path.split("->")
        print (menu_list)
        automation.ToolBarControl(AutomationId="23001").SplitButtonControl(Name="New").Click(30,16)
        
        for i in menu_list:
            print (i)
            control=automation.MenuItemControl(Name=i)
            automation.WaitForExist(control, 25)
            automation.MenuItemControl(Name=i).MoveCursor(50, 10)
            time.sleep(4)
            #automation.MenuItemControl(Name=i).DoubleClick()
            automation.SendKey(automation.Keys.VK_ENTER, waitTime=3)
            
    def select_run_request_actions_menu(self, menu_path):
        """      
        Description:- From Tasks and Animations panel, select run request and the path.
        :param menu_path= 'Run request->Run'
        :Usage: select_run_request_actions_menu('Run request->run') 
        """
        menu_list=menu_path.split("->")
        
        automation.SplitButtonControl(Name="Requests  selections").Click(30,10)
        
        for i in menu_list:
            print (i)
            time.sleep(4)
            automation.MenuItemControl(Name=i).MoveCursor(50, 8)
            time.sleep(4)
            automation.MenuItemControl(Name=i).Click(50, 8)
            
    def verify_web_url(self,url_text,msg):
        """      
        Description:- verify the web url of the current browser window
        :param url_text= the URL to verify
        :param msg= msg to pass for verification.
        :Usage: verify_web_url('http://www.google.com', 'Step x. Verify the  url') 
        """
        print (self.se_driver)
        actual_url=self.se_driver.current_url
        utillity.UtillityMethods.asin(self,url_text,actual_url,msg)
    
    def expand_doc_pane(self,pane_name):
        """      
        Description:- expand the dock pane containing settings, properties etc.
        :param pane_name= the pane name to click and expand
        :Usage: expand_doc_pane('Properties') 
        """
        options = { 'Tasks & Animations': '202',
                   'Requests & Data Sources': '191',
                   'Settings': '160',
                   'Properties': '150',
                   'File/Folder Properties': '1548'}
        panes=self.driver.find_element_by_id("59421").find_elements_by_id("1")
        time.sleep(2) 
        count=1
        for pane in panes:
            pane.click()
            time.sleep(2) 
            try:
                if self.driver.find_element_by_id(options[pane_name]).is_enabled():
#                     opane=self.driver.find_element_by_id(options[pane_name])
#                     w_x=opane.size['width']
#                     ac_pane=ActionChains(self.driver)
#                     ac_pane.move_to_element_with_offset(opane, w_x-25, 10).click().perform()
#                     del ac_pane
                    break
                else:
                    count=count + 1
                    print(count)
            except:pass
        if count==len(panes):
            print("Specified Pane ", pane_name, " does not exist")
        time.sleep(5)
        
    def settings_pane_select_data_source(self,folder,master_file):
        """      
        Description:- to select datasource name from setting pane.
        :param folder= the folder name to search for the datasource
        :param master_file= the master file to select.
        :Usage: settings_pane_select_data_source('ibisamp', 'car.mas) 
        """
        pane=automation.PaneControl(Name="Settings")
        automation.WaitForExist(pane, 30)
        
        radio=automation.RadioButtonControl(Name="Dynamic")
        radio.MoveCursor()
        radio.Click()
        
        time.sleep(3)
        
        '''Click on datasource button'''
        btn=automation.ButtonControl(AutomationId='15078')
        btn.MoveCursor()
        btn.Click()
        
        win_control=automation.WindowControl(Name="Select Data Source")
        AS_Utillity_Methods.wait_for_UI_object(self,win_control,30)
          
        AS_Utillity_Methods.select_UI_tree_view_item(self,folder, master_file, 'OK')
        
    def settings_pane_select_value_from(self,field_name):
        """      
        Description:- to select value from setting pane.
        :param field_name= the field name to doubleclick and select
        
        :Usage: settings_pane_select_value_from('COUNTRY') 
        """
        btn=automation.ButtonControl(AutomationId='15080')
        btn.MoveCursor()
        btn.Click()
        
        
        tree=automation.TreeControl(AutomationId='5445')
        AS_Utillity_Methods.wait_for_UI_object(self,tree, 30)
          
        tree_item=automation.TreeItemControl(Name=field_name)
        tree_item.ScrollIntoView()
        time.sleep(3)
        tree_item.MoveCursor()
        tree_item.DoubleClick()
    def settings_pane_select_value_from_list(self,field_name):
        """      
        Description:- to select value from setting pane.
        :param field_name= the field name to doubleclick and select
        
        :Usage: settings_pane_select_value_from('COUNTRY') 
        """
        btn=automation.ButtonControl(AutomationId='15080')
        btn.MoveCursor()
        btn.Click()
        
        
        tree=automation.TreeControl(AutomationId='5445')
        AS_Utillity_Methods.wait_for_UI_object(self,tree, 30)
          
        tree_item=automation.ListItemControl(Name=field_name)
        tree_item.ScrollIntoView()
        time.sleep(3)
        tree_item.MoveCursor()
        tree_item.DoubleClick()
    def click_picture_using_sikuli(self,image_name, xoffset,yoffset, click_type="single"):
        """      
        Description:- to click on a image pattern with offset values
        :param image_name='image1.png'
        :param xoffset= x offset value. The default is the left top of the object , to add x offset accordingly
        :param yoffset= y offset value. The default is the left top of the object, to add x offset accordingly
        :param click_type='single'
        :Usage: click_picture_using_sikuli('image1.png', '20','33','single') 
        """
        siklui_loc=os.getcwd() +"\\common\\lib\\sikuli\\runsikulix.cmd"
        image_name=os.getcwd() +"\\data\\"+image_name
        src_loc=os.getcwd() +"\\common\\lib\\sikuli\\click_picture --arg " + image_name + " " + xoffset + " " + yoffset + " " +click_type
        p=subprocess.Popen(siklui_loc+' -r '+ src_loc, shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)
    
    def click_picture_from_region_using_sikuli(self,image_name, xoffset,yoffset, click_type="single", screen_area='right'):
        """      
        Description:- to click on a image pattern with offset values
        :param image_name='image1.png'
        :param xoffset= x offset value. The default is the left top of the object , to add x offset accordingly
        :param yoffset= y offset value. The default is the left top of the object, to add x offset accordingly
        :param click_type='single'
        :Usage: click_picture_using_sikuli('image1.png', '20','33','single') 
        """
        siklui_loc=os.getcwd() +"\\common\\lib\\sikuli\\runsikulix.cmd"
        image_name=os.getcwd() +"\\data\\"+image_name
        arguments="{0} {1} {2} {3} {4}".format(image_name, xoffset, yoffset, click_type, screen_area)
        print (arguments)
        src_loc=os.getcwd() +"\\common\\lib\\sikuli\\click_picture_region --arg " + arguments
        p=subprocess.Popen(siklui_loc+' -r '+ src_loc, shell=True,stdout=subprocess.PIPE,stderr=subprocess.PIPE)    
            
    def add_css_url_by_keyboard(self,url):
        """      
        Description:- To add css/js by typing the text  in Settings pane
        :param offset= url to type in the editbox
        :Usage: add_css_url_by_keyboard('/approot/baseapp/blue.css') 
        """
        css_edit=automation.EditControl(AutomationId="23521")
        css_edit.Click()
        time.sleep(3)
        text=url
        pyautogui.typewrite(text)
        time.sleep(3)
        pyautogui.hotkey("enter")
        time.sleep(3)
    
    def add_css_url_by_button(self,button_type,css_name):
        """      
        Description:- To add css/js by using the button
        :param button_type: css or js
        :param css_name : the css file name
        :Usage: add_css_url_by_button('css','airblue.css')
        """
        if button_type=="css":
            btn_name="CSS..."
        else:
            btn_name="JavaScript"
        btn=automation.ButtonControl(Name=btn_name)
        btn.Click()
        win_control=automation.WindowControl(Name="Open File")
        AS_Utillity_Methods.wait_for_UI_object(self,win_control, 30)
   
        automation.TextControl(Name=css_name).DoubleClick(40,10)
        time.sleep(4)
    
    def select_tooltip_item(self, parent,item_name):
        '''
        Desc: This function is used to select an item from the single tooltip box. 
        '''
        tooltip_css= parent+" span[id*='tdgchart-tooltip']>div>ul>li [class*='tdgchart'][class*='tooltip'][class*='label']"
        elems=self.se_driver.find_elements_by_css_selector(tooltip_css)
        e1=elems[0]
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name)]
#         utillity.UtillityMethods.click_on_screen(self, e1, coordinate_type='middle', click_type=3)
#         utillity.UtillityMethods.click_on_screen(self, tooltip_item_to_be_selected, coordinate_type='middle', click_type=0)
        tooltip_item_to_be_selected.click()
        
        #coreutillobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
        time.sleep(10)
        
    def select_report_autolink_tooltip(self, parent_table_css, row, col, tooltip_path, verify_tooltip=None, msg=None, **kwargs):
        """
        Example Usage : iarun.select_report_autolink_tooltip("table[summary= 'Summary']", 6,2, "Auto Links->AUTOLINK_TARGET01")
        """
        time.sleep(4)
        browser = AS_Utillity_Methods.parseinitfile(self,'browser')
        #obj_move_cursor=self.se_driver.find_element_by_css_selector(parent_table_css)
        #utillity.UtillityMethods.click_type_using_pyautogui(self, obj_move_cursor, leftClick=True, cord_type='start', **kwargs)
        cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"   
        obj_cell_css=self.se_driver.find_element_by_css_selector(cell_css)
        if browser.lower()=="ie":
            utillity.UtillityMethods.click_type_using_pyautogui(self, obj_cell_css, leftClick=True, cord_type='top_middle', x_offset=40, y_offset=56, **kwargs)
        else:
            obj_cell_css.click()
        time.sleep(4)
        tooltips=tooltip_path.split('->')
        tooltip_css="#my_tooltip_id ul[class='tdgchart-tooltip-list']>li[class*='tdgchart-tooltip-pad']"
        tooltip_obj=self.se_driver.find_element_by_css_selector(tooltip_css)
        #utillity.UtillityMethods.click_on_screen(self, tooltip_obj, 'top_middle', y_offset=5, mouse_duration=0)
        for count, tooltip in enumerate(tooltips) :
            tooltip_elements=self.se_driver.find_elements_by_css_selector(tooltip_css)
            tooltip_text_list=[tooltip.text.strip().replace('>', '').replace('\n', '') for tooltip in tooltip_elements]
            tooltip_element=tooltip_elements[tooltip_text_list.index(tooltip)]
            if tooltips[count] == tooltips[-1] :
                if verify_tooltip!=None :
                    utillity.UtillityMethods.asequal(self, verify_tooltip, tooltip_text_list, msg)
                tooltip_element.click()
                #utillity.UtillityMethods.click_on_screen(self, tooltip_element, coordinate_type='middle', click_type=0)
            else :
                utillity.UtillityMethods.click_on_screen(self, tooltip_element, coordinate_type='left', x_offset=2)
                tooltip_element.click()
                tooltip_arrow=tooltip_element.find_element_by_css_selector("span[class='tdgchart-tooltip-arrow']")
                #utillity.UtillityMethods.click_on_screen(self, tooltip_arrow, coordinate_type='middle')
                tooltip_arrow.click()
                time.sleep(2)
        time.sleep(10)
   
    def move_mouse_to_chart_component(self, riser_or_marker_element, use_marker_enable=False, move_to_tooltip=True):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        x=0
        y=0
        riser_or_marker_element=self.se_driver.find_element_by_css_selector(riser_or_marker_element)
        coreutillobject.move_mouse_to_offset(self, x_offset=x, y_offset=y)
        time.sleep(1)
        if use_marker_enable==True:
            (x, y)=utillity.UtillityMethods.marker_enable(self, riser_or_marker_element)
            coreutillobject.python_move_to_offset(self, x_offset=x, y_offset=y, mouse_move_duration=2.5)
        else:
            coreutillobject.python_move_to_element(self, riser_or_marker_element)
        
    def get_elem_obj(self,element_css, pos):    
        element=self.se_driver.find_element_by_css_selector(element_css)
        a=utillity.UtillityMethods.get_object_screen_coordinate(self, element, coordinate_type=pos)
        return(a)
        
    def click_on_elem(self, element_css, x_offset, y_offset):
        
            
        element=self.se_driver.find_element_by_css_selector(element_css)
        a=utillity.UtillityMethods.get_object_screen_coordinate(self, element, coordinate_type='bottom_middle')
        print (x_offset)
        print (y_offset)
        x_offset=x_offset+a['x']
        y_offset=y_offset+a['y']
        print (x_offset)
        print (y_offset)
        print (a)
        coreutillobject.python_move_to_offset(self, x_offset=x_offset, y_offset=y_offset, mouse_move_duration=0.5)
        
#         tooltip_css="span[id*='tdgchart-tooltip']:not([style*='hidden'])"
#         tooltip_elem=self.se_driver.find_elements_by_css_selector(tooltip_css + " li")[0]
#         a=utillity.UtillityMethods.get_object_screen_coordinate(self, tooltip_elem, coordinate_type='start')
#         x_offset=a['x']
#         y_offset=a['y']
#         coreutillobject.python_move_to_offset(self, x_offset=x_offset, y_offset=y_offset, mouse_move_duration=0.5)
        
    def double_click_web_element(self, element_css):
        to_element=self.se_driver.find_element_by_css_selector(element_css)
        ac = ActionChains(self.se_driver)
        ac.move_to_element(to_element)
        ac.double_click(to_element).perform()

                  
    def click_on_webelement_default_click(self,select_css):
        to_element=self.se_driver.find_element_by_css_selector(select_css)
        ac = ActionChains(self.se_driver)
        ac.move_to_element(to_element)
        ac.click(to_element).perform()
        
    def verify_xml_xls(self, base_file_name, actual_file_name, msg):
        
        base_file_name = os.getcwd() +"\\data\\" + base_file_name
        actual_file_name = os.getcwd() +"\\data\\" + actual_file_name
        base_file_name = base_file_name
        verify_xml_xls_cmd="cscript.exe " +os.getcwd() +"\\common\\lib\\verify_xml_xls.vbs "+ actual_file_name + " " + base_file_name
#         try:
#             if bool(pywinauto.findwindows.find_window(title_re='Windows Script Host')):
#                 print ("VBA error message")
#                 app = pywinauto.application.Application().Connect(title_re='Windows Script Host', class_name='#32770')
#                 window = app.Dialog
#                 button = window.OK
#                 button.Click()
#         except:
#             pass
        resp=subprocess.call(verify_xml_xls_cmd, shell=True)
        AS_Utillity_Methods.asequal(self, 9, resp, msg)
        
    def save_file_from_browser(self,filename,**kwargs):
        '''
        Desc: This function is to save files from the browser . Example Excel or pptx files
        :param : filename= the file name to save to with the extension
        :kwargs: custom_ie_re = if more than one IE window the custom title name for a new IE window which contains the save dialog.
        :Usage : save_file_from_browser('one.xlsx', custom_ie_re='.*BIP_RUN.*')
        '''
        
        custom_ie_re=kwargs['custom_ie_re'] if 'custom_ie_re' in kwargs else ".*Internet Explore.*"
        custom_cr_re=kwargs['custom_cr_re'] if 'custom_cr_re' in kwargs else ".*Chrome"
        br = AS_Utillity_Methods.parseinitfile(self,'browser')
        
        if br=='Chrome':
            time.sleep(5)
            automation.WindowControl(RegexName="Save As.*").SetFocus()
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").Click()
            time.sleep(2)
            pyautogui.hotkey('ctrl','a')
            time.sleep(2)
            pyautogui.hotkey('del')
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").SetValue(os.getcwd() + "\data\\" + filename)
            automation.ButtonControl(Name="Save").SetFocus()
            automation.ButtonControl(Name="Save").Click(10, 10, False, 10)
            time.sleep(2)
            if automation.WindowControl(RegexName="Confirm Save.*").Exists():
                automation.WindowControl(RegexName="Confirm Save.*").SetFocus()
                automation.ButtonControl(Name="Yes").Click(10, 10, False, 10)
            time.sleep(6) 
            
            
        elif br=='Firefox':
            automation.WindowControl(RegexName="Opening.*").Exists(10, 1)
            automation.WindowControl(RegexName="Opening.*").SetFocus()
            time.sleep(2)
            automation.RadioButtonControl(Name="Save File").Select()
            time.sleep(2)
            automation.ButtonControl(Name="OK").SetFocus()
            automation.ButtonControl(Name="OK").Click(10, 10, False, 10)
            #automation.ButtonControl(Name="OK").Click()
            time.sleep(5)
            automation.WindowControl(RegexName="Enter name of file.*").SetFocus()
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").Click()
            time.sleep(2)
            pyautogui.hotkey('ctrl','a')
            time.sleep(2)
            pyautogui.hotkey('del')
            time.sleep(2)
            automation.EditControl(RegexName="File name.*").SetValue(os.getcwd() + "\data\\" + filename)
            automation.ButtonControl(Name="Save").SetFocus()
            automation.ButtonControl(Name="Save").Click(10, 10, False, 10)
            time.sleep(2)
            if automation.WindowControl(RegexName="Confirm Save.*").Exists():
                automation.WindowControl(RegexName="Confirm Save.*").SetFocus()
                automation.ButtonControl(Name="Yes").Click(10, 10, False, 10)
            time.sleep(6) 
        else:
                app = pywinauto.application.Application().connect(class_name="IEFrame", title_re=custom_ie_re)
                ieframe = app.IEFrame
                ieframe.Wait('ready')
                time.sleep(8)
                if bool(ieframe[u'Frame Notification Bar'].Exists()):
                    
                    framenotificationbar = ieframe[u'Frame Notification Bar']
                    framenotificationbar.Maximize()
                    framenotificationbar.Restore()
                    time.sleep(3)
                    pyautogui.hotkey('alt','n')
                    time.sleep(2)
                    pyautogui.press('tab')
                    time.sleep(2)
                    pyautogui.press('up')
                    time.sleep(2)
                    pyautogui.press('down')
                    time.sleep(2)
                    pyautogui.press('enter')
                    time.sleep(2)
                    dlg=app.window(title = 'Save As')
                    
                    dlg[u'Edit'].ClickInput()
                    time.sleep(2)
                    pyautogui.hotkey('ctrl','a')
                    time.sleep(2)
                    pyautogui.hotkey('del')
                    time.sleep(2)
                    #dlg[u'Edit'].set_text(os.getcwd() + "\data\\" + filename)
                    pyautogui.typewrite(os.getcwd() + "\data\\" + filename, interval=0.3, pause=3)
                    time.sleep(2)
                    dlg['Button1'].ClickInput()
                    time.sleep(2)
                       
                    if app.window(title='Confirm Save As').Exists()==True:
                        window1 = app.window(title='Confirm Save As')
                        window1.Maximize()
                        window1.Restore()
                        button = window1[u'Yes']
                        button.ClickInput()
                    
                    time.sleep(6)

    def navigate_to_env_tree_item(self, tree_path, item, **kwargs):
        
        """
        Usage: select_item_from_tree_view("driver.find_element_by_name("Static"), "jawin7->Domains->P20")
        params:parent_obj="driver.find_element_by_name("Static" );"jawin7->Domains->P20"
        """ 
        
        host= AS_Utillity_Methods.parseinitfile(self,'clientid')
        path1="Configured Environments->"+host
        menu_list=path1.split("->")
        for i in menu_list:
            if automation.ExpandCollapseState.Collapsed==automation.TreeItemControl(Name=i).CurrentExpandCollapseState():
                automation.TreeItemControl(Name=i).ScrollIntoView()
                time.sleep(3)
                automation.TreeItemControl(Name=i).Expand(waitTime=2)
                time.sleep(3)
                automation.TreeItemControl(Name=i).Click()
            if (automation.TreeItemControl(Name=i).CurrentExpandCollapseState()==3):
                time.sleep(3)
                automation.TreeItemControl(Name=i).DoubleClick()
                time.sleep(3)
        time.sleep(5)
        menu_list=tree_path.split("->")
        for i in menu_list:
            automation.TreeItemControl(Name=i).ScrollIntoView()
            time.sleep(3)
           
            if (automation.ExpandCollapseState.Collapsed==automation.TreeItemControl(Name=i).CurrentExpandCollapseState()):
                print (i)
                
                time.sleep(3)
                automation.TreeItemControl(Name=i).Expand(waitTime=2)
                time.sleep(3)
                automation.TreeItemControl(Name=i).Click()
            if (automation.TreeItemControl(Name=i).CurrentExpandCollapseState()==3):
                print (i)
                
                time.sleep(3)
                automation.TreeItemControl(Name=i).DoubleClick()

            
        automation.TreeItemControl(Name=item).Click()
        time.sleep(3)
    def click_tab_button_UI(self,tab_name,button_name):
        tab_item=automation.TabItemControl(Name=tab_name)
        btn=automation.ButtonControl(Name=button_name)
        tab_item.MoveCursor()
        time.sleep(3)
        tab_item.Click()
        time.sleep(5)
        btn.MoveCursor()
        time.sleep(3)
        btn.Click()
        time.sleep(5)
    
    def change_page_layout_from_properties(self, page_layout_name):
        automation.WindowControl(RegexName="App*").Resize(1024,800)
        
        AS_Utillity_Methods.expand_doc_pane(self,'Properties')
        time.sleep(2)
        automation.PaneControl(Name="Properties").ComboBoxControl(AutomationIs="841").Click()
        
        automation.ListItemControl(Name="Page layout 1").Click()
        time.sleep(2)
        automation.WindowControl(RegexName="App*").Maximize(waitTime=10)
        
    def compare_excel_sheet(self, base_file, actual_file, sheet_name):
        import xlrd
        wb1 = xlrd.open_workbook(os.getcwd() + "\data\\" + base_file)
        ws1 = wb1.sheet_by_name(sheet_name)
        wb2 = xlrd.open_workbook(os.getcwd() + "\data\\" + actual_file)
        ws2 = wb2.sheet_by_name(sheet_name)
        for r in range(0, ws1.nrows):
            for c in range(0, ws1.ncols):
                if ws1.cell(r, c).value == ws2.cell(r, c).value:
                    status=[0]
                    continue
                else:
                    status=[r+1,c+1]
                    return (status)
        return (status)
    
    def verify_excel_sheet(self, base_file, actual_file, sheet_name, msg):
        """
        base_file='C2053851_Base.xlsx'
        actual_file='C2053851_actual.xlsx'
        sheet_name='AHTML'
        usage: utillobj.verify_excel_sheet('C2053851_Base.xlsx', 'C2053851_actual.xlsx', 'AHTML', 'Step 05: Verify that all the records')
        @author : Niranjan
        """
        x= self.compare_excel_sheet(base_file, actual_file, sheet_name)
        self.asequal(len(x),1,msg+ ' Row,Column -'+ str(x))

