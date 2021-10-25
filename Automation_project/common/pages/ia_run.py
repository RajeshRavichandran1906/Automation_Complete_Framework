import sys
from common.lib.utillity import UtillityMethods as utillityobject
from common.lib.core_utility import CoreUtillityMethods as coreutillityobject
from common.lib.base import BasePage
from selenium.webdriver import ActionChains
from selenium.webdriver.support import expected_conditions as EC
import time, os,re,operator,pyautogui
from openpyxl import Workbook
from openpyxl import load_workbook
from functools import reduce
from selenium.webdriver.support.color import Color
from bs4 import BeautifulSoup
from common.lib.global_variables import Global_variables
from selenium.common.exceptions import NoSuchElementException
from common.locators.infoassist_locators import InfoassistLocators
from common.lib.javascript import JavaScript
if sys.platform == 'linux':
    from pykeyboard import PyKeyboard
    pykeyboard = PyKeyboard()
else:
    from uisoup import uisoup
    import keyboard as local_keyboard

class IA_Run(BasePage):
    """ Inherit attributes of the parent class = Baseclass """

    def __init__(self, driver):
        super(IA_Run, self).__init__(driver)

    def _validate_page(self, locator):
        self.longwait.until(EC.visibility_of_element_located(locator))
    
    def move_mouse_to_chart_component(self, riser_or_marker_element, use_marker_enable=False, move_to_tooltip=True):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        x=Global_variables.current_working_area_browser_x
        y=Global_variables.current_working_area_browser_y
        coreutillityobject.move_mouse_to_offset(self, x_offset=x, y_offset=y)
        time.sleep(1)
        if use_marker_enable==True:
            (x, y)=utillityobject.marker_enable(self, riser_or_marker_element)
            coreutillityobject.python_move_to_offset(self, x_offset=x, y_offset=y, mouse_move_duration=2.5)
        else:
            coreutillityobject.python_move_to_element(self, riser_or_marker_element)
        tooltip_css="span[id='tdgchart-tooltip']:not([style*='hidden'])"
        utillityobject.synchronize_with_number_of_element(self,tooltip_css,1, 25, pause_time=0.2)
        tooltip_elem=self.driver.find_elements_by_css_selector(tooltip_css + " li")[0]
        if move_to_tooltip :
            coreutillityobject.python_move_to_element(self, tooltip_elem)
    
    def select_chart_component(self, riser_or_marker_element, use_marker_enable=False):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        x=Global_variables.current_working_area_browser_x
        y=Global_variables.current_working_area_browser_y
        coreutillityobject.move_mouse_to_offset(self, x, y)
        time.sleep(1)
        if use_marker_enable==True:
            (x, y)=utillityobject.marker_enable(self, riser_or_marker_element)
            coreutillityobject.python_click_with_offset(self, x, y)
        else:
            coreutillityobject.left_click(self, riser_or_marker_element)
        tooltip_elem=self.driver.find_element_by_css_selector("div[class='tdgchart-tooltip']")
        coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_middle', yoffset=5)
    
    def multiselect_chart_component(self, riser_or_marker_element_list, use_marker_enable=False):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        x=Global_variables.current_working_area_browser_x
        y=Global_variables.current_working_area_browser_y
        coreutillityobject.move_mouse_to_offset(self, x, y)
        time.sleep(1)
        if sys.platform == 'linux':
            pykeyboard.press_key(pykeyboard.control_key)
        else:
            local_keyboard.press('ctrl')
        for riser_or_marker_element in riser_or_marker_element_list:
            if use_marker_enable==True:
                (x, y)=utillityobject.marker_enable(self, riser_or_marker_element)
                coreutillityobject.python_click_with_offset(self, x, y)
            else:
                coreutillityobject.python_left_click(self, riser_or_marker_element, mouse_move_duration=3.5)
            time.sleep(1)
        if sys.platform == 'linux':
            pykeyboard.release_key(pykeyboard.control_key)
        else:
            local_keyboard.release('ctrl') 
        time.sleep(1)
        #tooltip_elem=self.driver.find_element_by_css_selector("div[class='tdgchart-tooltip']")
        #coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_middle', yoffset=5)
                
    def verify_tooltip(self, expected_tooltip_list, msg='Step X: Verify default Chart tooltip'):
        '''
        Desc: This function will verify the tooltip. First mouse will move to the top left corner of the current working area,
                Then it will move to the required location to bring the tooltip disply and then read and compare.
        '''
        tooltip_css="span[id='tdgchart-tooltip']"
        raw_tooltip_list=self.driver.find_element_by_css_selector(tooltip_css).text.split('\n')
        actual_list=utillityobject.get_actual_tooltip_list(self, raw_tooltip_list)
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
        
    def verify_lasso_tooltip(self, expected_tooltip_list, msg='Step X: Verify default Chart tooltip'):
        '''
        Desc: This function will verify the tooltip. First mouse will move to the top left corner of the current working area,
                Then it will move to the required location to bring the tooltip disply and then read and compare.
        '''
        tooltip_css="div[class='tdgchart-tooltip']"
        raw_tooltip_list=self.driver.find_element_by_css_selector(tooltip_css).text.split('\n')
        actual_list=utillityobject.get_actual_tooltip_list(self, raw_tooltip_list)
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
        
    def hover_to_required_arrow_in_tooltip(self, item_name):
        first_level_tooltip_css = "#tdgchart-tooltip li[class*='tdgchart-tooltip-hover']"
        elems=self.driver.find_elements_by_css_selector(first_level_tooltip_css)
        first_level_tooltip=elems[[elem.text.strip() for elem in elems].index(item_name)]
        arrow_to_be_selected=first_level_tooltip.find_element_by_css_selector("span.tdgchart-tooltip-arrow")
        coreutillityobject.python_move_to_element(self, arrow_to_be_selected)
        tooltip_elem=self.driver.find_element_by_css_selector("span[id='tdgchart-tooltip'] #tdgchart-submenu")
        coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_middle', yoffset=5)
        
    def select_tooltip_item(self, item_name):
        '''
        Desc: This function is used to select an item from the single tooltip box. 
        '''
        if 'item_name' in ['Filter Chart', 'Exclude from Chart', 'Remove Filter']:
            tooltip_css= "#tdgchart-tooltip span.tdgchart-tooltip-label>div"
        else:
            tooltip_css= "#tdgchart-tooltip span.tdgchart-tooltip-label"
        elems=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name)]
        coreutillityobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
        time.sleep(10)
    
    def create_laso(self, source_element, target_element, source_element_location, source_xoffset, source_yoffset, target_element_location, target_xoffset, target_yoffset):
        '''
        Desc: This function is used to create a lasso tooltip by dragging from source to target element provided.
        '''
        source_dict=coreutillityobject.get_web_element_coordinate(self, source_element, element_location=source_element_location, xoffset=source_xoffset, yoffset=source_yoffset)
        x1=source_dict['x']
        y1=source_dict['y']
        target_dict=coreutillityobject.get_web_element_coordinate(self, target_element, element_location=target_element_location, xoffset=target_xoffset, yoffset=target_yoffset)
        x2=target_dict['x']
        y2=target_dict['y']
        coreutillityobject.drag_and_drop(self, x1, y1, x2, y2)
        time.sleep(2)
        
    def select_lasso_tooltip_item(self, item_name):
        '''
        Desc: This function is used to select an item from the single lasso tooltip box.
        '''
        tooltip_css= "div[class='tdgchart-tooltip'] span[class='tdgchart-tooltip-pad']"
        elems=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name)]
        coreutillityobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
        time.sleep(10)
        
    def select_bilevel_tooltip_item(self, item_name1, item_name2):
        '''
        Desc: This function is used to select an item from the second tooltip box. Mostly used for drill down pop up.
        '''
        IA_Run.hover_to_required_arrow_in_tooltip(self, item_name1)
        second_level_tooltip_css = "#tdgchart-tooltip #tdgchart-submenu span.tdgchart-tooltip-label"
        elems=self.driver.find_elements_by_css_selector(second_level_tooltip_css)
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name2)]
        coreutillityobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
        time.sleep(10)
    
    """=====================OLD FUNCTIONS==================================""" 
    def verify_document_objects(self, parent_css, component_type, msg, expected_value_list=None, expected_image_name=None):
        '''
        Desc :- This function is used to verify Document Objects. We have two type of objects 'textbox' OR 'image'
        parent_css= [id^='LOBJText'] OR [id^='LOBJPageItemImage']
        '''
        if component_type == 'textbox':
            iter_values=(val for val in (val.text.strip() for val in self.driver.find_elements_by_css_selector(parent_css + " table > tbody > tr"))  if val != '')
            verify_status= True
            for expected_value in expected_value_list:
                if expected_value == next(iter_values):
                    verify_status= True
                else:
                    verify_status=False
                    break
            del iter_values
            utillityobject.asequal(self, verify_status, True, msg)
        elif component_type == 'image':
            parent_css = parent_css + " img"
            utillityobject.verify_object_visible(self, parent_css, True, msg)
            utillityobject.verify_picture_using_sikuli(self,expected_image_name, msg)
            
    def verify_active_dashboard_prompts(self, component_type, parent_css, expected_value_list, msg, default_selected_check=None):
        '''
        Desc :- This function is used to verify Active Dashboard Prompts
        :param component_type=verify dashboard prompts component like checkbox, list, etc...
        :param parent_css=element css value
        :param expected_value_list=['[All]', 'Coffee', 'Food', 'Gifts']
        :param msg=output message to display in console
        :param default_selected_check=default selected option in dashboard prompts
        :usage verify_active_dashboard_prompts('drop_down', "#PROMPT_3_cs", ['[All]', 'Coffee', 'Food', 'Gifts'], "step 5: verify prompts", default_selected_check='[All]')
        '''
        if component_type.lower() == 'text':
            field_elem = self.driver.find_element_by_css_selector(parent_css + " input").get_attribute('value')
            verify_status= True if field_elem == expected_value_list[0] else False
            utillityobject.asequal(self, verify_status, True, msg)
        else:
            page_source_text=self.driver.page_source
            soup = BeautifulSoup(page_source_text, 'html.parser')
            iter_values=[val for val in [val.get_text(strip=True) for val in soup.select(parent_css + " table > tbody > tr")] if val != '']
            verify_status= True
            for checkbox_value in expected_value_list:
                if checkbox_value in iter_values:
                    verify_status= True
                else:
                    verify_status=False
                    break
            utillityobject.asequal(self, verify_status, True, msg)
            if default_selected_check != None:
                iter_items=(elem for elem in soup.select(parent_css + " table > tbody > tr"))
                for item in iter_items:
                    if item.get_text(strip=True) == default_selected_check:
                        status="checked" in str(item.select("input")[0])
                        break
                    else:
                        status=False
                utillityobject.asequal(self, status, True, msg)
    
    def select_active_dashboard_prompts(self, component_type, parent_css, selection_list, scroll_down_times=0, type_speed=0.5):
        '''
        Desc :- This function is used to select Active Dashboard Prompts
        :param component_type=select dashboard prompts component like checkbox, list, etc...
        :param parent_css=element css value
        :param selection_list=option need to select in dashboard prompt
        :usage select_active_dashboard_prompts('drop_down', "#PROMPT_3_cs", ['Biscotti', 'Capuccino'])
        '''
        time.sleep(3)
        if component_type.lower() == 'text':
            field_elem = self.driver.find_element_by_css_selector(parent_css + " input")
            utillityobject.click_on_screen(self, field_elem, 'left')
            time.sleep(1)
            if sys.platform == 'linux':
                pykeyboard.press_key(pykeyboard.control_key)
                pykeyboard.tap_key(character=u'\u0061')
                pykeyboard.release_key(pykeyboard.control_key)
                pykeyboard.tap_key(pykeyboard.delete_key)
                pykeyboard.type_string(str(selection_list[0]), interval=int(type_speed))
                time.sleep(1)
                pykeyboard.tap_key(pykeyboard.enter_key)
            else:
                local_keyboard.press('ctrl')
                local_keyboard.press('a')
                local_keyboard.release('a')
                local_keyboard.release('ctrl')
                local_keyboard.send('del')
                local_keyboard.write(selection_list[0], delay=int(type_speed))
                time.sleep(1)
                local_keyboard.send('enter')
        else:
            if sys.platform == 'linux':
                pykeyboard.press_key(pykeyboard.control_key)
            else:
                local_keyboard.press('ctrl')
            time.sleep(1)
            for selection_item_name in selection_list:
                iter_items=(elem for elem in self.driver.find_elements_by_css_selector(parent_css + " table > tbody > tr"))
                for item in iter_items:
                    if item.text.strip() == selection_item_name:
                        count=0
                        while scroll_down_times > count:
                            prompt_elem=self.driver.find_element_by_css_selector(parent_css)
                            utillityobject.click_on_screen(self, prompt_elem, 'bottom_right', click_type=0, x_offset=-10, y_offset=-10)
                            time.sleep(1)
                            count=count+1
                        if component_type == 'listbox':
                            item.click()
                        else:
                            item.find_element_by_css_selector("input").click()
                        break
            if sys.platform == 'linux':
                pykeyboard.release_key(pykeyboard.control_key)
            else:
                local_keyboard.release('ctrl')
            time.sleep(1)
    
    def verify_selected_value_in_active_dashboard_prompts(self, component_type, parent_css, expected_value_list, msg):
        '''
        Desc :- This function is used to verify Active Dashboard Prompts after selection at run time
        :param component_type=verify dashboard prompts component like checkbox,radio, list, etc...
        :param parent_css=element css value
        :param expected_value_list=['Coffee']
        :param msg=output message to display in console
        :usage verify_active_dashboard_prompts('radio', "#PROMPT_3_cs", ['Coffee'], "step 5: verify radio box")
        '''
        _ = component_type
        elems = self.driver.find_elements_by_css_selector(parent_css + " table > tbody > tr")
        for item in expected_value_list:
            ind=elems[[item in el1 for el1 in [el.text.strip() for el in elems]].index(True)]
            try:
                status=ind.find_element_by_css_selector("input").is_selected()
            except:
                status=False
                break
        utillityobject.asequal(self, status, True, msg)
        
    def verify_default_selected_listbox_value(self, parent_css, default_value, default_value_color, msg='Step X'):
        '''
        Desc:- Verify the selected list box value
        :param: parent_css = "#list_dPROMPT_1 table tr"
        :default_value = '[All]' 
        :default_value_color = 'gray8'
        :msg = "Step 1:"
        verify_selected_listbox_value("#list_dPROMPT_1 table tr", '[All]', 'gray8',msg="Step 1:")
        '''
        custom_msg=msg + ": Verify default selected list box value."
        expected_default_selected_value=utillityobject.color_picker(self, default_value_color, 'rgba')
        list_items_obj=utillityobject.validate_and_get_webdriver_objects(self, parent_css, 'Listbox')
        actual_field_selected_color = Color.from_string(list_items_obj[[x.text.strip() for x in list_items_obj].index(default_value)].value_of_css_property('color')).rgba
        utillityobject.asequal(self,expected_default_selected_value,actual_field_selected_color, custom_msg)
        
    def create_table_data_set(self,table_css,file_name, **kwargs):
        """    
        @param: table_css= ".arPivot tr:nth-child(1) table" or "table[summary='Summary']"   =  Need to provide the full parent path till table
        @param: file_name: "test1.xlsx" 
        Usage: create_table_data_set("table[summary='Summary']", "test1.xlsx" )
        """
        
        rows_css = table_css + " > tbody > tr"
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        if 'starting_rownum' in kwargs:
            start_rownum=kwargs['starting_rownum']
        else:
            start_rownum=0
            
        for r in range(start_rownum,no_of_rows):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text
                s.cell(row=r + 1, column=c + 1).value = str(value).replace(' ','').strip()
        wb.save(os.path.join(os.path.join(os.getcwd(), "data"), file_name))
    
    def verify_table_data_set(self,table_css,file_name,msg, **kwargs):
        """
        Usage: utillobj.verify_table_data_set("table[summary='Summary']", "test1.xlsx","Step 10: fail data set")
        """
        x= IA_Run.compare_table_data_set(self, table_css,file_name, **kwargs)
        utillityobject.asequal(self,len(x),1,msg+ ' Row,Column -'+ str(x))
        
    def compare_table_data_set(self, table_css, file_name, **kwargs):
        wb1 = load_workbook(os.path.join(os.path.join(os.getcwd(), "data"), file_name))
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        rows_css = table_css + " > tbody > tr"
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        if 'starting_rownum' in kwargs:
            start_rownum=kwargs['starting_rownum']
        else:
            start_rownum=0
        for r in range(start_rownum, no_of_rows):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(0, len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text
                if s1.cell(row=r + 1, column=c + 1).value != None and str(value) != '':
                    if s1.cell(row=r + 1, column=c + 1).value.replace(' ','').lower().strip() == str(value).replace(' ','').lower().strip():
                        status=[0]
                        continue
                    else:
                        status=[r+1,c, 'expected',s1.cell(row=r + 1, column=c + 1).value.replace(' ','').lower().strip(), 'actual',str(value).replace(' ','').lower().strip()]
                        return (status)
        return (status)
    
    def select_and_verify_drilldown_report_field(self, table_css, rownum, colnum, **kwargs):
        """
        :param: table_css= ".arPivot tr:nth-child(1) table" or "table[summary='Summary']"   =  Need to provide the full parent path till table.
        :param: rownum=1,2,3,....
        :param: colnum=1,2,3,...
        :param: expected_drill_down_tooltip='Drill Down 1'
        :param: underline=True
        :usage: select_and_verify_drilldown_report_field('table[summary='Summary']', 2, 2, expected_drill_down_tooltip='Drill Down 1', underline=True):
        """
        linked_cell_css=table_css + " > tbody > tr:nth-child(" + str(rownum) + ") > td:nth-child(" + str(colnum) + ") a"
        linked_cell_obj=self.driver.find_element_by_css_selector(linked_cell_css)
        if 'expected_drill_down_tooltip' in kwargs:
            actual_drill_down_tooltip=linked_cell_obj.get_attribute("title").strip()
            utillityobject.asequal(self,kwargs['expected_drill_down_tooltip'], actual_drill_down_tooltip , kwargs['msg'] + " Verification of drilldown tooltip")
        time.sleep(2)
        if 'underline' in kwargs:
                actual_decoration=True if 'underline' in linked_cell_obj.value_of_css_property("text-decoration") else False
                utillityobject.asequal(self, kwargs['underline'], actual_decoration, kwargs['msg'] + " Verification of Cell is underlined.")
        linked_cell_obj.click()        
#         utillityobject.default_left_click(self, object_locator=linked_cell_obj, **kwargs)
        time.sleep(3)

    def verify_table_cell_data_bars(self, table_css, msg, **kwargs):
        """
        Params: table_css = "table[summary='Summary']"
        Params: msg = 'Step 9:'
        kwargs: expected_number_of_bars=1, 2, 3, 4....(Integer value to verify the number of bars present.)
        Usage: verify_table_cell_data_bars("table[summary='Summary']", expected_number_of_bars=20, msg='Step 15.4:')
        Author: Aftab
        """
        data_bar_css=table_css + " > tbody > tr > td[class^='x'] table td table td[style*='background-color']"
        bar_items=self.driver.find_elements_by_css_selector(data_bar_css)
        time.sleep(1)
        if 'expected_number_of_bars' in kwargs:
            actual_number_of_bars=len(bar_items)
            time.sleep(1)
            utillityobject.asequal(self, kwargs['expected_number_of_bars'], actual_number_of_bars, msg + "a. Verify Number of data bars.")
            print(actual_number_of_bars)
        if 'expected_color' in kwargs:
            expected_bar_color=utillityobject.color_picker(self, kwargs['expected_color'], 'rgba')
            actual_bar_color=Color.from_string(bar_items[kwargs['expected_color_index']-1].value_of_css_property("background-color")).rgba
            print(expected_bar_color)
            print(actual_bar_color)
            utillityobject.asequal(self, actual_bar_color, expected_bar_color , msg+ "b.. Verification of Cell Background color.")
            time.sleep(1)
        
        
    def verify_table_cell_property(self, table_css, rownum, colnum, **kwargs):
        """
        Params: table_css = "table[summary='Summary']"
        cell_no = 1, 2, 3...any integer value for cell number, including heading cells.
        font_color = 'black'
        bg_color='red'
        text='CAR'
        font_name='Arial'
        font_size='10pt'
        bold=True
        italics=True
        underline=True
        text_align='left', OR, 'center' Or 'right'
        Usage: verify_report_cell_property("TableChart_1", 1, bg_color='Cyan', font_color='magenta', text_value='CAR', font_name='Arial', font_size='10pt', bold=True, italic=True, underline=True, text_align='Center')
        """
        if 'custom_table_css' in kwargs:
            table_css=kwargs['custom_table_css']
        else:
            table_css=table_css + " > tbody > tr:nth-child(" + str(rownum) + ") > td:nth-child(" + str(colnum) + ")"
        cell_obj=self.driver.find_element_by_css_selector(table_css)
        for key in kwargs:
            if 'bg_color' in key:
                expected_background_color=utillityobject.color_picker(self, kwargs['bg_color'], 'rgba')
                actual_background_color=Color.from_string(cell_obj.value_of_css_property("background-color")).rgba
                utillityobject.asin(self, actual_background_color, expected_background_color , kwargs['msg'] + ". Verification of Cell Background color.")
            if 'font_color' in key:
                expected_text_color=utillityobject.color_picker(self, kwargs['font_color'], 'rgba')
                actual_text_color=Color.from_string(cell_obj.value_of_css_property("color")).rgba
                utillityobject.asin(self, actual_text_color, expected_text_color, kwargs['msg'] + ". Verification of Cell Text color.")
            if 'text_value' in key:
                actual_text=cell_obj.text.strip()
                utillityobject.asequal(self, kwargs['text_value'], actual_text, kwargs['msg'] + ". Verification of Cell Text.")
            if 'font_name' in key:
                actual_font=cell_obj.value_of_css_property("font-family").strip('"').upper()
                utillityobject.asequal(self, kwargs['font_name'].upper(), actual_font, kwargs['msg'] + ". Verification of Cell Text Font name.")
            if 'font_size' in key:
                expected_font_size=round(float(1.333333*int(kwargs['font_size'][:-2])))
                actual_font_size=cell_obj.value_of_css_property("font-size")
                actual_font_size=round(float(actual_font_size[:-2]))
                utillityobject.asequal(self, expected_font_size, actual_font_size, kwargs['msg'] + ". Verification of Cell Text Font Size.")
            if 'bold' in key:
                actual_weight=True if cell_obj.value_of_css_property("font-weight") in ['700', 'bold'] else False
                utillityobject.asequal(self, kwargs['bold'], actual_weight, kwargs['msg'] + ". Verification of Cell Text is Bold.")
            if 'italic' in key:
                actual_style=True if cell_obj.value_of_css_property("font-style")=='italic' else False
                utillityobject.asequal(self, kwargs['italic'], actual_style, kwargs['msg'] + ". Verification of Cell Text is Italics.")
            if 'underline' in key:
                actual_decoration=True if 'underline' in cell_obj.value_of_css_property("text-decoration") else False
                utillityobject.asequal(self, kwargs['underline'], actual_decoration, kwargs['msg'] + ". Verification of Cell is underlined.")
            if 'text_align' in key:
                actual_font=cell_obj.value_of_css_property("text-align")
                utillityobject.asequal(self, kwargs['text_align'], actual_font, kwargs['msg'] + ". Verification of Cell Text alignment.")
            if 'title_popup' in key:
                obj_table_css=self.driver.find_element_by_css_selector(table_css)
                if Global_variables.browser_name in ['firefox']:
                    utillityobject.click_type_using_pyautogui(self, obj_table_css,move=True, **kwargs)
                else:
                    action = ActionChains(self.driver)
                    action.move_to_element(self.driver.find_element_by_css_selector(table_css)).perform()
                    del action
                time.sleep(1)
                actual_popup=self.driver.find_element_by_css_selector("#IBI_popupHere table").text.strip()
                utillityobject.asequal(self, kwargs['title_popup'], actual_popup, kwargs['msg'] + ". Verification of Title Popup.")

    
    def verify_header_footer_stying(self, header_footer_index, **kwargs):
        '''
        :params : header_footer_index=1,2,3...
        :params : kwargs['bg_color']='cyan'
        :params : kwargs['font_color']='blue'
        :params : kwargs['text_value']='Report Header' or 'Report Footer'...
        :params : kwargs['font_name']='Arial'
        :params : kwargs['font_size']='14pt'
        :params : kwargs['bold']=True
        :params : kwargs['italic']=True
        :params : kwargs['underline']=True
        :params : kwargs['text_align']='left' or 'right' or 'center'
        :Usage  : verify_header_footer_stying(1, bg_color='cyan', font_color='blue', text_value='Report Header', font_name='Arial', font_size='14pt', bold=True, 
                                             italic=True, underline=True, text_align='left')
        '''
        table_css="table[summary='Summary'] td[style^='padding'] td[class^='x']"
        actual_cell_elems=self.driver.find_elements_by_css_selector(table_css)
        if 'bg_color' in kwargs:
            expected_background_color=utillityobject.color_picker(self, kwargs['bg_color'], 'rgba')
            actual_background_color=Color.from_string(actual_cell_elems[header_footer_index-1].value_of_css_property("background-color")).rgba
            utillityobject.asin(self, actual_background_color, expected_background_color , kwargs['msg'] + ". Verification of Cell Background color.")
            time.sleep(1)
        if 'font_color' in kwargs:
            expected_text_color=utillityobject.color_picker(self, kwargs['font_color'], 'rgba')
            actual_text_color=Color.from_string(actual_cell_elems[header_footer_index-1].value_of_css_property("color")).rgba
            utillityobject.asin(self, actual_text_color, expected_text_color, kwargs['msg'] + ". Verification of Cell Text color.")
            time.sleep(1)
        if 'text_value' in kwargs:
            actual_text=actual_cell_elems[header_footer_index-1].text.strip().lower()
            utillityobject.asequal(self, kwargs['text_value'].lower(), actual_text, kwargs['msg'] + ". Verification of Cell Text.")
            time.sleep(1)
        if 'font_name' in kwargs:
            expected_font_name=kwargs['font_name'].lower()
            actual_font_name=actual_cell_elems[header_footer_index-1].value_of_css_property("font-family").strip('"').lower()
            utillityobject.asequal(self, expected_font_name, actual_font_name, kwargs['msg'] + ". Verification of Cell Text Font name.")
            time.sleep(1)
        if 'font_size' in kwargs:
            expected_font_size=(round(1.333333*int(kwargs["font_size"][:-2])))
            actual_size=round(float(actual_cell_elems[header_footer_index-1].value_of_css_property('font-size')[:-2]))
            utillityobject.asequal(self, expected_font_size, actual_size, kwargs['msg'] + ". Verification of Cell Text Font Size.")
            time.sleep(1)
        if 'bold' in kwargs:
            actual_weight=True if actual_cell_elems[header_footer_index-1].value_of_css_property("font-weight") in ['700', 'bold'] else False
            utillityobject.asequal(self, kwargs['bold'], actual_weight, kwargs['msg'] + ". Verification of Cell Text is Bold.")
            time.sleep(1)
        if 'italic' in kwargs:
            actual_style=True if actual_cell_elems[header_footer_index-1].value_of_css_property("font-style")=='italic' else False
            utillityobject.asequal(self, kwargs['italic'], actual_style, kwargs['msg'] + ". Verification of Cell Text is Italics.")
            time.sleep(1)
        if 'underline' in kwargs:
            actual_decoration=True if 'underline' in actual_cell_elems[header_footer_index-1].value_of_css_property("text-decoration") else False
            utillityobject.asequal(self, kwargs['underline'], actual_decoration, kwargs['msg'] + ". Verification of Cell is underlined.")
            time.sleep(1)
        if 'text_align' in kwargs:
            expected_text_align=kwargs['text_align'].lower()
            actual_text_align=actual_cell_elems[header_footer_index-1].value_of_css_property("text-align").strip('"').lower()
            utillityobject.asequal(self, expected_text_align, actual_text_align, kwargs['msg'] + ". Verification of Cell Text alignment.")
            time.sleep(1)
    
    
    def verify_default_amper_value(self, field_name, expected_value, msg):
        """
        Params: field_name: COUNTRY
        Params: expected_value: ENGLAND or All Values
        Usage: verify_default_amper_value('COUNTRY','ENGLAND',"Step 10: Verify Defualt amper values in run window")
        """
        field_css="div[class='autop-amper-ctrl-container'] div[class^='autop-amper']"
        fields=self.driver.find_elements_by_css_selector(field_css)
        reobj=re.match(r'.*:\n(.*)', fields[[re.sub(r':[\s\S]+', '', el.text) for el in fields].index(field_name)].text)
        utillityobject.asequal(self, expected_value, reobj.group(1), msg)
    
    def enter_text_in_amper_value_inputbox(self, field_name, input_text):
        '''
        Syntax: ia_resultobj.enter_text_in_TextBox('Text_1', "This is simple text input")
        '''
        inputbox_css="div[class='autop-amper-ctrl-container'] div[class^='autop-amper'][title^='"+field_name+"'] div[class^='ui-input-text'] input"
        inputbox_obj=utillityobject.validate_and_get_webdriver_object(self, inputbox_css, field_name+"in amper value input box")
        utillityobject.click_on_screen(self, inputbox_obj, coordinate_type='middle')
        utillityobject.click_on_screen(self, inputbox_obj, coordinate_type='middle', click_type=0)
        time.sleep(1)
        pyautogui.typewrite(input_text)
        time.sleep(4)
        
    def select_amper_value(self, field_name, value_list, long_value_set=True, verify_default_radio_button = True, **kwargs):
        """
        Params: field_name: COUNTRY
        Params: value_list=['ENGLAND','ITALY']
        params: long_value_set
        kwargs: All=True or None=True or Search=True applicable only for long_value_set
        kwargs: verify_small_value_list=['ENGLAND','ITALY','JAGUAR'] applicable for small value set
        Syntax:select_amper_value('COUNTRY', ['ENGLAND','ITALY'],False)
        Syntax:select_amper_value('Vendor Name', ['All'],All=True)
        Syntax:select_amper_value('Vendor Name', ['Audiovox','BOSE','Canon'],Search=True)
        @Author: Niranjan
        """
        field_css="div[class='autop-amper-ctrl-container'] div[class^='autop-amper'][title^='"+field_name+"'] [id$='button']"
        self.driver.find_element_by_css_selector(field_css).click()
        time.sleep(5)
        if long_value_set:
            radio_btn_css="div[class*='popup-container'][style*='max-width'] div[class*='ui-radio']"
            radios=self.driver.find_elements_by_css_selector(radio_btn_css)
            if verify_default_radio_button :
                all_value_radio_status=radios[0].find_element_by_css_selector("input")
                actual_all_value_radio_status = all_value_radio_status.get_attribute('checked')
                utillityobject.asequal(self, actual_all_value_radio_status, 'true', 'Step X: All Value button is checked by default')
            radios[1].click()
            time.sleep(1)
            if 'All' in kwargs:
                self.driver.find_element_by_css_selector("#av_btn_sel_all").click()
            if 'None' in kwargs:
                self.driver.find_element_by_css_selector("#av_btn_sel_none").click()
            if 'Search' in kwargs:
                for value in value_list:
                    element = self.driver.find_element_by_css_selector("div[class*='popup-container'][style*='max-width'] input[id='av_search']")
                    element.clear()
                    time.sleep(2)
                    element.click()
                    time.sleep(2)
                    element.send_keys(value)
                    time.sleep(2)
                    values=self.driver.find_elements_by_css_selector("div[class*='popup-container'][style*='max-width'] #av_grp_values .ui-checkbox")
                    values[[el.text for el in values].index(value)].click()
                    time.sleep(1)
            try:
                popup_close_btn_css="div[class*='popup-container'][style*='max-width'] a[data-rel='back']"            
                self.driver.find_element_by_css_selector(popup_close_btn_css).click()
                time.sleep(2)
            except:
                status=True
        else:
            values=self.driver.find_elements_by_css_selector("div[id*='ui-id'][style*='max-width'] li")
            if 'verify_small_value_list' in kwargs:
                print([el.text for el in values])
                utillityobject.asequal(self, kwargs['verify_small_value_list'], [el.text for el in values], "Step X: Verify small value list values at run time")
            for value in value_list:
                values[[el.text for el in values].index(value)].click()
                time.sleep(1)
            try:
                popup_close_btn_css="div[id*='ui-id'][style*='max-width'] div[class^='ui-header'] a[class^='ui-btn']"            
                self.driver.find_element_by_css_selector(popup_close_btn_css).click()
                time.sleep(2)
            except:
                status=True
        time.sleep(1)
        
    def select_amper_menu(self, item_name):
        """
        params: item_name: 'Close' OR 'Reset', OR 'Save' OR 'Run'
        Syntax: select_amper_menu('Run')
        """
        menu_btn_css=".autop-pane div[class^='autop-navbar'] a[title^='"+item_name+"']"
        self.driver.find_element_by_css_selector(menu_btn_css).click()
        time.sleep(2)
    
    
    def verify_autolink(self, table_css, field_value, rownum, colnum, expected_no_of_hlinks, msg, color_name='cerulean_blue_2'):
        """
        table_css = "table[summary='Summary']"
        field_value = "MODEL"
        rownum = 2 #tr = start with 2 [excludes heading]
        colnum = 2   #td = start with 1
        expected_no_of_hlinks = 18
        Usage: verify_autolink("table[summary='Summary']","JAGUAR",4,1,2,"Step 09.2: Verify Auto Drill applied in value V12XKE AUTO")
        """
        link_css=table_css+" > tbody > tr > td:nth-child(" + str(colnum) + ") a[href]"
        table_css=table_css + " > tbody > tr:nth-child(" + str(rownum) + ") > td:nth-child(" + str(colnum) + ") a[href]"
        actual_no_of_hlinks=len(self.driver.find_elements_by_css_selector(link_css))
        print(actual_no_of_hlinks)
        utillityobject.asequal(self, expected_no_of_hlinks, actual_no_of_hlinks, msg + 'a: verify number of hyper-link displayed in a column')
        
        value_href=self.driver.find_element_by_css_selector(table_css).text
        print(value_href)
        utillityobject.asequal(self, field_value, value_href, msg + 'b: verify hyper-link displayed')

        link_style=self.driver.find_element_by_css_selector(table_css).value_of_css_property("text-decoration")
        utillityobject.asin(self, 'underline', link_style, msg + 'c: verify hyper-link style')
        
        link_color=self.driver.find_element_by_css_selector(table_css)
        expected_text_color=utillityobject.color_picker(self, color_name, 'rgba')
        actual_text_color=Color.from_string(link_color.value_of_css_property("color")).rgba
        utillityobject.asin(self, actual_text_color, expected_text_color, msg + 'd: verify hyper-link color')   
    
    def verify_activereport_autolink(self, table_css, colnum, expected_autolink_rows, msg, color_name ='cerulean_blue_2'):
        """
        Desc:- This function is used to verify autolink rows which is present in the active report
        Usage: verify_activereport_autolink("#ITableData0",1, 57, "Step x:Verify autolinks present in the report")
        """
        expected_text_color=utillityobject.color_picker(self, color_name, 'rgba')
        total_rows_css = table_css+" > tbody > tr > td:nth-child(" + str(colnum) + ")"
        total_rows_list = self.driver.find_elements_by_css_selector(total_rows_css)
        act_autolink_rows = 0
        for row in total_rows_list:
            actaul_text_dec = row.value_of_css_property("text-decoration")
            actual_text_color = Color.from_string(row.value_of_css_property("color")).rgba
            if  'underline' in actaul_text_dec and actual_text_color == expected_text_color :
                act_autolink_rows = act_autolink_rows+1
        utillityobject.asequal(self, act_autolink_rows, expected_autolink_rows, msg)
        
    def verify_autolink_tooltip_values(self,parent_id, row, col, expected_tooltip_list, msg, **kwargs):
        """
        :param parent_id = "table[summary='Summary']"
        :param : row = 2
        :param : col = 1
        :param : expected_tooltip_list = ['Drill down to Report', 'Drill Down 2', 'Drill Down 3', 'Auto Links']
        Usage: verify_autolink_tooltip_values("table[summary='Summary']",3,2, expected_tooltip_list, "Step 8: verify the default Autolink tooltip values list")
        """
        cell_css=parent_id + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"
        obj_cell_obj=self.driver.find_element_by_css_selector(cell_css)
        coreutillityobject.left_click(self, obj_cell_obj, action_chain_click=True)
        #utillityobject.default_left_click(self, object_locator=obj_cell_css,action_move=True, **kwargs)
        tooltip_span_css="span[class*='tdgchart-tooltip'][style*='visibility']:not([style*='hidden'])"
        utillityobject.synchronize_with_number_of_element(self, tooltip_span_css, 1, 19)
        tooltip_css="{0}>div>ul>li".format(tooltip_span_css)  
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_list=[]
        for i in range(len(tooltips)):
            tooltip_list.append((tooltips[i].text.strip()).split('\n'))
        actual_tooltip_list=[line for line in reduce(operator.add, tooltip_list) if len(line)>1]
        actual_list = []
        for line in actual_tooltip_list:
            if bool(re.match(r'.*:\s.*', line)):
                reqobj = re.match('(.*):\s{1,}(.*)', line)
                new_element = str(reqobj.group(1)) + ":" + str(reqobj.group(2))
            elif bool(re.match(r'^>', line)):
                new_element=re.sub('>', '', line)
            else:
                new_element=line
            actual_list.append(new_element)
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
    
    def select_autolink_tooltip_menu(self, parent_table_css, row, col, menu_path, msg, **kwargs):
        '''
        parent_table_css= "table[summary='Summary']"
        Usage: select_autolink_tooltip_menu("table[summary='Summary']",2,1,'Drill down to CAR', "Step 09.1: Select the Auto Drill menu Drill down to CAR")
        '''
        move_cursor=parent_table_css
        obj_move_cursor=self.driver.find_element_by_css_selector(move_cursor)
        if Global_variables.browser_name in ['firefox', 'edge', 'ie']:
            #utillityobject.click_type_using_pyautogui(self, obj_move_cursor, move=True, **kwargs)
            coreutillityobject.python_move_to_element(self, obj_move_cursor)
        else:  
            action1 = ActionChains(self.driver)
            action1.move_to_element(obj_move_cursor).perform()
            del action1
        menus=menu_path.split('->')
        cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"
        tooltip_span_css="span[class*='tdgchart-tooltip'][style*='visibility']:not([style*='hidden'])"
        tooltip_css="{0}>div>ul>li".format(tooltip_span_css)  
        obj_cell_obj=self.driver.find_element_by_css_selector(cell_css)
        #self.driver.execute_script("arguments[0].scrollIntoView(true);", obj_cell_css);
        #utillityobject.default_left_click(self,object_locator=obj_cell_css,action_move=True, **kwargs)
        coreutillityobject.left_click(self, obj_cell_obj, action_chain_click=True)
        utillityobject.synchronize_with_number_of_element(self, tooltip_span_css, 1, 19)
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        for i in range(len(tooltips)):
            line1=tooltips[i].text.strip()
            if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                tooltips[i].click()
                break
        tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
        tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
        if len(menus) > 1:
            for i in range(len(tooltips1)):
                if(tooltips1[i].text == menus[1]):
                    utillityobject.asequal(self, True, True, msg)
                if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                    tooltips1[i].click()
                    break
#                     booln = (tooltips1[i].text == menus[1])
#                     print("check")
#                     utillityobject.asequal(self, True, booln, msg)

    def select_autolink_tooltip_menu_using_pyautogui(self, parent_table_css, row, col, menu_path, msg, **kwargs):
        '''
        parent_table_css= "table[summary='Summary']"
        Usage: select_autolink_tooltip_menu_using_pyautogui("table[summary='Summary']",2,1,'Drill down to CAR', "Step 09.1: Select the Auto Drill menu Drill down to CAR", x_offset=15, y_offset=-7, browser_height=80)
        '''
        move_cursor=parent_table_css
        obj_move_cursor=self.driver.find_element_by_css_selector(move_cursor)
        utillityobject.click_type_using_pyautogui(self, obj_move_cursor, move=True, **kwargs)
        time.sleep(4)
        menus=menu_path.split('->')
        cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"
        obj_cell_css=self.driver.find_element_by_css_selector(cell_css)
        utillityobject.click_type_using_pyautogui(self, obj_cell_css, leftClick=True, **kwargs)
        time.sleep(1)
        tooltip_css="span[class*='tdgchart-tooltip']>div>ul>li" 
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        for i in range(len(tooltips)):
            line1=tooltips[i].text.strip()
            print(line1)
            if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                utillityobject.default_left_click(self, object_locator=tooltips[i], **kwargs)
                break
        tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
        tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
        if len(menus) > 1:
            for i in range(len(tooltips1)):
                if(tooltips1[i].text == menus[1]):
                    utillityobject.asequal(self, True, True, msg)
                if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                    utillityobject.default_left_click(self, object_locator=tooltips1[i], **kwargs)
                    break
                
    
    
    def select_report_autolink_tooltip_menu_pyautogui(self,parent_table_css, row, col, menu_path, **kwargs):
        '''
        parent_table_css= "table[summary='Summary']"
        Usage: select_report_autolink_tooltip_menu_pyautogui("table[summary='Summary']",2,1,'Drill down to CAR', "Step 09.1: Select the Auto Drill menu Drill down to CAR", x_offset=15, y_offset=-7, browser_height=80)
        '''
        obj_move_cursor=self.driver.find_element_by_css_selector(parent_table_css)
        utillityobject.click_type_using_pyautogui(self, obj_move_cursor, leftClick=True, **kwargs)
        time.sleep(3)
        cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"   
        obj_cell_css=self.driver.find_element_by_css_selector(cell_css)
        utillityobject.click_type_using_pyautogui(self, obj_cell_css, leftClick=True, cord_type='right', x_offset=-6, **kwargs)
        time.sleep(1)
        tooltip_css="span[class*='tdgchart-tooltip']>div>ul>li"
        menus=menu_path.split('->')
        if len(menus)>1:
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    #utillityobject.click_type_using_pyautogui(self, tooltips[i], leftClick=True, cord_type='right', x_offset=-10, **kwargs)
                    utillityobject.click_type_using_pyautogui(self, tooltips[i], move=True, cord_type='left', x_offset=10, **kwargs)
                    break
            tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips1)):
                if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                    utillityobject.click_type_using_pyautogui(self, tooltips[i], leftClick=True, **kwargs)
                    break
        else:
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    #utillityobject.click_type_using_pyautogui(self, tooltips[i], leftClick=True, cord_type='right', x_offset=-10, **kwargs)
                    utillityobject.click_type_using_pyautogui(self, tooltips[i], leftClick=True, cord_type='left', x_offset=10, **kwargs)
                    break
    
    def select_report_autolink_tooltip_menu(self,parent_table_css, row, col, menu_path, msg, **kwargs):
        '''
        parent_table_css= "table[summary='Summary']"
        Usage: select_autolink_tooltip_menu_using_pyautogui("table[summary='Summary']",2,1,'Drill down to CAR', "Step 09.1: Select the Auto Drill menu Drill down to CAR", x_offset=15, y_offset=-7, browser_height=80)
        '''
        obj_move_cursor=self.driver.find_element_by_css_selector(parent_table_css)
        if Global_variables.browser_name in ['firefox']:
            # Move the cursor to top left location
            utillityobject.click_type_using_pyautogui(self, obj_move_cursor, move=True, **kwargs)
            time.sleep(2)
            #moves the cursor to the specified cell number and click on it
            cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"
            tooltip_css="span[class*='tdgchart-tooltip']>div>ul>li"   
            obj_cell_css=self.driver.find_element_by_css_selector(cell_css)
            utillityobject.default_left_click(self,object_locator=obj_cell_css,action_move=True, **kwargs)
            time.sleep(1)
            menus=menu_path.split('->')
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    utillityobject.click_type_using_pyautogui(self, tooltips[i], leftClick=True, **kwargs)
                    break
            if len(menus)>1:
                tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
                tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
                tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
                for i in range(len(tooltips1)):
                    if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                        utillityobject.click_type_using_pyautogui(self, tooltips[i], leftClick=True, **kwargs)
                        break
            time.sleep(4)
        elif Global_variables.browser_name in ['ie', 'edge']:
            # Move the cursor to top left location
            action1 = ActionChains(self.driver)
            action1.move_to_element(obj_move_cursor).perform()
            del action1
            time.sleep(2)
            #moves the cursor to the specified cell number and click on it
            cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"
            tooltip_css="span[class*='tdgchart-tooltip']>div>ul>li"   
            obj_cell_css=self.driver.find_element_by_css_selector(cell_css)
            action1=ActionChains(self.driver)
            action1.move_to_element_with_offset(obj_cell_css,xoffset=15,yoffset=8).click().perform()
            del action1
            #utillityobject.default_left_click(self,object_locator=obj_cell_css,action_move=True, **kwargs)
            time.sleep(1)
            menus=menu_path.split('->')
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    action1=ActionChains(self.driver)
                    action1.move_to_element_with_offset(tooltips[i],xoffset=25,yoffset=8).click().perform()
                    del action1
                    break
            if len(menus)>1:
                tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
                tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
                time.sleep(1)
                tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
                for i in range(len(tooltips1)):
                    if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                        action1=ActionChains(self.driver)
                        action1.move_to_element_with_offset(tooltips1[i],xoffset=25,yoffset=8).click().perform()
                        del action1
                        break
            time.sleep(4)
        else:
            # Move the cursor to top left location
            action1 = ActionChains(self.driver)
            action1.move_to_element(obj_move_cursor).perform()
            del action1
            time.sleep(2)
            #moves the cursor to the specified cell number and click on it
            cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"
            tooltip_css="span[class*='tdgchart-tooltip']>div>ul>li"   
            obj_cell_css=self.driver.find_element_by_css_selector(cell_css)
            action1=ActionChains(self.driver)
            action1.move_to_element_with_offset(obj_cell_css,xoffset=15,yoffset=8).click().perform()
            del action1
            #utillityobject.default_left_click(self,object_locator=obj_cell_css,action_move=True, **kwargs)
            time.sleep(1)
            menus=menu_path.split('->')
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    #tooltips[i].click()
                    action1=ActionChains(self.driver)
                    action1.move_to_element_with_offset(tooltips[i],xoffset=25,yoffset=8).click().perform()
                    del action1
                    break
            time.sleep(1)
            if len(menus)>1:
                tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
                tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
                tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
                for i in range(len(tooltips1)):
                    if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                        tooltips1[i].click()
                        break
            time.sleep(4) 
    
    def verify_report_autolink_tooltip_submenu(self,parent_table_css, row, col, menu_path, expected_tooltip_list, msg, **kwargs):
        """
        param parent_id = 'MAINTABLE_wbody1'         
        param : raiser_class = 'riser!s4!g4!mbar!'
        param : menu_path = 'Drill up to'
        param : expected_tooltip_list = ['Store Business Sub Region', 'Sale Year/Quarter'] ==> tootip submenu  
        Usage: verify_autolink_tooltip_submenu('MAINTABLE_wbody0','riser!s5!g0!mbar!r0!c0!', "Drill down to", a, "Step 05a: Verify the drill menu shows for only Region and Year.")
        Author: Niranjan
        """
        obj_move_cursor=self.driver.find_element_by_css_selector(parent_table_css)
        utillityobject.click_type_using_pyautogui(self, obj_move_cursor, leftClick=True, **kwargs)
        time.sleep(3)
        cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"   
        obj_cell_css=self.driver.find_element_by_css_selector(cell_css)
        utillityobject.click_type_using_pyautogui(self, obj_cell_css, leftClick=True, **kwargs)
        time.sleep(1)
        tooltip_css="span[class*='tdgchart-tooltip']>div>ul>li"
        menus=menu_path.split('->')
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        for i in range(len(tooltips)):
            line1=tooltips[i].text.strip()
            if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                coreutillityobject.left_click(self, tooltips[i])
                break
        actual_list = []
        tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        if Global_variables.browser_name in ('ie', 'firefox'):
            time.sleep(1)
        tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
        for i in range(len(tooltips1)):
            if tooltips1[i].text.strip()!='':
                actual_list.append(tooltips1[i].text.strip())
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
        time.sleep(4) 
    
    def verify_autolink_tooltip_values_usng_pyautogui(self,parent_id, row, col, expected_tooltip_list, msg, **kwargs):
        """
        :param parent_id = "table[summary='Summary']"
        :param : row = 2
        :param : col = 1
        :param : expected_tooltip_list = ['Drill down to Report', 'Drill Down 2', 'Drill Down 3', 'Auto Links']
        Usage: verify_autolink_tooltip_values("table[summary='Summary']",3,2, expected_tooltip_list, "Step 8: verify the default Autolink tooltip values list")
        """
        parent_table_css=parent_id  
        time.sleep(2)
        obj_move_cursor=self.driver.find_element_by_css_selector(parent_table_css)
        utillityobject.click_type_using_pyautogui(self, obj_move_cursor, leftClick=True, **kwargs)
        time.sleep(3)
        cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"   
        obj_cell_css=self.driver.find_element_by_css_selector(cell_css)
        utillityobject.click_type_using_pyautogui(self, obj_cell_css, leftClick=True, **kwargs)
        time.sleep(2)
        tooltip_css="span[class*='tdgchart-tooltip']>div>ul>li"  
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_list=[]
        for i in range(len(tooltips)):
            tooltip_list.append((tooltips[i].text.strip()).split('\n'))
        actual_tooltip_list=[line for line in reduce(operator.add, tooltip_list) if len(line)>1]
        actual_list = []
        for line in actual_tooltip_list:
            if bool(re.match(r'.*:\s.*', line)):
                reqobj = re.match('(.*):\s{1,}(.*)', line)
                new_element = str(reqobj.group(1)) + ":" + str(reqobj.group(2))
            elif bool(re.match(r'^>', line)):
                new_element=re.sub('>', '', line)
            else:
                new_element=line
            actual_list.append(new_element)
        print(actual_list)
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
    
    def verify_header_footer_doc_cell(self,parent_id,cell_no, **kwargs):
        """
        verify_header_footer_doc_cell('ITableData0', 1, font_color='magenta', msg='Step 12.1: Report Header'
        """
        table_css="#" + parent_id + " tr td[id^='THEAD'] span"
        cell_obj=self.driver.find_elements_by_css_selector(table_css)
        if 'font_color' in kwargs:
            expected_font_color=utillityobject.color_picker(self,kwargs['font_color'], 'rgba')
            actual_font_color=Color.from_string(cell_obj[cell_no-1].value_of_css_property("color")).rgba
            utillityobject.asequal(self,actual_font_color, expected_font_color , kwargs['msg'] + ". Verification of Font color")
        if 'bg_color' in kwargs:
            table_css="#" + parent_id + " tr td[id^='THEAD']"
            cell_obj=self.driver.find_elements_by_css_selector(table_css)
            expected_font_color=utillityobject.color_picker(self,kwargs['bg_color'], 'rgba')
            actual_font_color=Color.from_string(cell_obj[cell_no-1].get_attribute("bgcolor")).rgba
            print(actual_font_color,expected_font_color)
            utillityobject.asequal(self,actual_font_color, expected_font_color , kwargs['msg'] + ". Verification of Background color")

    def verify_header_footer_report_cell(self,parent_id,cell_no, **kwargs):
        """
        verify_header_footer_doc_cell('table[summary='Summary']', 1, font_color='magenta', msg='Step 12.1: Report Header'
        """
        table_css=parent_id + " td td"
        cell_obj=self.driver.find_elements_by_css_selector(table_css)
        if 'font_color' in kwargs:
            expected_font_color=utillityobject.color_picker(self,kwargs['font_color'], 'rgba')
            actual_font_color=Color.from_string(cell_obj[cell_no-1].value_of_css_property("color")).rgba
            utillityobject.asequal(self,actual_font_color, expected_font_color , kwargs['msg'] + ". Verification of Font color")
        if 'bg_color' in kwargs:
            expected_font_color=utillityobject.color_picker(self,kwargs['bg_color'], 'rgba')
            actual_font_color=Color.from_string(cell_obj[cell_no-1].value_of_css_property("background-color")).rgba
            print(actual_font_color,expected_font_color)
            utillityobject.asequal(self,actual_font_color, expected_font_color , kwargs['msg'] + ". Verification of Background color")

    def verify_report_table_margin(self, table_css, **kwargs):
        """
        Params: table_css = "table[summary='Summary']"
        Params: margin=True     
        Params: expected_table_elem='72pt',
        Params: margin_left='72pt', 
        Params: margin_right='72pt' 
        Params: margin_top='72pt' 
        Params: margin_bottom='72pt'
        Usage: verify_report_table_margin(table_elem, margin_left='72pt',margin_right='72pt', margin_top='72pt', margin_bottom='72pt')
                            or
        Usage: verify_report_table_margin(table_elem, margin=True, expected_table_elem="96px")
        """
        if 'margin' in kwargs:
            actual_table_elem=self.driver.find_element_by_css_selector(table_css).value_of_css_property('margin')
            utillityobject.asequal(self, kwargs['expected_table_elem'], actual_table_elem, 'Step X: Verify the following report (with the proper margin) is displayed. ')
            time.sleep(1)
        elif 'style_margin' in kwargs:
            actual_table_elem=self.driver.find_element_by_css_selector(table_css).get_attribute("style")
            utillityobject.asequal(self, kwargs['expected_table_elem'], actual_table_elem, 'Step X: Verify the following report (with the proper margin) is displayed. ')
            time.sleep(1)
        else:
            table_elem=self.driver.find_element_by_css_selector(table_css)
            margins=table_elem.get_attribute("style").strip().split(";")
            margin_dict={}
            for margin in margins:
                ob=re.match(r'(.*):(.*)', margin)
                margin_dict[ob.group(1)]=ob.group(2)
            if 'margin_left' in kwargs:
                utillityobject.asequal(self, kwargs['margin-left'], margin_dict['margin-left'], "Step X: Verification of Left Margin.")
            if 'margin_right' in kwargs:
                utillityobject.asequal(self, kwargs['margin-right'], margin_dict['margin-right'], "Step X: Verification of Right Margin.")
            if 'margin_top' in kwargs:
                utillityobject.asequal(self, kwargs['margin-top'], margin_dict['margin-top'], "Step X: Verification of Top Margin.")
            if 'margin_bottom' in kwargs:
                utillityobject.asequal(self, kwargs['margin-bottom'], margin_dict['margin-bottom'], "Step X: Verification of Bottom Margin.")
    
    
    def expand_toc(self, folder_path):
        """
        :Usage - 
        folder_path="EMEA"
        iarun_obj.expand_toc('EMEA')
        """ 
        folder_list=folder_path.split('->')
        #folder_list=folder_path
        for folder_name in folder_list:
            toc_items=self.driver.find_elements_by_css_selector("#divtocDHTML table")
            for item in toc_items:
                if item.text.strip()==folder_name:    
                    item.find_element_by_css_selector("img[id^='node']").click()
                    time.sleep(2)
                    break
                
    def select_toc_item(self, item_path):
        """
        :Usage - 
        item_path="EMEA->Stereo Systems"
        iarun_obj.select_toc_item('EMEA->Stereo Systems')
        """
        first_item=item_path.split('->')[:-1]
        folder_path=''.join(first_item)
        self.expand_toc(folder_path)
        item_name=item_path.split('->')[-1]
        toc_items=self.driver.find_elements_by_css_selector("#divtocDHTML table")
        for item in toc_items:
            if item.text.strip()==item_name:
                item.find_element_by_css_selector("img[id^='item']").click()
                time.sleep(2)
                break
            
    def verify_toc_item(self, item_list, msg):
        toc_items=self.driver.find_elements_by_css_selector("#divtocDHTML table")
        my_iter=(i.text.strip() for i in toc_items)
        for item in toc_items:
            if item_list[0]==next(my_iter):
                break
        status=True
        for item in item_list[1:]:
            if item==next(my_iter):
                status=True
            else:
                status=False
                break
        del my_iter
        utillityobject.asequal(self, status, True, msg)
    
    def verify_report_title_popup(self,table_css,title_column_no,expected_popup,msg):
        title_css=table_css+">tbody>tr:nth-child(1)>td:nth-child("+str(title_column_no)+")"
        title=self.driver.find_element_by_css_selector(title_css)
        if Global_variables.browser_name=='ie':
            action1 = ActionChains(self.driver)
            action1.move_to_element(title).perform()
            del action1
        else:
            utillityobject.click_on_screen(self,title,'middle')
        time.sleep(3)
        actual_popup=self.driver.find_element_by_css_selector("#IBI_popupHere table").text.strip()
        utillityobject.asequal(self,actual_popup,expected_popup,msg)
        print(actual_popup)
        
    def verify_active_document_page_layout_menu(self,table_css,expected_list,msg):
        """
        table_css="table[id='iLayTB$']"
        expected_list=['Layouts','Page 1','Page 2']
        Usage: verify_document_page_layout_menu("table[id='iLayTB$']",['Layouts','Page 1','Page 2'], "Step02: Verify")
        """
        popup_css="form[name='mergeform'] "+str(table_css)
        bipopups=self.driver.find_elements_by_css_selector(popup_css)
        menu_items=bipopups[len(bipopups)-1].find_elements_by_css_selector(" tr td")
        actual_list=[el.text.strip() for el in menu_items]
        utillityobject.asequal(self,actual_list,expected_list,msg)
    
    def select_active_document_page_layout_menu(self, value_to_select):
        """
        :param value_to_select='Regional Report'
        :Usage select_active_document_page_layout_menu('Regional Report')
        """
        popup_css="form[name='mergeform'] table[id^='iLayTB'] [id^='iLay']"
        bipopups=self.driver.find_elements_by_css_selector(popup_css)
        menu_item_index=bipopups[[el.text.strip() for el in bipopups].index(str(value_to_select))]
        utillityobject.click_on_screen(self, menu_item_index, 'middle', click_type=0)  
        
    '''    Old function can be deleted on the next commit/clean up 
    def select_report_autolink_tooltip(self, parent_table_css, row, col, tooltip_path, verify_tooltip=None, msg=None, **kwargs):
        """
        Example Usage : iarun.select_report_autolink_tooltip("table[summary= 'Summary']", 6,2, "Auto Links->AUTOLINK_TARGET01")
        """
        obj_move_cursor=self.driver.find_element_by_css_selector(parent_table_css)
        utillityobject.click_type_using_pyautogui(self, obj_move_cursor, leftClick=True, cord_type='start', **kwargs)
        cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"   
        obj_cell_css=self.driver.find_element_by_css_selector(cell_css)
        utillityobject.click_type_using_pyautogui(self, obj_cell_css, leftClick=True, cord_type='right', x_offset=-6, **kwargs)
        utillityobject.synchronize_with_number_of_element(self, "[id='my_tooltip_id'][style*='visible']", 1, expire_time=4)
        tooltips=tooltip_path.split('->')
        tooltip_css="#my_tooltip_id ul[class='tdgchart-tooltip-list']>li[class*='tdgchart-tooltip-pad']"
        tooltip_obj=self.driver.find_element_by_css_selector(tooltip_css)
        utillityobject.click_on_screen(self, tooltip_obj, 'top_middle', y_offset=5)
        for count, tooltip in enumerate(tooltips) :
            tooltip_elements=self.driver.find_elements_by_css_selector(tooltip_css)
            tooltip_text_list=[tooltip.text.strip().replace('>', '').replace('\n', '') for tooltip in tooltip_elements]
            tooltip_element=tooltip_elements[tooltip_text_list.index(tooltip)]
            if tooltips[count] == tooltips[-1] :
                if verify_tooltip!=None :
                    utillityobject.asequal(self, verify_tooltip, tooltip_text_list, msg)
                utillityobject.click_on_screen(self, tooltip_element, coordinate_type='middle', click_type=0)
            else :
                utillityobject.click_on_screen(self, tooltip_element, coordinate_type='left', x_offset=2)
                tooltip_arrow=tooltip_element.find_element_by_css_selector("span[class='tdgchart-tooltip-arrow']")
                utillityobject.click_on_screen(self, tooltip_arrow, coordinate_type='middle')
                time.sleep(2)'''
    def select_report_autolink_tooltip(self, parent_table_css, row, col, tooltip_path, verify_tooltip=None, msg=None, **kwargs):
        """
        Example Usage : iarun.select_report_autolink_tooltip("table[summary= 'Summary']", 6,2, "Auto Links->AUTOLINK_TARGET01")
        """
        obj_move_cursor=self.driver.find_element_by_css_selector(parent_table_css)
        utillityobject.click_type_using_pyautogui(self, obj_move_cursor, leftClick=True, cord_type='start', **kwargs)
        cell_css=parent_table_css + " > tbody > tr:nth-child(" + str(row) + ") > td:nth-child(" + str(col) + ") > a"   
        obj_cell_css=self.driver.find_element_by_css_selector(cell_css)
        utillityobject.click_type_using_pyautogui(self, obj_cell_css, leftClick=True, cord_type='top_middle', x_offset=6, y_offset=6, **kwargs)
        utillityobject.synchronize_with_number_of_element(self, "[id='my_tooltip_id'][style*='visible']", 1, expire_time=4)
        tooltips=tooltip_path.split('->')
        tooltip_css="#my_tooltip_id ul[class='tdgchart-tooltip-list']>li[class*='tdgchart-tooltip-pad']"
        tooltip_obj=self.driver.find_element_by_css_selector(tooltip_css)
        if Global_variables.browser_name in ['edge'] : # uisoup Mouse move not working for autolink tooltip in Edge browser. Workaround pyautogui implemented 
            coord = coreutillityobject.get_web_element_coordinate(self, tooltip_obj, element_location='top_middle', yoffset=5)
            pyautogui.moveTo(coord['x'], coord['y'])
        else :
            utillityobject.click_on_screen(self, tooltip_obj, 'top_middle', y_offset=5, mouse_duration=0)
        for count, tooltip in enumerate(tooltips) :
            tooltip_elements=self.driver.find_elements_by_css_selector(tooltip_css)
            tooltip_text_list=[tooltip.text.strip().replace('>', '').replace('\n', '') for tooltip in tooltip_elements]
            tooltip_element=tooltip_elements[tooltip_text_list.index(tooltip)]
            if tooltips[count] == tooltips[-1] :
                if verify_tooltip!=None :
                    if kwargs['verify_type']=='asin':
                        tooltip_css="#my_tooltip_id ul[class='tdgchart-tooltip-list']>li[class*='tdgchart-tooltip-pad'] span" 
                        tooltip_elements=self.driver.find_elements_by_css_selector(tooltip_css)
                        actual_tooltip_list=[tooltip.text.strip() for tooltip in tooltip_elements]
                        for item in tooltips:
                            if item in actual_tooltip_list:
                                status=True
                            else:
                                status=False
                                break
                        utillityobject.asequal(self, True, status, msg+"- Tooltips items in list") 
                    else:
                        utillityobject.asequal(self, verify_tooltip, tooltip_text_list, msg)
                utillityobject.click_on_screen(self, tooltip_element, coordinate_type='middle', click_type=0)
            else :
                utillityobject.click_on_screen(self, tooltip_element, coordinate_type='left', x_offset=2)
                tooltip_arrow=tooltip_element.find_element_by_css_selector("span[class='tdgchart-tooltip-arrow']")
                utillityobject.click_on_screen(self, tooltip_arrow, coordinate_type='middle')
                time.sleep(2)
        time.sleep(10)
    
    def switch_to_ia_run_window(self,wndnum, pause=15, **kwargs):
        """
        Usage: switch_to_window(1)
        """
        time.sleep(5)     
        utillityobject.switch_to_window_handle_test(self, wndnum, **kwargs)
        time.sleep(pause)
        self.driver.maximize_window()
        window_widht = self.driver.execute_script("return window.innerWidth|| document.documentElement.clientWidth|| document.body.clientWidth || document.body.scrollWidth;")
        if window_widht < self.driver.execute_script("return screen.availWidth;"):
            self.driver.maximize_window()
        time.sleep(pause/2)
        br=utillityobject.get_browser_height(self)
        innerHeight = self.driver.execute_script("return window.innerHeight|| document.documentElement.clientHeight|| document.body.clientHeight;")
        availHeight = self.driver.execute_script("return screen.availHeight")
        get_outer_height = 0
        if Global_variables.browser_name in ['ie', 'edge'] and wndnum > 0:
            get_outer_height = br['browser_height'] - br['outer_height']
        browser_height = availHeight - innerHeight
        Global_variables.current_working_area_browser_x = br['browser_width']
        Global_variables.current_working_area_browser_y = browser_height - get_outer_height
        
    def verify_table_cell_property_(self, rownum, colnum, table_css=None, **kwargs):#a is added in the table_css for report samples verification. That is the only change in this compared to the old one.
        """
        Params: hyperlink_css = "table[summary='Summary'] > tbody > tr:nth-child(4) > td:nth-child(2) a" to check underline in a tag
        cell_no = 1, 2, 3...any integer value for cell number, including heading cells.
        font_color = 'black'
        bg_color='red'
        text='CAR'
        font_name='Arial'
        font_size='10pt'
        bold=True
        italics=True
        underline=True
        text_align='left', OR, 'center' Or 'right'
        Usage: verify_table_cell_property("table[summary='Summary']", 1, bg_color='Cyan', font_color='magenta', text_value='CAR', font_name='Arial', font_size='10pt', bold=True, italic=True, underline=True, text_align='Center')
        """
        if table_css==None:
            table_css="table[summary='Summary'] > tbody > tr:nth-child(" + str(rownum) + ") > td:nth-child(" + str(colnum) + ") a"
        else:
            table_css=table_css
        cell_obj=self.driver.find_element_by_css_selector(table_css)
        for key in kwargs:
            if 'bg_color' in key:
                expected_background_color=utillityobject.color_picker(self, kwargs['bg_color'], 'rgba')
                actual_background_color=Color.from_string(cell_obj.value_of_css_property("background-color")).rgba
                utillityobject.asin(self, actual_background_color, expected_background_color , kwargs['msg'] + ". Verification of Cell Background color.")
            if 'font_color' in key:
                expected_text_color=utillityobject.color_picker(self, kwargs['font_color'], 'rgba')
                actual_text_color=Color.from_string(cell_obj.value_of_css_property("color")).rgba
                utillityobject.asin(self, actual_text_color, expected_text_color, kwargs['msg'] + ". Verification of Cell Text color.")
            if 'text_value' in key:
                actual_text=cell_obj.text.strip()
                utillityobject.asequal(self, kwargs['text_value'], actual_text, kwargs['msg'] + ". Verification of Cell Text.")
            if 'font_name' in key:
                actual_font=cell_obj.value_of_css_property("font-family").strip('"').upper()
                utillityobject.asequal(self, kwargs['font_name'].upper(), actual_font, kwargs['msg'] + ". Verification of Cell Text Font name.")
            if 'font_size' in key:
                expected_font_size=round(float(1.333333*int(kwargs['font_size'][:-2])))
                actual_font_size=cell_obj.value_of_css_property("font-size")
                actual_font_size=round(float(actual_font_size[:-2]))
                utillityobject.asequal(self, expected_font_size, actual_font_size, kwargs['msg'] + ". Verification of Cell Text Font Size.")
            if 'bold' in key:
                actual_weight=True if cell_obj.value_of_css_property("font-weight") in ['700', 'bold'] else False
                utillityobject.asequal(self, kwargs['bold'], actual_weight, kwargs['msg'] + ". Verification of Cell Text is Bold.")
            if 'italic' in key:
                actual_style=True if cell_obj.value_of_css_property("font-style")=='italic' else False
                utillityobject.asequal(self, kwargs['italic'], actual_style, kwargs['msg'] + ". Verification of Cell Text is Italics.")
            if 'underline' in key:
                actual_decoration=True if 'underline' in cell_obj.value_of_css_property("text-decoration") else False
                utillityobject.asequal(self, kwargs['underline'], actual_decoration, kwargs['msg'] + ". Verification of Cell is underlined.")
            if 'text_align' in key:
                actual_font=cell_obj.value_of_css_property("text-align")
                utillityobject.asequal(self, kwargs['text_align'], actual_font, kwargs['msg'] + ". Verification of Cell Text alignment.")
            if 'title_popup' in key:
                obj_table_css=self.driver.find_element_by_css_selector(table_css)
                if Global_variables.browser_name in ['firefox', 'ie', 'edge']:
                    coreutillityobject.python_move_to_element(self, obj_table_css)  
                else:
                    action = ActionChains(self.driver)
                    action.move_to_element(obj_table_css).perform()
                    del action
                time.sleep(1)
                actual_popup=self.driver.find_element_by_css_selector("#IBI_popupHere table").text.strip()
                utillityobject.asequal(self, kwargs['title_popup'], actual_popup, kwargs['msg'] + ". Verification of Title Popup.")
                
    def verify_visualize_bar_added_in_htmlreport(self, column_name, visualize_bar_color, expected_total_visual_bars, msg, table_css="table[summary='Summary']", column_position=1):
        '''
        Desc:-Verify visualization bar is added in the report column
        column_name : if column is wrapped into two lines then user should pass the fullname without space: example:Product Category:Syntax : ProductCategory.
        ia_runobj.verify_visualize_bar_added_in_htmlreport('Gross Profit', 'light_gray', 7, "Step 03:05: Verify visualization bar are added in the report")
        report_obj.verify_visualize_bar_added_in_htmlreport('AVEMargin', 'light_gray', 7, "Step 03:03: Verify visualization bar are added in the report")[AVE Margin is wrapped into two row]
        '''
        error_msg = "The given [{0}] column is not available in the report column title".format(column_name)
        total_visualize=0
        table_rows_css=table_css + ">tbody>tr"
        table_columns_title_css = table_rows_css + ':nth-child(1)>td'
        table_row_objs = self.driver.find_elements_by_css_selector(table_rows_css)
        columns_title_obj = self.driver.find_elements_by_css_selector(table_columns_title_css)
        column_name_index = [index for index in range(len(columns_title_obj)) if columns_title_obj[index].text.strip().replace('\n', '') == column_name]
        if column_name_index == [] :
            raise KeyError(error_msg)
        column_index = column_name_index[column_position-1]
        expected_visualize_color=utillityobject.color_picker(self, visualize_bar_color, 'rgba')
        for row in table_row_objs :
            column_visual_css ="td:nth-child({0}) + td>table table td".format(column_index+1)
            visualize_obj=row.find_elements_by_css_selector(column_visual_css)
            if len(visualize_obj) != 0 :
                actaul_visualize_color=Color.from_string(visualize_obj[0].value_of_css_property('background-color')).rgba
                if actaul_visualize_color==expected_visualize_color :
                    total_visualize+=1
        utillityobject.asequal(self, expected_total_visual_bars, total_visualize, msg)
        
    def expand_colapse(self, node_path, expand_colapse_toggle):
        tr_css = 'collapsed' if expand_colapse_toggle == 'expand' else 'expanded'
        node_list = node_path.split('->') if expand_colapse_toggle == 'expand' else node_path.split('->')[::-1]
        for node in node_list:
            all_colapsed_rows=self.driver.find_elements_by_css_selector("#divTreeTable > table > tbody > tr[class='branch {0}']".format(tr_css))
            for colapsed_row in all_colapsed_rows:
                node_elem=colapsed_row.find_element_by_css_selector('td:nth-child(1)')
                if node_elem.text.strip() == node:
                    plus_icon=node_elem.find_element_by_css_selector('a')
                    coreutillityobject.left_click(self, plus_icon)
                    
    def verify_column_heading(self, table_css, expected_title, msg):
        """
        table_css="#divTreeTable #treetable thead"
        expected_list="GrossProfitDiscountMSRPCOGSQty"
        msg="Step x:Verify report heading"
        self.verify_column_heading(table_css, expected_list, msg)
        """
        try:
            act_title_elem=self.driver.find_element_by_css_selector(table_css).text.strip()
            actual_title=act_title_elem.replace('\n', ' ')
            utillityobject.asequal(self, expected_title, actual_title, msg)
        except NoSuchElementException:
            raise LookupError("Verify column heading is not present in the table {0}".format(table_css))
        
    def verify_slider_labels(self, parent_css, expected_label_list, msg):
        """
        parent_css="#jschartHOLD_0"
        expected_label_list=['2013', '2014', '2015', '2016', '2017', '2018']
        msg="Step X: Verify labels
        IA_Run.verify_slider_labels("#jschart_HOLD_0", expected_label_list, "Step X: Verify slider labels")
        """
        css=parent_css+" [class^='slider-labels']"
        try :
            actual_label_elems=self.driver.find_elements_by_css_selector(css)
            actual_label_list=[elem.text.strip() for elem in actual_label_elems]
            utillityobject.asequal(self, expected_label_list, actual_label_list, msg)
        except NoSuchElementException():
            raise LookupError("Verify slider labels not present in the webpage{0}".format(css))
        
    def move_chart_slider(self, slider_value_to_select, parent_css="#jschart_HOLD_0"):
        """
        Descriptions : This method used to move slider to specific value by drag and drop slider marker.
        :arg - slider_marker : If you want to select slider value by using slider marker 1 (left side) then pass slider_marker=1 else slider_marker=2 (right side)
        example usage : move_filter_slider('Move Slider to 5011', 5020, 2)
        """
        select_value=int(slider_value_to_select)
        slider_marker_css = parent_css + " [class='sliderHandle']"
        slider_lables = self.driver.find_elements_by_css_selector(parent_css +" [class^='slider-labels']")
        slider_min_val = int(slider_lables[0].text.strip())
        slider_max_val = int(slider_lables[-1].text.strip())
        if select_value in range(slider_min_val, slider_max_val+1) :
            slider_body = self.driver.find_element_by_css_selector(parent_css + " [class='sliderBody']")
            slider_body_width = slider_body.size['width']
            slider_body_xoffset = coreutillityobject.get_web_element_coordinate(self, slider_body, element_location='middle_left')['x']
            slider_range = slider_max_val - slider_min_val
            slider_value_distance = slider_body_width / slider_range
            slider_value_range = select_value - slider_min_val
            slider_value_xoffset = slider_value_distance * slider_value_range
            slider_marker = self.driver.find_element_by_css_selector(slider_marker_css)
            slider_marker_mid_loc = coreutillityobject.get_web_element_coordinate(self, slider_marker)
            source_x = slider_marker_mid_loc['x']
            source_y = slider_marker_mid_loc['y']
            target_x = slider_body_xoffset + slider_value_xoffset
            target_y = source_y
            uisoup.mouse.move(source_x,source_y)
            time.sleep(1)
            uisoup.mouse.drag(source_x,source_y, target_x, target_y)
        else :
            error_msg = "Slider select value {0} is not in range({1}, {2})".format(select_value, slider_min_val, slider_max_val)
            raise ValueError(error_msg)
        
    def verify_added_text_in_textbox(self, expected_value_list, msg):
        """
        This function is to verify text added in the document page in IA tool
        self.verify_added_text_in_textbox(['Create a Chart Multi-page Dashboard Page 1'], "Step x: Verify text is added in the document page 1")
        """
        css="div[id*='LOBJText']"
        try :
            actual_elem=self.driver.find_elements_by_css_selector(css)
            actual_value_list = [t.text.strip() for t in actual_elem if t.text.strip()!='']
        except NoSuchElementException:
            raise LookupError("The actual text element is not available in the page{0}".format(css))
        utillityobject.asequal(self, expected_value_list, actual_value_list, msg)
     
    def verify_chart_title(self, chart_title_keyname, expected_title, msg="Step X : ", project_object=None):
        '''
        Desc: This function will verify chart title
        Usage: verify_chart_title("Current Month - Daily Sales TrendFor All MODEL", 'run', "Step05: Verify chart title in run window")
        '''
        chart_title_css=InfoassistLocators.__dict__[chart_title_keyname]
        element_object = utillityobject.validate_and_get_webdriver_object(self, chart_title_css[1], chart_title_keyname, project_object)
        actual_title_elem=element_object.text.replace('\n','')
        utillityobject.asequal(self, expected_title, actual_title_elem, msg)
        
    def select_month_in_calendardatepicker_dialog(self, month_to_select):
        """
        :param month_to_select='Mar'
        :Usage select_month_in_calendardatepicker_dialog('Mar')
        """
        month_dropdown_css="div.ui-datepicker .ui-datepicker-header .ui-datepicker-month"
        month_dropdown_obj=utillityobject.validate_and_get_webdriver_object(self, month_dropdown_css, 'Month dropdown')
        month_dropdown_obj.click()
        month_value_css="div.ui-datepicker .ui-datepicker-header .ui-datepicker-month option"
        month_values=utillityobject.validate_and_get_webdriver_objects(self, month_value_css, 'Month values')
        month_value=month_values[[elem.text.strip() for elem in month_values].index(month_to_select)]
        month_value.click()
        
    def select_year_in_calendardatepicker_dialog(self, year_to_select):
        """
        :param year_to_select='2015'
        :Usage select_year_in_calendardatepicker_dialog('2015')
        """
        year_dropdown_css="div.ui-datepicker .ui-datepicker-header .ui-datepicker-year"
        year_dropdown_obj=utillityobject.validate_and_get_webdriver_object(self, year_dropdown_css, 'Year dropdown')
        year_dropdown_obj.click()
        year_value_css="div.ui-datepicker .ui-datepicker-header .ui-datepicker-year option"
        year_values=utillityobject.validate_and_get_webdriver_objects(self, year_value_css, 'Year values')
        year_value=year_values[[elem.text.strip() for elem in year_values].index(year_to_select)]
        year_value.click()    
        
    def select_date_in_calendardatepicker_dialog(self, date_to_select):
        """
        :param date_to_select='15'
        :Usage select_date_in_calendardatepicker_dialog('15')
        """
        date_value_css="div.ui-datepicker table.ui-datepicker-calendar td"
        date_values=utillityobject.validate_and_get_webdriver_objects(self, date_value_css, 'Date values')
        date_value=date_values[[elem.text.strip() for elem in date_values].index(date_to_select)]
        date_value.click()
        
    """ ***********************************************************************AUTO PROMPT********************************************************************** """
        
    def select_field_filter_values_dropdown_in_auto_prompt(self, field_name):
        '''
        This function is used to click on  filter values drop down
        field_name='Business Sub region' or 'COUNTRY'.
        '''
        field_css="div[class='autop-amper-ctrl-container'] div[class^='autop-amper'][title^='"+field_name+"'] [id$='button']"
        elem_obj=utillityobject.validate_and_get_webdriver_object(self, field_css, "Filter Values Dropdown in Autoprompt")
        coreutillityobject.left_click(self, elem_obj)
        
    def select_radio_button_in_auto_prompt_values(self, radio_button_name):
        '''
        This function is used to select radio button in auto prompt value selection prompt
        radio_button_name=['All Values','Select Values']
        '''
        radio_btn_css="div[class*='popup-container'][style*='max-width'] div[class*='ui-radio']"
        elem_object=utillityobject.validate_and_get_webdriver_objects(self, radio_btn_css, "Radio buttons in all values selection prompt")
        all_value_radio=elem_object[[i.text.strip() for i in elem_object].index(radio_button_name)]
        coreutillityobject.left_click(self, all_value_radio)
    
    def select_value_button_in_auto_prompt(self, button_name):
        '''
        This function is used to select all or none button from the auto prompt values
        radio_button_name=['All','None']
        '''
        
        button_css="#av_btn_sel_all" if button_name=="All" else "#av_btn_sel_none"
        elem_object=utillityobject.validate_and_get_webdriver_object(self, button_css, "Value button CSS")
        coreutillityobject.left_click(self, elem_object)
        
    def select_auto_prompt_value_back_button(self):
        '''
        This function is used to click on back button in the prompt values
        '''
        back_btn_css="div[class*='popup-container'][style*='max-width'] a[data-rel='back']" 
        elem_obj=utillityobject.validate_and_get_webdriver_object(self, back_btn_css, 'Back button in select values prompt')
        coreutillityobject.left_click(self, elem_obj)
    
    def verify_filter_chained_group_icon_is_visible_in_autoprompt(self, msg):
        """
        This function is used to verify filter chained group icon is displayed in auto prompt window
        """
        icon_css = "#promptPanel .autop-amper-ctrl-container div[class*='autop-icon-chain-group'][title]" 
        icon_obj = utillityobject.validate_and_get_webdriver_object(self, icon_css, 'Filter chained group icon')
        icon_path = JavaScript.get_element_after_style_properties(self, icon_obj, 'background-image')
        icon_display = JavaScript.get_element_after_style_properties(self, icon_obj, 'display')
        visible_status = True if 'png_images/link.png' in icon_path and icon_display=='block' else False
        utillityobject.asequal(self, True, visible_status, msg)
            
    def select_field_filter_values_in_auto_prompt(self, value_list, selection_type):
        '''
        This function is used to select value from the auto prompt value by using check box selection
        value_list['Canada','East','Europe']
        '''
        if selection_type=="check":
            value_css = "div[class*='popup-container'][style*='max-width'] .ui-checkbox-off"
        elif selection_type=="list":
            value_css=".ui-popup-active .ui-selectmenu ul li a:not([class*='checkbox'])"
        elif selection_type=="listbox":
            value_css = "div[class*='popup-container'][style*='max-width'] #av_grp_values div"
            
        for value in value_list:
            utillityobject.scroll_down_and_find_item_using_mouse(self, value_css, value, pause_time=1)
            elem_objects=utillityobject.validate_and_get_webdriver_objects(self, value_css, "Field filter values in auto prompt")
            elm=elem_objects[[el.text.strip() for el in elem_objects].index(value)]
            JavaScript.scrollIntoView(self, elm, 2)
            elem_objects=utillityobject.validate_and_get_webdriver_objects(self, value_css, "Field filter values in auto prompt")
            elm=elem_objects[[el.text.strip() for el in elem_objects].index(value)]
            coreutillityobject.left_click(self, elm)
    
    def select_field_filter_values_in_auto_prompt_(self, value_list, selection_type):
        '''
        This function is used to select value from the auto prompt value by using check box selection
        value_list['Canada','East','Europe']
        '''
        if selection_type=="check":
            value_css = "div[class*='popup-container'][style*='max-width'] ul li a[class*='ui-checkbox']"
        elif selection_type=="input":
            value_css = "div[class*='popup-container'][style*='max-width'] #av_grp_values div"
        else:
            value_css=".ui-popup-active .ui-selectmenu ul li"
        for value in value_list:
            elem_objects=utillityobject.validate_and_get_webdriver_objects(self, value_css, "Field filter values in auto prompt values prompt")
            
            elm=elem_objects[[el.text.strip() for el in elem_objects].index(value)]
            print (elm.text.strip())
            count=0
        time.sleep(2)
        coreutillityobject.python_left_click(self, elm)
        time.sleep(0.5)
            
    def verify_field_filter_values_in_auto_prompt(self, expected_value_list, msg, verify_type):
        '''
        This function is used to select value from the auto prompt value by using check box selection
        value_list['Canada','East','Europe']
        '''
        value_css ="div[class*='popup-container'][style*='max-width'] li[role='option']" if verify_type == 'option' else "div[class*='popup-container'][style*='max-width'] #av_grp_values div label"
        elem_objects=utillityobject.validate_and_get_webdriver_objects(self, value_css, "Field values in auto prompt values prompt")
        actual_value_list=[el.text.strip() for el in elem_objects if el.is_displayed()]
        utillityobject.asequal(self, expected_value_list, actual_value_list, msg)
        
    def verify_field_filter_values_checked_property_in_auto_prompt(self, expected_value_list, msg, verify_type):
        '''
        This function is used to select value from the auto prompt value by using check box selection
        value_list['Canada','East','Europe']
        '''
        value_css ="div[class*='popup-container'][style*='max-width'] li[role='option']" if verify_type == 'option' else "div[class*='popup-container'][style*='max-width'] #av_grp_values div label"
        elem_objects=utillityobject.validate_and_get_webdriver_objects(self, value_css, "Field values in auto prompt values prompt")
        for elem in elem_objects:
            search_box_state=utillityobject.get_element_attribute(self, elem, 'class')
            if verify_type=="checked":
                utillityobject.asin(self, 'ui-checkbox-on',search_box_state, msg+" for "+elem.text.strip())
            else:
                utillityobject.as_notin(self, 'ui-checkbox-on',search_box_state, msg+" for "+elem.text.strip())
                        
    def select_close_button_in_field_filter_values_in_auto_prompt(self):
        '''
        This function will click on close button which is available in field filter values
        '''
        css=".ui-icon-delete"
        elem=utillityobject.validate_and_get_webdriver_object(self, css, 'AUTO Prompt Run Report Panel')
        coreutillityobject.left_click(self, elem)
        
    def get_autoprompt_field_labels(self):
        '''
        This function will get return field labels in auto prompt
        '''
        autoprompt_field_labels_css = ".autop-amper-ctrl-container .autop-chain-group .autop-amper .autop-amper-label"
        autoprompt_field_labels_obj = utillityobject.validate_and_get_webdriver_objects(self, autoprompt_field_labels_css, "Autoprompt fieled labels")
        autoprompt_field_labels = [label.text.strip() for label in autoprompt_field_labels_obj]
        return autoprompt_field_labels
    
    def get_autoprompt_field_object(self, field_name):
        '''
        This function will return auto prompt field objects
        get_autoprompt_field_object('CAR')
        '''
        autoprompt_filed_css = ".autop-amper-ctrl-container .autop-chain-group .autop-amper"
        autoprompt_filed_labels = IA_Run.get_autoprompt_field_labels(self)
        if field_name in autoprompt_filed_labels :
            autoprompt_field_obj = utillityobject.validate_and_get_webdriver_objects(self, autoprompt_filed_css, "Autoprompt field")
            field_index = autoprompt_filed_labels.index(field_name)
            field_object = autoprompt_field_obj[field_index]
            return field_object
        else :
            error_msg = "'{0}' field doesn't display in auto prompt window".format(field_name)
            raise KeyError(error_msg)
        
    def verify_selected_field_dropdown_value_in_autoprompt(self, field_name, expected_selected_value, msg):
        '''
        This function will verify the selected field dropdown value in autoprompt
        verify_selected_field_dropdown_value_in_autoprompt('CAR', 'BMW', "Step x: Verify selected value is BMW in the CAR field")
        '''
        field_object = IA_Run.get_autoprompt_field_object(self, field_name)
        selected_value_css = ".autop-control-wrapper span.autop-control"
        selected_value_obj = utillityobject.validate_and_get_webdriver_object(self, selected_value_css, 'Selected field dropdown label', field_object)
        actual_selected_value = selected_value_obj.text.strip()
        utillityobject.asequal(self, expected_selected_value,actual_selected_value, msg)
    
    def verify_selected_field_dropdown_value_count_in_autoprompt(self, field_name, expected_selected_value, msg):
        '''
        This function will verify selected field dropdown count in auto prompt
        verify_selected_field_dropdown_value_count_in_autoprompt('MODEL', '4')
        '''
        field_object = IA_Run.get_autoprompt_field_object(self, field_name)
        selected_value_count_css = ".autop-control-wrapper .ui-li-count"
        selected_value_obj = utillityobject.validate_and_get_webdriver_object(self, selected_value_count_css, 'Selected field dropdown label', field_object)
        actual_selected_value = selected_value_obj.text.strip()
        utillityobject.asequal(self, expected_selected_value,actual_selected_value, msg)
    
    def verify_autoprompt_field_labels(self, expected_field_label_list, msg, assert_type):
        '''
        This function will verify autoprompt field labels in auto prompt
        verify_autoprompt_field_labels(['CAR','MODEL','MINMPG','MAXMPG','COUNTRY', "Step x: Verify auto prompt field labels", 'asin')
        '''
        actual_field_label = IA_Run.get_autoprompt_field_labels(self)
        utillityobject.verify_list_values(self, expected_field_label_list, actual_field_label, msg, assert_type)
    
    def verify_field_textbox_value_in_autoprompt(self, field_name, expected_selected_value, msg):
        '''
        This function will verify field textbox value in autoprompt
        verify_field_textbox_value_in_autoprompt('COUNTRY', 'COUNTRY', "Step X:Verify field textbox value in autoprompt"):
        '''
        field_object = IA_Run.get_autoprompt_field_object(self, field_name)
        fiel_textbox_value_css = ".autop-control-wrapper input"
        fiel_textbox_value_obj = utillityobject.validate_and_get_webdriver_object(self, fiel_textbox_value_css, 'Auto prompt textbox', field_object)
        palceholder_value = fiel_textbox_value_obj.get_attribute('placeholder')
        text_value = fiel_textbox_value_obj.get_attribute('value')
        actual_selected_value = text_value if text_value != '' else palceholder_value
        utillityobject.asequal(self, expected_selected_value,actual_selected_value, msg)
        
    def enter_value_field_textbox_in_auto_prompt(self, field_name, input_value):
        '''
        This function will type value in the inputbox based on the provided field name
        '''
        field_object = IA_Run.get_autoprompt_field_object(self, field_name)
        fiel_textbox_value_css = ".autop-control-wrapper input"
        fiel_textbox_value_obj = utillityobject.validate_and_get_webdriver_object(self, fiel_textbox_value_css, 'Auto prompt textbox', field_object)
        fiel_textbox_value_obj.clear()
        utillityobject.set_text_to_textbox_using_keybord(self, input_value, text_box_elem=fiel_textbox_value_obj)
    
    def enter_value_search_textbox_popup_in_auto_prompt(self, input_value):
        '''
        This function will type value in the inputbox based on the provided field name
        '''
        fiel_textbox_value_css = "#av_search"
        fiel_textbox_value_obj = utillityobject.validate_and_get_webdriver_object(self, fiel_textbox_value_css, 'Auto prompt search popup textbox')
        fiel_textbox_value_obj.clear()
        utillityobject.set_text_to_textbox_using_keybord(self, input_value, text_box_elem=fiel_textbox_value_obj)
        
    def verify_all_or_select_values_button_selected_in_auto_prompt(self, selected_button, step_num):
        """
        Description : Verify All Values or Select Values button is selected in auto prompt filter filed popup.
        :uage : verify_all_or_select_values_button_selected_in_auto_prompt(['All Values'], "01.02")
        """
        actual_selected_button = [btn.text.strip() for btn in self.driver.find_elements_by_css_selector(".autop-sav-popup label[for='av_rb_foc_null_on']")]
        msg = "Step {0} : Verify {1} is selected in auto prompt filter fields popup".format(step_num, selected_button)
        utillityobject.asequal(self, selected_button, actual_selected_button, msg)
        
    def select_report_autolink_tooltip_using_actionchain(self, row, col, tooltip_path, parent_table_css="table[summary='Summary']"):
        """
        Example Usage : iarun.select_report_autolink_tooltip("table[summary= 'Summary']", 6,2, "Auto Links->AUTOLINK_TARGET01")
        :usage - select_report_autolink_tooltip_using_actionchain(5, 1, "Drill up to Store Business Sub Region")
        """
        parent_element = self.driver.find_element_by_css_selector(parent_table_css)
        coreutillityobject.python_move_to_element(self, parent_element, element_location="top_left")
        table_cell_css = "{0}>tbody>tr:nth-child({1})>td:nth-child({2})>a".format(parent_table_css, row, col)   
        table_cell_obj = self.driver.find_element_by_css_selector(table_cell_css)
        JavaScript.scrollIntoView(self, table_cell_obj)
        table_cell_obj.click()
        utillityobject.synchronize_with_number_of_element(self, "[id='my_tooltip_id'][style*='visible']", 1, expire_time=4)
        tooltips=tooltip_path.split('->')
        tooltip_css = "#my_tooltip_id ul[class='tdgchart-tooltip-list']>li[class*='tdgchart-tooltip-pad']"
        for count, tooltip in enumerate(tooltips) :
            tooltip_elements=self.driver.find_elements_by_css_selector(tooltip_css)
            tooltip_text_list=[tooltip.text.strip().replace('>', '').replace('\n', '') for tooltip in tooltip_elements]
            tooltip_element=tooltip_elements[tooltip_text_list.index(tooltip)]
            if tooltips[count] == tooltips[-1] :
                tooltip_element.click()
            else :
                ActionChains(self.driver).move_to_element(tooltip_element).perform()
                
    def verify_chart_autodrill_breadcrumb_text(self, expected_breadcrumb, step_num, parent_css="#jschart_HOLD_0"):
        """
        Description : Verify auto drill chart bread crumb text as list
        :Usage : chart_obj.verify_chart_autodrill_breadcrumb_text(['Home->NorthAmerica->West->UnitedStates->California->SanDiego->92101'], "01.01")
        """
        breadcrumb_css = parent_css + " foreignObject"
        breadcrumb_object = utillityobject.validate_and_get_webdriver_object(self, breadcrumb_css, "Chart auto drill breadcrumb")
        actual_breadcrumb = breadcrumb_object.text.strip().replace("\u2192", "->").replace(" ", "").split("\n")
        msg = "Step {0} : Verify chart auto drill {1} bread crumb displayed".format(step_num, expected_breadcrumb)
        utillityobject.asequal(self,expected_breadcrumb,actual_breadcrumb,msg)  