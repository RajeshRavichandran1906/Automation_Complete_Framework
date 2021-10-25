from common.lib import utillity
from common.lib.base import BasePage
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook
from openpyxl import load_workbook
import time,os, re

class Wf_Administration_Console(BasePage):
    def __init__(self, driver):
        super(Wf_Administration_Console, self).__init__(driver)

    def _validate_page(self, locator):
        self.longwait.until(EC.visibility_of_element_located(locator))
        
    def select_or_verify_top_toolbar_links(self, menu_path, **kwargs):
        """
        menu_path = 'licenses->WebFOCUS Client'
        verify_list=['WebFOCUS Client', 'User Audit', 'Third Party']
        Syntax: select_or_verify_top_toolbar_links('licenses', verify_list=list)
        Syntax: select_or_verify_top_toolbar_links('clearcache')
        @author = Niranjan
        """
        links={"configuration":"#topToolBarBox #console_configuration_id", 
               "security":"#topToolBarBox #console_web_security_id", 
               "reportcaster":"#topToolBarBox #console_report_caster_id", 
               "diagnostics":"#topToolBarBox #console_diagnostics_id", 
               "licenses":"#topToolBarBox span[id^='BiMenuButton']", 
               "clearcache":"#topToolBarBox #Clear_Cache", 
               "close":"#topToolBarBox #close_console_window", 
               "help_img":"#topToolBarBox #ConsoleHelpPage"}
        utillobj = utillity.UtillityMethods(self.driver)
        menu_list=menu_path.split('->')
        main_menu_item=self.driver.find_element_by_css_selector(links[menu_list[0]])
        main_menu_item.click()
        time.sleep(2)
        if 'verify_list' in kwargs:
            utillobj.select_or_verify_bipop_menu(verify='true',expected_popup_list=kwargs['verify_list'], msg=kwargs['msg'])
        if len(menu_list)>1:
            utillity.UtillityMethods.select_or_verify_bipop_menu(self, menu_list[1])
            time.sleep(2)
    
    def create_license_information(self,file_name):
        '''
        Syntax : create_license_information(licence_key+'.xlsx')
        '''
        css="#LicenseManagementDialog div[id^='BiVBox']:nth-child(1) div[id^='BiHBox']"
        elems=[el for el in self.driver.find_elements_by_css_selector(css) if len(el.text)!=0]
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        r=0
        for e in elems:
            act=e.text.strip().split("\n")
            if len(act)==1:
                self.driver.implicitly_wait= 0
                try:
                    if act[0]=='License Key' :
                        input_ele=e.find_elements_by_css_selector("input[id^='BiTextField']")[0]
                        act.append(input_ele.get_attribute('value').strip())
                    else :
                        e.find_element_by_css_selector("img[src$='checkmark_blk_16.png'][style*='inherit']")
                        act.append('checked')
                except:
                    pass
            c=0
            for v in act:
                s.cell(row=r + 1, column=c + 1).value = str(v)
                c=c+1
            r=r+1
        wb.save(os.getcwd() + "\data\\" + file_name)
    
    def verify_license_information(self,file_name,msg):
        ''''
        Syntax : verify_license_information(licence_key+'.xlsx','Step 05 : Verify that the frame displays the relevant information')
        '''
        status=self.compare_license_information(file_name)
        utillity.UtillityMethods.asequal(self,len(status),1,msg+ ' Row,Column -'+ str(status))
    
    def compare_license_information(self,file_name) :
        
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        css="#LicenseManagementDialog div[id^='BiVBox']:nth-child(1) div[id^='BiHBox']"
        elems=[el for el in self.driver.find_elements_by_css_selector(css) if len(el.text)!=0]
        r=0
        for e in elems:
            act=e.text.strip().split("\n")
            if len(act)==1:
                self.driver.implicitly_wait= 0
                try:
                    if act[0]=='License Key' :
                        input_ele=e.find_elements_by_css_selector("input[id^='BiTextField']")[0]
                        act.append(input_ele.get_attribute('value').strip())
                    else :
                        e.find_element_by_css_selector("img[src$='checkmark_blk_16.png'][style*='inherit']")
                        act.append('checked')
                except:
                    pass
            c=0
            for v in act:
                if (s1.cell(row=r + 1, column=c + 1).value)!=None :
                    if str(s1.cell(row=r + 1, column=c + 1).value).strip() == v:
                        status=[0]
                        c=c+1
                        continue
                    else:
                        print("Actual : "+str(s1.cell(row=r + 1, column=c + 1).value).strip()+" == Expected : "+v)
                        status=[r+1,c]
                        return (status)
                #print(status)
            r=r+1
        return (status)
    def expand_admin_console_tree(self, parent_tree_css, tree_abs_path):
        '''
        :param parent_tree_css='#console_tree_Configuration_Settings .bi-tree-view-body-content'
        :param tree_abs_path='Application Settings'
        :usage expand_admin_console_tree('#console_tree_Configuration_Settings .bi-tree-view-body-content', 'Application Settings')
        '''
        tree_rows=parent_tree_css + " table tr"
        folder_list=tree_abs_path.split('->')
        for item in folder_list:
            tree_items = self.driver.find_elements_by_css_selector(tree_rows)
            for i in range(len(tree_items)):
                if tree_items[i].text.strip() == item:
                    self.driver.implicitly_wait= 0
                    try:
                        expand_img = tree_items[i].find_element_by_css_selector("img[src*='triangle_collapsed']")
                        expand_img.click()
                        time.sleep(5)
                        break
                    except NoSuchElementException:
                        pass
            
    def select_admin_console_tree(self, parent_tree_css, tree_abs_path):
        '''
        :param parent_tree_css='#console_tree_Configuration_Settings .bi-tree-view-body-content'
        :param tree_abs_path='Application Settings'
        :usage select_admin_console_tree('#console_tree_Configuration_Settings .bi-tree-view-body-content', 'Application Settings')
        '''
        tree_rows=parent_tree_css + " table tr"
        item_list=tree_abs_path.split('->')
        Wf_Administration_Console.expand_admin_console_tree(self, parent_tree_css, tree_abs_path)
        tree_items = self.driver.find_elements_by_css_selector(tree_rows)
        for i in range(len(tree_items)):
            if tree_items[i].text.strip() == item_list[-1]:
                tree_items[i].find_element_by_css_selector("img[src$='png']").click()
                time.sleep(3)
                
    def save_cancel_restoredefaultvalues_button(self, scroll_type, btn_name, number_of_times=0):
        '''
        @param : scroll_type='up' or 'down'
        @param : btn_name = 'Save', or 'Cancel' or 'Restore Default Values'
        @param : number_of_times = 4
        Function Call : save_cancel_restoredefaultvalues_button('down', 'Save', 4)
        '''
        utillity.UtillityMethods.mouse_scroll(self, scroll_type, number_of_times)
        utillity.UtillityMethods.click_dialog_button(self, "#idInfoassistPropertiesPage div[id^='BiHBox']", btn_name)
        utillity.UtillityMethods.verify_js_alert(self, 'Successful', 'Step X: Verify Alert')
                
    def input_bihbox(self, parent_css, bihbox_item_name, **kwargs):
        '''
        parent_css='#idApplicationSettingsPageView' 
        input_control='checkbox/radoibutton/combobox/textbox'
        '''
        rows=self.driver.find_elements_by_css_selector(parent_css + " div[id^='BiHBox']:not([style*='background']")
        for row in rows:
            if bihbox_item_name in row.text.split('\n'):
                current_obj=row
                break
        if 'input_control' in kwargs:
            if kwargs['input_control'] in ('checkbox','radoibutton'):
                current_obj.find_element_by_css_selector("input").click()
            if kwargs['input_control'] == 'combobox':
                elem=current_obj.find_element_by_css_selector("[class$='combo-box'] div[id^='BiButton']")
                utillity.UtillityMethods.select_any_combobox_item(self, elem, kwargs['combobox_value'])
            if kwargs['input_control'] == 'textbox':
                elem=current_obj.find_element_by_css_selector("[class$='text-field']")
                utillity.UtillityMethods.set_text_field_using_actionchains(self, elem, kwargs['textbox_value'])
        if 'option_button' in kwargs:
            btn_id = 'Save' if kwargs['option_button'] == 'save' else 'Cancel'
            btn_css=parent_css + " div[id$='" + btn_id + "Button']"
            self.driver.find_element_by_css_selector(btn_css).click()
            time.sleep(1)
            utillity.UtillityMethods.verify_js_alert(self, 'Successful', 'Step X: Verify Alert')
        time.sleep(2)
    
    def verify_bihbox(self, parent_css, bihbox_item_name, input_control, expected_visible_symbol_or_text, msg, position=1):
        '''
        parent_css='#idApplicationSettingsPageView' 
        input_control='checkbox/radoibutton/combobox/textbox'
        expected_visible_symbol_or_text='checked/unchecked/any text"
        '''
        rows=self.driver.find_elements_by_css_selector(parent_css + " div[id^='BiHBox']:not([style*='background']")
        for row in rows:
            if bihbox_item_name in row.text.split('\n'):
                current_obj=row
                if input_control in ('checkbox','radoibutton'):
                    expected_status = True if expected_visible_symbol_or_text == 'checked' else False
                    total_check_box=current_obj.find_elements_by_css_selector("input[type='checkbox']")
                    actual_status=total_check_box[position-1].is_selected()
                    utillity.UtillityMethods.asequal(self, expected_status, actual_status, msg)
                if input_control == 'combobox':
                    actual_visible_symbol_or_text=current_obj.find_element_by_css_selector("[class*='combo-box']").text.strip()
                    utillity.UtillityMethods.asequal(self, expected_visible_symbol_or_text, actual_visible_symbol_or_text, msg)
                    print(actual_visible_symbol_or_text)
                if input_control == 'textbox':
                    actual_visible_symbol_or_text=current_obj.find_element_by_css_selector("[class$='text-field']").get_attribute("value")
                    utillity.UtillityMethods.asequal(self, expected_visible_symbol_or_text, actual_visible_symbol_or_text, msg)
                break
        time.sleep(2)