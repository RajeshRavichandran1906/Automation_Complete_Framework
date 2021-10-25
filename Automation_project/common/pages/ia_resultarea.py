from common.lib.utillity import UtillityMethods as utillityobject
from common.lib.core_utility import CoreUtillityMethods as coreutillityobject
from common.lib.base import BasePage
from selenium.webdriver import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver.support.color import Color
import time,re, os, pyautogui
from common.pages import visualization_resultarea, visualization_ribbon
from common.lib.global_variables import Global_variables
if sys.platform == 'linux':
    from pykeyboard import PyKeyboard
    pykeyboard = PyKeyboard()
else:
    import keyboard

class IA_Resultarea(BasePage):
    """ Inherit attributes of the parent class = Baseclass """

    def __init__(self, driver):
        super(IA_Resultarea, self).__init__(driver)
    
    def move_mouse_to_chart_component(self, riser_or_marker_element, initial_move_xy_dict=None, element_location='middle', xoffset=0, yoffset=0, use_marker_enable=False, move_to_tooltip=True):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        if initial_move_xy_dict != None:
            x=initial_move_xy_dict['x']
            y=initial_move_xy_dict['y']
        else:
            x=Global_variables.current_working_area_browser_x
            y=Global_variables.current_working_area_browser_y
        coreutillityobject.move_mouse_to_offset(self, x_offset=x, y_offset=y)
        time.sleep(1)
        if use_marker_enable==True:
            (x, y)=utillityobject.marker_enable(self, riser_or_marker_element)
            coreutillityobject.python_move_to_offset(self, x_offset=x, y_offset=y+yoffset, mouse_move_duration=2.5)
        else:
            coreutillityobject.python_move_to_element(self, riser_or_marker_element, element_location=element_location, xoffset=xoffset, yoffset=yoffset)
        tooltip_css="span[id*='tdgchart-tooltip']:not([style*='hidden'])"
        time.sleep(self.tooltip_wait_time)
        
#         utillityobject.synchronize_with_number_of_element(self,tooltip_css,1, 25, pause_time=0.2)
        tooltip_elem=self.driver.find_elements_by_css_selector(tooltip_css + " li")[0]
        if move_to_tooltip :
            coreutillityobject.python_move_to_element(self, tooltip_elem)
    
    def select_chart_component(self, riser_or_marker_element, use_marker_enable=False):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        x=Global_variables.current_working_area_browser_x
        y=Global_variables.current_working_area_browser_y
        coreutillityobject.move_mouse_to_offset(self, x, y)
        time.sleep(1)
        if use_marker_enable==True:
            (x, y)=utillityobject.marker_enable(self, riser_or_marker_element)
            coreutillityobject.python_click_with_offset(self, x, y)
        else:
            coreutillityobject.left_click(self, riser_or_marker_element)
#         tooltip_elem=self.driver.find_element_by_css_selector("div[class='tdgchart-tooltip']")
#         coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_middle', yoffset=5)
    
    def multiselect_chart_component(self, riser_or_marker_element_list, use_marker_enable=False):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        x=Global_variables.current_working_area_browser_x
        y=Global_variables.current_working_area_browser_y
        coreutillityobject.move_mouse_to_offset(self, x, y)
        time.sleep(1)
        if sys.platform == 'linux':
            pykeyboard.press_key(pykeyboard.control_key)
        else:
            keyboard.press('ctrl')
        for riser_or_marker_element in riser_or_marker_element_list:
            if use_marker_enable==True:
                (x, y)=utillityobject.marker_enable(self, riser_or_marker_element)
                coreutillityobject.python_click_with_offset(self, x, y)
            else:
                coreutillityobject.python_left_click(self, riser_or_marker_element, mouse_move_duration=3.5)
            time.sleep(1)
        if sys.platform == 'linux':
            pykeyboard.release_key(pykeyboard.control_key)
        else:
            keyboard.release('ctrl') 
        time.sleep(1)   
        #tooltip_elem=self.driver.find_element_by_css_selector("div[class='tdgchart-tooltip']")
        #coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_middle', yoffset=5)
                
    def verify_tooltip(self, expected_tooltip_list, msg='Step X: Verify default Chart tooltip'):
        '''
        Desc: This function will verify the tooltip. First mouse will move to the top left corner of the current working area,
                Then it will move to the required location to bring the tooltip disply and then read and compare.
        '''
        tooltip_css="span[id*='tdgchart-tooltip']"
        raw_tooltip_list=self.driver.find_element_by_css_selector(tooltip_css).text.replace(u'\xa0\n', '').replace(u'\xa0', ' ').split('\n')
        actual_list=utillityobject.get_actual_tooltip_list(self, raw_tooltip_list)
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
        
    def verify_lasso_tooltip(self, expected_tooltip_list, msg='Step X: Verify default Chart tooltip'):
        '''
        Desc: This function will verify the tooltip. First mouse will move to the top left corner of the current working area,
                Then it will move to the required location to bring the tooltip disply and then read and compare.
        '''
        tooltip_css="div[class='tdgchart-tooltip']"
        raw_tooltip_list=self.driver.find_element_by_css_selector(tooltip_css).text.replace(u'\xa0\n', '').replace(u'\xa0', ' ').split('\n')
        actual_list=utillityobject.get_actual_tooltip_list(self, raw_tooltip_list)
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
        
    def hover_to_required_arrow_in_tooltip(self, item_name):
        first_level_tooltip_css = "#tdgchart-tooltip li[class*='tdgchart-tooltip-hover']"
        elems=self.driver.find_elements_by_css_selector(first_level_tooltip_css)
        first_level_tooltip=elems[[item_name in elem for elem in [elem.text.strip().replace('\n','').replace('>','') for elem in elems if elem.text.strip() != '']].index(True)]
        arrow_to_be_selected=first_level_tooltip.find_element_by_css_selector("span.tdgchart-tooltip-arrow")
        coreutillityobject.python_move_to_element(self, arrow_to_be_selected)
        tooltip_elem=self.driver.find_element_by_css_selector("span[id*='tdgchart-tooltip'] [class*='tdgchart-submenu']")
        coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_middle', yoffset=5)
        
    def select_tooltip_item(self, item_name):
        '''
        Desc: This function is used to select an item from the single tooltip box. 
        '''
        if item_name in ['Filter Chart', 'Exclude from Chart', 'Remove Filter']:
            tooltip_css= "[id*='tdgchart-tooltip'] span.tdgchart-tooltip-label>div"
        else:
            tooltip_css= "[id*='tdgchart-tooltip'] span.tdgchart-tooltip-label"
        elems=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name)]
        coreutillityobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
#         time.sleep(10)
    
    def select_lasso_tooltip_item(self, item_name):
        '''
        Desc: This function is used to select an item from the single lasso tooltip box.
        '''
        tooltip_css= "div[class='tdgchart-tooltip'] span[class='tdgchart-tooltip-pad']"
        elems=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name)]
        coreutillityobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
#         time.sleep(10)
        
    def select_bilevel_tooltip_item(self, item_name1, item_name2):
        '''
        Desc: This function is used to select an item from the second tooltip box. Mostly used for drill down pop up.
        '''
        IA_Resultarea.hover_to_required_arrow_in_tooltip(self, item_name1)
        second_level_tooltip_css = "#tdgchart-tooltip [class*='tdgchart-submenu'] span.tdgchart-tooltip-label"
        elems=self.driver.find_elements_by_css_selector(second_level_tooltip_css)
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name2)]
        coreutillityobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
#         time.sleep(10)
    
    def verify_xy_axis_title(self, expected_title_list, parent_css, x_or_y_axis_title_css, x_or_y_axis_title_length=None, msg='Step X'):
        '''
        Desc: This function is used to verify X or Y axis title at a time.
        '''
        x_or_y_axis_title_css=parent_css+" " + x_or_y_axis_title_css
        actual_title_list=[title.text.strip() for title in self.driver.find_elements_by_css_selector(x_or_y_axis_title_css)]
        for expected_title, actual_title in zip(expected_title_list, actual_title_list):
            if expected_title[:x_or_y_axis_title_length] == actual_title[:x_or_y_axis_title_length]:
                status= True
            else:
                status=False
                break
        failure_message='' if status else " Actual {0} ,Expected {1}".format(actual_title_list,expected_title_list)
        utillityobject.asequal(self, status, True, msg+failure_message) 
    
    def verify_xyz_labels(self, expected_label_list, parent_css, xyz_axis_label_css, xyz_axis_label_length=None, msg='Step X'):
        '''
        Desc: This function is used to verify X or Y axis title at a time.
        '''
        xyz_axis_label_css=parent_css+" " + xyz_axis_label_css
        x=self.driver.find_elements_by_css_selector(xyz_axis_label_css) 
        my_iter=(i.text for i in x)
        for label in expected_label_list:
            if label[:xyz_axis_label_length] == next(my_iter)[:xyz_axis_label_length]:
                status= True
            else:
                status=False
                break
        del my_iter
        failure_message='' if status else " Actual {0} ,Expected {1}".format([i.text for i in x],expected_label_list)
        utillityobject.asequal(self, status, True, msg+failure_message) 
    
    def verify_xyz_labels_and_length(self, expected_label_list, parent_css, xyz_axis_label_css, msg='Step X'):
        '''
        Desc: This function is used to verify X or Y axis title at a time.
        '''
        xyz_axis_label_css=parent_css+" " + xyz_axis_label_css
        x=self.driver.find_elements_by_css_selector(xyz_axis_label_css)
         
        actual_length_of_label = len(x) 
        print(actual_length_of_label)
        if expected_label_list[:4] == [i.text for i in x][:4]:
            status = True
            print(len(expected_label_list))
            if actual_length_of_label == len(expected_label_list):
                status = True
            else: 
                status = False
        else:
            status = False
        failure_message='' if status else " Actual {0} ,Expected {1}".format([i.text for i in x],expected_label_list)
        utillityobject.asequal(self, status, True, msg+failure_message) 
        
    def verify_legends(self, expected_legend_list, parent_css, legend_length=None, msg='Step X'):
        '''
        Desc: This function is used to verify legends.
        '''
        legend_css=parent_css + " .legend text"
        legends=self.driver.find_elements_by_css_selector(legend_css)
        my_iter_x=(i.text for i in legends)
        for label_x in expected_legend_list:
            if label_x[:legend_length] in next(my_iter_x):
                statex= True
            else:
                statex=False
                break
        del my_iter_x
        failure_message='' if statex else " Actual {0} ,Expected {1}".format([i.text for i in legends],expected_legend_list)
        utillityobject.asequal(self, statex, True, msg+failure_message)
        
    def verify_pie_label_in_single_group(self, expected_label_list, parent_css, label_css, msg):
        '''
        Desc: This function is used to verify pie labels within a single group.
        '''
        pie_label_css=parent_css + " svg > g " + label_css
        labels=self.driver.find_elements_by_css_selector(pie_label_css)
        actual_label_list=[label.text.strip() for label in labels]
        utillityobject.asequal(self,expected_label_list, actual_label_list, msg)    
    
    def verify_pie_label_and_value_in_multiple_groups(self, expected_label_list, expected_total_label_list, parent_css, label_css, msg1, msg2):
        '''
        Desc: This function is used to verify pie labels and total labels in multiple groups.
        '''
        pie_label_css=parent_css + " svg > g " + label_css
        actual_label_list=[]
        labels=self.driver.find_elements_by_css_selector(pie_label_css)
        for i in range(0,len(labels)):
            label=self.driver.find_element_by_css_selector(parent_css + " text[class='pieLabel!g" + str(i) + "!mpieLabel!']").text.strip()
            actual_label_list.append(label)
        utillityobject.asequal(self,expected_label_list, actual_label_list, actual_label_list, msg1)
        actual_total_label_list=[]
        labels=self.driver.find_elements_by_css_selector(parent_css + " text[class^='totalLabel!g']")
        for i in range(0,len(labels)):
            total_label=self.driver.find_element_by_css_selector(parent_css + " text[class='totalLabel!g" + str(i) + "!mtotalLabel!']").text.strip()
            actual_total_label_list.append(total_label)
        utillityobject.asequal(self, expected_total_label_list, actual_total_label_list, msg2)
    
    def verify_number_of_risers(self, parent_css_with_tagname, risers_per_segment, expected_number, msg):
        '''
        Desc: This function is used to verify the number of risers. As this covers wide range of risers with different tags, user must 
        pass the parent css upto tag. Also we need to provide the number of risers per segment. It means if there is a simple bar chart,
        then number of risers per segment is '1', for stacked bar it will differ. 
        '''
        riser_css=parent_css_with_tagname + "[class^='riser']"
        total_risers=len(self.driver.find_elements_by_css_selector(riser_css))
        actual_number=int(total_risers/risers_per_segment)
        utillityobject.asequal(self, expected_number, actual_number, msg)
        
    def verify_number_of_markers(self, parent_css, markers_per_segment, expected_number, msg):
        '''
        Desc: This function is used to verify the number of markers in line chart. 
        '''
        marker_css=parent_css + " circle[class^='marker']"
        total_markers=len(self.driver.find_elements_by_css_selector(marker_css))
        actual_number=int(total_markers/markers_per_segment)
        utillityobject.asequal(self, expected_number, actual_number, msg)

    def verify_number_of_pie_segment(self, parent_css, risers_per_segment, expected_number, msg):
        '''
        Desc: This function is used to verify the number of pie segments.
        '''
        parent=parent_css + " svg.rootPanel>g.chartPanel"
        total_risers=len(self.driver.find_elements_by_css_selector(parent + " >g>g>g>path[class^='riser']"))
        actual_number=int(total_risers/risers_per_segment)
        utillityobject.asequal(self, expected_number, actual_number, msg)
    
    def verify_number_of_circles(self, parent_css, minimum_number, maximum_number, msg):
        '''
        Desc: This function is used to verify the number of circles. As number of circles differs in a less margin from resolution to resolution
        user will have to pass a range.
        '''
        parent=parent_css + " svg g"
        actual_number=len(self.driver.find_elements_by_css_selector(parent + " circle[class^='riser']"))
        if actual_number in range(minimum_number , maximum_number):
            status=True
        utillityobject.asequal(self, True, status, msg)
        
    def verify_number_of_connector_lines(self, parent_css, connector_line_per_segment, expected_number, msg):
        '''
        This function is used to verify number of connector lines
        '''
        connector_line_css=parent_css+ " path[class*='connectorLine']"
        total_risers=len(self.driver.find_elements_by_css_selector(connector_line_css))
        actual_number=int(total_risers/connector_line_per_segment)
        utillityobject.asequal(self, expected_number, actual_number, msg)
                
    def verify_ia_row_and_column_labels(self, expected_label, parent_css, matrix_type, label_length=None, msg='Step X'):
        '''
        Desc: This function used to verify Rows or Columns labels
        '''
        if matrix_type == 'rows':
            label_css = parent_css + " svg text[class^='row']"
        elif matrix_type == 'column':
            label_css = parent_css + " svg text[class^='col']"
        actual_val_list=(el.text.strip() for el in self.driver.find_elements_by_css_selector(label_css) if el.text.strip()!='')
        for label in expected_label:
            if label[:label_length]==next(actual_val_list)[:label_length] :
                status= True
            else:
                status=False
                break
        failure_message='' if status else " Actual {0} ,Expected {1}".format([i.text for i in self.driver.find_elements_by_css_selector(label_css) if i.text.strip()!=''],expected_label)
        utillityobject.asequal(self, status, True, msg+failure_message)
    
    """====================================================OLD FUNCTIONS=================================================="""
    
    def _validate_page(self, locator):
        self.longwait.until(EC.visibility_of_element_located(locator))
    
    def verify_document_objects(self, parent_css, component_type, msg, expected_value=None, expected_image_name=None):
        '''
        Desc :- This function is used to verify Document Objects. We have two type of objects 'textbox' OR 'image'
        parent_css= 'PageItemImage_1' OR 'Text_1'
        '''
        if component_type == 'textbox':
            utillityobject.switch_to_frame(self, pause=2, frame_css=parent_css + " iframe[id^='BiRichEdit']")
            get_text=self.driver.find_element_by_css_selector("html body").text.strip()
            utillityobject.switch_to_default_content(self, pause=3)
            utillityobject.asin(self, expected_value, get_text, msg)
        elif component_type == 'image':
            parent_css = parent_css + " img"
            utillityobject.verify_object_visible(self, parent_css, True, msg)
            utillityobject.verify_picture_using_sikuli(self,expected_image_name, msg)
    
    def verify_active_dashboard_prompts(self, component_type, parent_css, expected_value_list, msg):
        '''
        Desc :- This function is used to verify Active Dashboard Prompts
        :param component_type=verify dashboard prompts component like checkbox, list, etc...
        :param parent_css=element css value
        :param expected_value_list=['[All]', 'Coffee', 'Food', 'Gifts']
        :param msg=output message to display in console
        :usage verify_active_dashboard_prompts('checkbox', "#Prompt_3", ['[All]', 'Coffee', 'Food', 'Gifts'], "step 5: verify prompts")
        '''
        component_type_dict={'checkbox':'BiCheckBox',
                            'listbox':'BiListItem',
                            'radiobutton':'BiRadioButton',
                            'dropdown':'BiComboBox',
                            'text':'input'}
        if component_type.lower() == 'text':
            actual_text_value = self.driver.find_element_by_css_selector(parent_css + " " + component_type_dict[component_type]).get_attribute('value')
            status= True if actual_text_value == expected_value_list[0] else False
        else:
            iter_values=(val for val in (val.text.strip() for val in self.driver.find_elements_by_css_selector(parent_css + " [id^='"+component_type_dict[component_type]+"']") if val != '') if val != '')
            status= True
            for checkbox_value in expected_value_list:
                if checkbox_value == next(iter_values):
                    status= True
                else:
                    status=False
                    break
            del iter_values
        utillityobject.asequal(self, status, True, msg)
    
    
    def drag_drop_document_component(self, source_element_css, target_element_css, x, y, target_drop_point='top_right', mouse_speed=25):
        """
        Desc:- This function is used to drag and drop component inside document page.
        :param element_css="#TableChart_1"
        :param x=target x coordinate
        :param y=target x coordinate
        :param mouse_speed=contorl mouse move speed
        :usage drag_drop_document_component('#TableChart_2', 350, 0)
        """
        try:
            elem = self.driver.find_element_by_css_selector(source_element_css)
            content_location1 = utillityobject.get_object_screen_coordinate(self, elem, coordinate_type='top_middle', x_offset=9, y_offset=15)
            utillityobject.click_on_screen(self, elem, 'offset', click_type=0, x_offset=content_location1['x'], y_offset=content_location1['y'])
            utillityobject.synchronize_with_number_of_element(self, source_element_css + " [id*='BiResizeHandle']", 8, 9)
            del elem
        except NoSuchElementException:
            raise AttributeError(source_element_css + " given css element not able to found.")
        try:
            elem1 = self.driver.find_element_by_css_selector(source_element_css)
            try:
                elem2 = self.driver.find_element_by_css_selector(target_element_css)
            except NoSuchElementException:
                raise AttributeError(target_element_css + " given css element not able to found.")
            content_location = utillityobject.get_object_screen_coordinate(self, elem1, coordinate_type='top_middle', x_offset=19, y_offset=7)
            content_location_ = utillityobject.get_object_screen_coordinate(self, elem2, coordinate_type=target_drop_point, x_offset=19, y_offset=7)
            utillityobject.drag_drop_on_screen(self, sx_offset=content_location['x'], sy_offset=content_location['y'], tx_offset=content_location_['x']+x, ty_offset=content_location_['y']+y, mouse_speed=mouse_speed)
            del elem1
        except NoSuchElementException:
            raise AttributeError(source_element_css + " given css element not able to found.")
        
    def create_hold_type(self, hold_type, **kwargs):
        btn_css='#createFromHoldButton #createFromHoldMenuBtn'
        self.driver.find_element_by_css_selector(btn_css).click()
        utillityobject.select_or_verify_bipop_menu(self, hold_type, **kwargs)
        time.sleep(1)
            
    def create_report_data_set(self, parent_id, rows, cols, file_name, **kwargs):
        """
        params: parent_id='TableChart_1'
        kwargs: no_of_cells=4 or 8 (8 if heading wrapped to 2 rows)
        params: rows=Number of rows you want to verify (starts from 1, no column heading included)
        params: cols=total number of columns (Starts from 1)
        Usage: create_report_data_set('TableChart_1',5,4,8 'C2203727_Ds01.xlsx')
        Usage: create_report_data_set('TableChart_1',5,4,'C2203727_Ds01.xlsx',no_of_cells=4)
        Author: Niranjan
        """        
        if 'no_of_cells' in kwargs:
            no_of_cells=kwargs['no_of_cells']
        else:
            no_of_cells=cols
        field_css="#" + parent_id + " div[class^='x']"
        total_items=self.driver.find_elements_by_css_selector(field_css)
        l=[total_items[i].value_of_css_property("left") for i in range(cols)]
        total_items_count=no_of_cells
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        for r in range(0, rows):
            for c in range(0, cols):
                if total_items[total_items_count].value_of_css_property("left")==l[c]:
                    s.cell(row=r+1, column=c+1).value = str(total_items[total_items_count].text)
                    total_items_count=total_items_count+1
        wb.save(os.path.join(os.getcwd(), "data", file_name))
          
        
    def compare_report_data_set(self, parent_id, rows, cols, file_name,**kwargs):
        """
        params: parent_id='TableChart_1'
        params: no_of_cells=4 or 8 (8 if heading wrapped to 2 rows)
        params: rows=Number of rows you want to verify
        params: cols=total number of columns
        Usage: compare_report_data_set('TableChart_1', 5,4,'C2203727_Ds01.xlsx')
        Usage: compare_report_data_set('TableChart_1',5,4,'C2203727_Ds01.xlsx',no_of_cells=8)
        Author: Niranjan
        """        
        if 'no_of_cells' in kwargs:
            no_of_cells=kwargs['no_of_cells']
        else:
            no_of_cells=cols
        field_css="#" + parent_id + " div[class^='x']"
        total_items=self.driver.find_elements_by_css_selector(field_css)
        l=[total_items[i].value_of_css_property("left") for i in range(cols)]
        total_items_count=no_of_cells
        wb1 = load_workbook(os.path.join(os.getcwd(), "data", file_name))
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        for r in range(0, rows):
            for c in range(0, cols):
                if total_items[total_items_count].value_of_css_property("left")==l[c]:
                    if str(s1.cell(row=r+1, column=c+1).value).strip() == str(total_items[total_items_count].text.strip()):
                        status=[0]
                        total_items_count=total_items_count+1
                        continue
                    else:
                        status=[r+1,c]
                        return (status)
        return (status)
        
    def verify_report_data_set(self, parent_id, rows, cols, file_name,msg,**kwargs):
        """
        params: no_of_cells=4 or 8 (8 if heading wrapped to 2 rows)
        Usage: verify_report_data_set('TableChart_1', 19,6,'C2203727_Ds01.xlsx',"Step 02; Verify report dataset")
        Usage: verify_report_data_set('TableChart_1', 19,6,'C2203727_Ds01.xlsx',"Step 02; Verify report dataset",no_of_cells=4)
        """
        if 'no_of_cells' in kwargs:
            no_of_cells1=kwargs['no_of_cells']
        else:
            no_of_cells1=cols
        x= IA_Resultarea.compare_report_data_set(self, parent_id, rows, cols,file_name,no_of_cells=no_of_cells1)
        utillityobject.asequal(self,len(x),1,msg+ ' Row,Column -'+ str(x))     
    
    def verify_column_total(self, parent_id, expected_column_total_value_list, msg):
        '''
        Description: This is used for verifying the column total in a report as we are not able to caputre the values using our existing function.
        Params: expected_column_total_value_list=['TOTAL', '', '143,794', '173,204']
        '''
        field_css="#" + parent_id + " div[class^='x']"
        total_items=self.driver.find_elements_by_css_selector(field_css)
        actual_column_total_value_list=[total_items[i].text.strip() for i in range(len(total_items)-len(expected_column_total_value_list), len(total_items))]
        utillityobject.asequal(self, expected_column_total_value_list, actual_column_total_value_list, msg)
    
    def verify_row_total_report_titles_on_preview(self,colnum,no_of_cells,table_id,expected_list, msg):
        """
        params: colnum: 5
        params: no_of_cells: 12
        params: table_id='TableChart_1'
        params: expected_list: ['Category','Product ID','Product','Unit Sales','Dollar Sales', 'TOTAL']
        Syntax: verify_row_total_report_titles_on_preview(5, 5,'TableChart_1',columns, 'Step 04: Verify preview pane')
        """
        
        actual_list = []
        temp_heading_css="#" + table_id + " div[class^='x']"
        heading_class=self.driver.find_elements_by_css_selector(temp_heading_css)
        report_heading_css="#"+table_id+" [class*='title']"
        total_items=self.driver.find_elements_by_css_selector(report_heading_css)
        tot=len(total_items)
        no_of_heading=tot/no_of_cells
        for i in range(colnum-1):
            tmp=""
            for j in range(i, int(tot/no_of_heading), colnum):
                tmp+=total_items[j].text.strip()
            actual_list.append(tmp)
        actual_list.append(heading_class[j+1].text.strip())
        utillityobject.asequal(self,expected_list, actual_list, msg)
        
        
    def select_autoprompt(self, field_name, item_name):
        filter_values_css="div.autop-pane div[class*='autop-amper-select']"
        filter_objs=self.driver.find_elements_by_css_selector(filter_values_css)
        filter_values=[el.get_attribute("title") for el in self.driver.find_elements_by_css_selector(filter_values_css)]
        filter_objs[filter_values.index(field_name)].find_element_by_css_selector("a[id^='ui-id']").click()
        time.sleep(1)
        if self.driver.find_element_by_css_selector("div[class^='autop-sav-panel'] input#av_search").is_displayed():
            self.driver.find_element_by_css_selector("div[class^='autop-sav-panel'] input#av_search").click()
            action = ActionChains(self.driver)
            action.send_keys(item_name)
            items=self.driver.find_elements_by_css_selector("div[class^='autop-sav-panel'] div[id='av_grp_values'] div[class^='ui-btn']")
            items[0].click()
        else:
            items=self.driver.find_elements_by_css_selector("div[id^='ui-id'][class*='selectmenu '] a[class^='ui-btn']")
            items_list=[el.text.strip() for el in self.driver.find_elements_by_css_selector("div[id^='ui-id'][class*='selectmenu '] a[class^='ui-btn']")]
            items[items_list.index(item_name)].click()
        time.sleep(2)
    
    def select_multiselect_autoprompt(self, field_name, expected_default_selected_text, item_list, **kwargs):
        """
        params: select_opt='all' or 'none'
        """
        filter_values_css="div.autop-pane div[class*='autop-amper-select']"
        filter_objs=self.driver.find_elements_by_css_selector(filter_values_css)
        filter_values=[el.get_attribute("title") for el in self.driver.find_elements_by_css_selector(filter_values_css)]
        filter_objs[filter_values.index(field_name)].find_element_by_css_selector("a[id^='ui-id']").click()
        time.sleep(1)
        radios=self.driver.find_elements_by_css_selector("div[class^='autop-sav-panel'] div[class*='ui-radio']")
        if 'msg' in kwargs:
            actual_default_selected_text=radios[0].text.strip()
            utillityobject.asequal(self, expected_default_selected_text, actual_default_selected_text, "Step X : The value " + expected_default_selected_text + " selected by default.")
        radios[1].click()
        time.sleep(1)
        if 'select_opt' in kwargs:
            self.driver.find_element_by_css_selector("div[class^='autop-sav-panel'] a[id='av_btn_sel_" + kwargs['select_opt'] + "']").click()
            time.sleep(1)
        if self.driver.find_element_by_css_selector("div[class^='autop-sav-panel'] input#av_search").is_displayed():
            self.driver.find_element_by_css_selector("div[class^='autop-sav-panel'] input#av_search").click()
            for required_item in item_list:
                action = ActionChains(self.driver)
                action.send_keys(required_item)
                items=self.driver.find_elements_by_css_selector("div[class^='autop-sav-panel'] div[id='av_grp_values'] div[class^='ui-btn']")
                items[0].click()
                del action

        time.sleep(2)  
    
    def select_or_verify_document_page_menu(self, item_name,default_page_name='Page 1', **kwargs):
        """
        param: item_name: 'Page 2'.
        param: kwargs['verify']: This will only verify the bipopup menu.
        param: kwargs['expected_popup_list']: This is a list to verify the menu list.
        param: kwargs['msg']: This is the message to be passed to asertion.
        Syntax: select_or_verify_document_page_menu('Page 2')
        Syntax: select_or_verify_document_page_menu('Page 2',verify='true',expected_popup_list=['', 'Page 1', 'New Page', 'Page Options...'],msg='Step 10: Verify popup menu')
        @author = Niranjan
        """
        IA_Resultarea.verify_current_document_page_name(self, default_page_name, 'Step X: Verify the default selected page name.')
        self.driver.find_element_by_css_selector("#iaPagesMenuBtn div[class$='tool-bar-menu-button-drop-down-arrow']").click()
        time.sleep(2)
        if 'verify' in kwargs:
            utillityobject.select_or_verify_bipop_menu(self, item_name, verify=True, expected_popup_list=kwargs['expected_popup_list'], msg=kwargs['msg'])
        else:
            utillityobject.select_or_verify_bipop_menu(self, item_name)
        time.sleep(3)

    
    def enter_text_in_Textbox(self, text_box, text_input, delete_count=20, type_speed=1):
        '''
        Syntax: ia_resultobj.enter_text_in_TextBox('Text_1', "This is simple text input")
        '''
        textbox_obj=self.driver.find_element_by_css_selector('#'+text_box)
        utillityobject.click_on_screen(self, textbox_obj, coordinate_type='middle', click_type=0)
        time.sleep(1)
        if sys.platform == 'linux':
            pykeyboard.tap_key(pykeyboard.home_key)
        else:
            keyboard.send('home')
        time.sleep(1)
        count=0
        while count<delete_count:
            if sys.platform == 'linux':
                pykeyboard.tap_key(pykeyboard.delete_key)
            else:
                keyboard.send('del')
            count+=1
        time.sleep(1)
        if sys.platform == 'linux':
            pykeyboard.type_string(str(text_input), interval=int(type_speed))
        else:
            keyboard.write(text_input, delay=int(type_speed))
        time.sleep(4)
    
    def verify_text_in_Textbox(self, text_box, exp_textbox_text, msg):
        '''
        Syntax: ia_resultobj.verify_text_in_Textbox('#Text_1', 'This is the 2nd page Textbox', "Step 32e: Verify Textbox text")
        '''
        frameid=self.driver.find_element_by_css_selector(text_box).find_element_by_tag_name('iframe').get_attribute('id')
        act_textbox_text=self.driver.execute_script("return document.getElementById('"+frameid+"').contentWindow.document.getElementsByTagName('body')[0].innerText").strip()
        utillityobject.asequal(self, act_textbox_text, exp_textbox_text, msg)
          
    def select_autoprompt_tool_button(self, btn_name):
        """
        params: btn_name='back' or 'refresh' or 'save' or 'submit'(This is for run)
        Usage: select_autoprompt_tool_button('submit',)
        Author: Niranjan
        """
        btn_css="div.autop-pane a[class^='autop-btn-" + btn_name + "']"
        self.driver.find_element_by_css_selector(btn_css).click()
        time.sleep(1)
      
    def select_field_on_canvas(self, table_id, click_expected_col, **kwargs):
        """
        params: table_id='TableChart_1'
        params: click_expected_col=1,2,3...
        Syntax: select_field_on_canvas("TableChart_1", 1)
        Usage : select_field_on_canvas("TableChart_1", 1, click_type=1)
        """
        
        table_css="#" + table_id + " div[id^='ActivePreviewItem']"
        total_items=self.driver.find_elements_by_css_selector(table_css)
        heading_cell_num=(click_expected_col*2)-1
        total_items[heading_cell_num].click()
        time.sleep(5)
        if 'click_type' in kwargs:
            table_item=total_items[heading_cell_num]
            utillityobject.click_on_screen(self, table_item, coordinate_type='middle',click_type=kwargs['click_type'])
            
    def click_preview_canvas(self, canvas_css,element_location='middle', xoffset=0, yoffset=0, action_chain_click=False, mouse_move_duration=0.5):
        '''
        Desc:- This function will click on preview canvas
        '''
        elem=utillityobject.validate_and_get_webdriver_object(self, canvas_css, 'Preview Canvas')
        coreutillityobject.left_click(self, elem,element_location=element_location, xoffset=xoffset, yoffset=yoffset, action_chain_click=action_chain_click,mouse_move_duration=mouse_move_duration)
    
    def verify_preview_field_value_property(self, table_css, field_value, property_name, property_value, msg):
        """
        table_css = "TableChart_1"
        field_value = "BMW"
        property_name = "color"
        property_value = 'rgba(255, 0, 128, 1)'
        Usage: verify_preview_field_value_property("TableChart_1","BMW","color", "rgba(255, 0, 128, 1)","Step 09.2: Verify color has been applied for field value BMW")
        """
        table_css="TableChart_1"
        items=self.driver.find_elements_by_css_selector("#" + table_css + " div[class^='x']")
        items_list=[el.text.strip() for el in self.driver.find_elements_by_css_selector("#" + table_css + " div[class^='x']")]
        actual_property_value=items[items_list.index(field_value)].value_of_css_property(property_name)
        utillityobject.asequal(self, property_value, actual_property_value, msg + " verify " + property_name + " applied for " + field_value)
    
    def get_preview_report_cell_object(self, cell_num, parent_css="#TableChart_1"):
        """
        Description : This method will return preview report cell object.
        Note : You can get cell number by search "#TableChart_1 div[class^='x']" css. 
        :Usage - get_preview_report_cell_object(5)
        """
        cell_css = parent_css + " div[class^='x']"
        cell_objects = utillityobject.validate_and_get_webdriver_objects(self, cell_css, "Preview report cell")
        return cell_objects[cell_num - 1]

    def select_preview_report_context_menu(self, cell_num, context_menu_path, parent_css="#TableChart_1"):
        """
        Description : This method will right click on preview report cell and select context menu
        :Note - cell_num should start from 1
        :Usage - select_preview_report_context_menu(2, "Sort->Rank->On")
        """
        cell_object = IA_Resultarea.get_preview_report_cell_object(self, cell_num, parent_css)
        coreutillityobject.python_left_click(self, cell_object)
        time.sleep(8)
        utillityobject.right_click_with_offset(self, cell_object)
        utillityobject.synchronize_until_element_is_visible(self, "div[id^='BiPopup'][style*='inherit']", 90)
        context_menu_list=context_menu_path.split('->')
        for context_menu in context_menu_list :
            utillityobject.select_bipopup_list_item(self, context_menu)
        
    def verify_chart_header_footer_property(self, parent_id, header_footer_index, **kwargs):
        """
        Params: parent_id = "TableChart_1"
        header_footer_index = 1,2...
        font_color = 'black'
        ================
        background-color
        bg_color='yellow'
        bg_cell_no=1,2.....     "This is only applicable for colored background cell number"
        ================
        text_value='CAR'
        font_name=Arial
        font_size='12pt'
        bold=True
        italics=True
        underline=True
        text_align='25px'
        msg=step 9:
        Usage: verify_chart_header_footer_property("TableChart_1", 1, bg_color='Cyan', font_color='magenta', text_value='CAR', font_name='Arial',font_size='12pt', bold=True, italic=True, underline=True, text_align='25px', msg='Step 6:')
        Usage for background-color:verify_chart_header_footer_property('TableChart_1', 1, bg_cell_no=1, bg_color='yellow', msg='Step 6:')
        """
        total_elems=self.driver.find_elements_by_css_selector("#" + parent_id + " div[style*='solid'] span")
        header_footer_elems=[el for el in total_elems if bool(re.match('\S', el.text.strip()))]
        for key in kwargs:
            if 'bg_color' in key:
                        expected_background_color=utillityobject.color_picker(self, kwargs['bg_color'], 'rgba')
                        actual_background_color=Color.from_string(total_elems[kwargs['bg_cell_no']-1].value_of_css_property("background-color")).rgba
                        utillityobject.asequal(self, actual_background_color, expected_background_color , kwargs['msg'] + ". Verification of Cell Background color.")
            if 'font_color' in key:
                expected_text_color=utillityobject.color_picker(self, kwargs['font_color'], 'rgba')
                actual_text_color=Color.from_string(header_footer_elems[header_footer_index-1].value_of_css_property("color")).rgba
                utillityobject.asequal(self, actual_text_color, expected_text_color, kwargs['msg'] + ". Verification of Cell Text color.")
            if 'text_value' in key:
                actual_text=header_footer_elems[header_footer_index-1].text.strip()
                utillityobject.asequal(self, kwargs['text_value'], actual_text, kwargs['msg'] + ". Verification of Cell Text.")
            if 'font_name' in key:
                actual_font=header_footer_elems[header_footer_index-1].value_of_css_property("font-family").strip('"')
                utillityobject.asequal(self, kwargs['font_name'].upper(), actual_font.upper(), kwargs['msg'] + ". Verification of Cell Text Font name.")
            if 'font_size' in key:
                expected_font_size=round(float(1.333333*int(kwargs['font_size'][:-2])))
                actual_font_size=header_footer_elems[header_footer_index-1].value_of_css_property("font-size")
                actual_font_size=round(float(actual_font_size[:-2]))
                utillityobject.asequal(self, expected_font_size, actual_font_size, kwargs['msg'] + ". Verification of Cell Text Font Size.")
            if 'bold' in key:
                actual_weight=True if header_footer_elems[header_footer_index-1].value_of_css_property("font-weight") in ['700', 'bold'] else False
                utillityobject.asequal(self, kwargs['bold'], actual_weight, kwargs['msg'] + ". Verification of Cell Text is Bold.")
            if 'italic' in key:
                actual_style=True if header_footer_elems[header_footer_index-1].value_of_css_property("font-style")=='italic' else False
                utillityobject.asequal(self, kwargs['italic'], actual_style, kwargs['msg'] + ". Verification of Cell Text is Italics.")
            if 'underline' in key:
                actual_decoration=True if header_footer_elems[header_footer_index-1].value_of_css_property("text-decoration")=='underline' else False
                utillityobject.asequal(self, kwargs['underline'], actual_decoration, kwargs['msg'] + ". Verification of Cell is underlined.")
            if 'text_align' in key:
                actual_align=header_footer_elems[header_footer_index-1].value_of_css_property("left")
                val1=int(kwargs['text_align'][:-2])
                utillityobject.asin(self, range(val1 - 5,val1 + 5), actual_align[:-2], kwargs['msg'] + ". Verification of Cell Text alignment.")
    
    
    
    def verify_report_header_footer_property(self, parent_id, header_footer_index, **kwargs):
        """
        Params: parent_id = "TableChart_1"
        header_footer_index = 1,2...
        font_color = 'black'
        ================
        background-color
        bg_color='yellow'
        cell_no=1,2.....    "This is only applicable for colored background cell number"
        ================
        text_value='CAR'
        font_name=Arial
        font_size='12pt'
        bold=True
        italics=True
        underline=True
        text_align='25px'
        msg=step 9:
        Usage: verify_report_header_footer_property("TableChart_1", 1,bg_color='Cyan', font_color='magenta', text_value='CAR', font_name='Arial',font_size='12pt', bold=True, italic=True, underline=True, text_align='25px', msg='Step 6:')
        Usage for background-color:verify_report_header_footer_property('TableChart_1', 1, cell_no=1, bg_color='yellow', msg='Step 6:')
        """
        total_elems=self.driver.find_elements_by_css_selector("#" + parent_id + " div[style*='solid'] span")
        header_footer_elems=[el for el in total_elems if bool(re.match('\S', el.text.strip()))]
        item_bgs=self.driver.find_elements_by_css_selector("#" + parent_id + " div[style*='solid'] div[style*='background'][style*='position']")
        for key in kwargs:
            if 'bg_color' in key:
                try:
                    if 'bg_color' in key:
                        expected_background_color=utillityobject.color_picker(self, kwargs['bg_color'], 'rgba')
                        actual_background_color=Color.from_string(item_bgs[kwargs['cell_no']-1].value_of_css_property("background-color")).rgba
                        utillityobject.asequal(self, actual_background_color, expected_background_color , kwargs['msg'] + ". Verification of Cell Background color.")
                except:
                    utillityobject.asequal(self, False, expected_background_color , kwargs['msg'] + ". Verification of Cell Background color.")
            if 'font_color' in key:
                expected_text_color=utillityobject.color_picker(self, kwargs['font_color'], 'rgba')
                actual_text_color=Color.from_string(header_footer_elems[header_footer_index-1].value_of_css_property("color")).rgba
                utillityobject.asequal(self, actual_text_color, expected_text_color, kwargs['msg'] + ". Verification of Cell Text color.")
            if 'text_value' in key:
                actual_text=header_footer_elems[header_footer_index-1].text.strip()
                utillityobject.asequal(self, kwargs['text_value'], actual_text, kwargs['msg'] + ". Verification of Cell Text.")
            if 'font_name' in key:
                actual_font=header_footer_elems[header_footer_index-1].value_of_css_property("font-family").strip('"')
                utillityobject.asequal(self, kwargs['font_name'].upper(), actual_font.upper(), kwargs['msg'] + ". Verification of Cell Text Font name.")
            if 'font_size' in key:
                if 'font_size_pixel' in kwargs:
                    expected_font_size=(round(float(kwargs['font_size'][:-2])))
                else: 
                    expected_font_size=round(float(1.333333*int(kwargs['font_size'][:-2])))
                actual_font_size=header_footer_elems[header_footer_index-1].value_of_css_property("font-size")
                actual_font_size=round(float(actual_font_size[:-2]))
                utillityobject.asequal(self, expected_font_size, actual_font_size, kwargs['msg'] + ". Verification of Cell Text Font Size.")
            if 'bold' in key:
                actual_weight=True if header_footer_elems[header_footer_index-1].value_of_css_property("font-weight") in ['700', 'bold'] else False
                utillityobject.asequal(self, kwargs['bold'], actual_weight, kwargs['msg'] + ". Verification of Cell Text is Bold.")
            if 'italic' in key:
                actual_style=True if header_footer_elems[header_footer_index-1].value_of_css_property("font-style")=='italic' else False
                utillityobject.asequal(self, kwargs['italic'], actual_style, kwargs['msg'] + ". Verification of Cell Text is Italics.")
            if 'underline' in key:
                actual_decoration=True if header_footer_elems[header_footer_index-1].value_of_css_property("text-decoration")=='underline' else False
                utillityobject.asequal(self, kwargs['underline'], actual_decoration, kwargs['msg'] + ". Verification of Cell is underlined.")
            if 'text_align' in key:
                actual_align=header_footer_elems[header_footer_index-1].value_of_css_property("left")
                val1=int(kwargs['text_align'][:-2])
                utillityobject.asin(self, int(actual_align[:-2]), range(val1 - 5,val1 + 5), kwargs['msg'] + ". Verification of Cell Text alignment.")
    
    
    def verify_report_cell_property(self, parent_id, cell_no, **kwargs):
        """
        Params: parent_id = "TableChart_1"
        cell_no = 1, 2, 3...any integer value for cell number, including heading cells.(excluding empty cells in the table report)
        font_color = 'black'
        bg_cell_no = 1,2,3... any integer value for cell number, including heading cells and empty calls in the table report
        bg_color='red'
        text='CAR'
        font_name='Arial'
        font_size='10pt'
        bold=True
        italic=True
        underline=True
        text_align='left', OR, 'center' Or 'right'
        Usage: verify_report_cell_property("TableChart_1", 1, bg_cell_no = 2,bg_color='Cyan', font_color='magenta', text_value='CAR', font_name='Arial', font_size='10pt', bold=True, italic=True, underline=True, text_align='Center')
        """
        items=self.driver.find_elements_by_css_selector("#" + parent_id + " div[class^='x']")
        item_bgs=self.driver.find_elements_by_css_selector("#" + parent_id + " div[style*='solid'] div[style*='background']")
        for key in kwargs:
            if 'bg_color' in key:
                expected_background_color=utillityobject.color_picker(self, kwargs['bg_color'], 'rgba')
                actual_background_color=Color.from_string(item_bgs[kwargs['bg_cell_no']+1].value_of_css_property("background-color")).rgba
                utillityobject.asin(self, actual_background_color, expected_background_color , kwargs['msg'] + ". Verification of Cell Background color.")
            if 'font_color' in key:
                expected_text_color=utillityobject.color_picker(self, kwargs['font_color'], 'rgba')
                actual_text_color=Color.from_string(items[cell_no-1].value_of_css_property("color")).rgba
                utillityobject.asin(self, actual_text_color, expected_text_color, kwargs['msg'] + ". Verification of Cell Text color.")
            if 'text_value' in key:
                actual_text=items[cell_no-1].text.strip()
                utillityobject.asequal(self, kwargs['text_value'], actual_text, kwargs['msg'] + ". Verification of Cell Text.")
            if 'font_name' in key:
                actual_font=items[cell_no-1].value_of_css_property("font-family").strip('"').upper()
                utillityobject.asequal(self, kwargs['font_name'].upper(), actual_font, kwargs['msg'] + ". Verification of Cell Text Font name.")
            if 'font_size' in key:
                expected_font_size=round(float(1.333333*int(kwargs['font_size'][:-2])))
                actual_font_size=items[cell_no-1].value_of_css_property("font-size")
                actual_font_size=round(float(actual_font_size[:-2]))
                utillityobject.asequal(self, expected_font_size, actual_font_size, kwargs['msg'] + ". Verification of Cell Text Font Size.")
            if 'bold' in key:
                actual_weight=True if items[cell_no-1].value_of_css_property("font-weight") in ['700', 'bold'] else False
                utillityobject.asequal(self, kwargs['bold'], actual_weight, kwargs['msg'] + ". Verification of Cell Text is Bold.")
            if 'italic' in key:
                actual_style=True if items[cell_no-1].value_of_css_property("font-style")=='italic' else False
                utillityobject.asequal(self, kwargs['italic'], actual_style, kwargs['msg'] + ". Verification of Cell Text is Italics.")
            if 'underline' in key:
                actual_decoration=True if 'underline' in items[cell_no-1].value_of_css_property("text-decoration") else False
                utillityobject.asequal(self, kwargs['underline'], actual_decoration, kwargs['msg'] + ". Verification of Cell is underlined.")
            if 'text_align' in key:
                actual_font=items[cell_no-1].value_of_css_property("text-align")
                utillityobject.asequal(self, kwargs['text_align'], actual_font, kwargs['msg'] + ". Verification of Cell Text alignment.")

    
    def verify_live_preview_data_bars(self, parent_id, msg, **kwargs):
        """
        params: parent_id='TableChart_1'
        kwargs: expected_number_of_bars=1, 2, 3, 4....(Integer value toverify the number of bars present.)
        kwargs: index=0, 1, 2....(Integer value for the position of the bars)
        kwargs: color='green', 'red' , etc...
        Usage: verify_live_preview_data_bars('TableChart_1',10, 'Step 10: xxxx', index=3, color='green')
        Author: Niranjan
        """
        data_bar_css="#" + parent_id + " div[style*='position:'][style*='background-color'][style*='border']"
        bar_items=self.driver.find_elements_by_css_selector(data_bar_css)
        if 'expected_number_of_bars' in kwargs:
            actual_number_of_bars=len(bar_items)
            utillityobject.asequal(self, kwargs['expected_number_of_bars'], actual_number_of_bars, msg + " Verify Number of data bars.") 
        if 'index' in kwargs:
            expected_color=utillityobject.color_picker(self, kwargs['color'], 'rgba')
            actual_color=Color.from_string(bar_items[kwargs['index']].value_of_css_property("background-color")).rgba
            utillityobject.asequal(self, expected_color, actual_color, msg + " Verify color of the given data bar.")
    

    def select_autolink_tooltip_menu(self,parent_id, raiser_class, menu_path, wait_time=2, **kwargs):
        """
        param parent_id = 'MAINTABLE_wbody1'
        param : raiser_class = 'riser!s4!g4!mbar!'
        param : menu_path = 'Drill up to->Product Category' (At this point, i am considering one level of sub menu. We can enhance, if the level increases.)
        Usage: select_autolink_tooltip_menu("MAINTABLE_wbody1","riser!s0!g0!mbar", 'Drill up to->Product Category')
        Usage: select_autolink_tooltip_menu("MAINTABLE_wbody1","riser!s0!g0!mbar", 'Filter Chart')
        Author: Niranjan
        """
        move = self.driver.find_element_by_css_selector("#"+ parent_id)
        utillityobject.click_on_screen(self, move, coordinate_type='start', click_type=3)
        if 'MAINTABLE' in parent_id:
            tooltip_css1="div[class='tdgchart-submenu']>div>ul>li>span"
        else:
            tooltip_css1="div[class='tdgchart-submenu']>div>ul>li"
        #browser = utillityobject.parseinitfile(self, 'browser')
        #utillityobject.click_on_screen(self, obj_locator, coordinate_type='middle', x_offset=20, click_type=0)
        action = ActionChains(self.driver)
        raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
        tooltip_css="span[id*='tdgchart-tooltip'][style*='visible']>div>ul>li"
        #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip'][style*='visible']>div>ul>li"  
        obj_locator=self.driver.find_element_by_css_selector(raiser_css)
        if Global_variables.browser_name=="firefox":
            if 'x_offset' in kwargs:
                utillityobject.click_on_screen(self, obj_locator, coordinate_type='start', x_offset=kwargs['x_offset']+20, y_offset=kwargs['y_offset'])
            else:
#                 utillityobject.click_on_screen(self, obj_locator, coordinate_type='middle')
                coreutillityobject.python_move_to_element(self, obj_locator)
        else:
            if 'action_x_offset' in kwargs:
                action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css), xoffset=kwargs['action_x_offset'], yoffset=kwargs['action_y_offset']).perform()
            else:
#                 action.move_to_element(obj_locator).perform()
                coreutillityobject.python_move_to_element(self, obj_locator)
        #Original function#
#         tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
#         for i in range(len(tooltips)):
#             line1=tooltips[i].text.strip()
#             if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
#                 tooltips[i].click()
#                 break
#         if len(menus)>1:
#             tooltips=self.driver.find_elements_by_css_selector(tooltip_css1)
#             if self.browser in ('IE', 'Firefox'):
#                 time.sleep(1)
#             tooltips1=self.driver.find_elements_by_css_selector(tooltip_css1)
#             for i in range(len(tooltips1)):
#                 if menus[1] in (tooltips1[i].text.strip()).split('\n'):
#                     tooltips1[i].click()
#                     break
#         time.sleep(4) 
        
        
        utillityobject.synchronize_with_number_of_element(self, "span[id*='tdgchart-tooltip'][style*='visible']", 1, expire_time=4)
        tooltips=menu_path.split('->')
        #tooltip_css="#my_tooltip_id ul[class='tdgchart-tooltip-list']>li[class*='tdgchart-tooltip-pad']"
        tooltip_obj=self.driver.find_element_by_css_selector(tooltip_css)
        utillityobject.click_on_screen(self, tooltip_obj, 'top_middle', y_offset=5, mouse_duration=0)
        for count, tooltip in enumerate(tooltips) :
            tooltip_elements=self.driver.find_elements_by_css_selector(tooltip_css)
            tooltip_text_list=[tooltip.text.strip().replace('>', '').replace('\n', '') for tooltip in tooltip_elements]
            tooltip_element=tooltip_elements[tooltip_text_list.index(tooltip)]
            if tooltips[count] == tooltips[-1] :
                coordinate_type='middle'
                utillityobject.click_on_screen(self, tooltip_element, coordinate_type=coordinate_type, click_type=0)
            else :
                utillityobject.click_on_screen(self, tooltip_element, coordinate_type='left', x_offset=2)
                tooltip_arrow=tooltip_element.find_element_by_css_selector("span[class='tdgchart-tooltip-arrow']")
                if 'coordinate_type' in kwargs:                    
                    coordinate_type=kwargs['coordinate_type']
                else:
                    coordinate_type='middle'          
                utillityobject.click_on_screen(self, tooltip_arrow, coordinate_type=coordinate_type)
                time.sleep(2)
            tooltip_css=tooltip_css1
        time.sleep(10)
        
        """ Original Function
        action1 = ActionChains(self.driver)
        #move = self.driver.find_element_by_css_selector("g[class='chartPanel']")
        move = self.driver.find_element_by_css_selector("#"+ parent_id)
        utillityobject.click_on_screen(self, move, coordinate_type='start', click_type=3)
        '''if self.browser == 'Firefox':
            utillityobject.click_type_using_pyautogui(self, move, move=True, source_x=kwargs['x_offset'], source_y=kwargs['y_offset'])
        else:
            action1.move_to_element_with_offset(move,1,1).perform()
        time.sleep(2)
        del action1'''
        #browser = utillityobject.parseinitfile(self, 'browser')
        action = ActionChains(self.driver)
        menus=menu_path.split('->')
        raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
        tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip']>div>ul>li"  
        obj_locator=self.driver.find_element_by_css_selector(raiser_css)
        #action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css),1,1).perform()
        #time.sleep(2)
        if self.browser=="Firefox":
            if 'x_offset' in kwargs:
                utillityobject.click_on_screen(self, obj_locator, coordinate_type='start', x_offset=kwargs['x_offset']+20, y_offset=kwargs['y_offset'], click_type=0)
                #utillityobject.click_on_screen(self, obj_locator, source_x=kwargs['x_offset'], source_y=kwargs['y_offset'])
            else:
                utillityobject.click_on_screen(self, obj_locator, coordinate_type='start', x_offset=20, click_type=0)
        else:
            if 'action_x_offset' in kwargs:
                action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css), xoffset=kwargs['action_x_offset'], yoffset=kwargs['action_y_offset']).perform()
            else:
                action.move_to_element(obj_locator).perform()
        time.sleep(1)
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        for i in range(len(tooltips)):
            #if menus[0] in (tooltips[i].text.strip()).split('\n'):
            line1=tooltips[i].text.strip()
            if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                tooltips[i].click()
                break
        if len(menus)>1:
            tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            #action1=ActionChains(self.driver)
            #action1.move_to_element(tooltips[0]).perform()
            #del action1
            if self.browser in ('IE', 'Firefox'):
                time.sleep(1)
            #tooltip_css1="div[class='tdgchart-submenu']>div>ul>li>span"
            tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips1)):
                if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                    tooltips1[i].click()
                    break
        time.sleep(4) """
    
    def select_autolink_tooltip_menu_with_offset(self,parent_id, raiser_class, menu_path, wait_time=2, **kwargs):
        """
        param parent_id = 'MAINTABLE_wbody1'
        param : raiser_class = 'riser!s4!g4!mbar!'
        param : menu_path = 'Drill up to->Product Category' (At this point, i am considering one level of sub menu. We can enhance, if the level increases.)
        Usage: select_autolink_tooltip_menu("MAINTABLE_wbody1","riser!s0!g0!mbar", 'Drill up to->Product Category')
        Usage: select_autolink_tooltip_menu("MAINTABLE_wbody1","riser!s0!g0!mbar", 'Filter Chart')
        Author: Niranjan
        """
        move = self.driver.find_element_by_css_selector("#"+ parent_id)
        utillityobject.click_on_screen(self, move, coordinate_type='start', click_type=3)
        '''if self.browser == 'Firefox':
            utillityobject.click_type_using_pyautogui(self, move, move=True, source_x=kwargs['x_offset'], source_y=kwargs['y_offset'])
        else:
            action1.move_to_element_with_offset(move,1,1).perform()
        time.sleep(2)
        del action1'''
        if 'MAINTABLE' in parent_id:
            tooltip_css1="div[class='tdgchart-submenu']>div>ul>li>span"
        else:
            tooltip_css1="div[class='tdgchart-submenu']>div>ul>li"
        #browser = utillityobject.parseinitfile(self, 'browser')
        action = ActionChains(self.driver)
        menus=menu_path.split('->')
        raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
        tooltip_css="span[id*='tdgchart-tooltip'][style*='visible']>div>ul>li"
        #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip'][style*='visible']>div>ul>li" 
        obj_locator=self.driver.find_element_by_css_selector(raiser_css)
        #action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css),1,1).perform()
        #time.sleep(2)
        if Global_variables.browser_name=="firefox":
            if 'x_offset' in kwargs:
                utillityobject.click_on_screen(self, obj_locator, coordinate_type='start', x_offset=kwargs['x_offset']+20, y_offset=kwargs['y_offset'], click_type=0)
                #utillityobject.click_on_screen(self, obj_locator, source_x=kwargs['x_offset'], source_y=kwargs['y_offset'])
            else:
                utillityobject.click_on_screen(self, obj_locator, coordinate_type='start', x_offset=20, y_offset=10)
        else:
            if 'action_x_offset' in kwargs:
                action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css), xoffset=kwargs['action_x_offset'], yoffset=kwargs['action_y_offset']).perform()
            else:
                action.move_to_element(obj_locator).perform()
        time.sleep(1)
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        utillityobject.click_on_screen(self, tooltips[1], coordinate_type='start', x_offset=20, click_type=0)
        if Global_variables.browser_name=="firefox":
            if 'x_offset' in kwargs:
                utillityobject.click_on_screen(self, obj_locator, coordinate_type='start', x_offset=kwargs['x_offset']+20, y_offset=kwargs['y_offset'], click_type=0)
                #utillityobject.click_on_screen(self, obj_locator, source_x=kwargs['x_offset'], source_y=kwargs['y_offset'])
            else:
                utillityobject.click_on_screen(self, obj_locator, coordinate_type='start', x_offset=20, click_type=0)
        else:
            if 'action_x_offset' in kwargs:
                action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css), xoffset=kwargs['action_x_offset'], yoffset=kwargs['action_y_offset']).perform()
            else:
                action.move_to_element(obj_locator).perform()
        time.sleep(1)
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        for i in range(len(tooltips)):
            #if menus[0] in (tooltips[i].text.strip()).split('\n'):
            line1=tooltips[i].text.strip()
            if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                utillityobject.click_on_screen(self, tooltips[i], coordinate_type='start', x_offset=20, y_offset=12, click_type=0)
                utillityobject.click_on_screen(self, tooltips[i], coordinate_type='middle', x_offset=20)
                #tooltips[i].click()
                break
        if len(menus)>1:
            #tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css1)
            #action1=ActionChains(self.driver)
            #action1.move_to_element(tooltips[0]).perform()
            #del action1
            if Global_variables.browser_name in ('ie', 'firefox', 'edge'):
                time.sleep(1)
            #tooltip_css1="div[class='tdgchart-submenu']>div>ul>li>span"
            tooltips1=self.driver.find_elements_by_css_selector(tooltip_css1)
            for i in range(len(tooltips1)):
                if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                    utillityobject.click_on_screen(self, tooltips1[i], coordinate_type='middle', x_offset=20, y_offset=15)
                    tooltips1[i].click()
                    break
        time.sleep(4) 
    
    def verify_infoassist_row_column_header_labels(self,parent_id,matrix_type,expected_header,expected_label,msg):
        """
        :param parent_id: MAINTABLE_wbody1
        :param matrix_type: 'Rows' or 'columns'
        :param expected_rowheader: 'Product Category'
        :param expected_rowlabel: ['Hazelnut', 'B141', 'French Roast', 'B142', 'Kona', 'B144', 'Scone', 'F101', 'Biscotti']
        :param msg: "Step 10"
        Usage: verify_infoassist_row_column_header_labels('MAINTABLE_wbody1','Product Category', ['Hazelnut', 'B141', 'French Roast', 'B142', 'Kona', 'B144', 'Scone', 'F101', 'Biscotti'], "Step 04:")
        @author: Niranjan
        """
        actual_val_list=[]
        if matrix_type == 'Rows':
            header_css="#" + parent_id + " g.scrollRowTitle"
            label_css="#" + parent_id + " g.scrollRowLbl > g"
        else:
            header_css="#" + parent_id + " g.scrollColTitle"
            label_css="#" + parent_id + " g.scrollColLbl > g"

        actual_rowheader=self.driver.find_element_by_css_selector(header_css).text.strip()
        utillityobject.asequal(self, expected_header, actual_rowheader, msg + " Verify Header tile for " + matrix_type)
        actual_val_list.extend([x for x in [el.text.strip() for el in self.driver.find_elements_by_css_selector(label_css)] if x!=''])
        for label_x in expected_label:
            if label_x in actual_val_list:
                state= True
            else:
                state=False
                break
        utillityobject.asequal(self,state, True, msg + " Verify Header labels for " + matrix_type) 
               
    def select_ia_autolink_tooltip_menu(self,parent_id, raiser_class, menu_path, wait_time=2, **kwargs):
        """
        param parent_id = 'MAINTABLE_wbody1'
        param : raiser_class = 'riser!s4!g4!mbar!'
        param : menu_path = 'Drill up to->Product Category' (At this point, i am considering one level of sub menu. We can enhance, if the level increases.)
        Usage: select_ia_autolink_tooltip_menu("MAINTABLE_wbody1","riser!s0!g0!mbar", 'Drill up to->Product Category')
        Usage: select_ia_autolink_tooltip_menu("MAINTABLE_wbody1","riser!s0!g0!mbar", 'Filter Chart')
        Author: Niranjan
        """
        time.sleep(10)
        if 'MAINTABLE' in parent_id:
            tooltip_css1="div[class='tdgchart-submenu']>div>ul>li>span"
        else:
            tooltip_css1="div[class='tdgchart-submenu']>div>ul>li"
        action1 = ActionChains(self.driver)
        move = self.driver.find_element_by_css_selector("#"+ parent_id)
        #utillityobject.click_on_screen(self, move, coordinate_type='start', click_type=3)
        if Global_variables.browser_name=="firefox":
            pyautogui.moveTo(Global_variables.current_working_area_browser_x + 15, Global_variables.current_working_area_browser_y + 15)
        else:
            action1.move_to_element_with_offset(move,1,1).perform()
        del action1
        time.sleep(4)
        action = ActionChains(self.driver)
        menus=menu_path.split('->')
        raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
        tooltip_css="span[id*='tdgchart-tooltip']>div>ul>li"
        #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip']>div>ul>li"   
        obj_locator=self.driver.find_element_by_css_selector(raiser_css)
        if Global_variables.browser_name in ('firefox', 'ie', 'edge'):
            coreutillityobject.python_move_to_element(self, obj_locator, element_location='middle', xoffset=10)
            time.sleep(1)
            #utillityobject.click_on_screen(self, obj_locator, coordinate_type='middle', click_type=3)    , x_offset=source_offset_x, y_offset=source_offset_y
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    coreutillityobject.python_left_click(self, tooltips[i])
                    break
            if len(menus)>1:
                tooltips=self.driver.find_elements_by_css_selector(tooltip_css1)
                tooltips1=self.driver.find_elements_by_css_selector(tooltip_css1)
                for i in range(len(tooltips1)):
                    if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                        coreutillityobject.python_left_click(self, tooltips1[i])
                        break
            time.sleep(4)
#         elif Global_variables.browser_name=="ie":
#             if 'action_x_offset' in kwargs:
#                 action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css), xoffset=kwargs['action_x_offset'], yoffset=kwargs['action_y_offset']).perform()
#             else:
#                 action.move_to_element(obj_locator).perform()
#             time.sleep(1)
#             tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
#             for i in range(len(tooltips)):
#                 line1=tooltips[i].text.strip()
#                 if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
#                     action1=ActionChains(self.driver)
#                     action1.move_to_element_with_offset(tooltips[i],xoffset=25,yoffset=8).click().perform()
#                     del action1
#                     break
#             if len(menus)>1:
#                 #tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
#                 tooltips=self.driver.find_elements_by_css_selector(tooltip_css1)
#                 time.sleep(1)
#                 tooltips1=self.driver.find_elements_by_css_selector(tooltip_css1)
#                 for i in range(len(tooltips1)):
#                     if menus[1] in (tooltips1[i].text.strip()).split('\n'):
#                         action1=ActionChains(self.driver)
#                         action1.move_to_element_with_offset(tooltips1[i],xoffset=25,yoffset=8).click().perform()
#                         del action1
#                         break
#             time.sleep(4)
        else:
            if 'action_x_offset' in kwargs:
                action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css), xoffset=kwargs['action_x_offset'], yoffset=kwargs['action_y_offset']).perform()
            else:
                action.move_to_element(obj_locator).perform()
            time.sleep(1)
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    tooltips[i].click() #ActionChains click not working hence reverting back to selenium click
#                     action1=ActionChains(self.driver)
#                     action1.move_to_element_with_offset(tooltips[i],xoffset=25,yoffset=8).click().perform()
#                     del action1
                    break
            time.sleep(1)
            if len(menus)>1:
                #tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
                tooltips=self.driver.find_elements_by_css_selector(tooltip_css1)
                tooltips1=self.driver.find_elements_by_css_selector(tooltip_css1)
                for i in range(len(tooltips1)):
                    if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                        tooltips1[i].click()
                        break
            time.sleep(4)  

    
    
    def verify_autolink_tooltip_submenu(self,parent_id, raiser_class, drill_menu_item, expected_tooltip_list, msg, wait_time=2, **kwargs):
        """
        param parent_id = 'MAINTABLE_wbody1'         
        param : raiser_class = 'riser!s4!g4!mbar!'
        param : menu_path = 'Drill up to'
        param : expected_tooltip_list = ['Store Business Sub Region', 'Sale Year/Quarter'] ==> tootip submenu  
        Usage: verify_autolink_tooltip_submenu('MAINTABLE_wbody0','riser!s5!g0!mbar!r0!c0!', "Drill down to", a, "Step 05a: Verify the drill menu shows for only Region and Year.")
        Author: Niranjan
        """
        
        #action1 = ActionChains(self.driver)
        move = self.driver.find_element_by_css_selector("#"+ parent_id)
        utillityobject.click_on_screen(self, move, coordinate_type='start', click_type=3)
        '''if self.browser == 'Firefox':
            x_fr=move.location['x']
            y_fr=move.location['y']
            pyautogui.moveTo(x_fr + kwargs['x_offset'], y_fr + kwargs['y_offset'])
        else:
            action1.move_to_element_with_offset(move,1,1).perform()
        time.sleep(2)
        del action1'''
        m_offset_x=kwargs['x_offset'] if 'x_offset' in kwargs else 0
        m_offset_y=kwargs['y_offset'] if 'y_offset' in kwargs else 0
#         s_offset_x=kwargs['x_offset_menu1'] if 'y_offset_menu1' in kwargs else 0
#         s_offset_y=kwargs['y_offset_menu1'] if 'y_offset_menu1' in kwargs else 0
        action = ActionChains(self.driver)
        menus=drill_menu_item.split('->')
        raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
        tooltip_css="span[id*='tdgchart-tooltip']>div>ul>li"
        #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip']>div>ul>li"    
        obj_locator=self.driver.find_element_by_css_selector(raiser_css)
        if Global_variables.browser_name=="firefox":
            #utillityobject.click_type_using_pyautogui(self, obj_locator, kwargs['x_offset'], kwargs['y_offset'])
            utillityobject.click_on_screen(self, obj_locator, coordinate_type='middle', x_offset=m_offset_x, y_offset=m_offset_y, click_type=3)
            tooltip_obj=self.driver.find_element_by_css_selector(tooltip_css)
            utillityobject.click_on_screen(self, tooltip_obj, 'top_middle', y_offset=5, mouse_duration=0)
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    #tooltips[i].click()
                    utillityobject.click_on_screen(self, tooltips[i], coordinate_type='left', x_offset=2)
                    tooltip_arrow=tooltips[i].find_element_by_css_selector("span[class='tdgchart-tooltip-arrow']")
                    utillityobject.click_on_screen(self, tooltip_arrow, coordinate_type='middle')
                    #utillityobject.click_on_screen(self, tooltips[i], coordinate_type='start', x_offset=s_offset_x, y_offset=s_offset_y, click_type=0)
                    #utillityobject.click_type_using_pyautogui(self, tooltips[i], kwargs['x_offset_menu1'], kwargs['y_offset_menu1'])
                    break
#             utillityobject.synchronize_with_number_of_element(self, "span[id*='tdgchart-tooltip'][style*='visible']", 1, expire_time=4)
#         tooltips=menu_path.split('->')
#         #tooltip_css="#my_tooltip_id ul[class='tdgchart-tooltip-list']>li[class*='tdgchart-tooltip-pad']"
#         tooltip_obj=self.driver.find_element_by_css_selector(tooltip_css)
#         utillityobject.click_on_screen(self, tooltip_obj, 'top_middle', y_offset=5, mouse_duration=0)
#         for count, tooltip in enumerate(tooltips) :
#             tooltip_elements=self.driver.find_elements_by_css_selector(tooltip_css)
#             tooltip_text_list=[tooltip.text.strip().replace('>', '').replace('\n', '') for tooltip in tooltip_elements]
#             tooltip_element=tooltip_elements[tooltip_text_list.index(tooltip)]
#             if tooltips[count] == tooltips[-1] :
#                 utillityobject.click_on_screen(self, tooltip_element, coordinate_type='middle', click_type=0)
#             else :
#                 utillityobject.click_on_screen(self, tooltip_element, coordinate_type='left', x_offset=2)
#                 tooltip_arrow=tooltip_element.find_element_by_css_selector("span[class='tdgchart-tooltip-arrow']")
#                 utillityobject.click_on_screen(self, tooltip_arrow, coordinate_type='middle')
#                 time.sleep(2)
#             tooltip_css=tooltip_css1
#         time.sleep(10)
        
        
        elif Global_variables.browser_name in ('ie', 'edge'):
            if 'action_x_offset' in kwargs:
                action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css), xoffset=kwargs['action_x_offset'], yoffset=kwargs['action_y_offset']).perform()
            else:
                action.move_to_element(obj_locator).perform()
            time.sleep(1)
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    action1=ActionChains(self.driver)
                    action1.move_to_element_with_offset(tooltips[i],xoffset=25,yoffset=8).click().perform()
                    del action1
                    break
        else:
            if 'action_x_offset' in kwargs:
                action.move_to_element_with_offset(self.driver.find_element_by_css_selector(raiser_css), xoffset=kwargs['action_x_offset'], yoffset=kwargs['action_y_offset']).perform()
            else:
                action.move_to_element(obj_locator).perform()
            time.sleep(1)
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            for i in range(len(tooltips)):
                line1=tooltips[i].text.strip()
                if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                    tooltips[i].click()
                    break
        time.sleep(1)
        actual_list = []
        tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        if Global_variables.browser_name in ('ie', 'firefox', 'edge'):
            time.sleep(1)
        tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
        for i in range(len(tooltips1)):
            if tooltips1[i].text.strip()!='':
                actual_list.append(tooltips1[i].text.strip())
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
        time.sleep(4) 
    
    def backup_autolink_tooltip_menu(self,elem, parent_css, raiser_class, menu_path, expected_no_of_risers, **kwargs):
            time.sleep(25)
            if len(self.driver.find_elements(By.CSS_SELECTOR, elem))!=expected_no_of_risers:
                pyautogui.hotkey('f11')
                time.sleep(10)
                pyautogui.hotkey('f11')
                time.sleep(20)
                IA_Resultarea.select_ia_autolink_tooltip_menu(self,parent_css, raiser_class, menu_path,wait_time=1, **kwargs)
                time.sleep(25)
            else:
                pass
    
    def verify_autolink(self, table_css, field_value, expected_no_of_hlinks, msg,**kwargs):
        """
        table_css = "TableChart_1"
        field_value = "Accessories"
        expected_no_of_hlinks = 7
        kwargs = link_color='magenta'
        Usage: verify_autolink("TableChart_1","Accessories",7,"Step 09.2: Verify Auto Drill applied in Accessories")
        """
#         temp_heading_css="#" + table_css + " div[class^='x']"
        link_css="#" + table_css + " div[class^='x'] a[href]"
        
        actual_no_of_hlinks=len(self.driver.find_elements_by_css_selector(link_css))
        utillityobject.asequal(self, expected_no_of_hlinks, actual_no_of_hlinks, msg + 'a: verify number of hyper-link displayed in a column')
        
        value_href_list=[el.text.strip() for el in self.driver.find_elements_by_css_selector(link_css)]
        utillityobject.asin(self, field_value, value_href_list, msg + 'b: verify hyper-link displayed')        
        
        if 'link_color' in kwargs:
            expected_link_color=utillityobject.color_picker(self, kwargs['link_color'], 'rgba')
        else:
            expected_link_color='rgba(51, 102, 255, 1)'        
        actual_link_color=Color.from_string(self.driver.find_elements_by_css_selector(link_css)[0].value_of_css_property("color")).rgba
        utillityobject.asequal(self, expected_link_color, actual_link_color, msg + 'd: verify hyper-link color')

    def expand_field_tree(self, folder_path):
        """
        :Param : folder_path='Store->Store->Store,Country->'
        Syntax: expand_field_tree('Store->Store->Store,Country->')
        @author = Nasir
        """
        Datatree_rows = "#iaMetaDataBrowser div[id^='QbMetaDataTree'] table>tbody>tr"
        folder_list=folder_path.split('->')[:-1]
        for item in folder_list:
            datetree_items = self.driver.find_elements_by_css_selector(Datatree_rows)
            for i in range(len(datetree_items)):
                folder_img = datetree_items[i].find_element_by_css_selector("td>img")
                img_src=folder_img.get_attribute("src")
                if datetree_items[i].text.strip() == item and 'path_arrow_tree_closed' in img_src:
                    try:
                        datetree_items = self.driver.find_elements_by_css_selector(Datatree_rows)
                        datetree_items[i].find_element_by_css_selector("td>img").click()
                        time.sleep(1)
                        break
                    except:
                        datetree_items = self.driver.find_elements_by_css_selector(Datatree_rows)
                        datetree_items[i].find_element_by_css_selector("td>img").click()
                        time.sleep(1)
                        break
    
    def verify_color_scale_esri_maps(self, popup_id, expected_color_scale, msg):
        """    Usage::: this function combines both title and labels
        verify_color_scale_esri_maps("pfjTableChart_1", ['Revenue', '0M', '136.5M', '273M', '409.4M', '545.8M',], "Step <no>")
        """
        actual_color_scale=[]
        sync_css="#"+popup_id+" [class^='colorScale-labels']"
        elem1=(By.CSS_SELECTOR, sync_css)
        self._validate_page(elem1)
        time.sleep(5)
        parent_css1="#"+popup_id+" [class^='colorScaleLegend-title']"
        actual_color_scale.append(self.driver.find_element_by_css_selector(parent_css1).text)
        #actual_color_scale.extend(self.driver.find_element_by_css_selector(parent_css1).text)
        parent_css="#"+popup_id+" [class^='colorScale-labels']"
        color_scale_title=self.driver.find_elements_by_css_selector(parent_css)
        actual_color_scale.extend([i.text for i in color_scale_title])
        utillityobject.asequal(self, expected_color_scale, actual_color_scale, msg + ": Verify the color scale title and labels")   

    def create_across_report_data_set(self, parent_id, header_rows, header_cols, data_rows, data_cols, file_name, **kwargs):
        """
        parent_id='TableChart_1'
        header_rows=The number of rows in header
        header_cols=The number of columns in header
        data_rows=Number of rows you want to verify (starts from 1, no column heading included)
        data_cols=total number of columns (Starts from 1)
        Usage: create_across_report_data_set('TableChart_1', 3, 8, 10, 8, 'C2203727_Ds01.xlsx')
        Author: Niranjan
        """        
        field_css="#" + parent_id + " div[class^='x']"
        total_items=self.driver.find_elements_by_css_selector(field_css)
        total_items_count=0
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        for r in range(0, header_rows):
            for c in range(0, header_cols):
                s.cell(row=r+1, column=c+1).value = str(total_items[total_items_count].text.strip())
                total_items_count=total_items_count+1
        for r in range(0, data_rows):
            for c in range(0, data_cols):
                s.cell(row=header_rows+r+1, column=c+1).value = str(total_items[total_items_count].text.strip())
                total_items_count=total_items_count+1
        wb.save(os.path.join(os.getcwd(), "data", file_name))  
        
    def compare_across_report_data_set(self, parent_id, header_rows, header_cols, data_rows, data_cols, file_name, **kwargs):
        """
        parent_id='TableChart_1'
        header_rows=The number of rows in header
        header_cols=The number of columns in header
        data_rows=Number of rows you want to verify (starts from 1, no column heading included)
        data_cols=total number of columns (Starts from 1)
        Usage: compare_across_report_data_set('TableChart_1', 3, 8, 10, 8, 'C2203727_Ds01.xlsx')
        Author: Niranjan
        """
        field_css="#" + parent_id + " div[class^='x']"
        total_items=self.driver.find_elements_by_css_selector(field_css)
        total_items_count=0
        wb = load_workbook(os.path.join(os.getcwd(), "data", file_name))
        status=[]
        s = wb.get_sheet_by_name('Sheet')
        for r in range(0, header_rows):
            for c in range(0, header_cols):
                if s.cell(row=r+1, column=c+1).value==None and str(total_items[total_items_count].text.strip())=='':
                    total_items_count=total_items_count+1
                    continue
                else:
                    if s.cell(row=r+1, column=c+1).value == str(total_items[total_items_count].text.strip()):
                        status=[0]
                        total_items_count=total_items_count+1
                    else:
                        status=[r+1,c]
                        return (status)
        for r in range(0, data_rows):
            for c in range(0, data_cols):
                if s.cell(row=header_rows+r+1, column=c+1).value==None and str(total_items[total_items_count].text.strip())=='':
                    total_items_count=total_items_count+1
                    continue
                else:
                    if s.cell(row=header_rows+r+1, column=c+1).value == str(total_items[total_items_count].text.strip()):
                        total_items_count=total_items_count+1
                    else:
                        status=[header_rows+r+1,c]
                        return (status)
        return (status)
    
    def select_chart_autolink_tooltip_menu_using_pyautogui(self, parent_id, raiser_class, menu_path, **kwargs):
        '''
        parent_table_css= "table[summary='Summary']"
        Usage: select_autolink_tooltip_menu_using_pyautogui("table[summary='Summary']",2,1,'Drill down to CAR', "Step 09.1: Select the Auto Drill menu Drill down to CAR", x_offset=15, y_offset=-7, browser_height=80)
        '''
        obj_move_cursor=self.driver.find_element_by_css_selector("#"+ parent_id)
        utillityobject.click_type_using_pyautogui(self, obj_move_cursor, move=True, **kwargs)
        time.sleep(4)
        menus=menu_path.split('->')
        raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
        obj_cell_css=self.driver.find_element_by_css_selector(raiser_css)
        utillityobject.click_type_using_pyautogui(self, obj_cell_css, leftClick=True, **kwargs)
        time.sleep(1)
        tooltip_css="span[class*='tdgchart-tooltip']>div>ul>li" 
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        for i in range(len(tooltips)):
            line1=tooltips[i].text.strip()
            if menus[0] in (re.sub('>', '', line1) if bool(re.match(r'^>', line1)) else tooltips[i].text.strip()).split('\n'):
                utillityobject.default_left_click(self, object_locator=tooltips[i], active_move_offset=True, ax_offset='25', ay_offset='8', cord_type='left', x_offset='25', **kwargs)
                break
        time.sleep(1)
        tooltip_css="div[class='tdgchart-submenu']>div>ul>li"
        tooltips1=self.driver.find_elements_by_css_selector(tooltip_css)
        if len(menus) > 1:
            for i in range(len(tooltips1)):
#                 line2=tooltips1[i].text.strip()
                if menus[1] in (tooltips1[i].text.strip()).split('\n'):
                    tooltips[i].click()
                    coreutillityobject.left_click(self, tooltips1[i])
                    break
    
    def verify_across_report_data_set(self, parent_id, header_rows, header_cols, data_rows, data_cols, file_name, msg, **kwargs):
        """
        parent_id='TableChart_1'
        header_rows=The number of rows in header
        header_cols=The number of columns in header
        data_rows=Number of rows you want to verify (starts from 1, no column heading included)
        data_cols=total number of columns (Starts from 1)
        Usage: verify_across_report_data_set('TableChart_1', 3, 8, 10, 8, 'C2203727_Ds01.xlsx',"Step 02; Verify report dataset")
        """
        x=IA_Resultarea.compare_across_report_data_set(self, parent_id, header_rows, header_cols, data_rows, data_cols, file_name)
        utillityobject.asequal(self,len(x),1,msg+ ' Row,Column -'+ str(x))    


    def verify_number_of_chart_segment(self, parent_id, expected_number, msg, **kwargs):
        """
        :param :parent_id = 'MAINTABLE_wbody1'
        :param :kwargs['custom_css'] = ".chartPanel [tdgtitle]"       (Other css value user pass)
        :param :expected_number = 1 or 2....                          (Any other Integer number) 
        :param :msg: 'Step 10: Verify number of pie chart segments displayed'
        :Usage :verify_number_of_chart_segment('TableChart_1', 1, 'Step 4: Expect to see the Series Selection set to Line.', custom_css=".chartPanel .groupPanel path[class='riser!s0!g0!mline!']")
                OR
        :Usage :verify_number_of_chart_segment('MAINTABLE_0', 1, 1, 'Step 4: Expect to see the Series Selection set to Line.')
        @author: AAkhan
        """
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else ".chartPanel [tdgtitle]"
        parent="#" + parent_id + " " + custom_css
        total_risers=len(self.driver.find_elements_by_css_selector(parent))
        actual_number=int(total_risers)
        utillityobject.asequal(self, expected_number, actual_number, msg)
        
    def verify_tagcloud_chart_text(self, parent_css, expected_text_list, msg):
        """
        :Param :parent_css ='#pfjTableChart_1'
        :Param :expected_text_list = '['Group 4', 'Group 3', 'Group 2', 'Group 1', 'Group 0']'
        :Param :msg ='Step x: Verify tag cloud chart text'
        """
        chart_css=parent_css + " g.groupPanel>text"
        act=self.driver.find_elements_by_css_selector(chart_css)
        actual=[]
        for i in range(len(act)):
            actual.append(act[i].text)
        utillityobject.asequal(self, actual, expected_text_list, msg)

    def verify_tagcloud_chart_text_color(self, parent_css, expected_text_color_dict, msg):
        """
        :Param :parent_css='PfjTableChart_1'
        :Param :expected_text_color_dict='{'Group 4':'bar_blue','Group 3':'Pale_green1'}'
        :Param :msg='Step x:Verify tagcloud chart text color'
        """
        chart_css=parent_css + " g.groupPanel > text"
        elems=self.driver.find_elements_by_css_selector(chart_css)
        text_list=[el.text.strip() for el in elems]
        for text_val, color_val in expected_text_color_dict.items():
            expected_font_color=utillityobject.color_picker(self, color_val, 'rgba')
            actual_font_color=Color.from_string(elems[text_list.index(text_val)].get_attribute("fill")).rgba
            utillityobject.asequal(self, expected_font_color, actual_font_color, "Verify text color of " + text_val + " as " + color_val)
            
    def verify_multicolor(self, parent_id, fill_type, expected_multicolor_list, msg):
        """
        :Param :parent_css ='TableChart_1'
        :Param : fill_type = 'stroke' OR 'fill'
        :Param : expected_multicolor_list='['sandy_brown', 'sorbus1','elf_green','cinnabar3','crusta1','cinnabar3','persian_red','sorbus1','cinnabar3','persian_red']'
        :Param : msg='Step x: Verify the created line chart colo'
        """
        elem=self.driver.find_elements_by_css_selector(parent_id+" [class*='riser!']")[0]
        fill_type = fill_type.lower() if Global_variables.browser_name in ['edge'] else fill_type #fill_type value changing to lower case letter for Edge browser because Edge driver can find only lowercase attribute
        txt=elem.get_attribute(fill_type)
        id1=re.match('url\((.*)\)', txt).group(1)
        color_list=self.driver.find_elements_by_css_selector("defs>"+id1+" > stop")
        expected=[]
        act=[]
        for i in range(len(expected_multicolor_list)):
            expected_color=utillityobject.color_picker(self, expected_multicolor_list[i], 'rgba')
            expected.append(expected_color)
            actual_color=Color.from_string(color_list[i].value_of_css_property('stop-color')).rgba
            act.append(actual_color)
        utillityobject.asequal(self, act, expected, msg)
    
    def verify_marker_tooltip(self,parent_id,marker_class,expected_tooltip,msg,**kwargs):
        
        '''
        :param : parent_id = 'iosTabs0_f' or 'wbody1_f'
        :param : marker_class='marker!s0!g0!mmarker!'
        :param : expected_tooltip=['Unit Sales: 421K', 'X: Biscotti']
        :param : kwargs['eventpanel_index']=2
        :usage : verify_line_chart_tooltip('iosTabs0_f','marker!s0!g0!mmarker!',['Unit Sales: 421K', 'X: Biscotti'],'Step 02 : Verify Line Chart tooltip')
        :usage : verify_marker_tooltip('MAINTABLE_wbody0_f','marker!s0!g9!mmarker!r0!c3!',['Region:  West', 'Product ID:  G121', 'Unit Sales:  47432', 'Filter Chart', 'Exclude from Chart'],'Step 03 : Verify line chart tooltip',eventpanel_index=4)
        '''
        
        browser=Global_variables.browser_name
        parent=self.driver.find_element_by_css_selector("#"+parent_id)        
        parent_x_offset=kwargs['parent_x_offset'] if 'parent_x_offset' in kwargs else 1
        parent_y_offset=kwargs['parent_y_offset'] if 'parent_y_offset' in kwargs else 1
        if browser in ('ie', 'edge'):
            ActionChains(self.driver).move_to_element_with_offset(parent,parent_x_offset,parent_y_offset).perform()
        else :
            utillityobject.click_on_screen(self, parent, coordinate_type='start', x_offset=parent_x_offset, y_offset= parent_y_offset)
            
        time.sleep(2)
        marker_css='#'+parent_id+" circle[class*='"+marker_class+"']"
        eventpanel_css="#"+parent_id+" .eventPanel>rect[pointer-events='all']"
        event_panels=self.driver.find_elements_by_css_selector(eventpanel_css)
        eventpanel=event_panels[int(kwargs['eventpanel_index'])-1] if 'eventpanel_index' in kwargs else event_panels[0]
        time.sleep(2)
        marker_transform=self.driver.find_element_by_css_selector( marker_css).get_attribute('transform').strip()[10:-1]
        marker_location=marker_transform.split(' ') if browser in ('ie', 'edge') else marker_transform.split(',')
        marker_x_offset=float(marker_location[0])+int(kwargs['marker_x_offset']) if 'marker_x_offset' in kwargs else float(marker_location[0])
        if len(marker_location)==1 :
            marker_y_offset=2+int(kwargs['marker_y_offset']) if 'marker_y_offset' in kwargs else 2
        else :
            marker_y_offset=2+float(marker_location[1])+float(kwargs['marker_y_offset']) if 'marker_y_offset' in kwargs else 2+float(marker_location[1])
        if browser in ('ie', 'edge'):
            ActionChains(self.driver).move_to_element_with_offset(eventpanel, xoffset=marker_x_offset, yoffset=marker_y_offset).perform()
        else :
            utillityobject.click_on_screen(self, eventpanel, coordinate_type='start',x_offset=marker_x_offset-3,y_offset=marker_y_offset+5)
        time.sleep(2)
        tooltip_css="span[id*='tdgchart-tooltip'] > div"
        #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip'] > div"
        tooltips=self.driver.find_element_by_css_selector(tooltip_css).text.replace(u'\xa0\n', u'  ').split('\n')
        actual_tooltip_list=[tooltip.strip() for tooltip in tooltips if tooltip.strip() != '']
        utillityobject.asequal(self,actual_tooltip_list,expected_tooltip,msg)
        
    def ia_exit_save(self, btn_name, parent_object=None):
        """
        :Param: btn_name='Yes' or 'OK', or 'Cancel'
        """
        btn_css="#saveAllDlg [class*='active'][class*='window'] #btn"+btn_name
        webdriver_object_name = btn_name + ' button in save prompt dialog' 
        elem=utillityobject.validate_and_get_webdriver_object(self, btn_css, webdriver_object_name, parent_object)
        elem.click()
        time.sleep(2)
        
    def verify_fexcode_syntax(self,expected_syntax_list,msg):
        '''
        :Description - verify fex syntax code in fex procedure code
        :param : expected_syntax_list=["DEFAULTH &WF_NODATA='None';","SET NODATA=&WF_NODATA"]
        :usage : verify_fexcode_syntax(expected_syntax_list,'Step 09.1 : Verify Syntax for Missing Value')
        '''
        visualization_ribbon.Visualization_Ribbon.select_top_toolbar_item(self,'toolbar_showfex')
        time.sleep(2)
        actual_fex_code = IA_Resultarea.get_fexcode_syntax(self)
        for syntax in expected_syntax_list :
            if syntax in actual_fex_code :
                status=True
            else :
                status=False
                break
        utillityobject.asequal(self,True,status,msg)
        ok_btton=self.driver.find_element_by_css_selector("#showFexOKBtn")
        utillityobject.click_on_screen(self,ok_btton, "middle", click_type=0, pause=1)
    
    def verify_syntax_not_in_fexcode(self,expected_syntax_list,msg):
        '''
        :Description - verify fex syntax code in fex procedure code
        :param : expected_syntax_list=["DEFAULTH &WF_NODATA='None';","SET NODATA=&WF_NODATA"]
        :usage : verify_fexcode_syntax(expected_syntax_list,'Step 09.1 : Verify Syntax for Missing Value')
        '''
        visualization_ribbon.Visualization_Ribbon.select_top_toolbar_item(self,'toolbar_showfex')
        time.sleep(2)
        actual_fex_code = IA_Resultarea.get_fexcode_syntax(self)
        for syntax in expected_syntax_list :
            if syntax in actual_fex_code :
                status=False
                break
            else :
                status=True
        utillityobject.asequal(self,True,status,msg)
        ok_btton=self.driver.find_element_by_css_selector("#showFexOKBtn")
        utillityobject.click_on_screen(self,ok_btton, "middle", click_type=0, pause=1)
    
    def get_fexcode_syntax(self):
        """
        Description : Return the fexcode 
        """
        frame_css="div[id^='BiDialog'] iframe[id^='BiRichEdit']"
        visualization_resultarea.Visualization_Resultarea.wait_for_property(self,frame_css,1,15)
        if Global_variables.browser_name in ['chrome']:
            script = '''return document.querySelector("{0}").contentDocument.querySelector("body").innerText'''.format(frame_css)
            return self.driver.execute_script(script)
        else:
            WebDriverWait(self.driver,10).until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, frame_css)))
            fexcode_syntax = self.driver.find_element_by_css_selector("body>div").text
            coreutillityobject.switch_to_default_content(self)
            return fexcode_syntax
    
    def verify_stop_color(self, parent_id, fill_type, expected_multicolor_list, msg):
        """
        :Param :parent_css ='#TableChart_1 svg>g>g>g .chartFrame'
        :Param : fill_type = 'stroke' OR 'fill'
        :Param : expected_multicolor_list='['sandy_brown', 'sorbus1','elf_green','cinnabar3','crusta1','cinnabar3','persian_red','sorbus1','cinnabar3','persian_red']'
        :Param : msg='Step x: Verify the created line chart colo'
        """
        elem=self.driver.find_elements_by_css_selector(parent_id)[0]
        txt=elem.get_attribute(fill_type)
        id1=re.match('url\((.*)\)', txt).group(1)
        color_list=self.driver.find_elements_by_css_selector("defs>"+id1+" > stop")
        expected=[]
        act=[]
        for i in range(len(expected_multicolor_list)):
            expected_color=utillityobject.color_picker(self, expected_multicolor_list[i], 'rgba')
            expected.append(expected_color)
            actual_color=Color.from_string(color_list[i].value_of_css_property('stop-color')).rgba
            act.append(actual_color)
        utillityobject.asequal(self, act, expected, msg)
        
    def verify_current_document_page_name(self, expected_page_name, msg):
        """
        Usage: verify_current_document_page_name('Page 1, 'Step X: Verify the default selected page name.')
        """
        actual_page_name = self.driver.find_element_by_css_selector("#iaPagesMenuBtn").text.strip()
        utillityobject.asequal(self, expected_page_name, actual_page_name, msg)
        
    def verify_document_page_options(self, page_list=None, enable_button_list=None, disable_button_list=None, msg='Step X', close_dialog_btn_name=None):
        """
        Usage: verify_document_page_options(page_list=['Page 1','Page 2'],msg="Step03: Verify pages")
        """
        dialog_css='#pageOptionsDlg'
        toolbar_btn_dict={'rename':dialog_css + ' #pageOptionsToolBar #renamePageBtn',
                        'duplicate':dialog_css + ' #pageOptionsToolBar #duplicatePageBtn',
                        'moveup':dialog_css + ' #pageOptionsToolBar #moveUpBtn',
                        'movedown':dialog_css + ' #pageOptionsToolBar #moveDownBtn',
                        'delete':dialog_css + ' #pageOptionsToolBar #deletePageBtn'}
        close_btn_dict={'ok':dialog_css + " div[id^='IABottomBar'] #pageOptionsOkBtn",
                        'cancel':dialog_css + " div[id^='IABottomBar'] #pageOptionsCancelBtn",
                        'cross':dialog_css + " div[id^='BiButton'][class*='window-close-button']"}
        
        caption_css="#pageOptionsDlg [class*='caption'] [class*='bi-label']"
        custom_msg=msg + ": Verify the caption of the dialog."
        utillityobject.verify_popup(self, dialog_css, custom_msg, caption_css=caption_css, caption_text='Page Options')
        custom_msg=msg + ": Verify the page list on display."
        if page_list != None:
            utillityobject.verify_items_in_dialog(self,dialog_css+ " #iaPageList", page_list, custom_msg)
        if enable_button_list != None:
            for btn in enable_button_list:
                custom_msg=msg + ": Verify " + btn + " button is enabled."
                btn_elem = self.driver.find_element_by_css_selector(toolbar_btn_dict[btn])
                status=utillityobject.verify_element_disable(self, element=btn_elem)
#                 status=bool(re.match('.*button-disabled', self.driver.find_element_by_css_selector(toolbar_btn_dict[btn]).get_attribute('class')))
                utillityobject.asequal(self, False, status, msg)
        if disable_button_list != None:
            for btn in enable_button_list:
                btn_elem = self.driver.find_element_by_css_selector(toolbar_btn_dict[btn])
                status=utillityobject.verify_element_disable(self, element=btn_elem)
#                 status=bool(re.match('.*button-disabled', self.driver.find_element_by_css_selector(toolbar_btn_dict[btn]).get_attribute('class')))
                utillityobject.asequal(self, True, status, msg)
        if close_dialog_btn_name != None:
            self.driver.find_element_by_css_selector(close_btn_dict[close_dialog_btn_name]).click()
            
    def select_page_in_document_page_options(self, page_name,click_type='left', close_dialog_btn_name=None):
        """
        Usage:select_page_in_document_page_options('Page 1',click_type='right')
        Usage:select_page_in_document_page_options('Page 1',close_dialog_btn_name='ok')
        """
        dialog_css='#pageOptionsDlg'
        close_btn_dict={'ok':dialog_css + " div[id^='IABottomBar'] #pageOptionsOkBtn",
                        'cancel':dialog_css + " div[id^='IABottomBar'] #pageOptionsCancelBtn",
                        'cross':dialog_css + " div[id^='BiButton'][class*='window-close-button']"}
        
        utillityobject.select_item_in_dialog(self,dialog_css+ " #iaPageList", page_name,click_type)
        time.sleep(1)
        if close_dialog_btn_name != None:
            self.driver.find_element_by_css_selector(close_btn_dict[close_dialog_btn_name]).click()
            
    def select_actions_in_document_page_options(self, toolbar_button_name, action_value=None, close_dialog_btn_name=None):
        """
        Usage: select_actions_in_document_page_options('duplicate')
        Usage: select_actions_in_document_page_options('rename',action_value="Rename Page 2") 
        """
        dialog_css='#pageOptionsDlg'
        toolbar_btn_dict={'rename':dialog_css + ' #pageOptionsToolBar #renamePageBtn',
                        'duplicate':dialog_css + ' #pageOptionsToolBar #duplicatePageBtn',
                        'moveup':dialog_css + ' #pageOptionsToolBar #moveUpBtn',
                        'movedown':dialog_css + ' #pageOptionsToolBar #moveDownBtn',
                        'delete':dialog_css + ' #pageOptionsToolBar #deletePageBtn'}
        close_btn_dict={'ok':dialog_css + " div[id^='IABottomBar'] #pageOptionsOkBtn",
                        'cancel':dialog_css + " div[id^='IABottomBar'] #pageOptionsCancelBtn",
                        'cross':dialog_css + " div[id^='BiButton'][class*='window-close-button']"}
        page_button_elem=self.driver.find_element_by_css_selector(toolbar_btn_dict[toolbar_button_name])
        utillityobject.default_click(self, page_button_elem)
        if toolbar_button_name == 'rename':
            field_elem=self.driver.find_element_by_css_selector("#iaPageList input")
            utillityobject.set_text_field_using_actionchains(field_elem, action_value)
        if close_dialog_btn_name != None:
            self.driver.find_element_by_css_selector(close_btn_dict[close_dialog_btn_name]).click()
            
    def verify_report_column_titles_on_preview(self, no_of_colnum, no_of_column_cells, expected_list, table_css='#TableChart_1', msg='Step X'):
        '''
        Desc: This function is used to verify the column title for report.
        params: no_of_cells: Total number of cells in table occupied by title. For long column name it expands over multiple line.
        '''
        actual_list = []
        temp_heading_css = table_css + " div[class^='x']"
        heading_class=self.driver.find_elements_by_css_selector(temp_heading_css)
        report_heading_css= table_css + " div[class='" + heading_class[0].get_attribute("class") + "']"
        total_items=self.driver.find_elements_by_css_selector(report_heading_css)
        tot=len(total_items)
        no_of_heading=tot/no_of_column_cells
        for i in range(no_of_colnum):
            tmp=""
            for j in range(i, int(tot/no_of_heading), no_of_colnum):
                tmp+=total_items[j].text.strip()
                actual_list.append(tmp)
        utillityobject.asequal(self,expected_list, actual_list, msg)
        
    def verify_visualize_bar_added_in_previewreport(self, visualize_bar_color, exp_visualize_bars, message, table_css="#TableChart_1"):
        '''
        Desc:-This function is used to verify visualize bar added in the report
        report_obj.visualize_bar_added_in_preview_report('TableChart_1', 'light_gray', "Step X: Verify visualize color")
        '''
        actual_total_visualize=0
        table_row_elem = table_css + " style ~ div[style*='background-color'][style*='border']"
        exp_visualize_bar_color=utillityobject.color_picker(self, visualize_bar_color, 'rgba')
        no_of_elems=self.driver.find_elements_by_css_selector(table_row_elem)
        for bar in no_of_elems :
            actual_visualize_color=Color.from_string(bar.value_of_css_property('background-color')).rgba
            if exp_visualize_bar_color==actual_visualize_color :
                actual_total_visualize+=1
        utillityobject.asequal(self, exp_visualize_bars, actual_total_visualize, message)
        
    def verify_arc_chart_group_labels(self,Group_label_css,expected_group_label,msg,text_alignment="middle"):
        '''
        Desc:- This function used to verify the group label value and alignment of group_label in ARC chart.
        usage:-verify_arc_chart_group_labels("g.group-value>text",expected_group_label,text_alignment="middle")
        '''
        try:
            group_label=self.driver.find_element_by_css_selector(Group_label_css).text
            text_anchor=self.driver.find_element_by_css_selector(Group_label_css).value_of_css_property("text-anchor")
            utillityobject.asequal(self,expected_group_label,group_label,msg)
            utillityobject.asequal(self,text_alignment,text_anchor,msg)    
        except NoSuchElementException:
            raise AttributeError("The GroupLabel element of arc_chart is not found")
    
    def verify_animate_button_position(self,slider_css="rect[class='sliderBody']", coordinate_type='top_left',msg="StepX verify animate button position"):
        '''Desc:-This function is used to verify the animate button is in first position of the slider.
           usage:-verify_animate_button_position()-it verifies the animate button is in start position
           usage:-verify_animate_button_position("rect[class='sliderBody']",'bottom_right'," StepX verfiy animate button is in end position")
        '''
        slider_body=self.driver.find_element_by_css_selector(slider_css)
        body_cord=coreutillityobject.get_web_element_coordinate(self, slider_body, element_location=coordinate_type)
        accurate_body_cord=int(round(body_cord['x']))
        slider_handle=self.driver.find_element_by_css_selector("rect[class='sliderHandle']")
        handle_cord=coreutillityobject.get_web_element_coordinate(self,slider_handle,element_location='middle')
        accurate_handle_cord=int(round(handle_cord['x']))    
        if accurate_handle_cord in range(accurate_body_cord-2 , accurate_body_cord+2):#accurate_handle_cord coordinate changes browser to browser so based on plus or minus two
            status=True
        utillityobject.asequal(self, True, status, msg)
    
    def click_animate_button(self,button_css="rect[class='animateButton']",element_location='middle', xoffset=0, yoffset=0, mouse_move_duration=0.5):
        '''
        Desc:-This function is used to click animation button of the slider
        usage:-click_animate_button(button_css="rect[class='animateButton']"):
        '''  
        try:
            animate_button_ele=self.driver.find_element_by_css_selector(button_css)
        except NoSuchElementException:
            raise AttributeError("Animate button is not present in the page {0}".format(button_css))
        coreutillityobject.python_left_click(self, animate_button_ele, element_location, xoffset, yoffset, mouse_move_duration)
        time.sleep(2)
        
    def right_click_on_barchart(self, parent_css, riser_css):
        '''
        Desc:-This function is used to right click on the bar in chart
        Usage : right_click_on_barchart("#TableChart_1", "rect[class='riser!s0!g1!mbar!']")
        '''
        try :
            css=parent_css+" "+riser_css
            obj_locator=self.driver.find_element_by_css_selector(css)
            coreutillityobject.right_click(self, obj_locator)
        except NoSuchElementException :
            raise AttributeError("The given riser element is not available in the page".format(css))
    
    def verify_chart_animate_slider(self,slider_css,handle_css,msg):
        try:
            self.driver.find_element_by_css_selector(slider_css)
            self.driver.find_element_by_css_selector(handle_css)
            status=True
        except NoSuchElementException:
            raise LookupError("The Animate slider is not found")
        utillityobject.asequal(self,status,True,msg)
    
    def right_click_on_chart_bar_or_line_in_perview_and_select_context_menu(self, parent_css, riser_css, context_menu_path):
        '''
        Desc:-This function is used to right click on the bar/line in chart preview and select context menu.
        Usage : right_click_on_chart_bar_or_line_in_perview_and_select_context_menu("#TableChart_1", "rect[class='riser!s0!g1!mbar!']", 'Visibility->Show')
        '''
        IA_Resultarea.right_click_on_barchart(self, parent_css, riser_css)
        for menu_item in context_menu_path.split('->'):
            utillityobject.select_bipopup_list_item(self, menu_item)
    
    def right_click_on_chart_bar_or_line_in_perview_and_verify_context_menu(self, parent_css, riser_css, expected_list, step_number, context_menu_path=None, compare_type='asequal'):
        '''
        Desc:-This function is used to right click on the bar/line in chart preview and verify context menu.
        Usage : right_click_on_chart_bar_or_line_in_perview_and_verify_context_menu("#TableChart_1", "rect[class='riser!s0!g1!mbar!']", ['Visibility'], 'Step 3')
        '''
        IA_Resultarea.right_click_on_barchart(self, parent_css, riser_css)
        if context_menu_path:
            for menu_item in context_menu_path.split('->'):
                utillityobject.select_bipopup_list_item(self, menu_item)
        utillityobject.verify_bipopup_list_item(self, expected_list, step_number, comparison_value=compare_type)
       
    def close_ia_without_save(self):
        """
        Description : This method click IA button and click on exit to close IA then click on No button.
        """
        visualization_ribbon.Visualization_Ribbon.select_tool_menu_item(self, "menu_close")
        utillityobject.synchronize_with_visble_text(self, "#saveAllDlg", "No", 15)
        IA_Resultarea.ia_exit_save(self, "No")
    
    def close_run_preview_window(self):
        """
        Description : Left click on run preview window close button to close run window
        """
        close_button = utillityobject.validate_and_get_webdriver_object(self, "#resultArea .window-close-button", "Run preview window close button")
        coreutillityobject.left_click(self, close_button)