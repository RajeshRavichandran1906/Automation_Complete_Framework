import sys
from common.lib.utillity import UtillityMethods as utillityobject
from common.lib.core_utility import CoreUtillityMethods as coreutillityobject
from common.lib.base import BasePage
from common.locators.visualization_resultarea_locators import VisualizationResultareaLocators
from selenium.webdriver import ActionChains
from common.pages.visualization_miscelaneous import Visualization_Miscelaneous as miscelaneousobject
from selenium.webdriver.support import expected_conditions as EC
import operator
from functools import reduce
from openpyxl import Workbook
from openpyxl import load_workbook
import time,re, os, pyautogui
from selenium.webdriver.support.color import Color
from selenium.common.exceptions import NoSuchElementException,\
    StaleElementReferenceException
from common.lib.global_variables import Global_variables
if sys.platform == 'linux':
    from pykeyboard import PyKeyboard
    pykeyboard = PyKeyboard()
    from pymouse import PyMouse
    mouse_=PyMouse()
else:
    from pywinauto.handleprops import parent
    from uisoup import uisoup
    import keyboard

class Visualization_Resultarea(BasePage):
    """ Inherit attributes of the parent class = Baseclass """

    def __init__(self, driver):
        super(Visualization_Resultarea, self).__init__(driver)
    
    def _validate_page(self, locator):
        self.longwait.until(EC.visibility_of_element_located(locator))
    
    def move_mouse_to_chart_component(self, riser_or_marker_element, initial_move_xy_dict=None, element_location='middle', xoffset=0, yoffset=0, use_marker_enable=False, move_to_tooltip=True):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        if initial_move_xy_dict != None:
            x=initial_move_xy_dict['x']
            y=initial_move_xy_dict['y']
        else:
            x=Global_variables.current_working_area_browser_x
            y=Global_variables.current_working_area_browser_y
        coreutillityobject.move_mouse_to_offset(self, x_offset=x, y_offset=y)
        time.sleep(1)
        if use_marker_enable==True:
            (x, y)=utillityobject.marker_enable(self, riser_or_marker_element)
            coreutillityobject.python_move_to_offset(self, x_offset=x, y_offset=y, mouse_move_duration=2.5)
        else:
            coreutillityobject.python_move_to_element(self, riser_or_marker_element, element_location=element_location, xoffset=xoffset, yoffset=yoffset)
        tooltip_css="span[id*='tdgchart-tooltip']:not([style*='hidden'])"
        utillityobject.synchronize_with_number_of_element(self,tooltip_css,1, 25, pause_time=0.2)
        tooltip_elem=self.driver.find_elements_by_css_selector(tooltip_css + " li")[0]
        if move_to_tooltip :
            coreutillityobject.python_move_to_element(self, tooltip_elem)
    
    def select_chart_component(self, riser_or_marker_element, use_marker_enable=False):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        x=Global_variables.current_working_area_browser_x
        y=Global_variables.current_working_area_browser_y
        coreutillityobject.move_mouse_to_offset(self, x, y)
        time.sleep(1)
        if use_marker_enable==True:
            (x, y)=utillityobject.marker_enable(self, riser_or_marker_element)
            coreutillityobject.python_click_with_offset(self, x, y)
        else:
            coreutillityobject.left_click(self, riser_or_marker_element)
#         tooltip_elem=self.driver.find_element_by_css_selector("div[class='tdgchart-tooltip']")
#         coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_middle', yoffset=5)
    
    def multiselect_chart_component(self, riser_or_marker_element_list, use_marker_enable=False):
        '''
        Desc: This function is used to move the mouse over to 
        '''
        x=Global_variables.current_working_area_browser_x
        y=Global_variables.current_working_area_browser_y
        coreutillityobject.move_mouse_to_offset(self, x, y)
        time.sleep(1)
        if sys.platform == 'linux':
            pykeyboard.press_key(pykeyboard.control_key)
        else:
            keyboard.press('ctrl')
        for riser_or_marker_element in riser_or_marker_element_list:
            if use_marker_enable==True:
                (x, y)=utillityobject.marker_enable(self, riser_or_marker_element)
                coreutillityobject.python_click_with_offset(self, x, y)
            else:
                coreutillityobject.python_move_to_element(self, riser_or_marker_element)
                time.sleep(1)
                coreutillityobject.python_left_click(self, riser_or_marker_element, mouse_move_duration=3.5)
            time.sleep(1)
        if sys.platform == 'linux':
            pykeyboard.release_key(pykeyboard.control_key)
        else:
            keyboard.release('ctrl') 
        time.sleep(1)   
        #tooltip_elem=self.driver.find_element_by_css_selector("div[class='tdgchart-tooltip']")
        #coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_middle', yoffset=5)
                
    def verify_tooltip(self, expected_tooltip_list, msg='Step X: Verify default Chart tooltip', parent_css='#MAINTABLE_wbody1'):
        '''
        Desc: This function will verify the tooltip. First mouse will move to the top left corner of the current working area,
                Then it will move to the required location to bring the tooltip disply and then read and compare.
        '''
        time.sleep(self.tooltip_wait_time)
        tooltip_css="span[id*='tdgchart-tooltip']:not([style*='hidden'])"
#         raw_tooltip_list=self.driver.find_element_by_css_selector(tooltip_css).text.replace(u'\xa0\n', '').replace(u'\xa0', ' ').split('\n')
        raw_tooltip_list=utillityobject.validate_and_get_webdriver_object(self, tooltip_css, "tooltip", pause_time=0.1).text.replace(u'\xa0\n', '').replace(u'\xa0', ' ').split('\n')
        actual_list=utillityobject.get_actual_tooltip_list(self, raw_tooltip_list)
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
        
    def verify_lasso_tooltip(self, expected_tooltip_list, msg='Step X: Verify default Chart tooltip'):
        '''
        Desc: This function will verify the tooltip. First mouse will move to the top left corner of the current working area,
                Then it will move to the required location to bring the tooltip disply and then read and compare.
        '''
        tooltip_css="[id^='ibi'][class*='tdgchart-tooltip']:not([style*='hidden'])"
        raw_tooltip_list=self.driver.find_element_by_css_selector(tooltip_css).text.replace(u'\xa0\n', '').replace(u'\xa0', ' ').split('\n')
        actual_list=utillityobject.get_actual_tooltip_list(self, raw_tooltip_list)
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)
        
    def hover_to_required_arrow_in_tooltip(self, item_name):
        first_level_tooltip_css = "[id*='tdgchart-tooltip']:not([style*='hidden']) li[class*='tdgchart-tooltip-hover']"
        time.sleep(self.tooltip_wait_time)
        elems=self.driver.find_elements_by_css_selector(first_level_tooltip_css)
#         print([tooltip.text.strip() for tooltip in elems])
        for tooltip in elems :
            if tooltip.text.strip().replace('\n','').replace('>','') == item_name :
                first_level_tooltip=tooltip
                break
            #else :
            #    raise KeyError("Unable to find {0} option in tooltip".format(item_name))
        tooltip_elem=self.driver.find_elements_by_css_selector("span[id*='tdgchart-tooltip']:not([style*='hidden']) li:not([class*='pointer'])")[-1]
        coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_left', xoffset=10, yoffset=10)
        arrow_to_be_selected=first_level_tooltip.find_element_by_css_selector("span.tdgchart-tooltip-arrow")
        coreutillityobject.python_move_to_element(self, arrow_to_be_selected)
#         time.sleep(self.wait_time)
        tooltip_elem=self.driver.find_element_by_css_selector("span[id*='tdgchart-tooltip'] [class*='tdgchart-submenu']:not([style*='hidden'])")
        coreutillityobject.python_move_to_element(self, tooltip_elem, element_location='top_middle', yoffset=5)
        
    def select_tooltip_item(self, item_name):
        '''
        Desc: This function is used to select an item from the single tooltip box. 
        '''
        if item_name in ['Filter Chart', 'Exclude from Chart', 'Remove Filter']:
            tooltip_css= "[id*='tdgchart-tooltip'] span.tdgchart-tooltip-label>div"
        else:
            tooltip_css= "[id*='tdgchart-tooltip'] span.tdgchart-tooltip-label"
        elems=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name)]
        coreutillityobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
        time.sleep(10)
    
    def select_lasso_tooltip_item(self, item_name):
        '''
        Desc: This function is used to select an item from the single lasso tooltip box.
        '''
        tooltip_css= ".tdgchart-tooltip span[class='tdgchart-tooltip-pad']:not([style*='hidden'])"
        elems=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name)]
        coreutillityobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
        time.sleep(10)
        
    def select_bilevel_tooltip_item(self, item_name1, item_name2):
        '''
        Desc: This function is used to select an item from the second tooltip box. Mostly used for drill down pop up.
        '''
        Visualization_Resultarea.hover_to_required_arrow_in_tooltip(self, item_name1)
        second_level_tooltip_css = "[id*='tdgchart-tooltip'] [class*='tdgchart-submenu'] span.tdgchart-tooltip-label"
        elems=self.driver.find_elements_by_css_selector(second_level_tooltip_css)
        tooltip_item_to_be_selected=elems[[elem.text.strip() for elem in elems].index(item_name2)]
        coreutillityobject.left_click(self, tooltip_item_to_be_selected, element_location='middle_left', xoffset=10)
        time.sleep(10)
    
    def verify_xy_axis_title(self, expected_title_list, parent_css, x_or_y_axis_title_css, x_or_y_axis_title_length=None, msg='Step X'):
        '''
        Desc: This function is used to verify X or Y axis title at a time.
        '''
        x_or_y_axis_title_css=parent_css+" " + x_or_y_axis_title_css
        actual_title_list=[title.text.strip() for title in self.driver.find_elements_by_css_selector(x_or_y_axis_title_css)]
        for expected_title, actual_title in zip(expected_title_list, actual_title_list):
            if expected_title[:x_or_y_axis_title_length] == actual_title[:x_or_y_axis_title_length]:
                status= True
            else:
                status=False
                break
        utillityobject.asequal(self, status, True, msg) 
    
    def verify_xyz_labels(self, expected_label_list, parent_css, xyz_axis_label_css, xyz_axis_label_length=None, msg='Step X'):
        '''
        Desc: This function is used to verify X or Y axis title at a time.
        '''
        xyz_axis_label_css=parent_css+" " + xyz_axis_label_css
        x=self.driver.find_elements_by_css_selector(xyz_axis_label_css) 
        my_iter=(i.text for i in x)
        for label in expected_label_list:
            if label[:xyz_axis_label_length] == next(my_iter)[:xyz_axis_label_length]:
                status= True
            else:
                status=False
                break
        del my_iter
        utillityobject.asequal(self, status, True, msg) 
    
    def verify_legends(self, expected_legend_list, parent_css, legend_length=None, msg='Step X'):
        '''
        Desc: This function is used to verify legends.
        '''
        legend_css=parent_css + " .legend text"
        legends=self.driver.find_elements_by_css_selector(legend_css)
        my_iter_x=(i.text for i in legends)
        for label_x in expected_legend_list:
            if label_x[:legend_length] in next(my_iter_x):
                statex= True
            else:
                statex=False
                break
        del my_iter_x
        utillityobject.asequal(self, statex, True, msg)
        
    def verify_pie_label_in_single_group(self, expected_label_list, parent_css, label_css, msg):
        '''
        Desc: This function is used to verify pie labels within a single group.
        '''
        pie_label_css=parent_css + " " + label_css
        labels=self.driver.find_elements_by_css_selector(pie_label_css)
        actual_label_list=[label.text.strip() for label in labels]
        utillityobject.asequal(self,expected_label_list, actual_label_list, msg)    
    
    def verify_pie_label_and_value_in_multiple_groups(self, expected_label_list, expected_total_label_list, parent_css, label_css, msg1, msg2):
        '''
        Desc: This function is used to verify pie labels and total labels in multiple groups.
        '''
        pie_label_css=parent_css + " " + label_css
        actual_label_list=[]
        labels=self.driver.find_elements_by_css_selector(pie_label_css)
        for i in range(0,len(labels)):
            label=self.driver.find_element_by_css_selector(parent + " text[class='pieLabel!g" + str(i) + "!mpieLabel!']").text.strip()
            actual_label_list.append(label)
        utillityobject.asequal(self,expected_label_list, actual_label_list, actual_label_list, msg1)
        actual_total_label_list=[]
        labels=self.driver.find_elements_by_css_selector(parent + " text[class^='totalLabel!g']")
        for i in range(0,len(labels)):
            total_label=self.driver.find_element_by_css_selector(parent + " text[class='totalLabel!g" + str(i) + "!mtotalLabel!']").text.strip()
            actual_total_label_list.append(total_label)
        utillityobject.asequal(self, expected_total_label_list, actual_total_label_list, msg2)
    
    def verify_number_of_risers(self, parent_css_with_tagname, risers_per_segment, expected_number, msg):
        '''
        Desc: This function is used to verify the number of risers. As this covers wide range of risers with different tags, user must 
        pass the parent css upto tag. Also we need to provide the number of risers per segment. It means if there is a simple bar chart,
        then number of risers per segment is '1', for stacked bar it will differ. 
        '''
        riser_css=parent_css_with_tagname + "[class^='riser']"
        total_risers=len(self.driver.find_elements_by_css_selector(riser_css))
        actual_number=int(total_risers/risers_per_segment)
        utillityobject.asequal(self, expected_number, actual_number, msg)

    def verify_number_of_pie_segment(self, parent_css, risers_per_segment, expected_number, msg):
        '''
        Desc: This function is used to verify the number of pie segments.
        '''
        parent=parent_css + " svg.rootPanel>g.chartPanel"
        total_risers=len(self.driver.find_elements_by_css_selector(parent + " path[class^='riser']"))
        actual_number=int(total_risers/risers_per_segment)
        utillityobject.asequal(self, expected_number, actual_number, msg)
    
    def verify_number_of_circles(self, parent_css, minimum_number, maximum_number, msg):
        '''
        Desc: This function is used to verify the number of circles. As number of circles differs in a less margin from resolution to resolution
        user will have to pass a range.
        '''
        parent=parent_css + " svg g"
        actual_number=len(self.driver.find_elements_by_css_selector(parent + " circle[class^='riser']"))
        if actual_number in range(minimum_number , maximum_number):
            status=True
        utillityobject.asequal(self, True, status, msg)
                
    def create_laso(self, source_element, target_element, source_element_location, source_xoffset, source_yoffset, target_element_location, target_xoffset, target_yoffset, enable_marker=False):
        '''
        Desc: This function is used to create a lasso tooltip by dragging from source to target element provided.
        '''
        if enable_marker == True:
            source_dict=utillityobject.enable_marker_using_javascript(self, source_element, coordinate_type=source_element_location, x_offset=source_xoffset, y_offset=source_yoffset)
            x1=source_dict['x']
            y1=source_dict['y']
            target_dict=utillityobject.enable_marker_using_javascript(self, target_element, coordinate_type=target_element_location, x_offset=target_xoffset, y_offset=target_yoffset)
            x2=target_dict['x']
            y2=target_dict['y'] 
        else :
            source_dict=coreutillityobject.get_web_element_coordinate(self, source_element, element_location=source_element_location, xoffset=source_xoffset, yoffset=source_yoffset)
            x1=source_dict['x']
            y1=source_dict['y']
            target_dict=coreutillityobject.get_web_element_coordinate(self, target_element, element_location=target_element_location, xoffset=target_xoffset, yoffset=target_yoffset)
            x2=target_dict['x']
            y2=target_dict['y']
        coreutillityobject.create_lasso(self, x1, y1, x2, y2)
        time.sleep(2)
    
    def verify_visualization_row_and_column_labels(self, expected_label, parent_css, matrix_type, label_length=None, msg='Step X'):
        '''
        Desc: This function used to verify Rows or Columns labels
        '''
        if matrix_type == 'rows':
            label_css = parent_css + " svg text[class^='row']"
        elif matrix_type == 'column':
            label_css = parent_css + " svg text[class^='col']"
        actual_val_list=(el.text.strip() for el in self.driver.find_elements_by_css_selector(label_css) if el.text.strip()!='')
        for label in expected_label:
            if label[:label_length]==next(actual_val_list)[:label_length] :
                status= True
            else:
                status=False
                break
        utillityobject.asequal(self, status, True, msg)
    
    def rename_or_verify_grouped_riser_name(self, new_name=None, default_verify_name=None, close_button_name=None, msg='Step X'):
        '''
        Desc: This is to verify and rename the Rename-Dialog for Group riser name.
        '''
        rename_dialog_css = "div[id^='BiDialog'] [class*='window-active']"
        name_text_box_css = rename_dialog_css + ' input'
        miscelaneousobject.wait_for_object(self, rename_dialog_css, 'number', expected_number=1)
        name_text_box_elem = self.driver.find_element_by_css_selector(name_text_box_css)
        if default_verify_name != None:
            custom_msg = msg + ' :Verify the Default name in the Name text box.'
            actual_name=utillityobject.get_element_attribute(self, name_text_box_elem, 'value')
            utillityobject.asequal(self, actual_name, default_verify_name, custom_msg)
        if new_name != None:
            utillityobject.set_text_to_textbox_using_keybord(self, new_name, text_box_css=name_text_box_css)
        if close_button_name != None:
            utillityobject.click_dialog_button(self,rename_dialog_css, close_button_name)
        time.sleep(1)
    
    
    def select_resultarea_panel_caption_btn(self, position, parent_css, select_button_name, drop_down_menu_item_name=None):
        '''
        Desc: This function is used to click on the right corner panel buttons. And also we can select the items from the drop down arrow.
        param: position: 0, 1, 2, ... This is the Chart/Grid panel position number starts/scans from top-left to bottom-right.
        param :parent_css = This is the css of the panel.(to selecet IA or Visualization css area)
        select_button_name: 'menu' OR 'maximize' OR 'restore' OR 'close'
        Usage: select_panel_caption_btn(0, 'maximize')
        Usage: select_panel_caption_btn(0, 'close', custom_css="[class*='window-caption']")
        @author = Niranjan
        '''
        panel_css="#resultArea "+parent_css
        btn_css="div[class$='" + select_button_name + "-button']"
        panels=self.driver.find_elements_by_css_selector(panel_css)
        panels[position].find_element_by_css_selector(btn_css).click()
        time.sleep(1)
        if drop_down_menu_item_name != None:
            utillityobject.select_bipopup_list_item(self, drop_down_menu_item_name)
            
                                    
    """=====================OLD FUNCTIONS==================================""" 
    
    """=====================OLD FUNCTIONS==================================""" 
    def verify_report_titles_on_preview(self,colnum,no_of_cells,table_id,expected_list, msg):
        """
        params: colnum: 5
        params: no_of_cells: 12
        params: table_id='TableChart_1'
        params: expected_list: ['Category','Product ID','Product','Unit Sales','Dollar Sales']
        Syntax: verify_column_titles_on_preview(5, 5,'TableChart_1',columns, 'Step 04: Verify preview pane')
        """
        actual_list = []
#         temp_heading_css="#" + table_id + " div[class^='x']"
#         heading_class=self.driver.find_elements_by_css_selector(temp_heading_css)
#         report_heading_css="#"+table_id+" div[class='" + heading_class[0].get_attribute("class") + "']"
        report_heading_css="#" + table_id + " div[class^='x'][class*='title']"
        total_items=self.driver.find_elements_by_css_selector(report_heading_css)
        tot=len(total_items)
        no_of_heading=tot/no_of_cells
        for i in range(colnum):
            tmp=""
            for j in range(i, int(tot/no_of_heading), colnum):
                tmp+=total_items[j].text.strip()
            actual_list.append(tmp)
#         print(actual_list)
        utillityobject.asequal(self,expected_list, actual_list, msg)
        
    def verify_grid_column_heading(self, parent_id, expected_column_list, msg):
        """
        param: expected_column_list=['Sale Year', 'Product Category', 'Revenue']
        parent_id: 'MAINTABLE_wbody1' 
        Syntax: verify_grid_column_heading(0, 'maximize')
        @author = Niranjan
        """
        dimension_css="#" + parent_id + " .tablePanel .rowTitle text"
        measure_css="#" + parent_id + " .tablePanel g[class^='colHeader'] text"
        dimension_list=[el1.text for el1 in self.driver.find_elements_by_css_selector(dimension_css)]
        measure_list=[el2.text for el2 in self.driver.find_elements_by_css_selector(measure_css)]
        actual_column_list=dimension_list+measure_list
        for x in range(len(expected_column_list)):
            status=True if actual_column_list[x][:4] in expected_column_list[x] else False
        utillityobject.asequal(self,True, status, msg)
        
    def verify_grid_row_val(self, parent_id, expected_row_list, msg):
        """
        This function is to verify the first row in the Grid table
        parent_id: 'MAINTABLE_wbody1'
        expected_row_list: ['Accessories', 'Charger', '$2,052,711.00', '$1,970,123.91']
        Syntax: verify_grid_row_val('MAINTABLE_wbody1', ['Accessories', 'Charger', '$2,052,711.00', '$1,970,123.91'], 'Step 10: verify grid row')
        @author = Niranjan
        """
        dimension_text_css="#" + parent_id + " g.rowLabels"
        dimension_first_elem_y_val=self.driver.find_element_by_css_selector(dimension_text_css + " > text").get_attribute("y")
        measure_row_css="#" + parent_id + " g.innerTable g.row0"
        measure_el=self.driver.find_elements_by_css_selector(measure_row_css + " text")
        measure_val = [el2.text for el2 in measure_el]
        if Global_variables.browser_name in  ['edge'] :
            dimension_el=self.driver.find_elements_by_css_selector(dimension_text_css + " text")
            dimension_val=[el1.text for el1 in dimension_el if el1.get_attribute("y") == dimension_first_elem_y_val]
        else :
            dimension_el=self.driver.find_elements_by_css_selector(dimension_text_css + " text[y='" + dimension_first_elem_y_val + "']")
            dimension_val=[el1.text for el1 in dimension_el]
        actual_list = dimension_val + measure_val
        utillityobject.asequal(self,expected_row_list, actual_list, msg)
        
    def verify_panel_caption_label(self, position, expected_label, msg):
        """
        param: position: 0, 1, 2, ... This is the Chart/Grid panel position number starts/scans from top-left to bottom-right.
        param: expected_label= 'Grid1' or 'line1'
        Syntax: verify_panel_caption_label(0, 'Grid1', 'Step 10: Verify Grid1 is displayed')
        @author = Niranjan
        """
        panel_css="#resultArea div[id^='BoxLayoutMiniWindow']"
        label_css="div[class$='dv-caption-label']"
        panels=self.driver.find_elements_by_css_selector(panel_css)
        actual_label=panels[position].find_element_by_css_selector(label_css).text.strip()
        utillityobject.asequal(self, expected_label, actual_label, msg)
    
    def select_panel(self, label_name):  
        """
        param: label_name= 'Grid1' or 'line1'
        Syntax: select_panel('Grid1')
        @author = Niranjan
        """
        panel_css="#resultArea div[id^='BoxLayoutMiniWindow']"
        label_css="div[class$='dv-caption-label']"
        panels=self.driver.find_elements_by_css_selector(panel_css)
        labels=[el.find_element_by_css_selector(label_css).text.strip() for el in panels]
        panels[labels.index(label_name)].find_element_by_css_selector(label_css).click()
        time.sleep(1)

    def select_panel_caption_btn(self, position, select_type='maximize', **kwargs):
        """
        param: position: 0, 1, 2, ... This is the Chart/Grid panel position number starts/scans from top-left to bottom-right.
        param :kwargs['custom_css'] = "div[id^='BoxLayoutMiniWindow']"       (to selecet IA or Visualization css area)
        kwargs['select_type']: 'menu' OR 'maximize' OR 'restore' OR 'close'
        Usage: select_panel_caption_btn(0, 'maximize')
        Usage: select_panel_caption_btn(0, 'close', custom_css="[class*='window-caption']")
        @author = Niranjan
        """
        if 'custom_css' in kwargs:
            custom_css=kwargs['custom_css']
        else:
            custom_css = "div[id^='BoxLayoutMiniWindow']"
        panel_css="#resultArea "+custom_css
        btn_css="div[class*='" + select_type + "-button']"
        panels=self.driver.find_elements_by_css_selector(panel_css)
        btn_csss = panels[position].find_element_by_css_selector(btn_css)
        coreutillityobject.python_left_click(self, btn_csss)
        time.sleep(1)
        if 'menu_item' in kwargs:
            utillityobject.select_or_verify_bipop_menu(self, kwargs['menu_item'])

    def verify_default_tooltip_values(self,parent_id, raiser_class, expected_tooltip_list, msg, **kwargs):#Need to Delete
        """
        :kwargs: default_move= False (if default_move in kwargs -> mouse not move)
        :param parent_id = 'MAINTABLE_wbody1'
        :param : raiser_class = 'riser!s4!g4!mbar!'
        :param : expected_tooltip_list = ['Product:French Roast', 'Unit Price:73,710.00', 'Ordered Units:285689', 'Product:French Roast', 'Order Number:1965387', 'Filter Chart', 'Exclude from Chart']
        :param : color = 'red' or 'green'....color should be added in the color.data file
        expected_tooltip=['Product:French Roast', 'Unit Price:73,710.00', 'Ordered Units:285689', 'Product:French Roast', 'Order Number:1965387', 'Filter Chart', 'Exclude from Chart']
        Usage: verify_default_tooltip_values("MAINTABLE_wbody1","riser!s0!g0!mbar",expected_tooltip, "red", "Step 8: verify the default tooltip values")
        """  
        parent_css="#"+ parent_id     
        if 'default_move' in kwargs:
            pass
        else:
            move_to_parentID = self.driver.find_element_by_css_selector("#"+ parent_id)
            fullpath_of_raiser_css=parent_css + " [class*='" + raiser_class + "']"
            raiser_elem=self.driver.find_element_by_css_selector(fullpath_of_raiser_css)
            if Global_variables.browser_name in ['firefox', 'ie', 'edge']:
                coreutillityobject.move_to_element(self, move_to_parentID)
                time.sleep(.03)
                coreutillityobject.move_to_element(self, raiser_elem)
                if Global_variables.browser_name =='firefox': pass
                else: time.sleep(2)
            else:
                action1 = ActionChains(self.driver)
                action1.move_to_element_with_offset(move_to_parentID,1,1).perform()
                action1.move_to_element(raiser_elem).perform()                
                del action1
        Visualization_Resultarea.verify_tooltip(self, expected_tooltip_list, msg=msg, parent_css=parent_css)
        '''
        
            
        if 'default_move' in kwargs:
            pass
        else:
            action = ActionChains(self.driver)
            raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
            obj_locator=self.driver.find_element_by_css_selector(raiser_css)
            if Global_variables.browser_name in ['firefox', 'ie']:
                utillityobject.click_type_using_pyautogui(self, obj_locator, **kwargs)
            else:
                action.move_to_element(obj_locator).perform()
        time.sleep(1)
        tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip']"
        tooltips=self.driver.find_element_by_css_selector(tooltip_css).text.split('\n')
#         act_list=[]
        actual_list=[]      
        for line in tooltips:
            if bool(re.match(r'.*:\s.*', line)):
                reqobj = re.match('(.*):\s{1,}(.*)', line)
                new_element = str(reqobj.group(1)) + ":" + str(reqobj.group(2))
            elif bool(re.match(r'^>', line)):
                new_element = re.sub('>', '', line)
            elif bool(line==''):
                continue
            else:
                new_element=line
            actual_list.append(new_element)
        actual_list=[list_value for list_value in actual_list if list_value not in '']
#         for item in act_list:
#             if item != '':
#                 actual_list.append(item)
        utillityobject.asequal(self, expected_tooltip_list, actual_list, msg)'''
        
    def verify_default_circle_tooltip_values(self, parent_id, raiser_class, expected_tooltip_list, msg, **kwargs):#Need to Delete
        """
        :param parent_id = 'MAINTABLE_wbody1'
        :param : raiser_class = 'riser!s4!g4!mbar!'
        :param : expected_tooltip="Fund Category:International" #this is fixed according to this test case need
        :Usage: verify_default_tooltip_values("MAINTABLE_wbody1","riser!s0!g0!mbar",expected_tooltip, "Step 8: verify the default tooltip values")
        """
        if 'default_move' in kwargs:
            pass
        else:
            action1 = ActionChains(self.driver)
            move1 = self.driver.find_element_by_css_selector("#"+ parent_id)
            if Global_variables.browser_name in ['firefox', 'ie', 'edge']:
                utillityobject.click_type_using_pyautogui(self, move1, move=True, **kwargs)
            else:
                action1.move_to_element_with_offset(move1,1,1).perform()
            time.sleep(5)
            del action1
        action = ActionChains(self.driver)
        raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
        tooltip_css="span[id*='tdgchart-tooltip']>div>ul>li"
        #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip']>div>ul>li"
        raiser_css_obj=self.driver.find_element_by_css_selector(raiser_css)
        if 'default_move' in kwargs:
            pass
        else:
            if Global_variables.browser_name in ['firefox', 'ie', 'edge']:
                utillityobject.click_type_using_pyautogui(self, raiser_css_obj, **kwargs)
            else:
                action.move_to_element(raiser_css_obj).perform()
        time.sleep(2)
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_list=[]
        for i in range(len(tooltips)):
            tooltip_list.append((tooltips[i].text.strip()).split('\n'))
        actual_tooltip_list=[line for line in reduce(operator.add, tooltip_list) if len(line)>1]
        actual_list = []
        for line in actual_tooltip_list:
            if bool(re.match(r'.*:\s.*', line)):
                reqobj = re.match('(.*):\s{1,}(.*)', line)
                new_element = str(reqobj.group(1)) + ":" + str(reqobj.group(2))
            elif bool(re.match(r'^>', line)):
                new_element=re.sub('>', '', line)
            else:
                new_element=line
            actual_list.append(new_element)
        utillityobject.asequal(self, expected_tooltip_list, actual_list[3], msg)

    
    def verify_chart_tooltip_values(self,parent_frame_no,riser_class,expected_list,text, **kwargs):#Need to Delete
        """
       :usage: verify_chart_tooltip_values(0, "riser!s1!g8!mbar!", ['Category:Coffee', 'Product:Latte', 'Unit Sales:878,063'],"Step VP: Verify Category, Product and Dollar Sales value from Bottom")
        :usage: verify_chart_tooltip_values(0, "riser!s0!g2!mbar!", ['Category:Gifts', 'Product:Mug', 'Dollar Sales:4,522,521'],"Step VP: Verify Category, Product and Unit Sales value from Top")
        @author : Nasir
        """
        actual_list = []
        parent_frame_no=kwargs['parent_frame_num'] if 'parent_frame_num' in kwargs else '0'
        css = "[id*='MAINTABLE_" + str(parent_frame_no) + "'] [class*='" + riser_class + "']"
        if 'default_move' in kwargs:
            pass
        else:
            action = ActionChains(self.driver)
            move1 = self.driver.find_element_by_css_selector("[id*='MAINTABLE_" + str(parent_frame_no) + "']")
            if Global_variables.browser_name in ['firefox', 'ie', 'edge']:
                utillityobject.click_type_using_pyautogui(self, move1, move=True, **kwargs)
            else:
                action.move_to_element_with_offset(move1,1,1).perform()
            time.sleep(2)
        riser = self.driver.find_element_by_css_selector(css)
        action1 = ActionChains(self.driver)
        if Global_variables.browser_name in ['firefox', 'ie', 'edge']:
            utillityobject.click_type_using_pyautogui(self, riser, **kwargs)
        else:
            action1.move_to_element(riser).perform()
        time.sleep(2)
        values_list = (self.driver.find_element_by_css_selector("#tdgchart-tooltip table").text).split('\n')
        for line in values_list:
            reqobj = re.match('(.*):\s{1,}(.*)', line)
            first_element = str(reqobj.group(1))
            second_element = str(reqobj.group(2))
            new_element = first_element + ":" + second_element
            actual_list.append(new_element)
        utillityobject.asequal(self,actual_list, expected_list,text)
    
    def select_default_tooltip_menu(self,parent_id, raiser_class, menu_path, wait_time=2, **kwargs):#Need to Delete
        """
        param parent_id = 'MAINTABLE_wbody1'
        param : raiser_class = 'riser!s4!g4!mbar!'
        kwargs: verify_menu1=['Customer State Province', 'Product Subcategory'](if given then it will just verify, if not given it will click on menu)
        param : menu_path = 'Drill up to->Product Category' (At this point, i am considering one level of sub menu. We can enhance, if the level increases.)
        Usage: select_default_tooltip_menu("MAINTABLE_wbody1","riser!s0!g0!mbar", 'Drill up to->Product Category')
        Usage: select_default_tooltip_menu("MAINTABLE_wbody1","riser!s0!g0!mbar", 'Filter Chart')
        Usage: select_default_tooltip_menu("MAINTABLE_wbody1", "riser!s0!g2!mbar",'Drill down to Product Subcategory',drilldown_menu='Computers')
        Usage:resultobj.select_default_tooltip_menu("MAINTABLE_wbody1", "riser!s0!g0!mmarker!r0!c0",'Drill down to->Product Subcategory',
        wait_time=1,x_offset=1,y_offset=-25,x_offset_menu1=0,y_offset_menu1=-25,verify_menu1=['Customer State Province', 'Product Subcategory'],msg="Step15: Verify Drill down to menu")
        Usage:resultobj.select_default_tooltip_menu("MAINTABLE_wbody1", "riser!s0!g0!mmarker!r0!c0",'Drill down to->Product Subcategory',
        wait_time=1,x_offset=1,y_offset=-25,x_offset_menu1=0,y_offset_menu1=-25)
        Author: Niranjan
        """
        if 'default_move' in kwargs:
            pass
        else:
            move1 = self.driver.find_element_by_css_selector("#"+ parent_id)
            utillityobject.click_type_using_pyautogui(self, move1, move=True, **kwargs)
            time.sleep(3)
        #action1 = ActionChains(self.driver)
        menus=menu_path.split('->')
        raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
        if len(menus)>1:
            tooltip_css="span[id*='tdgchart-tooltip']>div>ul>li"
            #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip']>div>ul>li"
            obj_locator=self.driver.find_element_by_css_selector(raiser_css)
            if 'default_move' in kwargs:
                pass
            else:
                utillityobject.click_type_using_pyautogui(self, obj_locator,**kwargs)
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            Visualization_Resultarea.hover_to_required_arrow_in_tooltip(self, menus[0])   
            if 'verify_menu1' in kwargs: 
                tooltip_css1="div[class='tdgchart-submenu']>div>ul>li>span"
                tooltips1=self.driver.find_elements_by_css_selector(tooltip_css1) 
#                 tooltip_div_css="#{0} div[class='tdgchart-submenu'][style*='visible']".format(parent_id)
                tooltip_div_css="div[class='tdgchart-submenu'][style*='visible']"
                utillityobject.synchronize_with_number_of_element(self, tooltip_div_css, 1, 15, pause_time=0.3)
                tooltip_css1="div[class='tdgchart-submenu']>div>ul>li>span"
                tooltips1=self.driver.find_elements_by_css_selector(tooltip_css1)                              
                menu1=[el.text.strip() for el in tooltips1]
                utillityobject.as_List_equal(self,menu1,kwargs['verify_menu1'],kwargs['msg'])
            else:
                second_level_tooltip_css = "[id*='tdgchart-tooltip'] [class*='tdgchart-submenu'] span.tdgchart-tooltip-label"
                elems=self.driver.find_elements_by_css_selector(second_level_tooltip_css)
                for elem in elems:
                    if elem.text.strip()==menus[1]:
                        coreutillityobject.left_click(self, elem, element_location='middle_left', xoffset=10)
                        break
            time.sleep(wait_time*5)             
        else:
            if 'drilldown_menu' in kwargs:
                option=kwargs['drilldown_menu']#removed .lower(), since 1st level not working in visualization
                tooltip_css="span[id*='tdgchart-tooltip'][style*='visible'] div span[onclick*='"+option+"']"
                #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip'][style*='visible'] div span[onclick*='"+option+"']"
                tooltip_div_css="span[id*='tdgchart-tooltip'][style*='visible']"
                #tooltip_div_css="#{0} span[id*='tdgchart-tooltip'][style*='visible']".format(parent_id)
            elif 'click_menu' in kwargs:
                tooltip_css="div[class='tdgchart-tooltip'][style*='visible'] span[class='tdgchart-tooltip-pad']"
                tooltip_div_css="div[class='tdgchart-tooltip'][style*='visible']"
            else:
                tooltip_css="span[id*='tdgchart-tooltip'][style*='visible'] [onclick^='ibiChart']"
                #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip'][style*='visible'] [onclick^='ibiChart']"
                tooltip_div_css="span[id*='tdgchart-tooltip'][style*='visible']"
                #tooltip_div_css="#{0} span[id*='tdgchart-tooltip'][style*='visible']".format(parent_id)
            if 'default_move' in kwargs:
                pass
            else:
                obj_locator=self.driver.find_element_by_css_selector(raiser_css)
                utillityobject.click_type_using_pyautogui(self, obj_locator,**kwargs)
            utillityobject.synchronize_with_number_of_element(self, tooltip_div_css, 1, 15, pause_time=0.3)
            tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
            tooltip_item_to_be_selected=tooltips[[elem.text.strip() for elem in tooltips].index(menus[0])]
            tooltip_item_to_be_selected.click()
    
    def choose_right_click_menu_item_for_prompt(self, prompt_css, item_name, **kwargs):
        """
        :params: prompt_css='Prompt_1'...
        :params: item_name='Properties' or 'Delete' or 'Size and Position'
        :syntax: choose_right_click_menu_item_for_prompt('Prompt_1', 'Properties')
        @author : Nasir
        """
        resize_handler_image_css = str(prompt_css)+" img[src*='resize_marker_16']"
        prompt_obj=self.driver.find_element_by_css_selector(str(prompt_css))
        utillityobject.default_click(self, prompt_obj)
        utillityobject.synchronize_with_number_of_element(self, resize_handler_image_css, 8, 25)
        resize_handler_images_obj = self.driver.find_elements_by_css_selector(resize_handler_image_css)
        utillityobject.default_click(self, resize_handler_images_obj[1], click_option=1, pause=1)
        time.sleep(4)
        utillityobject.select_or_verify_bipop_menu(self, item_name, **kwargs)
    
    def customize_active_dashboard_properties(self, prompts=None, source=None, targets=None, msg="Step X", btn_type=None):
        """
        :params: source={'report':'Report 1', 'field':'Cost of Goods', 'condition':'Equals to', 'sort':'Ascending', 'includeall':'yes|no'}
        :params: targets={'select_candidate_report|select_target_name':'Report1', 'add_to_target':'yes|no', 'remove_from_target':'yes|no','verify_target_name|verify_candidate_report':['Report1']}
        :syntax: resultobj.customize_active_dashboard_properties('None', source, 'None')    # to select only Source 
        @author : Nasir
        """
        utillityobject.synchronize_with_number_of_element(self, '#btnADPOK', 1, 25)
        if prompts != None:
            if 'select_prompts' in prompts:
                utillityobject.select_item_in_dialog(self, "#gridADPrompts", prompts['select_prompts'])
            if 'verify_prompts' in prompts:
                utillityobject.verify_items_in_dialog(self, "#gridADPrompts div[class$='content']", prompts['verify_prompts'], msg + ".a: Verify " + str(prompts['verify_prompts']) + " in 'prompts' section")
        if source != None:
            if 'select_report' in source:
                utillityobject.select_combobox_item(self, 'comboSourceReports', source['select_report'])
            if 'verify_report' in source:
                elem1=self.driver.find_element_by_css_selector("#comboSourceReports>div")
                d=utillityobject.get_attribute_value(self, elem1, "text")
                d['text']=d['text'].replace(u'\xa0', u' ')
                utillityobject.asequal(self, d['text'], source['verify_report'], msg + ".b: Verify report_type - "+ source['verify_report'] + " displayed")
            if 'select_field' in source:
                utillityobject.select_combobox_item(self, 'comboSourceFields', source['select_field'])
            if 'verify_field' in source:
                elem1=self.driver.find_element_by_css_selector("#comboSourceFields>div")
                d=utillityobject.get_attribute_value(self, elem1, "text")
                d['text']=d['text'].replace(u'\xa0', u' ')
                utillityobject.asequal(self, d['text'], source['verify_field'], msg + ".c: Verify report_type - "+ source['verify_field'] + " displayed")
            if 'select_condition' in source:
                utillityobject.select_combobox_item(self, 'comboConditions', source['select_condition'])
            if 'verify_condition' in source:
                elem1=self.driver.find_element_by_css_selector("#comboConditions>div")
                d=utillityobject.get_attribute_value(self, elem1, "text")
                utillityobject.asequal(self, d['text'], source['verify_condition'], msg + ".d: Verify report_type - "+ source['verify_condition'] + " displayed")
            if 'select_sort' in source:
                utillityobject.select_combobox_item(self, 'comboSortValues', source['select_sort'])
            if 'verify_sort' in source:
                elem1=self.driver.find_element_by_css_selector("#comboSortValues>div")
                d=utillityobject.get_attribute_value(self, elem1, "text")
                utillityobject.asequal(self, d['text'], source['verify_sort'], msg + ".e: Verify report_type - "+ source['verify_sort'] + " displayed")
            if 'select_includeall' in source:
                self.driver.find_element_by_css_selector("#checkShowAll input").click()
            if 'verify_includeall' in source:
                ocheck=self.driver.find_element_by_css_selector("#checkShowAll input").is_selected()
                utillityobject.asequal(self,source['verify_includeall'], ocheck, msg + '.g: Verifying Include_All checkbox box status - ' + str(source['verify_includeall']))
        if targets != None:
            if 'select_candidate_report' in targets:
                utillityobject.select_item_in_dialog(self, "#gridADPCandidates", targets['select_candidate_report'])
            if 'verify_candidate_report' in targets:
                utillityobject.verify_items_in_dialog(self, "#gridADPCandidates div[class$='content']", targets['verify_candidate_report'], msg + ".h: Verify " + str(targets['verify_candidate_report']) + " in 'Targets' section")
            if 'add_to_target' in targets:
                elem=self.driver.find_element_by_css_selector("#btnAddToTargetList img")
                coreutillityobject.left_click(self, elem)
                time.sleep(2)
            if 'remove_from_target' in targets:
                elem=self.driver.find_element_by_css_selector("#btnRemoveFromTargetList img")
                coreutillityobject.left_click(self, elem)
                time.sleep(2)
            if 'select_target_name' in targets:
                utillityobject.select_item_in_dialog(self, "#gridADPTargets", targets['select_target_name'])
            if 'verify_target_name' in targets:
                utillityobject.verify_items_in_dialog(self, "#gridADPTargets div[class$='content']", targets['verify_target_name'], msg + ".i: Verify " + str(targets['verify_target_name']) + " in 'Targets' section")
        'Click OK'
        time.sleep(1)
        btn=btn_type if btn_type != None else "ok" 
        button_obj = {'ok':'btnADPOK','cancel':'btnADPCancel','Apply':'btnADPApply'}
        button_css = "#adpPropertiesDlg [id=" + button_obj[btn] + "] img"
        self.driver.find_element_by_css_selector(button_css).click()
        time.sleep(3)
        
    def add_and_verify_prompts_in_cascade(self, cascades=None, available_prompts=None, selected_prompts=None, msg="Step x"):
        """
        @Param :cascades={'add_cascade':'True','select_cascade':'list_1','delete_cascade':'True','verify_cascades':['Cascasde_1']}
        @Param :available_prompts={'select_promps':'list_1','add_prompt':'True', 'remove_prompt':'True', 'verify_prompts':['list_1','checkbox_2']}  
        @Param :selected_prompts={'select_selected_prompts':'list_1', 'verify_selected_items':['list_1', 'checkbox_2']}
        @Param :msg="Verify added prompts are available in selected prompts"
        Syntax:add_and_verify_prompts_in_cascade(cascades=None, available_prompts=None, selected_prompts=None, msg="Step x")
        @author:Bhagavathi
        """
        cascade_css=self.driver.find_element_by_css_selector("#btnCascadesOpt img")
        utillityobject.default_click(self, cascade_css, click_option=0)
        time.sleep(5)
        if cascades!=None:
            if 'add_cascade' in cascades:
                self.driver.find_element_by_css_selector("#btnNewCacsade img").click()
                time.sleep(2)
            if 'select_cascade' in cascades:
                cascade_names=self.driver.find_elements_by_css_selector("#listADPCascades div[id^='BiListItem']")
                cascade_names_list=[el.text.strip() for el in cascade_names]
                cascade_names[cascade_names_list.index(cascades['select_cascade'])].click()
            if 'delete_cascade' in cascades:
                self.driver.find_element_by_css_selector("#btnDeleteCascade img").click()
                time.sleep(5)
            if 'verify_cascades' in cascades:
                act_cascades=self.driver.find_elements_by_css_selector("#listADPCascades div[id^='BiListItem']")
                act_cascades_list=[el.text.strip() for el in act_cascades]
#                 print(act_cascades_list)
                utillityobject.asequal(self, act_cascades_list, cascades['verify_cascades'], msg + "Cascade Name"+ str(act_cascades_list) +" is visible in cascade tab")   
        if available_prompts!=None:
            if 'select_prompts' in available_prompts:
                cascade_available_prompt_css=self.driver.find_elements_by_css_selector("#gridAvailableADP div[class$='content'] table tbody tr td:nth-child(1)")
                actual_prompt_list=[el.text.strip() for el in cascade_available_prompt_css  if bool(re.match('\S+', el.text.strip()))]
#                 print("select_prompts", actual_prompt_list)
                cascade_available_prompt_css[actual_prompt_list.index(available_prompts['select_prompts'])].click()
            if 'add_prompt' in available_prompts:
                add_image_css=self.driver.find_element_by_css_selector("#btnAddToCascade img")
                utillityobject.default_click(self, add_image_css, click_option=0)
                time.sleep(2)
            if 'remove_prompt' in available_prompts:
                remove_img_css=self.driver.find_element_by_css_selector("#btnRemoveFromCascade img")
                utillityobject.default_click(self, remove_img_css, click_option=0)
                time.sleep(2)
            if 'verify_prompts' in available_prompts:
                cascade_available_prompt_css=self.driver.find_elements_by_css_selector("#gridAvailableADP div[class$='content'] table tbody tr td:nth-child(1)")
                actual_prompt_list=[el.text.strip() for el in cascade_available_prompt_css  if bool(re.match('\S+', el.text.strip()))]
#                 print("verify_prompts", actual_prompt_list)
                utillityobject.asequal(self, actual_prompt_list, available_prompts['verify_prompts'], msg + "The following prompts"+ str(actual_prompt_list) +" are visible in Available prompts")
        if selected_prompts!=None:
            if 'select_selected_prompts' in selected_prompts:
                selected_prompts=self.driver.find_elements_by_css_selector("#gridSelectedPrompts div[class$='content'] table tbody tr td:nth-child(1)")
                actual_selected_prompt_list=[el.text.strip() for el in selected_prompts  if bool(re.match('\S+', el.text.strip()))]
                selected_prompts[actual_selected_prompt_list.index(selected_prompts['select_selected_prompts'])].click()
            if 'verify_selected_items' in selected_prompts:
                selected__prompt_css=self.driver.find_elements_by_css_selector("#gridSelectedPrompts div[class$='content'] table tbody tr td:nth-child(1)")
                actual_selected_prompt_list=[el.text.strip() for el in selected__prompt_css  if bool(re.match('\S+', el.text.strip()))]
#                 print("verify_selected_items", actual_selected_prompt_list)
                utillityobject.asequal(self, actual_selected_prompt_list, selected_prompts['verify_selected_items'], msg + "Selected prompts"+ str(actual_selected_prompt_list) +" are visible in selected prompts")

    def verify_riser_chart_XY_labels(self, parent_id, expected_xval_list, expected_yval_list, msg, x_axis_label_length=4, **kwargs):#Need to Delete
        """
        parent_id: 'wall1' OR 'MAINTABLE0'
        params: expected_xval_list=['Accessories', 'Camcorder', 'Computers', 'Media Player', 'Stereo Sys...', 'Televisions', 'Video Prod...']
        params: expected_yval_list=['0', '20M', '40M', '60M', '80M', '100M']
        Syntax.1: verify_riser_chart_XY_labels('wall1', expected_xval_list, expected_yval_list, 'Step <no.>')
        Syntax.2: verify_riser_chart_XY_labels('wall1', expected_xval_list, expected_yval_list, 'Step <no.>', y_custom_css="text[class^='zaxisOrdinal-labels']")
        Syntax.3: verify_riser_chart_XY_labels('wall1', expected_xval_list, expected_yval_list, 'Step <no.>', x_custom_css="text[class^='zaxisOrdinal-labels']")
        Syntax.3: verify_riser_chart_XY_labels('wall1', expected_xval_list, expected_yval_list, 'Step <no.>', z_expected_labels=['Coffee', 'Food'])
        @author: Niranjan
        """
        # debuggin
        if parent_id.startswith('#'):
            parent=parent_id
        else:
            parent="#" + parent_id
        if expected_xval_list !=[]:
            if 'x_custom_css' in kwargs:
                x_custom_css=kwargs['x_custom_css']
            else:
                x_custom_css="svg > g text[class^='xaxis'][class*='labels']"
            x_sync_css=parent+" "+x_custom_css
            utillityobject.synchronize_until_element_is_visible(self, x_sync_css, 290)
            time.sleep(5)
            x=self.driver.find_elements_by_css_selector(x_sync_css)
            x = [label for label in x if label.is_displayed()]
            my_iter_x=(i.text for i in x)
            for label_x in expected_xval_list:
                if label_x[:x_axis_label_length] == next(my_iter_x)[:x_axis_label_length]:
                    statex= True
                    state=True
                else:
                    statex=expected_xval_list
                    state=[i.text for i in x]
                    break
            del my_iter_x
            utillityobject.asequal(self, statex, state, msg + " Verify the x-Axis labels.") 
        if expected_yval_list !=[]:
            if 'y_custom_css' in kwargs:
                y_custom_css=kwargs['y_custom_css']
            else:
                y_custom_css="svg > g text[class^='yaxis-labels']"
            y_sync_css=parent+" "+ y_custom_css
            utillityobject.synchronize_until_element_is_visible(self, y_sync_css, 290)
            time.sleep(5) 
            y=self.driver.find_elements_by_css_selector(y_sync_css)
            my_iter_y=(l.text for l in y)
            for label_y in expected_yval_list:
                if label_y == next(my_iter_y):
                    statey= True
                    state= True
                else:
                    statey=expected_yval_list 
                    state= [l.text for l in y] 
                    break
            del my_iter_y
            utillityobject.asequal(self, statey, state, msg + " Verify the y-Axis labels.")
        if 'z_expected_labels' in kwargs:
            if 'z_custom_css' in kwargs:
                z_custom_css=kwargs['z_custom_css']
            else:
                z_custom_css="svg > g text[class^='zaxis'][class*='labels']"
            z_sync_css=parent+" "+ z_custom_css
            utillityobject.synchronize_until_element_is_visible(self, z_sync_css, 290)
            time.sleep(5) 
            z=self.driver.find_elements_by_css_selector(z_sync_css) 
            my_iter_z=(m.text for m in z)
            for label_z in kwargs['z_expected_labels']:
                if label_z == next(my_iter_z):
                    statez= True
                    state= True
                else:
                    statez=kwargs['z_expected_labels']
                    state= [m.text for m in z]  
                    break
            del my_iter_z
            utillityobject.asequal(self, statez, state, msg + " Verify the z-Axis labels.")            
            
            
    def verify_riser_pie_labels_and_legends(self, parent_id, expected_label_list, msg,**kwargs):#Need to Delete
        """
        parent_id: 'wall1' OR 'MAINTABLE0'
        :kwargs same_group=True  (Want to verify same groups labels)
        :kwargs custom_css=True  (dynamic css path)
        params: expected_label_list=['Unit Sales','Dollar Sales']
        params: expected_total_label_list=['3.7M','46.2M']
        Syntax: verify_riser_pie_labels_and_legends('wall2', ['Unit Sales','Dollar Sales'], ['Coffee','Food'],"Step:10",expected_total_label_list=['3.7M','46.2M'])
        Syntax: verify_riser_pie_labels_and_legends('MAINTABLE_wbody0', ['Unit Sales','Unit Sales'], 'Unit Sales'],"Step:10",custom_css="text[class^='pieLabel!g']",same_group=True)
        @author: Nasir
        """
        parent="#" + parent_id + " svg > g"
        actual_label_list=[]
        if 'same_group' in kwargs:
            custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else "text[class^='pieLabel!g']"
            labels=self.driver.find_elements_by_css_selector(parent + " " + custom_css)
            for i in range(0,len(labels)):
                label=labels[i].text.strip()
                actual_label_list.append(label)
        else:
            if 'expected_total_label_list' in kwargs:
                actual_total_label_list=[]
                labels=self.driver.find_elements_by_css_selector(parent + " text[class^='totalLabel!g']")
                for i in range(0,len(labels)):
                    total_label=self.driver.find_element_by_css_selector(parent + " text[class='totalLabel!g" + str(i) + "!mtotalLabel!']").text.strip()
                    actual_total_label_list.append(total_label)
                utillityobject.asequal(self,kwargs['expected_total_label_list'], actual_total_label_list, msg + " Verify the pie total labels.")
            labels=self.driver.find_elements_by_css_selector(parent + " text[class^='pieLabel!g']")
            for i in range(0,len(labels)):
                label=self.driver.find_element_by_css_selector(parent + " text[class='pieLabel!g" + str(i) + "!mpieLabel!']").text.strip()
                actual_label_list.append(label)
        utillityobject.asequal(self,expected_label_list, actual_label_list, msg + " Verify the pie labels.")

    def verify_riser_legends(self, parent_id, expected_legend_list, msg,**kwargs):#Need to delete
        """
        parent_id: 'wall1' OR 'MAINTABLE0'
        params: expected_legend_list=['Coffee','Food']        
        Syntax: verify_riser_pie_labels_and_legends('wall2', ['Coffee','Food'],"Step 04: Verify Chart Legend")
        @author: Nasir
        """
        legend_css="#" + parent_id + " .legend text"
        legends=self.driver.find_elements_by_css_selector(legend_css)
        actual_legend_list = [legend.text.strip() for legend in legends]
        for i in range(0, len(expected_legend_list)):
            try:
                if expected_legend_list[i][:4] in actual_legend_list[i]:
                    statex= True
                else:
                    statex=False
                    break
            except IndexError:
                statex=False
        if statex==True:
            utillityobject.asequal(self, statex, True, msg + "Verify the legends.")
        else:
            utillityobject.asequal(self, expected_legend_list, actual_legend_list, msg + "Verify the legends.")
    
    def create_lasso(self, parent_panel_id, source_tag, source_riser, **kwargs):#Need to delete

        """
        parent_panel_id= 'MAINTABLE_wbody1'
        source_tag='rect'
        source_riser='riser!s0!g0!mbar!'
        kwargs['target_tag']=We need to provide it if we want to create lasso by drag.
        kwargs['target_riser']=We need to provide it if we want to create lasso by drag.
        Usage:create_lasso("0",'rect', 'riser!s0!g0!mbar!')
        Usage:create_lasso("0",'rect', 'riser!s0!g0!mbar!', target_tag='rect', target_riser='riser!s0!g2!mbar!')
        Usage:create_lasso("MAINTABLE_wbody1","rect",'riser!s0!g1!mbar!',offsetx='-300',offsety='-10')
        @author: Gobizen, Niranjan
        """
        # Nasir:::: Added flat codes to check the lasso is working in IE or not, Need to update the flat codes using any available function.
        action = ActionChains(self.driver)
        selector = "#"+parent_panel_id + " "+str({0})+"[class*='"+str({1})+"']"
        source = self.driver.find_element_by_css_selector(selector.format(source_tag,source_riser))
        if 'target_riser' in kwargs:
            target = self.driver.find_element_by_css_selector(selector.format(kwargs['target_tag'],kwargs['target_riser']))
            #if Global_variables.browser_name == 'firefox':
            if Global_variables.browser_name in ['firefox','chrome']:
                utillityobject.drag_to_using_pyautogui(self, source, target,**kwargs)
#                 if 'x_offset' in kwargs:
#                     utillityobject.drag_to_using_pyautogui(self, source, target, 0, kwargs['x_offset'], 0, kwargs['y_offset'])
#                 else:
#                     utillityobject.drag_to_using_pyautogui(self, source, target)
            else:
                cord_type=kwargs['cord_type'] if 'cord_type' in kwargs else 'middle' 
                source_obj=utillityobject.get_object_screen_coordinate(self, source, coordinate_type=cord_type, **kwargs)
                source_obj_x=source_obj['x']
                source_obj_y=source_obj['y']
                target_obj=utillityobject.get_object_screen_coordinate(self, target, coordinate_type=cord_type, **kwargs)
                target_obj_x=target_obj['x']
                target_obj_y=target_obj['y']
                time.sleep(2)
                if sys.platform == 'linux':
                    mouse_.press(int(source_obj_x), int(source_obj_y))
                    time.sleep(3)
                    pyautogui.moveTo(int(target_obj_x), int(target_obj_y), duration=3)
                    time.sleep(3)
                    mouse_.release(int(target_obj_x), int(target_obj_y))
                else:
                    mouse_obj=uisoup.mouse
                    time.sleep(1)
                    mouse_obj.drag(source_obj_x,source_obj_y, target_obj_x, target_obj_y)
#             else:
#                 action.drag_and_drop(source, target).perform()
        elif 'offsetx' in kwargs:
            if 'cord_type' in kwargs:
                cord_type = kwargs['cord_type']
            else:
                cord_type = 'middle'
            #if Global_variables.browser_name == 'firefox':
            if Global_variables.browser_name in ['firefox']:
                source_obj = utillityobject.click_on_screen(self, source,coordinate_type=cord_type)
                time.sleep(2)
                pyautogui.dragRel(kwargs['offsetx'], kwargs['offsety'], button='left')
                time.sleep(5)
            elif Global_variables.browser_name in ['ie', 'edge']:
                cord_type=kwargs['cord_type'] if 'cord_type' in kwargs else 'middle' 
                source_obj=utillityobject.get_object_screen_coordinate(self, source, coordinate_type=cord_type, **kwargs)
                source_obj_x=source_obj['x']
                source_obj_y=source_obj['y']
                target_obj_x=source_obj['x'] + (kwargs['offsetx'])
                target_obj_y=source_obj['y'] + (kwargs['offsety'])
                time.sleep(2)
                if sys.platform == 'linux':
                    mouse_.press(int(source_obj_x), int(source_obj_y))
                    time.sleep(3)
                    pyautogui.moveTo(int(target_obj_x), int(target_obj_y), duration=3)
                    time.sleep(3)
                    mouse_.release(int(target_obj_x), int(target_obj_y))
                else:
                    mouse_obj=uisoup.mouse
                    time.sleep(1)
                    mouse_obj.drag(source_obj_x,source_obj_y, target_obj_x, target_obj_y)
            else:
                action1 = ActionChains(self.driver)
                action1.drag_and_drop_by_offset(source,int(kwargs['offsetx']),int(kwargs['offsety'])).perform()
                time.sleep(5)
        else:
            if Global_variables.browser_name in ['ie', 'edge']:
                time.sleep(2)
                cord_type=kwargs['cord_type'] if 'cord_type' in kwargs else 'middle' 
                source_obj=utillityobject.get_object_screen_coordinate(self, source, coordinate_type=cord_type, **kwargs)
                source_obj_x=source_obj['x']
                source_obj_y=source_obj['y']
                target_obj_x=source_obj['x'] + (3)
                target_obj_y=source_obj['y'] + (3)
                time.sleep(2)
                if sys.platform == 'linux':
                    mouse_.press(int(source_obj_x), int(source_obj_y))
                    time.sleep(3)
                    pyautogui.moveTo(int(target_obj_x), int(target_obj_y), duration=3)
                    time.sleep(3)
                    mouse_.release(int(target_obj_x), int(target_obj_y))
                else:
                    mouse_obj=uisoup.mouse
                    time.sleep(1)
                    mouse_obj.drag(source_obj_x,source_obj_y, target_obj_x, target_obj_y)
                #ac=ActionChains(self.driver)
                #ac.drag_and_drop_by_offset(source, 3, 3).perform()                
            else:
                source.click()
        time.sleep(2)     
    
    def create_lasso_using_action_chain(self, source_riser_css, target_riser_css, parent_css="#MAINTABLE_wbody1"):
        """
        Description : Drag mouse from one riser to another riser to create lasso in chart
        :param - source_riser_css : Css value of source riser. Example : rect[class='riser!s0!g1!mbar']
        :param - target_riser_css : Css value of target riser. Example : rect[class='riser!s0!g4!mbar']
        :Usage - create_lasso_using_action_chain("rect[class*='riser!s0!g0!mbar']", "rect[class*='riser!s1!g1!mbar")
        """
        source_riser_css = parent_css + " " + source_riser_css
        target_riser_css = parent_css + " " + target_riser_css
        source_riser_object = utillityobject.validate_and_get_webdriver_object(self, source_riser_css, "Source chart riser")
        target_riser_object = utillityobject.validate_and_get_webdriver_object(self, target_riser_css, "Target chart riser")
        ActionChains(self.driver).drag_and_drop(source_riser_object, target_riser_object).perform()
        
    def select_or_verify_lasso_filter(self, **kwargs):#Need to Delete
        """
        param kwargs['select'] = 'item to be verified'. This must be string.
        param : kwargs['verify'] = This must be a list
        param: kwargs['msg'] = If you need to verify the list, then use this msg.
        Usage: select_or_verify_lasso_filter(verify=['1 point','Filter Chart','Exclude from Chart'],msg='Step 04: Expect to see the left-click options appear')
        Author: Niranjan
        """
        tooltip_popup_css="[id^='ibi'][class*='tdgchart-tooltip']:not([style*='hidden'])"
        menu_css=tooltip_popup_css+" .tdgchart-tooltip-pad"
        utillityobject.synchronize_with_number_of_element(self, tooltip_popup_css, 1, 100, pause_time=0.3)
        tooltip_popup_obj = self.driver.find_element_by_css_selector(tooltip_popup_css)
        try:
            menus=tooltip_popup_obj.find_elements_by_css_selector(menu_css)
            coreutillityobject.move_to_element(self, menus[0])
        except NoSuchElementException:
            raise AttributeError("Add Comment editor not found in DOM")
        menus=self.driver.find_elements_by_css_selector(menu_css)
        menu_list=[el.text.strip() for el in menus]
        if 'select' in kwargs:
            filter_obj=menus[menu_list.index(kwargs['select'])]
            utillityobject.default_click(self, filter_obj)
        if 'verify' in kwargs:
            utillityobject.asequal(self, kwargs['verify'], menu_list, kwargs['msg'])
        time.sleep(1)

    def verify_number_of_riser(self, parent_id, riser_per_segment, expected_number, msg,**kwargs):#Need to Delete
        """
        parent_id: 'MAINTABLE_wbody2'
        riser_per_segment: 1 or 2 ..(Any integer value. This is the number of segments in one bar)
        expected_number: This is also a number. 
        syntax verify_number_of_riser('MAINTABLE_wbody2', 3, 4, 'Step 10: Verify number of risers displayed')
        @author: Kiruthika 8Jun2016
        """
        if 'custome_css' in kwargs :
            riser_css="#" + parent_id + kwargs['custome_css']
        else :
            riser_css="#" + parent_id + " svg g.risers >g>rect[class^='riser']"
        total_risers=len(self.driver.find_elements_by_css_selector(riser_css))
        actual_number=int(total_risers/riser_per_segment)
        utillityobject.asequal(self, expected_number, actual_number, msg)

    def verify_number_of_pie_segments(self, parent_id, riser_per_segment, expected_number, msg):#Need to delete
        """
        parent_id: 'MAINTABLE_wbody1'
        riser_per_segment: 1 or 2 ..(Any integer value. This is the number of segments in pie chart)
        expected_number: This is also a number. 
        syntax verify_number_of_pie_segments('MAINTABLE_wbody2', 3, 4, 'Step 10: Verify number of pie chart segments displayed')
        @author: Magesh
        """
        parent="#" + parent_id + " svg.rootPanel>g.chartPanel"
        total_risers=len(self.driver.find_elements_by_css_selector(parent + " >g>g>g>path[class^='riser']"))
        actual_number=int(total_risers/riser_per_segment)
        utillityobject.asequal(self, expected_number, actual_number, msg)
    
    def verify_number_of_circle(self, parent_id, start_range, end_number, msg):#Need to Delete
        """
        :param: parent_id='MAINTABLE_wbody1'
        :param: start_range=1 or 2 or 3 or ...
        :param: end_number=1 or 2 or 3 or ...  
        :usage: verify_number_of_riser('MAINTABLE_wbody1', 3, 4, 'Step 10: Verify number of Circle displayed')
        :note: if you want check only '3' then pass start_range=3 and end_range=4
        @author: AAkhan
        """
        parent="#" + parent_id + " svg g"
        actual_number=len(self.driver.find_elements_by_css_selector(parent + " circle[class^='riser']"))
        if actual_number in range(start_range , end_number):
            actual_number=True
        utillityobject.asequal(self, True, actual_number, msg)

    def create_grid_data_set(self,parent_id,rows, file_name):
        """
        params: parent_id='MAINTABLE_wbody1'
        Usage: create_grid_data_set('MAINTABLE_wbody1',5,'C2140681.xlsx')
        Author: Niranjan
        """
        wb = Workbook()
        row_css="#" + parent_id + " g.innerTable g[class^=row]"
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(row_css)
        for r in range(0, rows):
            columns = table_rows[r].find_elements_by_css_selector("text")
            for c in range(0, len(columns)):
                s.cell(row=r + 1, column=c + 1).value = str(columns[c].text)
        wb.save(os.getcwd() + "\data\\" + file_name)   
        
    def compare_grid_data_set(self, parent_id,rows, file_name):
        """
        params: parent_id='MAINTABLE_wbody1'
        Usage: create_grid_data_set('MAINTABLE_wbody1',5,'C2140681.xlsx')
        Author: Niranjan
        """
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        status=[]
        row_css="#" + parent_id + " g.innerTable g[class^=row]"
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(row_css)
        for r in range(0, rows):
            columns = table_rows[r].find_elements_by_css_selector("text")
            for c in range(0, len(columns)):
                if s1.cell(row=r + 1, column=c + 1).value == str(columns[c].get_attribute("text")):
                    status=[0]
                    continue
                else:
                    status=[r+1,c]
                    return (status)
        return (status)     


    def verify_grid_data_set(self,parent_id,rows, file_name, msg):
        """
        Usage: verify_grid_data_set('MAINTABLE_wbody1',5,'C2140681.xlsx',"Step 10: fail data set")
        """
        x= self.compare_grid_data_set(parent_id,rows, file_name)
        utillityobject.asequal(self,len(x),1,msg+ ' Row,Column -'+ str(x))
        
    def verify_xaxis_title(self,popup_id,expected_title,msg,**kwargs):#Need to be deleted.
        """
        Params: table_id='MAINTABLE_wbody0_f'
        Params: expected_title='RCOUNTRY : CAR'
        Syntax: verify_xaxis_title('MAINTABLE_wbody0_f', 'COUNTRY : CAR', 'Step 03: Verify Title')
        @Author: Kiruthika 
        """
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else "text[class^='xaxis'][class$='title']"
        parent_css="#"+popup_id+" " + custom_css
        utillityobject.synchronize_with_visble_text(self, parent_css, expected_title, 290)
        xtitle=self.driver.find_element_by_css_selector(parent_css).text.strip()
        utillityobject.asequal(self,xtitle,expected_title,msg)
    
    def verify_yaxis_title(self,popup_id,expected_title,msg, **kwargs):#Need to be deleted.
        """
        Params: table_id='MAINTABLE_wbody0_f'
        Params: expected_title='SALES'
        Syntax: verify_yaxis_title('MAINTABLE_wbody0_f', 'SALES', 'Step 03: Verify Title')
        @Author: Kiruthika 
        """
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else "text[class='yaxis-title']"
        parent_css="#"+popup_id+" " + custom_css
        utillityobject.synchronize_with_visble_text(self, parent_css, expected_title, 290)
        ytitle=self.driver.find_element_by_css_selector(parent_css).text.strip()
        utillityobject.asequal(self,ytitle,expected_title,msg)
                    
#Function 45: Drag and Drop Visualization to required position
    def drag_and_drop_visualization(self,source_chart,target_chart, position, **kwargs):
        """

        :param source_chart: BubbleMap1
        :param target_chart: Grid1
        :param position: left_most,left,top_most,top,centre,bottom,bottom_most,right,right_most
        :Usage: ia.drag_and_drop_visualization("BubbleMap1","Grid1","bottom")
        :author: Sindhuja Date: June 9
        """
        action = ActionChains(self.driver)
        if 'Prompts' in kwargs:
            Source= self.driver.find_element_by_xpath(VisualizationResultareaLocators.Prompts.format(source_chart))
        else:
            Source= self.driver.find_element_by_xpath(VisualizationResultareaLocators.chart_type.format(source_chart))
        Target= self.driver.find_element_by_xpath(VisualizationResultareaLocators.chart_type.format(target_chart))
        if Global_variables.browser_name == 'firefox':
            utillityobject.mouse_action_using_pyautogui(self, Source, 25, -7, mouse_down=True)
        else:
            action.click_and_hold(Source)
        if Global_variables.browser_name == 'firefox':
            utillityobject.mouse_action_using_pyautogui(self, Target, mouse_move=True)
        else:
            action.move_to_element(Target)
        if position=="left_most":
            left_most_elem=self.driver.find_element(*VisualizationResultareaLocators.Left_most)
            if Global_variables.browser_name == 'firefox':
                utillityobject.mouse_action_using_pyautogui(self, left_most_elem, mouse_up=True)
            else:
                action.release(left_most_elem)
        if position=="left":
            left_elem=self.driver.find_element(*VisualizationResultareaLocators.Left)
            if Global_variables.browser_name == 'firefox':
                utillityobject.mouse_action_using_pyautogui(self, left_elem, mouse_up=True)
            else:
                action.release(left_elem)
        if position=="top_most":
            top_most_elem=self.driver.find_element(*VisualizationResultareaLocators.Top_most)
            if Global_variables.browser_name == 'firefox':
                utillityobject.mouse_action_using_pyautogui(self, top_most_elem, mouse_up=True)
            else:
                action.release(top_most_elem)
        if position=="top":
            top_elem=self.driver.find_element(*VisualizationResultareaLocators.Top)
            if Global_variables.browser_name == 'firefox':
                utillityobject.mouse_action_using_pyautogui(self, top_elem, mouse_up=True)
            else:
                action.release(top_elem)
        if position=="centre":
            centre_elem=self.driver.find_element(*VisualizationResultareaLocators.Centre)
            if Global_variables.browser_name == 'firefox':
                utillityobject.mouse_action_using_pyautogui(self, centre_elem, mouse_up=True)
            else:
                action.release(centre_elem)
        if position=="bottom":
            bottom_elem=self.driver.find_element(*VisualizationResultareaLocators.Bottom)
            if Global_variables.browser_name == 'firefox':
                utillityobject.mouse_action_using_pyautogui(self, bottom_elem, mouse_up=True)
            else:
                action.release(bottom_elem)
        if position=="bottom_most":
            bottom_most_elem=self.driver.find_element(*VisualizationResultareaLocators.Bottom_most)
            if Global_variables.browser_name == 'firefox':
                utillityobject.mouse_action_using_pyautogui(self, bottom_most_elem, mouse_up=True)
            else:
                action.release(bottom_most_elem)
        if position=="right":
            right_elem=self.driver.find_element(*VisualizationResultareaLocators.Right)
            if Global_variables.browser_name == 'firefox':
                utillityobject.mouse_action_using_pyautogui(self, right_elem, mouse_up=True)
            else:
                action.release(right_elem)
        if position=="right_most":
            right_most_elem=self.driver.find_element(*VisualizationResultareaLocators.Right_most)
            if Global_variables.browser_name == 'firefox':
                utillityobject.mouse_action_using_pyautogui(self, right_most_elem, mouse_up=True)
            else:
                action.release(right_most_elem)
        action.perform()
 
    def verify_visualization_row_column_header_labels(self,parent_id,matrix_type,expected_header,expected_label,msg, label_length=4):#Need to Delete
        """
        :param parent_id: MAINTABLE_wbody1
        :param matrix_type: 'Rows' or 'columns'
        :param expected_rowheader: 'Product Category'
        :param expected_rowlabel: ['Hazelnut', 'B141', 'French Roast', 'B142', 'Kona', 'B144', 'Scone', 'F101', 'Biscotti']
        :param msg: "Step 10"
        Usage: verify_matrix_marker_row_header_labels('MAINTABLE_wbody1','Product Category', ['Hazelnut', 'B141', 'French Roast', 'B142', 'Kona', 'B144', 'Scone', 'F101', 'Biscotti'], "Step 04:")
        @author: Niranjan
        """
        if matrix_type == 'Rows':
            header_css="#" + parent_id + " g.scrollRowTitle"
            label_css="#" + parent_id + " g.scrollRowLbl > g text"
        else:
            header_css="#" + parent_id + " g.scrollColTitle"
            label_css="#" + parent_id + " g.scrollColLbl > g text"

        actual_rowheader=self.driver.find_element_by_css_selector(header_css).text.strip().replace('\n','')
        utillityobject.asin(self, expected_header, actual_rowheader, msg + " Verify Header tile for " + matrix_type)
        actual_val_list=(el.text.strip() for el in self.driver.find_elements_by_css_selector(label_css) if el.text.strip()!='')
        for label_x in expected_label:
            if label_x[:label_length]==next(actual_val_list)[:label_length] :
                state= True
            else:
                state=False
                break
        utillityobject.asequal(self,state, True, msg + " Verify Header labels for " + matrix_type) 


    def wait_for_property(self, parent_css, expected_number, expire_time=250, text_option='dom_visible_text', **kwargs):
        """
        @author: Aftab
        :param parent_css = "#MAINTABLE_wbody1 svg > g text[class^='xaxis'][class*='labels']"     '''This is css locator'''
        :param expected_number = 1,2,3.....                                                       '''Only Interger Number'''
        :param expire_time = 50 or 100                                               '''Only Interger Number to wait time unitil condition not that specific time''
        :param string_value='Country'        this need to verify the exact string visible in Dome
        :param with_regular_exprestion=True        This will remove ' '(space) and '\n' from string
        :usage wait_time("#MAINTABLE_wbody1 svg > g text[class^='xaxis'][class*='labels']", 2)
            OR
        :Usage wait_for_property("#TableChart_1 [align='justify']", 1, string_value="Draganddropfieldsontothecanvasorintothequerypanetobeginbuildingyourreport.")
            OR
        :Usage wait_for_property("#TableChart_1 [align='justify']", 1, string_value="Draganddropfieldsontothecanvasorintothequerypanetobeginbuildingyourreport.", with_regular_exprestion=True)
            OR
        :Usage wait_for_property("#TableChart_1 [align='justify']", 1, expire_time= 50, string_value="Draganddropfieldsontothecanvasorintothequerypanetobeginbuildingyourreport.", with_regular_exprestion=True)
        """
        timeout=0
        run_ = True
        while run_:
            timeout+=1
            if timeout == expire_time+1:
                raise ValueError('The Provided CSS - [ {0} ] for syncronization mismatched'.format(parent_css))
            if 'string_value' in kwargs:
                if 'parent_object' in kwargs:
                    temp_str_value_elem=kwargs['parent_object']
                    temp_str_value=utillityobject.get_attribute_value(self, temp_str_value_elem, text_option)
                else:
                    try:
                        temp_str_value_elem=self.driver.find_element_by_css_selector(parent_css)
                        temp_str_value=utillityobject.get_attribute_value(self, temp_str_value_elem, text_option)
                    except NoSuchElementException:
                        time.sleep(1)
                        continue
                    except StaleElementReferenceException:
                        time.sleep(1)
                        continue
                str_value = re.sub(' ','',temp_str_value[text_option]).replace('\n','')
                if str(str(kwargs['string_value']).replace(' ','')) in str_value:
                    run_=False
                    break
            else:
                try:
                    temp_obj = self.driver.find_elements_by_css_selector(parent_css)
                    check_obj_exist = temp_obj[0]
                    del check_obj_exist
                except IndexError:
                    time.sleep(1)
                    continue
                if len(temp_obj) == expected_number:
                    run_=False
                    break
                else:
                    time.sleep(1)
            time.sleep(1)
            
            
    def verify_data_labels(self, parent_id, expected_datalabel, msg, **kwargs):
        """
        :param :parent_id = 'MAINTABLE_wbody1'
        :param :kwargs['custom_css'] = ".chartPanel text[class^='mdataLabels']"       (Other css value user pass)
        :param :expected_datalabel: ['1,546', '1,446', '1,646']
        :param :msg: 'Step 10: Verify number of pie chart segments displayed'
        :Usage :verify_data_labels('TableChart_1', ['1,546', '1,446', '1,646'], 'Step 4: Verify data labels', custom_css=".chartPanel .groupPanel text[class^='mdataLabels']")
                OR
        :Usage :verify_data_labels('MAINTABLE_0', ['1,546', '1,446', '1,646'], 'Step 4: Verify data labels')
        @author: AAkhan
        """
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else ".chartPanel text[class^='mdataLabels']"
        parent="#" + parent_id + " " + custom_css
        actual_datalabel=[]
        total_datalabel=self.driver.find_elements_by_css_selector(parent)
        if 'data_label_length' in kwargs:
            my_iter_x=(i.text for i in total_datalabel)
            for label_x in expected_datalabel:
                if label_x[:kwargs['data_label_length']] == next(my_iter_x)[:kwargs['data_label_length']]:
                    statex= True
                    exp_statex=True
                else:
                    statex=[i.text for i in total_datalabel]
                    exp_statex=expected_datalabel
                    break
            del my_iter_x
            utillityobject.asequal(self, exp_statex, statex, msg)
        else:
            for i in range(len(total_datalabel)):
                actual_datalabel.append(total_datalabel[i].text.strip())
            utillityobject.as_List_equal(self, expected_datalabel, actual_datalabel, msg)
            
    def verify_data_labels_(self, parent_id, expected_datalabel, msg, **kwargs):
        """
        :param :parent_id = 'MAINTABLE_wbody1'
        :param :kwargs['custom_css'] = ".chartPanel text[class^='mdataLabels']"       (Other css value user pass)
        :param :expected_datalabel: ['1,546', '1,446', '1,646']
        :Usage :verify_data_labels_('TableChart_1', ['1,546', '1,446', '1,646'], 'Step4: Verify data labels', custom_css=".chartPanel .groupPanel text[class^='mdataLabels']")
                OR
        :Usage :verify_data_labels_('MAINTABLE_0', ['1,546', '1,446', '1,646'], 'Step 4: Verify data labels', data_label_length=2)
        @author: AAkhan
        """
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else ".chartPanel text[class^='mdataLabels']"
        parent="#" + parent_id + " " + custom_css
        actual_datalabel_=[]
        total_datalabel=self.driver.find_elements_by_css_selector(parent)
        actual_datalabel_ = [e for e in [i.text.strip() for i in total_datalabel] if e != '']      
        if 'data_label_length' in kwargs:
            for i in range(len(actual_datalabel_)):
                if expected_datalabel[i][:kwargs['data_label_length']] == actual_datalabel_[i][:kwargs['data_label_length']]:
                    statex= True
                else:
                    statex=False
                    break
            utillityobject.asequal(self, statex, True, msg)
        else:
            utillityobject.as_List_equal(self, expected_datalabel[:len(actual_datalabel_)], actual_datalabel_, msg)
    
    def verify_header_footer_text(self, foreign_object_class, **kwargs):
        page_css="#MAINTABLE_wbody1 foreignObject[class='" + foreign_object_class + "'] span"
        page_obj=self.driver.find_element_by_css_selector(page_css)
        if 'font_color' in kwargs:
            expected_font_color=utillityobject.color_picker(self,kwargs['font_color'], 'rgba')
            actual_font_color=Color.from_string(page_obj.value_of_css_property("color")).rgba
            utillityobject.asequal(self,actual_font_color, expected_font_color , kwargs['msg'] + ". Verification of Font color")
        if 'bold' in kwargs:
                actual_weight=True if page_obj.value_of_css_property("font-weight") in ['700', 'bold'] else False
                utillityobject.asequal(self, kwargs['bold'], actual_weight, kwargs['msg'] + ". Verification of Cell Text is Bold.")
    
    def select_drilldown_tooltip_menu(self, parent_id, raiser_class, menu_path, wait_time=2, **kwargs):
        """
        :param :parent_id = 'MAINTABLE_wbody1'
        :param :raiser_class = "riser!s0!g2!mbar" 
        :param :menu_path = 'Drill Down_1' #to select the drilldown value from tooltip
        :Usage :resultobj.select_drilldown_tooltip_menu("MAINTABLE_wbody1", "riser!s0!g2!mbar", 'Drill Down_1')
        """
        if 'default_move' in kwargs:
            pass
        else:
            move_to_parentID = self.driver.find_element_by_css_selector("#"+ parent_id)
            fullpath_of_raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
            raiser_elem=self.driver.find_element_by_css_selector(fullpath_of_raiser_css)
            if Global_variables.browser_name in ['firefox', 'ie', 'edge']:
                utillityobject.click_type_using_pyautogui(self, move_to_parentID, move=True, **kwargs)
                utillityobject.click_type_using_pyautogui(self, raiser_elem, **kwargs)
            else:
                action1 = ActionChains(self.driver)
                action1.move_to_element_with_offset(move_to_parentID,1,1).perform()
                action1.move_to_element(raiser_elem).perform()
                del action1
        time.sleep(1)
        tooltip_css="span[id*='tdgchart-tooltip']>div>ul>li [class*='tdgchart'][class*='tooltip'][class*='label']"
        #tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip']>div>ul>li [class*='tdgchart'][class*='tooltip'][class*='label']"
        tooltips=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_item_to_be_selected=tooltips[[elem.text.strip() for elem in tooltips].index(menu_path)]
        tooltip_item_to_be_selected.click()
        time.sleep(wait_time) 

            
    def rename_grouped_riser(self, new_name, **kwargs):#Need to Delete
        """
        :param :new_name = "Access, Comp, TV & Video"
        :param :kwargs['verify_name'] = "Accessories and Computers and Televisions and 1 more."
        :Usage :resultobj.rename_grouped_riser("Access, Comp, TV & Video", verify_name="Accessories and Computers and Televisions and 1 more.", msg = 'Step03')
        """
        btn_css = "div[id^='BiDialog'] [class*='window-active']"
        Visualization_Resultarea.wait_for_property(self, btn_css, 1)
        edit_title_obj=self.driver.find_element_by_css_selector(btn_css)
        if 'verify_name' in kwargs:
            actual_name=edit_title_obj.find_element_by_css_selector("input").get_attribute('value')
            utillityobject.asequal(self, actual_name, kwargs['verify_name'], kwargs['msg']+": Verify current name.")
        time.sleep(3)
        text_field = edit_title_obj.find_element_by_css_selector("input")
        utillityobject.set_text_field_using_actionchains(self, text_field, new_name)
        utillityobject.click_dialog_button(self,btn_css,"OK")
        time.sleep(1)
    
    def verify_drilldown_navigation_menu(self,parent_css,expected_menu_list,msg):
        '''
        :Param : parent_css='foreignObject div'
        :Param : expected_menu_list=['Home', '->', 'Stereo Systems']
        :Usage : verify_drilldown_navigation_menu('foreignObject div', ['Home','->','ENGLAND'], 'Step 21.6 : Verify drilldown navigation menu')
        '''
        navigation_css=parent_css+" span"
        actual_menu_list=[]
        menu_elements=self.driver.find_elements_by_css_selector(navigation_css)
        for menu in menu_elements :
            if '\u2192' in menu.text.strip() :
                actual_menu_list.append('->')
            else :
                actual_menu_list.append(menu.text.strip())
        utillityobject.asequal(self,actual_menu_list,expected_menu_list,msg)  
        
    def verify_header_footer_property(self, parent_id, header_footer_index, verify_style, step_num):
        """
        parent_id = "MAINTABLE_1"     
        header_footer_index = 1
        verify_style={'font_color':'red','bold':True, 'text_value':'100.00','text_align':'754','bg_color':'yellow'}
        step_num='11.1'
        font_name='Arial',font_size='12pt',italic=True, underline=True
        Usage:verify_header_footer_property('MAINTABLE_1', 1, verify_style,"11")
        """
        
        total_elems=self.driver.find_elements_by_css_selector("#" + parent_id + " div span")
        header_footer_elems=[el for el in total_elems if bool(re.match('\S', el.text.strip()))]
        item_bgs=self.driver.find_elements_by_css_selector("#" + parent_id + " div span[style*='background']")
        text_align_elems=self.driver.find_elements_by_css_selector("#" + parent_id + " [class='title']")
        for key in verify_style:
            if 'bg_color' in key:
                expected_background_color=utillityobject.color_picker(self, verify_style[key], 'rgba')
                actual_background_color=Color.from_string(utillityobject.get_element_css_propery(self, item_bgs[header_footer_index-1],"background-color")).rgba
                utillityobject.asequal(self, actual_background_color, expected_background_color , "Step"+str(step_num) + ": Verification of Background color.")
            if 'font_color' in key:
                expected_text_color=utillityobject.color_picker(self, verify_style[key], 'rgba')
                actual_text_color=Color.from_string(utillityobject.get_element_css_propery(self, header_footer_elems[header_footer_index-1],"color")).rgba
                utillityobject.asequal(self, actual_text_color, expected_text_color, "Step"+str(step_num) + ": Verification of Text color.")
            if 'text_value' in key:
                actual_text=header_footer_elems[header_footer_index-1].text.strip()
                utillityobject.asequal(self, verify_style[key], actual_text, "Step"+str(step_num) + ": Verification of Text.")
            if 'font_name' in key:
                actual_font=utillityobject.get_css_value(self, header_footer_elems[header_footer_index-1],"font-family").strip('"')
                utillityobject.asequal(self, verify_style[key].upper(), actual_font.upper(), "Step"+str(step_num) + ": Verification of Text Font name.")
            if 'font_size' in key:
                expected_font_size=round(float(1.333333*int(verify_style[key][:-2])))
                actual_font_size=utillityobject.get_element_css_propery(self, header_footer_elems[header_footer_index-1],"font-size")
                actual_font_size=round(float(actual_font_size[:-2]))
                utillityobject.asequal(self, expected_font_size, actual_font_size, "Step"+str(step_num) + ": Verification of Text Font Size.")
            if 'bold' in key:
                actual_weight=True if utillityobject.get_element_css_propery(self, header_footer_elems[header_footer_index-1],"font-weight") in ['700', 'bold'] else False
                utillityobject.asequal(self, verify_style[key], actual_weight, "Step"+str(step_num) + ": Verification of Cell Text is Bold.")
            if 'italic' in key:
                actual_style=True if utillityobject.get_element_css_propery(self, header_footer_elems[header_footer_index-1],"font-style")=='italic' else False
                utillityobject.asequal(self, verify_style[key], actual_style, "Step"+str(step_num) + ": Verification of Text is Italics.")
            if 'underline' in key:
                actual_decoration=True if utillityobject.get_element_css_propery(self, header_footer_elems[header_footer_index-1],"text-decoration")=='underline' else False
                utillityobject.asequal(self, verify_style[key], actual_decoration, "Step"+str(step_num) + ": Verification of Text is underlined.")
            if 'text_align' in key:                
                if Global_variables.browser_name in ['ie', 'edge']:
                    actual=utillityobject.get_element_attribute(self,text_align_elems[header_footer_index-1],"left")
                    actual_align=re.match('\w+\(([0-9]+).*',actual).group(1)
                    val1=int(verify_style[key])
                else:
                    actual=utillityobject.get_element_attribute(self,text_align_elems[header_footer_index-1],"transform")
                    actual_align=re.match('\w+\(([0-9]+).*',actual).group(1)
                    val1=int(verify_style[key])
                utillityobject.asin(self, int(actual_align), range(val1 - 5,val1 + 5), "Step"+str(step_num) + ": Verification of Text alignment.")
    
    def click_on_marker(self, parent_css,riser_or_marker_element):
        
        """Description:This used to click on marker in chart
        @requires: For drilldown functionality to click on marker,it navigates the webpage in new window
        @params: parent_css:#jschart_HOLD_0
        @params: riser_or_marker_element:marker!s0!g3!mmarker
        usage:click_on_marker("#jschart_HOLD_0","marker!s0!g3!mmarker")
        """
        marker_css = parent_css + " [class*='" + riser_or_marker_element + "']"
        marker_obj=self.driver.find_element_by_css_selector(marker_css)
        utillityobject.click_on_screen(self, marker_obj, 'middle', javascript_marker_enable=True, click_type=0)
    
    def verify_sparkline_chart_tooltip(self, tooltip, step_num, parent_css="#jschart_HOLD_0", chart_location='middle', xoffset=0, yoffset=0):
        """
        Description : Verify Spark Line chart tooltip
        Usage : verify_sparkline_chart_tooltip('14.02', "01.02")
        """
        chart_css = parent_css  + " div[id*='sparkline'] canvas"
        sparkline_chart = utillityobject.validate_and_get_webdriver_object(self, chart_css, "Spark line chart canvas")
        coreutillityobject.python_move_to_element(self, sparkline_chart, chart_location, xoffset, yoffset)
        tooltip_css = "[id='jqstooltip'][style*='visible']"
        utillityobject.synchronize_until_element_is_visible(self, tooltip_css, 20)
        tooltip_obj = utillityobject.validate_and_get_webdriver_object(self, tooltip_css, "Spark line chart tooltip")
        msg = "Step {0} : Verify Spak Line chart tooltip".format(step_num)
        actual_tooltip = tooltip_obj.text.replace("\u25cf", "").strip()
        utillityobject.asequal(self, tooltip, actual_tooltip, msg)