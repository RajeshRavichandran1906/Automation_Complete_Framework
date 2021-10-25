from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
import time
from common.lib.base import BasePage
from common.lib.utillity import UtillityMethods
from common.lib.core_utility import CoreUtillityMethods
from common.pages.page_designer_miscelaneous import PageDesignerMiscelaneous
from common.locators.page_designer_design import Designer_chart_workbook_data_tab
from common.locators.page_designer_design import Designer
from common.lib.javascript import JavaScript
from common.pages.wf_mainpage import Wf_Mainpage
from common.locators.designer_chart_locators import DesignerChart as dc_locators
from common.locators.designer_chart_locators import DesignerInsight as insight_locators
from selenium.webdriver.common.keys import Keys
import re
import sys
import pyautogui
from common.lib.webfocus import poptop_dialog
from common.lib.global_variables import Global_variables
from common.pages.dataformat_dialog import DataFormat_Dialog
from common.pages.visualization_resultarea import Visualization_Resultarea
from builtins import object
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import keyboard

if sys.platform == 'linux':
    from pykeyboard import PyKeyboard
    from pymouse import PyMouse
    pykeyboard = PyKeyboard()
    mouse_=PyMouse() 
else:
    from uisoup import uisoup

class Designer_Metadata(BasePage):
    """ Inherit attributes of the parent class = Baseclass """
    
    def __init__(self, driver):
        super(Designer_Metadata, self).__init__(driver)
        self._utillity = UtillityMethods(self.driver)
        self._coreutility = CoreUtillityMethods(self.driver)
        self._page_design_misc = PageDesignerMiscelaneous(self.driver)
        self._j_script = JavaScript(self.driver)
        self._wf_main_page_obj = Wf_Mainpage(self.driver)
        self._popup_dialog = poptop_dialog.Poptop_Dialog(self.driver)
        self._vislization_result_area = Visualization_Resultarea(self.driver)
        self.MEASURE_PARENT_CSS = None
        self.DIMENSIONS_PARENT_CSS = None
        self.QUERYBOX_PARENT_CSS = ".wfc-bucket-acc-page"
        self.QUERYTOOL_PARENT_CSS = self.QUERYBOX_PARENT_CSS + " [data-ibx-name='bucketToolbar']"
        self.QUERYBUCKET_PARENT_CSS = self.QUERYBOX_PARENT_CSS + " [data-ibx-name='bucketGroup']"
        self.CHART_PICKER_PARENT_CSS = '.chart-picker-box'
        self.FILTER_BAR_PARENT_CSS = '.filter-bar'
        
        
    def get_dimension_field_element(self, field_path):
        pass
    
    def get_measures_field_element(self, field_path):
        pass
    
    def get_querytool_item(self, toolitem_name):
        
        querytool_items_title_dict = {'Change chart orientation' : 'swapChart',
                               'Clear buckets content' : 'deleteAllFields',
                               'Stacked' : 'layoutBlaStacked',
                               'Side-by-Side' : 'layoutBlaSideBySide', 
                               'Absolute' : 'layoutBlaAbsolute',
                               'Percent' : 'layoutBlaPercent'}
        querytool_item_css = self.QUERYTOOL_PARENT_CSS + ' [data-ibx-name=' + querytool_items_title_dict[toolitem_name] + ']'
        querytool_item_description = toolitem_name + ' inside query tool box'
        required_querytool_element=self._utillity.validate_and_get_webdriver_object(querytool_item_css, querytool_item_description)
        return required_querytool_element
    
    def get_querybucket_group_box_element(self, header_name):
        
        '''
        header_name :- Examples:- 'Horizontal' , 'Vertical'..
        '''
        querybucket_group_box_css = self.QUERYBUCKET_PARENT_CSS + ' .wfc-bucket'
        querybucket_group_box_description='query bucket groups'
        querybucket_group_box_items=self._utillity.validate_and_get_webdriver_objects(querybucket_group_box_css, querybucket_group_box_description)
        get_item = False 
        for required_querybucket_group_box_element in querybucket_group_box_items:
            querybucket_group_box_header_css = '.wfc-bucket-header'
            querybucket_group_box_header_description='query bucket group headers'
            temp_obj=self._utillity.validate_and_get_webdriver_object(querybucket_group_box_header_css, querybucket_group_box_header_description, parent_object=required_querybucket_group_box_element)
            if temp_obj.text.strip().lower() == header_name.strip().lower():
                get_item = True
                return required_querybucket_group_box_element
        if get_item == False:
            raise ValueError('Provided [ ' + header_name + ' ] is not available within the query bucket.')
        
    def get_querybucket_field_header_item(self, header_name):
        
        required_querybucket_group_box_element = Designer_Metadata.get_querybucket_group_box_element(header_name)
        querybucket_group_box_header_css = '.wfc-bucket-header'
        querybucket_group_box_header_description='query bucket ' + header_name + ' field header'
        required_querybucket_header_element=self._utillity.validate_and_get_webdriver_object(querybucket_group_box_header_css, querybucket_group_box_header_description, parent_object=required_querybucket_group_box_element)
        return required_querybucket_header_element 
    
    def get_querybucket_field_box_item(self, header_name):
        
        required_querybucket_group_box_element = Designer_Metadata.get_querybucket_group_box_element(header_name)
        querybucket_field_box_css = '.wfc-bucket-pills'
        querybucket_field_box_description='query bucket ' + header_name + ' field box'
        required_querybucket_field_box_element=self._utillity.validate_and_get_webdriver_object(querybucket_field_box_css, querybucket_field_box_description, parent_object=required_querybucket_group_box_element)
        return required_querybucket_field_box_element
    
    
    def get_chart_type_elements(self, chart_title=None, parent_css=None, return_dict_value=False):
        """
        Description: This method is used to get chart type elements
        :Usage: get_chart_type_elements('vertical_stacked_bar', dc_locators.CHART_PICKER_EXPANDED)
        """
        chart_title_dict = {'vertical_stacked_bar' : 'Vertical Stacked Bar', 'horizontal_bar' : 'Horizontal Bar',
                            'vertical_side_by_side_bar' : 'Vertical Side-by-Side Bar', 'ring_pie' : 'Ring Pie',
                            'absolute_line' : 'Absolute Line', 'vertical_stacked_area' : 'Vertical Stacked Area',
                            'scatter_bubble' : 'Scatter/Bubble', 'circle_plot' : 'Circle Plot', 'tree_map' : 'TreeMap',
                            'datagrid' : 'DataGrid', 'matrix_marker' : 'Matrix Marker', 'proportional_symbol_map' : 'Proportional Symbol Map',
                            'choropleth_map' : 'Choropleth Map', 'waterfall' : 'Waterfall', 'gauge' : 'Gauge', 'funnel_chart' : 'Funnel Chart',
                            'mekko_chart' : 'Mekko Chart', 'tagcloud' : 'Tagcloud', 'streamgraph' : 'Streamgraph', 'arc_chart' : 'Arc Chart',
                            'calendar_heat_map_chart' : 'Calendar Heat Map Chart', 'usa_state_cartogram' : 'USA State Cartogram',
                            'chord_diagram' : 'Chord Diagram', 'cluster_diagram' : 'Cluster Diagram', 'compare_2_bars' : 'Compare 2 Bars',
                            'datatables_grid' :'Datatables Grid', 'force_directed_chart' : 'Force Directed Chart',
                            'hexagon_bin_scatter_chart' : 'Hexagon Bin Scatter Chart', 'histogram_chart' : 'Histogram Chart',
                            'sparkline_kpi' : 'Sparkline KPI', 'kpi_with_sparkline_large' : 'KPI with Sparkline Large',
                            'liquid_gauge_chart' : 'Liquid Gauge Chart', 'choropleth_usa_map_chart' : 'Choropleth USA map Chart',
                            'world_choropleth_&_bubble_map' : 'World Choropleth & Bubble Map', 'marker_chart' : 'Marker Chart',
                            'organization_chart' : 'Organization Chart', 'pack_chart' : 'Pack Chart', 'population_pyramid' : 'Population Pyramid',
                            'range_chart' : 'Range Chart', 'ratio_chart' : 'Ratio Chart', 'sankey_flow_chart' : 'Sankey Flow Chart',
                            'simple_bar_chart' : 'Simple Bar Chart', 'sunburst_chart' : 'Sunburst Chart', 'timeline_chart' : 'Timeline Chart',
                            'us_map_chart' : 'US Map Chart', 'what_if_assist' : 'What-If Assist'}
        if return_dict_value:
            return chart_title_dict[chart_title.lower()]
        chart_type_css = parent_css + " .wfc-chartpicker-type-button[title^='{0}']".format(chart_title_dict[chart_title.lower()])
        chart_type_description = chart_title + ' inside chart picker box' 
        required_chart_type_element=self._utillity.validate_and_get_webdriver_object(chart_type_css, chart_type_description)
        return required_chart_type_element
    
    def select_line_style_options(self, line_style):
        """
        Description: This method is used to select line_style in series format option
        :Usage: select_line_style_options('Short_Dash_Dot_Dot')
        """
        lines_dict = {'Solid': "solid", "Dash" : "dash", "Dash_dot":"dash_dot","Dash_dot_dot":"dash_dot_dot", "Long_dash" :"long_dash",
            "Dot":"dot", "Long_Dash_Dot":"long_dash_dot", "Short_Dash":"short_dash",
            "Short_Dash_Dot":"short_dash_dot","Short_Dash_Dot_Dot":"short_dash_dot_dot", "Short_Dot":"short_dot"}
        if line_style in lines_dict :
            Parent_css=".wfc-line-select-menu.pop-top"
            self._coreutility.left_click(self._utillity.validate_and_get_webdriver_object(dc_locators.LINE_DROPDOWN_CSS, 'Line_style_dropdown_css'))
            line_style_css= Parent_css + " div[style*='images/{0}.svg']".format(lines_dict[line_style])
            line_style_elem=self._utillity.validate_and_get_webdriver_object(line_style_css,"Line_style_css")
            self._coreutility.left_click(line_style_elem)
        else :
            msg = "Currently [{0}] line style is not available  ".format(line_style)
            raise KeyError(msg)
        
    def select_chart_picker_element(self, chart_name, expand=False):
        """
        Description: This method is used to select the chart item
        :Usage: select_chart_picker_element('range_chart')
        """
        if expand:
            self._coreutility.left_click(self._utillity.validate_and_get_webdriver_object(dc_locators.CHART_EXPAND_BUTTON, 'expand'))
            self._utillity.synchronize_until_element_is_visible(dc_locators.CHART_PICKER_EXPANDED, 20)
            self._coreutility.left_click(self.get_chart_type_elements(chart_title=chart_name, parent_css=dc_locators.CHART_PICKER_EXPANDED))
            self._coreutility.left_click(self._utillity.validate_and_get_webdriver_object(dc_locators.CHART_EXPAND_BUTTON, 'collapse'))
        else :
            self._coreutility.left_click(self.get_chart_type_elements(chart_title=chart_name, parent_css=dc_locators.CHART_PICKER_DEFAULT))
        
    def verify_chart_index(self, chart_name, chart_index, msg):
        """
        Description: This method is used to verify the chart index
        :Usage : verify_chart_index('vertical_stacked_bar', 0 , 'Step 4.1')
        """
        chart_picker_elements = self._utillity.validate_and_get_webdriver_objects(dc_locators.CHART_PICKER_DEFAULT + ' .wfc-chartpicker-type-button', 'chart-picker')
        verify_text = self._utillity.get_element_attribute(chart_picker_elements[chart_index], 'title')
        self._utillity.asin(self.get_chart_type_elements(chart_title=chart_name, return_dict_value=True), verify_text, msg)
            
    def get_expand_collapse_element_in_chart_picker_box(self):
        
        expand_collapse_css = self.CHART_PICKER_PARENT_CSS + ' .wfc-chartpicker-next-button > .ibx-label-glyph'
        expand_collapse_description = 'expand collapse arrow icon inside chart picker box'
        expand_collapse_element = self._utillity.validate_and_get_webdriver_object(expand_collapse_css, expand_collapse_description)
        return expand_collapse_element
    
    def get_filter_bar_drop_point_element(self):
        
        filter_bar_drop_point_css = self.FILTER_BAR_PARENT_CSS + ' .wfc-fb-drop-label'
        filter_bar_drop_point_description = 'Drop a Filter or a Field Here'
        filter_bar_drop_point_element = self._utillity.validate_and_get_webdriver_object(filter_bar_drop_point_css, filter_bar_drop_point_description)
        return filter_bar_drop_point_element
    
    def invoke_designer_tool_using_api(self, tool='chart', master='baseapp', mrid=None, mrpass=None, folder_path=None):
        '''
        Desc: This function will invoke the designer for tools like report, chart, workbook.
        '''
        setup_url = self._utillity.get_setup_url()
        if folder_path != None:
            folder=folder_path
        else:
            project = self._utillity.parseinitfile('project_id')
            suite = self._utillity.parseinitfile('suite_id')
            group_id=self._utillity.parseinitfile('group_id')
            folder = project + '_' + suite + '/' + group_id
        api_url = setup_url.replace('home8206', '') + 'designer?tool=' + tool + '&master=' + master + '&item=IBFS%3A%2FWFC%2FRepository%2F' + folder
        self.driver.get(api_url)
        self._utillity.wf_login(mrid=mrid, mrpass=mrpass, add_home_path=False)
        
    def invoke_designer_using_api(self, master_file, mrid=None, mrpass=None, folder_path=None):
        '''
        Desc: This function will invoke the designer for tools like report, chart, workbook.
        '''
        setup_url = self._utillity.get_setup_url()
        if folder_path != None:
            folder=folder_path
        else:
            project = self._utillity.parseinitfile('project_id')
            suite = self._utillity.parseinitfile('suite_id')
            group_id=self._utillity.parseinitfile('group_id')
            folder = project + '_' + suite + '/' + group_id
        api_url = setup_url.replace('home8206', '') + 'designer?&master={0}&item=IBFS:/WFC/Repository/{1}&tool=framework&startlocation=IBFS:/WFC/Repository/{1}'.format(master_file, folder)
        self.driver.get(api_url)
        self._utillity.wf_login(mrid=mrid, mrpass=mrpass, add_home_path=False)
        self._utillity.wait_for_page_loads(60)
        
    def edit_fex_with_designer_using_api(self, fex_file, mrid=None, mrpass=None, folder_path=None):
        '''
        Desc: This function will invoke the designer for tools like report, chart, workbook.
        '''
        setup_url = self._utillity.get_setup_url()
        if folder_path != None:
            folder=folder_path
        else:
            project = self._utillity.parseinitfile('project_id')
            suite = self._utillity.parseinitfile('suite_id')
            group_id=self._utillity.parseinitfile('group_id')
            folder = project + '_' + suite + '/' + group_id
        api_url = setup_url.replace('home8206', '') + 'designer?&item=IBFS:/WFC/Repository/{0}/{1}.fex&startlocation=IBFS:/WFC/Repository/{0}'.format(folder, fex_file.lower())
        self.driver.get(api_url)
        self._utillity.wf_login(mrid=mrid, mrpass=mrpass, add_home_path=False)
    
    def edit_page_with_designer_using_api(self, page_name, mrid=None, mrpass=None, folder_path=None):
        '''
        Desc: This function will invoke the designer for tools like report, chart, workbook.
        '''
        setup_url = self._utillity.get_setup_url()
        if folder_path != None:
            folder=folder_path
        else:
            project = self._utillity.parseinitfile('project_id')
            suite = self._utillity.parseinitfile('suite_id')
            group_id=self._utillity.parseinitfile('group_id')
            folder = project + '_' + suite + '/' + group_id
        api_url = setup_url.replace('home8206', '') + 'designer?&item=IBFS:/WFC/Repository/{0}/{1}&startlocation=IBFS:/WFC/Repository/{0}'.format(folder, page_name.lower())
        self.driver.get(api_url)
        self._utillity.wf_login(mrid=mrid, mrpass=mrpass, add_home_path=False)
        self._utillity.wait_for_page_loads(60)
        
    def run_insight_procedure_tool_using_api(self, fex,mrid=None, mrpass=None, folder_path=None): 
        '''
        Desc: This function will run the designer_chart saved fex
        usage:run_designer_tool_using_api("C2345603")
        '''
        setup_url = self._utillity.get_setup_url()
        if folder_path != None:
            folder=folder_path
        else:
            project = self._utillity.parseinitfile('project_id')
            suite = self._utillity.parseinitfile('suite_id')
            user_id=self._utillity.parseinitfile('mrid')
            folder = project + '_' + suite + '/' + '~'+user_id
            api_url = setup_url.replace('home8206', '') + 'run.bip?BIP_REQUEST_TYPE=BIP_LAUNCH&BIP_folder' + '=IBFS%3A%2FWFC%2FRepository%2F' + folder  + '&BIP_item='+fex+'.fex'
        self.driver.get(api_url)
        self._utillity.wf_login(mrid=mrid, mrpass=mrpass, add_home_path=False)
        
    def invoke_ia_tool_using_api(self, tool='chart', master='baseapp', mrid=None, mrpass=None, folder_path=None):
        '''
        Desc: This function will invoke the designer for tools like report, chart, workbook.
        '''
        setup_url = self._utillity.get_setup_url()
        if folder_path != None:
            folder=folder_path
        else:
            project = self._utillity.parseinitfile('project_id')
            suite = self._utillity.parseinitfile('suite_id')
            group_id=self._utillity.parseinitfile('group_id')
            folder = project + '_' + suite + '/' + group_id
        api_url = setup_url.replace('home8206', '') + 'ia?tool=' + tool + '&master=' + master + '&item=IBFS%3A%2FWFC%2FRepository%2F' + folder
        self.driver.get(api_url)
        self._utillity.wf_login(mrid=mrid, mrpass=mrpass, add_home_path=False)
    
    def invoke_designer_tool_in_edit_mode_using_api(self, fex, tool='chart', mrid=None, mrpass=None, folder_path=None):
        '''
        Desc: This function will invoke the info assist tools in edit mode for a particular fex in tools like report, chart, visualization.
        '''
        setup_url = self._utillity.get_setup_url()
        if folder_path != None:
            folder=folder_path
        else:
            project = self._utillity.parseinitfile('project_id')
            suite = self._utillity.parseinitfile('suite_id')
            group_id=self._utillity.parseinitfile('group_id')
            folder = project + '_' + suite + '/' + group_id
        if tool == 'workbook':
            api_url = setup_url.replace('home8206', '') + 'designer?&item=IBFS%3A%2FWFC%2FRepository/'+folder+'/'+fex + '&tool='+tool
        else:
            api_url = setup_url.replace('home8206', '') + 'designer?&item=IBFS%3A%2FWFC%2FRepository/'+folder+'/'+fex+'.fex&tool='+tool
        self.driver.get(api_url)
        self._utillity.wf_login(mrid=mrid, mrpass=mrpass, add_home_path=False)
    
    def invoke_ia_tool_in_edit_mode_using_api(self, fex, tool='chart', mrid=None, mrpass=None, folder_path=None):
        '''
        Desc: This function will invoke the info assist tools in edit mode for a particular fex in tools like report, chart, visualization.
        '''
        setup_url = self._utillity.get_setup_url()
        if folder_path != None:
            folder=folder_path
        else:
            project = self._utillity.parseinitfile('project_id')
            suite = self._utillity.parseinitfile('suite_id')
            group_id=self._utillity.parseinitfile('group_id')
            folder = project + '_' + suite + '/' + group_id
        api_url = setup_url.replace('home8206', '')+ 'ia?&item=IBFS%3A%2FWFC%2FRepository/'+folder+'/'+fex+'.fex&tool='+tool
        self.driver.get(api_url)
        self._utillity.wf_login(mrid=mrid, mrpass=mrpass, add_home_path=False)

    def run_designer_tool_using_api(self, fex,mrid=None, mrpass=None, folder_path=None, tool=None): 
        '''
        Desc: This function will run the designer_chart saved fex
        usage:run_designer_tool_using_api("C2345603")
        '''
        fex = fex if tool == 'workbook' else fex + '.fex' 
        setup_url = self._utillity.get_setup_url()
        if folder_path != None:
            folder=folder_path
        else:
            project = self._utillity.parseinitfile('project_id')
            suite = self._utillity.parseinitfile('suite_id')
            group_id=self._utillity.parseinitfile('group_id')
            folder = project + '_' + suite + '/' + group_id
        api_url = setup_url.replace('home8206', '') + 'run.bip?BIP_REQUEST_TYPE=BIP_LAUNCH&BIP_folder=IBFS%253A%252FWFC%252FRepository%252F' + folder +'%252F&BIP_item='+ fex
        self.driver.get(api_url)
        self._utillity.wf_login(mrid=mrid, mrpass=mrpass, add_home_path=False)
    
    def double_click_dimension_field_in_metadata_tree(self, field_name):
        """
        Description : This method will double click on dimension field.
        :usage : double_click_on_dimension_field('CAR') or double_click_on_dimension_field('Product->Product->Product,Category')
        """
        field_name_list = field_name.split('->')
        element_to_click = self.get_recursive_tree_object(field_name_list, dc_locators.DIMENSIONS_TAB_CSS, 0)
        self._coreutility.double_click(element_to_click)
    
    def double_click_measures_field_in_metadata_tree(self, field_name):
        """
        Description : This method will double click on measures field.
        :usage : double_click_on_dimension_field('CAR')
        """
        field_name_list = field_name.split('->')
        element_to_click = self.get_recursive_tree_object(field_name_list, dc_locators.MEASURES_TAB_CSS, 0)
        self._coreutility.double_click(element_to_click)
        
    def verify_show_title_in_font_styling(self, msg, present=False, verify_text=None, parent_css=None):
        """
        Description: This method is used to verify the font title in format window
        :Usage :  verify_show_title_in_font_styling('Step 4.5: Verify ', present=True, parent_css='div[data-ibx-name='axisTitlePage']')
        """
        font_obj = Designer_FontStyle(self.driver, parent_css=parent_css, msg=msg)
        font_obj.verify_show_title(present=present, verify_text=verify_text)
        
    def right_click_on_datapane(self,field_name, context_menu_item_path, datapane='measures'):
        """
        Description : This method will right on measures field or dimensions field and select the item.
        :usage : double_click_on_dimension_field('Product->Product->Product,Category', 'Properties', datapane='dimensions')
        """
        field_name_list = field_name.split('->')
        if datapane == 'measures':
            element_to_click = self.get_recursive_tree_object(field_name_list, dc_locators.MEASURES_TAB_CSS, 0)
        elif datapane == 'dimensions':
            element_to_click = self.get_recursive_tree_object(field_name_list, dc_locators.DIMENSIONS_TAB_CSS, 0)
        self._coreutility.right_click(element_to_click)
        self._wf_main_page_obj.select_context_menu_item(context_menu_item_path, row_css="[role='menuitem']")
    
    def field_expand(self, parent_object):
        """
        Description : This function is used to expand the fields on the data pane
        usage : field_expand("css_object")
        """
        try:
            self._utillity.validate_and_get_webdriver_object(".tnode-btn-expanded", 'collapse-button', parent_object)
        except AttributeError:
            expand_button = self._utillity.validate_and_get_webdriver_object(".tnode-btn-collapsed", 'expand-button', parent_object)
            self._coreutility.left_click(expand_button)
    
    def get_recursive_tree_object(self, field_name_list, css, depth):
        """
        Description: This function is used to expand the tree and double click it
        usage : get_recursive_tree_object(field_name_list, dimension_tree_css, 0)
        """
        if depth > 0:
            field_obj = self._utillity.validate_and_get_webdriver_objects(css + " + [class*='tnode-children'] [class*='tnode-label']", 'dimension_tree_description')
        else:
            field_obj = self._utillity.validate_and_get_webdriver_objects(css , 'dimension_tree_description')
        field_text = self._j_script.get_elements_text(field_obj)
        for item in field_name_list:
            try:
                text_index = field_text.index(item) 
            except ValueError:
                raise Exception("Element {0} is not found".format(item))
            object_to_click = field_obj[text_index]
            self._j_script.scrollIntoView(object_to_click)
            if len(field_name_list) == 1:
                return object_to_click
            else:
                self.field_expand(object_to_click)
            field_name_list.pop(0)
            object_id = "#" + self._utillity.get_element_attribute(object_to_click, 'id')
            return self.get_recursive_tree_object(field_name_list, object_id, depth+1)
    
    def search_fields_in_search_box(self, text_to_enter=None, send_keys=None):
        """
        Descriptions : This method is used to enter string in the field search text box
        :Usage : search_fields_in_search_box('Product')
        """
        search_box_css = ".wfc-mdfp-search-box div[data-ibx-type=\"ibxTextField\"] input"
        if send_keys:
            text_box_obj = self._utillity.validate_and_get_webdriver_object(search_box_css, 'search_box')
            text_box_obj.send_keys(send_keys)
        else:    
            self._utillity.set_text_to_textbox_using_keybord(text_to_enter, search_box_css)
    
    def verify_dimensions_field_is_empty(self,msg):
        """
        Descriptions : This method is used to verify the dimensions is empty
        :Usage : verify_dimensions_field_is_empty("Step:4.1")
        """ 
        observed_field=self._utillity.validate_and_get_webdriver_object(dc_locators.DIMENSIONS_FIELD_AREA_CSS, 'dimensions-field_area')
        empty_list=observed_field.text.strip()
        self._utillity.asequal(empty_list,'',msg)
    
    def verify_measures_field_is_empty(self,msg):
        """
        Descriptions : This method is used to verify the measures is empty
        :Usage : verify_measures_field_is_empty("Step:4.1)
        """ 
        observed_field=self._utillity.validate_and_get_webdriver_object(dc_locators.MEASURES_FIELD_AREA_CSS, 'measures-field_area')
        empty_list=observed_field.text.strip()
        self._utillity.asequal(empty_list,'',msg)
    
    def verify_dimensions_fields(self, expected_field_list, msg,compare_type='asequal',field_path=None):
        """
        Descriptions : This method is used to verify the fields present under dimensions box
        :Usage : verify_dimensions_fields(['Product Name','ID Product' ], 'Step 4.1: Verify the dimensions_list')
        """
        if field_path!= None:
            field_name = field_path.split('->')
            self.get_recursive_tree_object(field_name, dc_locators.DIMENSIONS_VERIFY_CSS,0)   
        observed_list_objects = self._utillity.validate_and_get_webdriver_objects(dc_locators.DIMENSIONS_VERIFY_CSS, 'dimensions-fields')
        observed_list = [element.text for element in observed_list_objects]
        self._utillity.verify_list_values(expected_field_list, observed_list, msg, assert_type=compare_type)
    
    def verify_measures_fields(self, expected_field_list, msg, compare_type='asequal',field_path=None):
        """
        Descriptions : This method is used to verify the fields present under dimensions box
        :Usage : verify_dimensions_fields(['Product Name','ID Product' ], 'Step 4.1: Verify the dimensions_list')
        """
        if field_path!= None:
            field_name = field_path.split('->')
            self.get_recursive_tree_object(field_name, dc_locators.MEASURES_VERIFY_CSS,0)
        observed_list_objects = self._utillity.validate_and_get_webdriver_objects(dc_locators.MEASURES_VERIFY_CSS, 'measures-fields')
        observed_list = [element.text for element in observed_list_objects]
        self._utillity.verify_list_values(expected_field_list, observed_list, msg, assert_type=compare_type)
    
    def verify_variables_list(self, expected_field_list, msg, compare_type='asequal'):
        """
        Descriptions : This method is used to verify the values under the variables
        :Usage : verify_variables_list(['Product Name','ID Product' ], 'Step 4.1: Verify the dimensions_list')
        """
        observed_list_objects = self._utillity.validate_and_get_webdriver_objects(dc_locators.VARIABLES_CSS, 'variables-fields')
        observed_list = [element.text for element in observed_list_objects]
        self._utillity.verify_list_values(expected_field_list, observed_list, msg, assert_type=compare_type)
    
    def select_layout_options(self, item_name):
        """
        Description : This method select the layout option from the layout menu bar
        :Usage select_layout_options('Heading')
        """
        layout_button = self._utillity.validate_and_get_webdriver_object(dc_locators.LAYOUT_BUTTON_CSS, 'layout-button')
        self._coreutility.left_click(layout_button)
        self._wf_main_page_obj.select_context_menu_item(item_name)
    
    def select_more_options(self, item_path):
        """
        Description : This method select the more option from the tool bar
        :Usage select_more_options('Run with Insight')
        """
        layout_button = self._utillity.validate_and_get_webdriver_object(dc_locators.MORE_BUTTON_CSS, 'more-button')
        self._coreutility.left_click(layout_button)
        self._wf_main_page_obj.select_context_menu_item(item_path)
    
    def verify_layout_options_if_checked(self, checked_item_list, msg):
        """
        Description : This method verify if the layout option is checked
        :Usage verify_layout_options_if_checked(['Heading'])
        """
        layout_button = self._utillity.validate_and_get_webdriver_object(dc_locators.LAYOUT_BUTTON_CSS, 'layout-button')
        self._coreutility.left_click(layout_button)
        self._utillity.synchronize_with_number_of_element(dc_locators.POPUP_CSS, 1, self._wf_main_page_obj.chart_short_timesleep)
        drop_down_list = self._utillity.validate_and_get_webdriver_objects(dc_locators.POPUP_CSS + " div[role='menuitemcheckbox']", 'checked-list')
        drop_down_text_list = [element.text for element in drop_down_list]
        for element in checked_item_list:
            if element in drop_down_text_list:
                element_index = drop_down_text_list.index(element)
                if_checked = self._utillity.get_element_attribute(drop_down_list[element_index], 'aria-checked')
                self._utillity.asequal(if_checked,'true', msg + " {0} is checked".format(element))
            else:
                raise Exception("Element {0} is not present".format(element))
        self._coreutility.left_click(self._utillity.validate_and_get_webdriver_object(dc_locators.SPACER_CSS, 'spacer_css'))
    
    def verify_all_layout_options(self, item_list, msg, compare_type='asequal'):
        """
        Description : This method verify all the options in the layout dropdown
        :Usage verify_all_layout_options(['Heading','Footing'])
        """
        layout_button = self._utillity.validate_and_get_webdriver_object(dc_locators.LAYOUT_BUTTON_CSS, 'layout-button')
        self._coreutility.left_click(layout_button)
        self._wf_main_page_obj.verify_context_menu_item(item_list, msg, comparision_type=compare_type)
        self._coreutility.left_click(self._utillity.validate_and_get_webdriver_object(dc_locators.SPACER_CSS, 'spacer_css'))
        
    def select_field_or_variables(self, select_option):
        """
        Description: This method is used to select variables or fields in data pane
        :Usage : select_field_or_variables('Fields')
        """
        tab_buttons = self._utillity.validate_and_get_webdriver_objects(".wfc-mdfp-tab-pane div[data-ibx-type='ibxTabButton']", 'tab-buttons')
        for element in tab_buttons:
            if element.text.strip() == select_option:
                self._coreutility.left_click(element)
    
    def select_query_or_format_tab_option(self, select_option='format'):
        """
        Description: This method is used to select query or format tab
        :Usage : select_query_or_format_tab(select_option='format')
        """
        tab_buttons = self._utillity.validate_and_get_webdriver_objects(".ibx-csl-items-container .ibxtool-pane-tab", 'tab-buttons')
        tab_to_click = tab_buttons[1] if select_option=='format' else tab_buttons[0]
        self._coreutility.left_click(tab_to_click)
    
    def click_heading_or_footing(self, layout_option, return_item=False, text_to_enter=None):
        """
        Description: This method is used to click on a header or footer
        :Usage : click_heading_or_footing('Heading', return_item=False)
        """
        if layout_option == 'Heading':
            click_object = self._utillity.validate_and_get_webdriver_object(dc_locators.HEADING_CSS, 'heading')
        elif layout_option == 'Footing':
            click_object = self._utillity.validate_and_get_webdriver_object(dc_locators.FOOTING_CSS, 'footing')
        else:
            raise Exception("Layout option is not given")
        self._coreutility.left_click(click_object)
        if text_to_enter:
            pyautogui.typewrite(text_to_enter)
        if return_item:
            return click_object
    
    def get_heading_or_footing_text_list(self, layout_option='Heading'):
        """
        Description: This method is used to get a header or footer text
        :Usage : get_heading_or_footing_text_list(layout_option='Heading')
        """
        if layout_option == 'Heading':
            verify_object_frame = dc_locators.HEADING_CSS + " iframe"
        elif layout_option == 'Footing':
            verify_object_frame = dc_locators.FOOTING_CSS + " iframe"
        else:
            raise Exception("Layout option is not given")
        self._coreutility.switch_to_frame(verify_object_frame)
        try:
            frame_text = self._utillity.validate_and_get_webdriver_object('body', 'frame-text').text.strip()
            frame_text_list = re.split('\s+', frame_text)
        finally:
            self._coreutility.switch_to_default_content()
        return frame_text_list if not None else None
    
    def verify_date_time_in_header_or_footer(self, index, msg, date_time_format='mm/dd/yy', layout_option='Heading'):
        """
        Description: This method is used to verify the date and time in header or footer
        :Usage : verify_date_time_in_header_or_footer(0, 'Step 4.1: Verify the mm/dd/yy', date_time_format='mm/dd/yy', layout_option='Heading')
        """
        observed_list = self.get_heading_or_footing_text_list(layout_option=layout_option)
        if date_time_format == 'mm/dd/yy':
            status = bool(re.match('^(0[1-9]|1[012])/(0[1-9]|[12][0-9]|3[01])/\d\d$', observed_list[index]))
        elif date_time_format == 'mm/dd/yyyy':
            status = bool(re.match('^(0[1-9]|1[012])/(0[1-9]|[12][0-9]|3[01])/(19|20)\d\d$', observed_list[index]))
        elif date_time_format == 'yyyy/mm/dd':
            status = bool(re.match('^(19|20)\d\d/(0[1-9]|1[012])/(0[1-9]|[12][0-9]|3[01])$', observed_list[index]))
        elif date_time_format == 'dd/mm/yyyy':
            status = bool(re.match('^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[012])/(19|20)\d\d$', observed_list[index]))
        elif date_time_format == 'hh.mm.ss':
            status = bool(re.match('^(0[0-9]|1[0-9]|2[0-3]).([0-5][0-9]).([0-5][0-9])', observed_list[index]))
        else:
            raise Exception("The {0} format entered is not supported".format(date_time_format))
        self._utillity.asequal(True, status, msg)
        
    def verify_text_in_heading_or_footing(self, expected_text_list, msg, layout_option='Heading', compare_type=None):
        """
        Description: This method is used to verify the text in heading or footing
        :Usage : verify_text_in_heading_or_footing(['Product'], 'Step 4.1: Verify', layout_option='Heading')
        """
        observed_list = self.get_heading_or_footing_text_list(layout_option=layout_option)
        self._utillity.verify_list_values(expected_text_list, observed_list, msg, assert_type=compare_type)
    
    def close_heading_or_footing_toolbar(self):
        """
        Description: This method is used to close the heading or footinf toolbae
        :Usage : close_heading_or_footing_toolbar()
        """
        close_button_css = dc_locators.HEADING_FOOTING_TOOLBAR_CSS +  " [class*='close-button']" 
        close_button_obj = self._utillity.validate_and_get_webdriver_object(close_button_css, 'close')
        self._coreutility.left_click(close_button_obj)
    
    def click_an_item_on_toolbar(self, item_name):
        """
        Description : This function is used to click an item from the toolbar
        :Usage : select_an_item_from_toolbar('preview')
        """
        button_css = item_name.upper() + '_BUTTON_CSS'
        clickable_obj =  self._utillity.validate_and_get_webdriver_object(dc_locators.__dict__[button_css], 'button_locator')
        self._coreutility.left_click(clickable_obj)
        
    def verify_toolbar_item_enabled_or_disabled(self, item_name, msg, disabled='true'):
        """
        Description : This function is used to verify if the button is enabled or disabled
        :Usage : verify_toolbar_item_enabled_disabled('preview, enabled=True, "Step 4.5: Verify preview button is enabled")
        """
        button_css = item_name.upper() + '_BUTTON_CSS'
        verify_obj =  self._utillity.validate_and_get_webdriver_object(dc_locators.__dict__[button_css], 'button_locator')
        obj_disabled_attributed = self._utillity.get_element_attribute(verify_obj, 'aria-disabled')
        self._utillity.asequal(disabled, obj_disabled_attributed, msg)
    
    def get_filter_shelf_fields(self):
        """
        Description: This function will get filter shelf
        :Usage : get_filter_shelf_fields([['Product Category','All'],['Product','All']], "Step 4.5: Verify the filter shelf)
        """
        filter_shelf_css = ".filter-bar .filter-pill .filter-field-box"
        filter_shelfs = self._utillity.validate_and_get_webdriver_objects(filter_shelf_css, 'filter_shelf_css')
        return filter_shelfs
    
    def verify_filter_shelf_field_right_click_options(self, field_name, expected_list, step_num, context_menu_path=None, comparision_type='asequal', xoffset=0, yoffset=0):
        """
        Description: This function verify right click option from field from filter shelf.
        :Usage : select_filter_shelf_field('Product Category', ['Edit'], '09.00', context_menu_path='Single')
        """
        filter_shelfs = self.get_filter_shelf_fields()
        shelf_field = None
        for shelf in filter_shelfs:
            if shelf.text.strip() == field_name:
                shelf_field = shelf
                break
        if shelf_field == None:
            raise LookupError(field_name + ' not able to found in filter shelf.')
        self._coreutility.python_right_click(shelf_field, xoffset=xoffset, yoffset=yoffset)
        if context_menu_path:
            self._wf_main_page_obj.select_context_menu_item(context_menu_path)
        self._wf_main_page_obj.verify_context_menu_item(expected_list, msg='Step '+str(step_num), comparision_type=comparision_type)
    
    def select_filter_shelf_field(self, field_name, context_menu_path=None, left_click=False, right_click=False, xoffset=0, yoffset=0):
        """
        Description: This function select field from filter shelf.
        :Usage : select_filter_shelf_field('Product Category', context_menu_path='Single', right_click=True, yoffset=-5)
        """
        filter_shelfs = self.get_filter_shelf_fields()
        shelf_field = None
        for shelf in filter_shelfs:
            if shelf.text.strip() == field_name:
                shelf_field = shelf
                break
        if shelf_field == None:
            raise LookupError(field_name + ' not able to found in filter shelf.')
        if left_click:
            self._coreutility.python_left_click(shelf_field, xoffset=xoffset, yoffset=yoffset)
        if right_click:
            self._coreutility.python_right_click(shelf_field, xoffset=xoffset, yoffset=yoffset)
            self._wf_main_page_obj.select_context_menu_item(context_menu_path)
    
    def verify_filter_shelf_fields_and_values(self, list_values, msg):
        """
        Description: This function is used to verify the values in filter shelf
        :Usage : verify_filter_shelf_fields_and_values([['Product Category','All'],['Product','All']], "Step 4.5: Verify the filter shelf)
        """
        observed_field_dict = {}
        if len(list_values) < 1:
            try:
                self._utillity.validate_and_get_webdriver_objects(dc_locators.FILTER_BAR + " .filter-pill", 'filter-pills')
            except AttributeError:
                self._utillity.asequal(True, True, msg)
                return
        filter_pills = self._utillity.validate_and_get_webdriver_objects(dc_locators.FILTER_BAR + " .filter-pill", 'filter-pills')
        for obj in filter_pills:
            filter_field = self._utillity.validate_and_get_webdriver_object(" .new-filter-field", 'filter-field', obj)
            filter_value = self._utillity.validate_and_get_webdriver_object(" .new-filter-value", 'filter-value', obj)
            observed_field_dict[filter_field.text.strip()] = filter_value.text.strip()
        for element in list_values:
            try:
                if observed_field_dict[element[0]] == element[1]:
                    status = True
            except KeyError:
                status = False
                raise Exception("{0} and {1} are not found in the given dictionary".format(element[0], element[1]))
        self._utillity.asequal(status, True, msg)
    
    def get_slider_in_filter_shelf_fields_popup(self, label_value=None, min_label_value=None, max_label_value=None, marker_value=None):
        """
        Description: This function will get slider in filter shelf field popup.
        :Usage : get_slider_in_filter_shelf_fields_popup(label_value='17', min_label_value='17', max_label_value='19', marker_value=19)
        """
        if label_value:
            locator_css = ".pop-top .ibx-slider-label-value .ibx-label-text"
            locator_name = 'slider label text'
        if min_label_value:
            locator_css = ".pop-top .ibx-slider-label-min .ibx-label-text"
            locator_name = 'slider minimum label text'
        if max_label_value:
            locator_css = ".pop-top .ibx-slider-label-max .ibx-label-text"
            locator_name = 'slider maximum label text'
        if marker_value:
            locator_css = ".pop-top .ibx-slider-marker[aria-valuenow='{0}']".format(str(marker_value).strip().replace(',',''))
            locator_name = 'slider marker'
        try:
            self._utillity.synchronize_until_element_is_visible(".pop-top .filter-slider-marker", self.home_page_medium_timesleep)
            slider_elem = self._utillity.validate_and_get_webdriver_object(locator_css, locator_name)
            return slider_elem
        except Exception as e:
            raise Exception(e)
    
    def verify_slider_in_filter_shelf_fields_popup(self, msg, label_value=None, min_label_value=None, max_label_value=None, marker_value=None, marker_status=None):
        """
        Description: This function will verify slider in filter shelf field popup.
        :Usage : verify_slider_in_filter_shelf_fields_popup('Step 09.00: Verify', label_value='17', min_label_value='17', max_label_value='19', marker_value=19, marker_status='Visible')
        """
        slider_obj = self.get_slider_in_filter_shelf_fields_popup(label_value=label_value, min_label_value=min_label_value, max_label_value=max_label_value, marker_value=marker_value)
        if marker_value:
            if marker_status not in ['Visible', 'Not visible']:
                raise LookupError("Please pass marker status from 'Visible', 'Not visible'")
            status_ = 'visible' if slider_obj.is_displayed() else 'not visible'
            marker_status_ = 'visible' if marker_status.lower() == 'visible' else 'not visible'
            self._utillity.asequal(marker_status_, status_, msg)
        elif label_value:
            expected_data = label_value
            actual_data = slider_obj.text.strip()
        elif min_label_value:
            expected_data = min_label_value
            actual_data = slider_obj.text.strip()
        elif max_label_value:
            expected_data = max_label_value
            actual_data = slider_obj.text.strip()        
        self._utillity.asequal(expected_data, actual_data, msg)
            
    def select_slider_marker_in_filter_shelf_fields_popup(self, marker_value=None, move_left_times=None, move_right_times=None, pause_=0.5):
        """
        Description: This function will select slider in filter shelf field popup.
        :Usage : select_slider_marker_in_filter_shelf_fields_popup(marker_value=19, move_left_times=19, move_right_times=19)
        """
        slider_obj = self.get_slider_in_filter_shelf_fields_popup(marker_value=marker_value)
        self._coreutility.left_click(slider_obj)
        if move_left_times:
            for _ in range(int(move_left_times)):
                if sys.platform == 'linux':
                    pykeyboard.press_key(pykeyboard.left_key)
                    pykeyboard.release_key(pykeyboard.left_key)
                else:
                    keyboard.send('left')
                time.sleep(pause_)
        if move_right_times:
            for _ in range(int(move_right_times)):
                if sys.platform == 'linux':
                    pykeyboard.press_key(pykeyboard.right_key)
                    pykeyboard.release_key(pykeyboard.right_key)
                else:
                    keyboard.send('right')
                time.sleep(pause_)
                
    def get_filter_shelf_fields_popup(self, property_type, text_box_title=None, button_name=None, checkbox_name=None, radio_name=None):
        """
        Description: This function will get filter shelf field popup.
        :Usage : get_filter_shelf_fields_popup('text_box', text_box_title='EMEA')
        """
        element_name = "filter shelf fields popup {0}"
        if property_type == 'text_box':
            text_box = ".pop-top [data-ibx-type='ibxTextField'][title='{0}']".format(text_box_title)
            text_box_elem = self._utillity.validate_and_get_webdriver_object(text_box, element_name.format('text_box'))
            return text_box_elem
        elif property_type == 'button':
            button = ".pop-top [role='button'][data-ibx-type='ibxButton']"
            button_elem = self._utillity.validate_and_get_webdriver_objects(button, element_name.format('button'))
            for button_obj in button_elem:
                if button_obj.text.strip() == button_name:
                    return button_obj
        elif property_type == 'checkbox':
            checkbox = ".pop-top [role='checkbox'][data-ibx-type='ibxSelectCheckItem']"
            checkbox_elem = self._utillity.validate_and_get_webdriver_objects(checkbox, element_name.format('checkbox'))
            for checkbox_obj in checkbox_elem:
                if checkbox_obj.text.strip() == checkbox_name:
                    return checkbox_obj
        elif property_type == 'radio':
            radio = ".pop-top [role='radio'][data-ibx-type='ibxRadioButton']"
            radio_elem = self._utillity.validate_and_get_webdriver_objects(radio, element_name.format('radio'))
            for radio_obj in radio_elem:
                if radio_obj.text.strip() == radio_name:
                    return radio_obj
    
    def verify_filter_shelf_fields_popup(self, property_type, msg, text_box_title=None, text_box_input=None, button_name=None, checkbox_name=None, radio_name=None, assert_type='asequal'):
        """
        Description: This function will verify filter shelf field popup.
        :Usage : verify_filter_shelf_fields_popup('text_box', 'Step 09.00: Verify', text_box_title='EMEA')
        """
        filter_shelf = self.get_filter_shelf_fields_popup(property_type, text_box_title=text_box_title, button_name=button_name, checkbox_name=checkbox_name, radio_name=radio_name)
        if property_type == 'text_box':
            input_box_elem = self._utillity.validate_and_get_webdriver_object('input', '{0} text_box'.format(text_box_title), parent_object=filter_shelf)
            actual_data = [self._utillity.get_element_attribute(input_box_elem, 'value')]
            expected_data = [text_box_input]
        elif property_type == 'button':
            actual_data = [self._utillity.validate_and_get_webdriver_object('.ibx-label-text', '{0} button'.format(button_name), parent_object=filter_shelf).text.strip()]
            expected_data = [button_name]
        elif property_type == 'checkbox':
            actual_data = [self._utillity.validate_and_get_webdriver_object('.ibx-label-text', '{0} checkbox'.format(checkbox_name), parent_object=filter_shelf).text.strip()]
            expected_data = [checkbox_name]
        elif property_type == 'radio':
            actual_data = [self._utillity.validate_and_get_webdriver_object('.ibx-label-text', '{0} radio'.format(radio_name), parent_object=filter_shelf).text.strip()]
            expected_data = [radio_name]
        self._utillity.verify_list_values(expected_data, actual_data, msg, assert_type=assert_type)
            
    def select_filter_shelf_fields_popup(self, property_type, text_box_title=None, text_box_input=None, button_name=None, checkbox_name=None, radio_name=None):
        """
        Description: This function will select filter shelf field popup.
        :Usage : select_filter_shelf_fields_popup('text_box', text_box_title='EMEA')
        """
        filter_shelf = self.get_filter_shelf_fields_popup(property_type, text_box_title=text_box_title, button_name=button_name, checkbox_name=checkbox_name, radio_name=radio_name)
        if property_type == 'text_box':
            self._utillity.set_text_to_textbox_using_keybord(text_box_input, text_box_elem=filter_shelf)
            self._utillity.synchronize_with_visble_text('input', text_box_input, self.home_page_long_timesleep, text_option='text_value')
        else:
            self._coreutility.left_click(filter_shelf)
    
    def change_format_access(self, access_item):
        """
        Description: This method is used to change the access item 
        :Usage : change_format_access('Axis')
        """
        drop_down_button = self._utillity.validate_and_get_webdriver_object(dc_locators.FORMAT_QUICK_ACCESS + " .ibx-select-open-btn", 'drop-down')
        self._coreutility.left_click(drop_down_button)
        self._wf_main_page_obj.select_context_menu_item(access_item, pop_up_css="div[class*='pop-top']", row_css=".wfc-chart-sel-item:not([style*='none'])")
        
    def change_theme_in_format_tab(self, theme_name, msg=None, verify_list=None, compare_type='asequal'):
        """
        Description: This method is used to change the theme name 
        :Usage : change_theme_in_format_tab('Light')
        """
        drop_down_button = self._utillity.validate_and_get_webdriver_object(dc_locators.CHART_THEME_INPUT + " .ibx-select-open-btn", 'drop-down')
        self._coreutility.left_click(drop_down_button)
        if verify_list:
            popup_css = "div[class*='pop-top'] .wfc-chart-sel-item:not([style*='none'])"
            self._utillity.synchronize_with_number_of_element("div[class*='pop-top']", 1, 20)
            observed_obj = self._utillity.validate_and_get_webdriver_objects(popup_css, 'theme-list')
            observed_list = [element.text for element in observed_obj]
            self._utillity.verify_list_values(verify_list, observed_list , msg, assert_type=compare_type)
        self._wf_main_page_obj.select_context_menu_item(theme_name, pop_up_css="div[class*='pop-top']", row_css=".wfc-chart-sel-item:not([style*='none'])")
    
    def change_marker_option(self, option_name):
        """
        Description: This function is used to select marker option from its drop down under quick->Marker.
        usage : change_marker_option('Tick')
        """
        marker_dropdown_elem = self._utillity.validate_and_get_webdriver_object(dc_locators.QUICK_MARKER_MARKER,"marker_dropdown_series")
        self._coreutility.left_click(marker_dropdown_elem)
        self._utillity.synchronize_until_element_is_visible(".pop-top",self._wf_main_page_obj.chart_long_timesleep)
        self._wf_main_page_obj.select_context_menu_item(option_name, pop_up_css="div[class*='pop-top']", row_css=" div[class*='button-simple ibx-select-item']")
    
    def create_data_set(self, file_name):
        """
        Description: Read the all data from Data grid and write data in workbook
        :Usage : create_data_set('test')
        """
        try:
            workbook = Workbook()
            sheet = workbook.get_sheet_by_name('Sheet')
            file_path = os.path.join(os.getcwd(), "data", file_name + ".xlsx" )
            if os.path.exists(file_path) :
                os.remove(file_path)
            dimensions_header_row = self._utillity.validate_and_get_webdriver_objects(dc_locators.DATA_GRID_ROW_HEADER1, "DATA GRID DIMENSIONS HEADER")
            measures_header_row = self._utillity.validate_and_get_webdriver_objects(dc_locators.DATA_GRID_ROW_HEADER2, "DATA GRID MEASURES HEADER")
            rows = dimensions_header_row + measures_header_row
            rows_data_ = [row for row in rows if row.is_displayed()]
            colmun_num = 1
            row_num = 1
            for row_data_ in rows_data_:
                sheet.cell(row=row_num, column=colmun_num).value = row_data_.text.strip().replace("\n", "")
                colmun_num += 1
            dimensions_cell_row = self._utillity.validate_and_get_webdriver_objects(dc_locators.DATA_GRID_ROW_CELL1, "DATA GRID DIMENSIONS CELL")
            cells = [row for row in dimensions_cell_row if row.is_displayed()]
            colmun_num = 1
            row_num = 2
            start_index = cells[0].get_attribute('y')
            for index_, row_data in enumerate(cells, 1):
                if row_data.get_attribute('y') == start_index and index_ != 1:
                    row_num = 2
                    colmun_num += 1
                    sheet.cell(row=row_num, column=colmun_num).value = row_data.text.strip().replace("\n", "")
                    row_num += 1
                    continue
                else:
                    sheet.cell(row=row_num, column=colmun_num).value = row_data.text.strip().replace("\n", "")
                    row_num += 1
            measures_cell_row = self._utillity.validate_and_get_webdriver_objects(dc_locators.DATA_GRID_ROW_CELL2, "DATA GRID MEASURES CELL")
            cells = [row for row in measures_cell_row if row.is_displayed()]
            colmun_num += 1
            column_flag = colmun_num
            row_num = 2
            start_index = cells[0].get_attribute('y')
            for row_data in cells:
                cells_data = self._utillity.validate_and_get_webdriver_objects('text', "DATA GRID MEASURES CELL", parent_object=row_data)
                cells_data_ = [cell for cell in cells_data if cell.is_displayed()]
                for data in cells_data_:
                    sheet.cell(row=row_num, column=colmun_num).value = data.text.strip().replace("\n", "")
                    colmun_num += 1
                colmun_num = column_flag
                row_num +=1
            workbook.save(file_path)
        except Exception as e:
            raise Exception(e)
        print('done')
    
    def __compare_data_set(self, file_name):
        """
        Description: Read the all data from Data grid and compare with exists data
        :Usage : __compare_data_set('test')
        """
        try:
            diff = []
            file_path = os.path.join(os.getcwd(), "data", file_name + ".xlsx" )
            workbook = load_workbook(file_path)
            sheet = workbook.get_sheet_by_name('Sheet')
            if os.path.exists(file_path) != True :
                error_msg = "[{0}] data set not exists.".format(file_path)
                raise FileNotFoundError(error_msg)
            dimensions_header_row = self._utillity.validate_and_get_webdriver_objects(dc_locators.DATA_GRID_ROW_HEADER1, "DATA GRID DIMENSIONS HEADER")
            measures_header_row = self._utillity.validate_and_get_webdriver_objects(dc_locators.DATA_GRID_ROW_HEADER2, "DATA GRID MEASURES HEADER")
            rows = dimensions_header_row + measures_header_row
            rows_data_ = [row for row in rows if row.is_displayed()]
            colmun_num = 1
            row_num = 1
            for row_data_ in rows_data_:
                if str(sheet.cell(row=row_num, column=colmun_num).value) == row_data_.text.strip().replace("\n", ""):
                    diff=[0]
                else:
                    diff=[row_num, colmun_num, 'expected : ', str(sheet.cell(row=row_num, column=colmun_num).value), 'actual : ', row_data_.text.strip().replace("\n", "")]
                    return diff
                colmun_num += 1
            dimensions_cell_row = self._utillity.validate_and_get_webdriver_objects(dc_locators.DATA_GRID_ROW_CELL1, "DATA GRID DIMENSIONS CELL")
            cells = [row for row in dimensions_cell_row if row.is_displayed()]
            colmun_num = 1
            row_num = 2
            start_index = cells[0].get_attribute('y')
            for index_, row_data in enumerate(cells, 1):
                if row_data.get_attribute('y') == start_index and index_ != 1:
                    row_num = 2
                    colmun_num += 1
                    if str(sheet.cell(row=row_num, column=colmun_num).value) == row_data.text.strip().replace("\n", ""):
                        diff=[0]
                    else:
                        diff=[row_num, colmun_num, 'expected : ', str(sheet.cell(row=row_num, column=colmun_num).value), 'actual : ', row_data.text.strip().replace("\n", "")]
                        return diff
                    row_num += 1
                    continue
                else:
                    if str(sheet.cell(row=row_num, column=colmun_num).value) == row_data.text.strip().replace("\n", ""):
                        diff=[0]
                    else:
                        diff=[row_num, colmun_num, 'expected : ', str(sheet.cell(row=row_num, column=colmun_num).value), 'actual : ', row_data.text.strip().replace("\n", "")]
                        return diff
                    row_num += 1
            measures_cell_row = self._utillity.validate_and_get_webdriver_objects(dc_locators.DATA_GRID_ROW_CELL2, "DATA GRID MEASURES CELL")
            cells = [row for row in measures_cell_row if row.is_displayed()]
            colmun_num += 1
            column_flag = colmun_num
            row_num = 2
            start_index = cells[0].get_attribute('y')
            for row_data in cells:
                cells_data = self._utillity.validate_and_get_webdriver_objects('text', "DATA GRID MEASURES CELL", parent_object=row_data)
                cells_data_ = [cell for cell in cells_data if cell.is_displayed()]
                for data in cells_data_:
                    if str(sheet.cell(row=row_num, column=colmun_num).value) == data.text.strip().replace("\n", ""):
                        diff=[0]
                    else:
                        diff=[row_num, colmun_num, 'expected : ', sheet.cell(row=row_num, column=colmun_num).value, 'actual : ', data.text.strip().replace("\n", "")]
                        return diff
                    colmun_num += 1
                colmun_num = column_flag
                row_num +=1
            return diff
        except Exception as e:
            raise Exception(e)

    def verify_data_grid_set(self, file_name, step_num):
        """
        Description: Read the all data from Data grid and compare with exists data
        :Usage : verify_data_grid_set('test', '09.00')
        """
        diff = self.__compare_data_set(file_name)
        msg = "Step {0} : Verify grid data set - {1}".format(step_num, diff)
        self._utillity.asequal(1, len(diff), msg)
    
    def verify_data_grid_tooltip(self, parent_css, cell_data, expected_tooltip_list, msg, row_index=None, col_index=None):
        """
        Description: This method only verify 'Measures' fields tool-tip.
        :Usage : verify_data_grid_tooltip('#jschart_HOLD_0', '$49,598,845.24', 'Step 09.00: Verify tool-tip value.')
        """
        field_row_css = "{0} .tablePanel .innerTable [class*='row']".format(str(parent_css))
        cell_row_objs = self._utillity.validate_and_get_webdriver_objects(field_row_css, 'Grid row cell')
        status=False
        for row_index_, rows_data in enumerate(cell_row_objs, 1):
            if str(cell_data) in rows_data.text.strip():
                if row_index != None:
                    if row_index_ == int(row_index):
                        status = True
                        break
                else:
                    status = True
                    break 
        if status:
            cell_data_obj = self._utillity.validate_and_get_webdriver_objects('text', str(cell_data) + 'text', parent_object=rows_data)
            for col_index_, text_data in enumerate(cell_data_obj, 1):
                if str(cell_data) in text_data.text.strip():
                    if col_index != None:
                        if col_index_ == int(col_index):
                            break 
                    else:
                        break
            self._vislization_result_area.move_mouse_to_chart_component(text_data, move_to_tooltip=False)
            self._vislization_result_area.verify_tooltip(expected_tooltip_list, msg=msg)
        else:
            raise LookupError("'{}' this text does not got in grid.".format(str(cell_data)))
        
    def verify_default_theme_in_format_tab(self,theme_name,msg):
        """
        Description: This method is used to verify the default theme selected under format tab
        :Usage : verify_default_theme_in_format_tab('Designer 2018',"verify the default theme selected")
        """
        theme_text_box_elem =self._utillity.validate_and_get_webdriver_object(dc_locators.CHART_THEME_INPUT + " input", 'theme_textbox')
        default_value=self._utillity.get_element_attribute(theme_text_box_elem,'aria-label')
        self._utillity.asequal(default_value.strip(),theme_name,msg)
        
    def change_format_labels(self, item_name, type_to_change):
        """
        Description: This method is used to change the labels
        :Usage : change_format_labels('Rotation', 'Normal')
        """
        if item_name =='Rotation':
            button_css = item_name.upper() + '_CSS'
        drop_down_button = self._utillity.validate_and_get_webdriver_object(dc_locators.__dict__[button_css] + " .ibx-select-open-btn", 'drop-down')
        self._coreutility.left_click(drop_down_button)
        self._wf_main_page_obj.select_context_menu_item(type_to_change, pop_up_css="div[class*='pop-top']", row_css=".wfc-chart-sel-item:not([style*='none'])")

    def select_new_calculation(self, field='dimensions'):
        """
        Description: This method is used to select the new dimensions
        :Usage:  select_new_calculation(field='dimensions')
        """
        menu_button_css = field.upper() + '_MENU_BUTTON'
        menu_button = self._utillity.validate_and_get_webdriver_object(dc_locators.__dict__[menu_button_css], 'menu-button')
        self._coreutility.left_click(menu_button)
        self._wf_main_page_obj.select_context_menu_item('New calculation...', pop_up_css="div[class*='pop-top']", row_css="div[data-ibx-type='ibxMenuItem']")
        
    def click_calculation_operator_button(self, operator):
        """
        Description: This method is used to select the operator in new calculation
        :Usage: click_calculation_operator_button('x')
        :Params : use * for multiplication and / for division (for others use the same as given in the new calculation table)
        """
        self._utillity.synchronize_with_number_of_element('.wfc-calculator-operators-pane', 1, 25)
        relational_dict = {'=':'EQ', '!=':'NE', '>':'GT', '<':'LT', '>=': 'GE', '<=': 'LE' }
        if operator in ['=', '!=', '<', '>', '<=', '>=']:
            operator = relational_dict[operator]
        operator_button = self._utillity.validate_and_get_webdriver_object(".wfc-calculator-operators-pane div[data-ibxp-user-value='{0}']".format(operator.upper()), 'operator-button')
        self._coreutility.left_click(operator_button)
    
    def double_click_on_calculation_dimensions(self, field_path):
        """
        Description : This method will double click on the calculations
        :usage : double_click_on_calculation_dimensions('COUNTRY')
        """
        field_name_list = field_path.split('->')
        element_to_click = self.get_recursive_tree_object(field_name_list, dc_locators.NEW_CALCULATIONS_DIMENSIONS, 0)
        self._coreutility.double_click(element_to_click)
    
    def verify_calculation_dialog(self, text_to_verify, msg, compare_type='asequal'):
        """
        Description : This method will verify the text box in the calculations
        :usage : verify_calculation_dialog('CAR.ORIGIN.COUNTRY', 'Step 5: Verify the text box')
        """
        self._coreutility.switch_to_frame(dc_locators.CALCULTION_FRAME)
        calculation_text_box = self._utillity.validate_and_get_webdriver_object('body', 'dialog-body').text.strip()
        if compare_type == 'asequal':
            self._utillity.asequal(text_to_verify, calculation_text_box, msg)
        elif compare_type =='asin':
            self._utillity.asin(text_to_verify, calculation_text_box, msg)
        elif compare_type =='asnotin':
            self._utillity.as_notin(text_to_verify, calculation_text_box, msg)
        self._coreutility.switch_to_default_content()
    
    def edit_calculation_dialog_editor(self, value_to_change):
        """
        Description : This method will change the value inside textbox on the calculations
        :usage : edit_calculation_dialog_editor('DISCOUNT_US *250')
        """
        self._utillity.synchronize_until_element_is_visible(dc_locators.CALCULTION_FRAME, self.home_page_long_timesleep)
        try:
            self._coreutility.switch_to_frame(dc_locators.CALCULTION_FRAME)
            calculation_editor_box = self._utillity.validate_and_get_webdriver_object('body', 'CALCULTION_EDITOR body')
            self._utillity.set_text_to_textbox_using_keybord(value_to_change, text_box_elem=calculation_editor_box)
            self._utillity.synchronize_with_visble_text('body', value_to_change, self.home_page_long_timesleep)
            self._coreutility.switch_to_default_content()
        except (AttributeError, TimeoutError, NoSuchElementException, ValueError, StaleElementReferenceException):
            self._coreutility.switch_to_default_content()
            
    def edit_calculation_dialog_name(self, name_to_change):
        """
        Description: This function is used to change the nae of new calculation
        :Usage : edit_calculation_dialog_name('Profit')
        """
        calculation_dialog_name = self._utillity.validate_and_get_webdriver_object(".wfc-expression-name input[role='textbox']", 'CALCULTION_DIALOG_NAME')
        self._utillity.set_text_to_textbox_using_keybord(name_to_change, text_box_elem=calculation_dialog_name)
    
    def click_edit_format_on_calculation_dialog(self):
        '''
        Description: This function is used click edit format button and perform some operation.
        :Usage : click_edit_format_on_calculation_dialog()
        '''
        self._utillity.synchronize_until_element_is_visible(".pop-top [title='Edit format'] .ibx-label-icon", self.home_page_long_timesleep)
        edit_format = self._utillity.validate_and_get_webdriver_object(".pop-top [title='Edit format'] .ibx-label-icon", 'Edit format')
        self._coreutility.left_click(edit_format)
    
    def edit_bin_values(self, bin_width_text=None, show_as=None, click_type='OK'):
        """
        Description: This method is used to edit the bin values
        :Usage : edit_bin_values(bin_width_text=1000, show_as=Value)
        """
        if bin_width_text:
            self._utillity.set_text_to_textbox_using_keybord(bin_width_text, ".wfc-binwidth-textfield input")
        if show_as:
            show_as_elements = self._utillity.validate_and_get_webdriver_objects('.wfc-binvalues-button', 'show-as-buttons')
            show_index = 0 if show_as == 'Value' else 1
            self._coreutility.left_click(show_as_elements[show_index])
        if click_type:
            button_element = self._utillity.validate_and_get_webdriver_object(".bin-values-dlg-menu-item[aria-hidden='false'] .dmi-btn-{0}".format(click_type.lower()), 'button')
            self._coreutility.left_click(button_element)
    
    def verify_element_added_in_datapane(self, msg, field_path=None, verify_value=None, datapane='dimensions'):
        """
        Description: This methos is used to verify the value added in teh data pane
        :Usage : verify_element_added_in_datapane('Step 4.1', field_path='CAR->COMP', verify_value='COMP')
        """
        field_name_list = field_path.split('->')
        if datapane == 'measures':
            element_to_verify = self.get_recursive_tree_object(field_name_list, dc_locators.MEASURES_TAB_CSS, 0)
        elif datapane == 'dimensions':
            element_to_verify = self.get_recursive_tree_object(field_name_list, dc_locators.DIMENSIONS_TAB_CSS, 0)
        self._utillity.asequal(element_to_verify.text.strip(), verify_value, msg)
        
    def edit_font_options(self, **kwargs):
        """
        Description: This method is used to change the labels font options
        :Usage : change_label_font_options(font_name = 'TIMES', format_style='bold', size ='12', unit='pt', color='#3f48cc' )
        :Params : pass the hex values from the DOM for the color
        """
        if kwargs['expand']:
            label_button = self._utillity.validate_and_get_webdriver_object(kwargs['parent_css'], 'title')
            self._coreutility.left_click(label_button)
        if 'font_name' in kwargs:
            font_name_obj = self._utillity.validate_and_get_webdriver_object(kwargs['parent_css']+  " div[data-ibx-name='fontName']", 'font-css')
            self._coreutility.left_click(font_name_obj)
            self._wf_main_page_obj.select_context_menu_item(kwargs['font_name'], pop_up_css="div[class*='pop-top']", row_css=".ibx-select-item:not([style*='none'])")
        if 'format_style' in kwargs:
            font_style_obj = self._utillity.validate_and_get_webdriver_object(kwargs['parent_css'] + " div[data-ibx-name^='{0}']".format(kwargs['format_style']), 'style')
            self._coreutility.left_click(font_style_obj)
        if 'size' in kwargs:
            size_name_obj = self._utillity.validate_and_get_webdriver_object(kwargs['parent_css'] +  " div[data-ibx-name='fontSizeSelector']", 'size-css')
            self._coreutility.left_click(size_name_obj)
            self._wf_main_page_obj.select_context_menu_item(kwargs['size'], pop_up_css="div[class*='pop-top']", row_css=".ibx-select-item:not([style*='none'])")
        if 'unit' in kwargs:
            unit_name_obj = self._utillity.validate_and_get_webdriver_object(kwargs['parent_css'] +  " div[data-ibx-name='fontSizeUnit']", 'size-css')
            self._coreutility.left_click(unit_name_obj)
            self._wf_main_page_obj.select_context_menu_item(kwargs['unit'], pop_up_css="div[class*='pop-top']", row_css=".ibx-select-item:not([style*='none'])")
        if 'color' in kwargs:
            color_change_button = self._utillity.validate_and_get_webdriver_object(kwargs['parent_css'] + " .wfc-color-swatch-btn", 'color button')
            self._coreutility.left_click(color_change_button)
            if kwargs['color_option'] == 'palette':
                color_palette = "div[data-ibx-name='wfcColorPopup'][aria-hidden='false']"
                self._utillity.synchronize_with_number_of_element(color_palette, 1, 40)
                self._utillity.verify_object_visible(color_palette, True, 'Step x : Color palette is visible')
                total_colours = self._utillity.validate_and_get_webdriver_objects(dc_locators.COLOURS_CSS, 'colours')
                for element in total_colours:
                    if self._utillity.get_element_attribute(element, 'data-pp-hex-value') == kwargs['color']:
                        self._coreutility.left_click(element)
            elif kwargs['color_option'] == 'More':
                #TODO : add functionality later depending on the test case
                pass        
        
    def verify_font_options(self, msg, **kwargs):
        """
        Description: This method is used to verify the labels font options
        :Usage : verify_font_options('Step 4.1:', verify_font_name = 'TIMES')
        """
        if 'expand' in kwargs:
            label_button = self._utillity.validate_and_get_webdriver_object(kwargs['parent_css'], 'title')
            self._coreutility.left_click(label_button)
        if 'verify_font_name' in kwargs:
            observed_font_name = self._utillity.get_element_attribute(self._utillity.validate_and_get_webdriver_object(kwargs['parent_css']+  " div[data-ibx-name='fontName'] input", 'font-css'), 'aria-label')
            self._utillity.asequal(kwargs['verify_font_name'], observed_font_name, msg)
        if 'verify_size' in kwargs:
            observed_size = self._utillity.get_element_attribute(self._utillity.validate_and_get_webdriver_object(kwargs['parent_css'] +  " div[data-ibx-name='fontSizeSelector'] input", 'size-css'), 'aria-label')
            self._utillity.asequal(kwargs['verify_size'], observed_size, msg)
            
    def verify_style_in_chart_preview(self, attribute, verify_text, msg, parent_css=".wfc-bc-output-div", child_css=' svg g[fill]', axis='x'):
        """
        Description: This method is used to verify the fontstyle in the chart axis 
        :Usage : verify_style_in_chart_preview('font-weight', 'bold, 'Step  4.1:')
        :Params:  attribute can be transform, fill(in hex code), font-style, font-weight, font-size, font-family
        """
        axis_index = 1 if axis == 'y' else 0
        axis_elements = self._utillity.validate_and_get_webdriver_objects(parent_css + child_css, 'axis')
        element_attribute = self._utillity.get_element_attribute(axis_elements[axis_index], attribute).strip('\'')
        self._utillity.asequal(verify_text, element_attribute, msg)
        
    def verify_color_in_chart_axis(self, color, msg, attribute='fill', parent_css=".wfc-bc-output-div", child_css=' svg g[fill]', axis='x'):
        """
        Description: This method is used to verify the x axis title color in chart
        :Usage : verify_color_in_chart_axis(blue, 'Step 4.1: Verify blue is in x axis')
        """
        axis_index = 1 if axis == 'y' else 0
        axis_elements = self._utillity.validate_and_get_webdriver_objects(parent_css + child_css, 'axis')
        self._utillity.verify_element_color_using_css_property("svg g[fill]", color, msg, attribute=attribute, element_obj=axis_elements[axis_index])
        
    def verify_axis_rotation_in_chart(self, verify_text, msg, parent_css=".wfc-bc-output-div", child_css=' svg g[fill]', axis='x', compare_type='asin'):
        """
        Description: This method is used to verify the rotation of the text element in the axis
        :Usage : verify_axis_rotation_in_chart('rotate(-45)', 'Step 6.1')
        """
        axis_index = 1 if axis == 'y' else 0
        axis_elements = self._utillity.validate_and_get_webdriver_objects(parent_css + child_css, 'axis')
        text_elements = self._utillity.validate_and_get_webdriver_objects('text', 'axis-text', axis_elements[axis_index])
        for index, element in enumerate(text_elements):
            rotate_attribute = self._utillity.get_element_attribute(element, 'transform')
            if compare_type == 'asin':
                self._utillity.asin(verify_text, rotate_attribute, "{0}.{1}: {2} has the rotate attribute {3}".format(msg, index+1, self._j_script.get_element_text(element), verify_text))
            elif compare_type == 'asnotin':
                self._utillity.as_notin(verify_text, rotate_attribute, "{0}.{1}: {2} has the no rotation attribute {3}".format(msg, index+1, self._j_script.get_element_text(element), verify_text))
    
    def verify_legend_style(self, attribute, verify_text, msg, parent_css=".wfc-bc-output-div", child_css=' .legend text'):
        """
        Description: This method is used to verify legend style
        :Usage : verify_legend_style('font-weight', 'BOOKMAN', 'Step 2.1')
        """
        legend_elements = self._utillity.validate_and_get_webdriver_objects(parent_css + child_css, 'legends')
        for index, element in enumerate(legend_elements):
            legend_attribute = self._utillity.get_element_attribute(element, attribute).strip('\'')
            self._utillity.asequal(verify_text, legend_attribute, "{0}.{1}: {2} has the attribute {3}".format(msg, index+1, self._j_script.get_element_text(element), verify_text))
    
    def verify_query_bucket_values(self, bucket_name, bucket_list, msg, compare_type='asequal'):
        """
        Description: This method is used to verify query bucket values
        :Usage : verify_query_bucket_values('Size', ['Sales'], "Step 2.1")
        """
        query_buckets = self._utillity.validate_and_get_webdriver_objects(dc_locators.QUERY_CSS + " div[data-ibx-type='bucket']", 'query-buckets')
        for element in query_buckets:
            if self._utillity.validate_and_get_webdriver_object('.wfc-bucket-header', 'header', element).text.strip() == bucket_name:
                bucket_items = self._utillity.validate_and_get_webdriver_objects('.wfc-bucket-pills div[data-ibx-type="bucketPill"]', 'bucket_list', element)
                observed_bucket_list = [item.text.strip() for item in bucket_items]
                self._utillity.verify_list_values(bucket_list, observed_bucket_list, msg, assert_type=compare_type)
                
    def click_menu_widget_on_query_buckets(self, bucket_name, context_path):
        """
        Description: This method is used to click the menu widget and select items
        :Usage : click_menu_widget_on_query_buckets('Size', 'Split axis')
        """
        query_buckets = self._utillity.validate_and_get_webdriver_objects(dc_locators.QUERY_CSS + " div[data-ibx-type='bucket']", 'query-buckets')
        for element in query_buckets:
            if self._utillity.validate_and_get_webdriver_object('.wfc-bucket-header', 'header', element).text.strip() == bucket_name:
                self._coreutility.left_click(self._utillity.validate_and_get_webdriver_object('.wfc-bucket-popup-btn', 'bucket_list', element))
                self._wf_main_page_obj.select_context_menu_item(context_path, row_css=" div[data-ibx-name^='wfc']:not([style*='none'])")
                
    def select_datatype_in_format_data(self, type_name):
        """
        Description: This method is used to select datatype in format data tab
        @param datatype_format:'alpha', 'numeric', 'date' or 'custom'.
        :usage select_datatype_in_format_data('alpha')
        """
        DataFormat_Dialog.datatype_parent_group_css = '.dialog-menu-item.ibx-popup .wfc-df-datatype-group'
        DataFormat_Dialog.select_datatype(self, type_name)
        DataFormat_Dialog.datatype_parent_group_css = '.ibx-dialog.ibx-popup .wfc-df-datatype-group'
        
    def select_checkbox_in_format_data(self, checkbox_name, toggle):
        """
        Description: This function will select or un_select check box under data format popup for any one of the section 'alpha', 'numeric', 'date' or 'custom'.
        @param label_name:'Variable length'.
        @param check_uncheck_toggle: True or False
        :usage select_checkbox_in_dataformat_popup('Variable length', True)
        """
        DataFormat_Dialog.data_format_dialog_css = '.dialog-menu-item.ibx-popup'
        DataFormat_Dialog.select_checkbox_in_dataformat_popup(self, checkbox_name, toggle)
        DataFormat_Dialog.data_format_dialog_css = '.ibx-dialog.ibx-popup'
        
    def verify_drag_and_verify_warning_logo_on_the_cursor(self, source_x, source_y, target_x, target_y, target_obj, msg):
        """
        Description: This method is used to drag and drop field into the object and verify its not draggable
        :Usage : verify_drag_and_verify_warning_logo_on_the_cursor(100, 615, 853, 404, target_obj, "Step 3")
        """
        if sys.platform == 'linux':
                mouse_.press(int(source_x), int(source_y))
                time.sleep(3)
                pyautogui.moveTo(int(target_x), int(target_y), duration=3)
                time.sleep(3)
                mouse_attribute = self._utillity.get_element_attribute(target_obj, 'style')
                self._utillity.asin('cursor: not-allowed', mouse_attribute, msg)
                mouse_.release(int(target_x), int(target_y))
        else:
            if Global_variables.browser_name=='chrome':
                pyautogui.mouseDown(source_x, source_y)
                time.sleep(3)
                pyautogui.moveTo(target_x, target_y)
                time.sleep(10)
                mouse_attribute = self._utillity.get_element_attribute(target_obj, 'style')
                self._utillity.asin('cursor: not-allowed', mouse_attribute, msg)
                time.sleep(5)
                pyautogui.mouseUp()
            else:
                mouse_obj=uisoup.mouse
                mouse_obj.click(source_x, source_y)
                time.sleep(Global_variables.longwait)
                mouse_obj.press_button(source_x, source_y)
                mouse_obj.move(target_x, target_y)
                mouse_attribute = self._utillity.get_element_attribute(target_obj, 'style')
                self._utillity.asin('cursor: not-allowed', mouse_attribute, msg)
                time.sleep(Global_variables.longwait)
                mouse_obj.release_button()
                time.sleep(2*Global_variables.longwait)
    
    def drag_variables_to_filter_pane(self, item_path):
        """
        Description: This method is used to drag variables to filter pane 
        :Usage : drag_variables_to_filter_pane('Query Variables->Filters and Variables->Store Name')
        """
        field_name_list = item_path.split('->')
        source_object = self.get_recursive_tree_object(field_name_list, dc_locators.VARIABLES_CSS, 0)
        source_axis = self._coreutility.get_web_element_coordinate(source_object)
        drag_to_object = self._utillity.validate_and_get_webdriver_object(dc_locators.FILTER_BAR, 'filter-bar')
        destination_axis = self._coreutility.get_web_element_coordinate(drag_to_object)
        if sys.platform == 'linux':
            mouse_.press(int(source_axis['x']), int(source_axis['y']))
            time.sleep(3)
            pyautogui.moveTo(int(destination_axis['x']), int(destination_axis['y']), duration=3)
            time.sleep(3)
            mouse_.release(int(destination_axis['x']), int(destination_axis['y']))
        else:
            pyautogui.mouseDown(source_axis['x'], source_axis['y'], duration=1)
            pyautogui.moveTo(destination_axis['x'], destination_axis['y'], duration=1)
            time.sleep(1)
            pyautogui.mouseUp(destination_axis['x'], destination_axis['y'])
    
    def drag_data_fields_to_filter_bar(self, field_path, field_type='Dimensions'):
        """
        Description: This method us used to drag and drop data field into the filter tab 
        :Usage : drag_data_fields_to_filter_bar('Product->Discount', field_type='Dimensions')
        """
        field_name_list = field_path.split('->')
        if field_type=='Measures':
            element_to_drag = self.get_recursive_tree_object(field_name_list, dc_locators.MEASURES_TAB_CSS, 0)
        elif field_type=='Dimensions':
            element_to_drag = self.get_recursive_tree_object(field_name_list, dc_locators.DIMENSIONS_TAB_CSS, 0)
        filter_bar_obj = self._utillity.validate_and_get_webdriver_object(dc_locators.FILTER_BAR, 'filter-bar')
        self._j_script.scrollIntoView(filter_bar_obj)
        source_axis = self._coreutility.get_web_element_coordinate(element_to_drag)
        destination_axis = self._coreutility.get_web_element_coordinate(filter_bar_obj)
        if sys.platform == 'linux':
            mouse_.press(int(source_axis['x']), int(source_axis['y']))
            time.sleep(3)
            pyautogui.moveTo(int(destination_axis['x']), int(destination_axis['y']), duration=3)
            time.sleep(3)
            mouse_.release(int(destination_axis['x']), int(destination_axis['y']))
        else:
            pyautogui.mouseDown(source_axis['x'], source_axis['y'], duration=1)
            pyautogui.moveTo(destination_axis['x'], destination_axis['y'], duration=1)
            time.sleep(1)
            pyautogui.mouseUp(destination_axis['x'], destination_axis['y'])
        
    def drag_and_drop_variables_to_heading_or_footing(self, item_to_drag, layout_option, text_to_enter=None):
        """
        Description: This method is used to drag and drop variables into the header or footer
        :Usage : drag_and_drop_variables_to_header_or_footer('Query Variables->Filters and Variables->Store Name', 'Heading', spaces_to_click=' ')
        """
        if text_to_enter:
            text_to_enter = Keys.END + text_to_enter
        drag_to_object = self.click_heading_or_footing(layout_option, return_item=True, text_to_enter=text_to_enter)
        field_name_list = item_to_drag.split('->')
        source_object = self.get_recursive_tree_object(field_name_list, dc_locators.VARIABLES_CSS, 0)
        source_axis = self._coreutility.get_web_element_coordinate(source_object)
        destination_axis = self._coreutility.get_web_element_coordinate(drag_to_object)
        if sys.platform == 'linux':
            mouse_.press(int(source_axis['x']), int(source_axis['y']))
            time.sleep(3)
            pyautogui.moveTo(int(destination_axis['x']), int(destination_axis['y']), duration=3)
            time.sleep(3)
            mouse_.release(int(destination_axis['x']), int(destination_axis['y']))
        else:
            self._coreutility.drag_and_drop(source_axis['x'], source_axis['y'], destination_axis['x'], destination_axis['y'])
        
    def drag_and_drop_field_to_bucket(self, field_name, bucket_name, field_location):
        """
        Descriptions : This method will drag and drop Dimensions/Measures fields on bucket under display.
        :Usage drag_and_drop_filed_to_bucket('CAR', 'Verticals', measure)
        """
        field_css = dc_locators.DIMENSIONS_TAB_CSS if field_location == 'dimension' else dc_locators.MEASURES_TAB_CSS
        field_name_list = field_name.split('->')
        source_object = self.get_recursive_tree_object(field_name_list, field_css, 0)
        source_cord = self._coreutility.get_web_element_coordinate(source_object)
        bucket_obj = [elem for elem in self._utillity.validate_and_get_webdriver_objects(dc_locators.QUERY_BUCKET_CSS, bucket_name) if elem.text.strip()==bucket_name][0]
        bucket_container_obj = self._utillity.validate_and_get_webdriver_object(dc_locators.QUERY_BUCKET_CONTAINER_CSS, bucket_name+' container', parent_object=bucket_obj)
        target_cord=self._coreutility.get_web_element_coordinate(bucket_container_obj)
        if sys.platform == 'linux':
            mouse_.press(int(source_cord['x']), int(source_cord['y']))
            time.sleep(3)
            pyautogui.moveTo(int(target_cord['x']), int(target_cord['y']), duration=3)
            time.sleep(3)
            mouse_.release(int(target_cord['x']), int(target_cord['y']))
        else:
            self._coreutility.drag_and_drop(source_cord['x'], source_cord['y'], target_cord['x'], target_cord['y'])
    
    def remove_query_bucket_field(self, bucket_name, filed_name):
        """
        Descriptions : This method will remove field's(hover over field and click 'x') in query bucket.
        :Usage remove_query_bucket_field('Vertical', 'Customers')
        """
        self._utillity.synchronize_until_element_is_visible(dc_locators.QUERY_BUCKET_CSS, self.home_page_long_timesleep)
        bucket_objs = self._utillity.validate_and_get_webdriver_objects(dc_locators.QUERY_BUCKET_CSS, 'query bucket')
        bucket_list = [query_bucket.text.strip() for query_bucket in bucket_objs]
        status_ = False
        for index_, bucket_name_  in enumerate(bucket_list):
            if bucket_name in bucket_name_:
                bucket_obj = bucket_objs[index_]
                status_ = True
                break
        if status_ == False:
            raise LookupError('{0} not found inside query panel'.format(bucket_name))
        field_objs = self._utillity.validate_and_get_webdriver_objects(dc_locators.QUERY_BUCKET_FIELD_CSS, 'query bucket field_name', parent_object=bucket_obj)
        query_bucket_field_list = [query_bucket_field.text.strip() for query_bucket_field in field_objs]
        status_ = False
        for field_index_, bucket_name_field_  in enumerate(query_bucket_field_list):
            if filed_name in bucket_name_field_:
                field_obj = field_objs[field_index_]
                status_ = True
                break
        if status_ == False:
            raise LookupError('{0} not found inside {1} query bucket'.format(filed_name, bucket_name))
        self._coreutility.python_move_to_element(field_obj)
        self._utillity.synchronize_until_element_is_visible_within_parent_object(field_obj, dc_locators.QUERY_BUCKET_FIELD_REMOVE_BUTTON_CSS, self.home_page_short_timesleep)
        remove_button = self._utillity.validate_and_get_webdriver_object(dc_locators.QUERY_BUCKET_FIELD_REMOVE_BUTTON_CSS, 'query bucket field_name', parent_object=field_obj)
        self._coreutility.left_click(remove_button)
    
    def get_display_toolbar_options_in_query_bucket(self):
        """
        Descriptions : This method will get display tool_bar options in query bucket and return display toolbar parent object and it's options list.
        :Usage get_display_toolbar_options_in_query_bucket()
        """
        self._utillity.synchronize_until_element_is_visible(dc_locators.QUERY_BUCKET_DISPLAY_TOOLBAR_CSS, self.home_page_long_timesleep)
        bucket_display_toolbar_obj = self._utillity.validate_and_get_webdriver_objects(dc_locators.QUERY_BUCKET_DISPLAY_TOOLBAR_CSS, 'QUERY_BUCKET_DISPLAY_TOOLBAR_CSS')
        bucket_display_toolbar_text = [dis_tool.get_attribute('title').strip() for dis_tool in bucket_display_toolbar_obj if dis_tool.get_attribute('title').strip() != '']
        return (bucket_display_toolbar_obj, bucket_display_toolbar_text)
    
    def select_display_toolbar_in_query_bucket(self, display_toolbar):
        """
        Descriptions : This method will select display tool_bar options in query bucket.
        :Usage select_display_toolbar_in_query_bucket('Stacked)
        """
        bucket_display_toolbar_obj, bucket_display_toolbar_text = self.get_display_toolbar_options_in_query_bucket()
        for index_, toolbar_ in enumerate(bucket_display_toolbar_text):
            if toolbar_.lower() == display_toolbar.lower():
                toolbar_obj = bucket_display_toolbar_obj[index_]
                break
        self._utillity.synchronize_until_element_is_visible_within_parent_object(toolbar_obj, '.ibx-label-icon', self.home_page_long_timesleep)
        self._coreutility.left_click(toolbar_obj)
    
    def verify_display_toolbar_in_query_bucket(self, display_toolbar_list, display_message, assert_type='asequal'):
        """
        Descriptions : This method will verify display tool_bar options as list in query bucket.
        :Usage verify_display_toolbar_in_query_bucket(['Stacked'], 'Step 09.00: verify', assert_type='asin')
        """
        bucket_display_toolbar_obj= self.get_display_toolbar_options_in_query_bucket()
        self._utillity.verify_list_values(display_toolbar_list, bucket_display_toolbar_obj[1], display_message, assert_type=assert_type)
    
    def verify_selected_display_toolbar_in_query_bucket(self, display_toolbar, display_message):
        """
        Descriptions : This method will verify selected display tool_bar options in query bucket.
        :Usage verify_selected_display_toolbar_in_query_bucket('Stacked', 'Step 09.00: verify')
        """
        bucket_display_toolbar_obj, bucket_display_toolbar_text = self.get_display_toolbar_options_in_query_bucket()
        for index_, toolbar_ in enumerate(bucket_display_toolbar_text):
            if toolbar_.lower() == display_toolbar.lower():
                toolbar_obj = bucket_display_toolbar_obj[index_]
                break
        self._utillity.synchronize_until_element_is_visible_within_parent_object(toolbar_obj, '.ibx-label-icon', self.home_page_long_timesleep)
        class_data = self._utillity.get_element_attribute(toolbar_obj, 'class')
        status_ = '{0} selected'.format(display_toolbar.lower()) if 'checked' in class_data else '{0} not selected'.format(display_toolbar.lower())
        self._utillity.asequal('{0} selected'.format(display_toolbar.lower()), status_, display_message)
    
    def click_toolbar_save(self):
        """
        Descriptions : This method used to click on toolbar save button 
        """
        save_btn_css="[class^='ibxtool-toolbar-group'] div[class^='ibxtool-toolbar-button'][class*='toolbar-save']"
        save_btn_description = 'Save button available in toolbar'
        save_btn_obj=self._utillity.validate_and_get_webdriver_object(save_btn_css, save_btn_description)
        self._coreutility.left_click(save_btn_obj)
    
    def enter_filename_in_open_dialog_resources(self, folder_path=None, title=None, name=None, ok_button=False, cancel_button=False):
        """
        Descriptions : This method used to perform different operation such as save page, save as page, open page, etc
        :Usage : page_designer_open_dialog_resources(folder_path='P116->S7074', title='Test')
        :Usage : page_designer_open_dialog_resources(ok_button=True)
        """
        parent_css="div[class*='open-dialog-resources']"
        cancel_btn_css=parent_css + " div[data-ibx-name='btnCancel']"
        self._utillity.synchronize_with_visble_text(cancel_btn_css, visble_element_text='Cancel', expire_time=15)
        if folder_path !=None :
            pass # Need to implement
        if title !=None :
            title_textbox_css = "div[data-ibx-name='sdtxtFileTitle']"
            title_textbox_obj = self._utillity.validate_and_get_webdriver_object(title_textbox_css, 'Save dialog title text box')
            title_textbox_input_obj = self._utillity.validate_and_get_webdriver_object('input', 'Save dialog title text box input', parent_object=title_textbox_obj)
            self._utillity.set_text_to_textbox_using_keybord(title, text_box_elem=title_textbox_input_obj)
            self._utillity.synchronize_with_visble_text_within_parent_object(title_textbox_obj, 'input', title, 90, text_option='text_value')
        if name !=None :
            pass # Need to implement
        if ok_button == True :
            ok_button_css=parent_css + " div[data-ibx-name='btnOK']"
            ok_button_obj=self._utillity.validate_and_get_webdriver_object(ok_button_css, 'Ok button')
            self._coreutility.left_click(ok_button_obj)
        if cancel_button == True :
            cancel_button_obj=self._utillity.validate_and_get_webdriver_object(cancel_btn_css, 'Cancel button')
            self._coreutility.left_click(cancel_button_obj)
        time.sleep(2)
        try : #Checking whether already file exits dialog box appear
            dialog_box_obs=self._utillity.validate_and_get_webdriver_objects("div[data-ibx-type='ibxDialog'][class*='pop-top']")
            if len(dialog_box_obs)>0 :
                self._page_design_misc.dialog_box(button_name_to_click='OK')
        except :
            pass    
        
    def save_chart_from_toolbar(self, title_to_save, wait_time=3):
        """
        Descriptions : This method used to click on save button on tool bar and save current page with default name  
        :Usage save_chart_from_toolbar('test')
        """
        self.click_toolbar_save()
        self.enter_filename_in_open_dialog_resources(title=title_to_save, ok_button=True)
        time.sleep(wait_time)
        
    def close_chart_from_application_menu(self, confirm_dialog='No'):
        '''
        Descriptions : This will click application button and then select close button
        :Usage close_chart_from_application_menu(confirm_dialog='Yes')
        '''
        self.click_application_button() 
        self.select_option_from_application_menu('Close')
        self.confirmation_dialog(confirm_dialog, confirm_dialog)
    
    def click_application_button(self):
        """
        Descriptions : This method will click on application button in ribbon tool_bar.
        :Usage click_application_button()
        """
        self._utillity.synchronize_until_element_is_visible(dc_locators.RIBBON_APPLICATION_BUTTON_CSS, self.home_page_long_timesleep)
        app_btn_obj=self._utillity.validate_and_get_webdriver_object(dc_locators.RIBBON_APPLICATION_BUTTON_CSS, 'application button available in toolbar')
        self._coreutility.left_click(app_btn_obj)
    
    def select_option_from_application_menu(self, option_name):
        """
        Descriptions : This method will click on application button and select option from menu.
        :Usage select_option_from_application_menu('Save as...')
        """
        self._wf_main_page_obj.select_context_menu_item(option_name, expire_time=self.home_page_long_timesleep)
    
    def verify_application_menu_options(self, expected_options_list, step_num, comparision_type='asequal', menu_item_path=None):
        """
        Descriptions : This method will click on application button and verify options listed in menu.
        :Usage verify_application_menu_options(['Chart'], '09.00', comparision_type='asin', menu_item_path='New')
        """
        if menu_item_path:
            self._wf_main_page_obj.select_context_menu_item(menu_item_path, expire_time=self.home_page_long_timesleep)
        self._wf_main_page_obj.verify_context_menu_item(expected_options_list, msg='Step {0}'.format(step_num), comparision_type=comparision_type)
    
    def save_as_from_application_menu(self, title=None, name=None):
        """
        Descriptions : This method will handle save as dialog.
        :Usage save_as_from_application_menu(title='Chart', name='chart')
        """
        self.click_application_button()
        self.select_option_from_application_menu('Save as...')
        self.resource_dialog('Save as', 'Save as', title, name)
        
    def resource_dialog(self, dialog_name, button_name, title=None, name=None):
        """
        Descriptions : This method will handle save as dialog.
        :Usage resour_dialog(title='Chart', name='chart')
        """
        self._utillity.synchronize_with_visble_text(dc_locators.POP_TOP_CSS, dialog_name, self.home_page_long_timesleep)
        if title:
            self._wf_main_page_obj.open_popup_dialog_from_action_bar('Title', 'text_box', property_value=title)
        if name:
            self._wf_main_page_obj.open_popup_dialog_from_action_bar('Name', 'text_box', property_value=name)
        self._wf_main_page_obj.button_in_popup_dialog_from_action_bar(button_name, 'click')
        self.confirmation_dialog('OK', 'OK')
        
    def confirmation_dialog(self, dialog_name, button_name):
        """
        Descriptions : This method will handle confirmation dialog.
        :Usage confirmation_dialog('OK', 'OK')
        """
        try: 
            self._utillity.synchronize_with_visble_text(dc_locators.POP_TOP_CSS, dialog_name, self.add_field_timesleep)
            self._wf_main_page_obj.button_in_popup_dialog_from_action_bar(button_name, 'click')
        except:
            pass
        
    def select_tab_button(self, button_name):
        """
        Descriptions : This method will select tab button(bottom left hand corner) from designer.
        :Usage select_tab_button('Data')
        """
        tab_list_obj = self._utillity.validate_and_get_webdriver_objects(Designer.TAB_SELECTION_CSS, 'Desinger bottom tab selection')
        tab_button_obj = [button for button in tab_list_obj if button.is_displayed() and (button.text.strip()).lower() == button_name.lower()][0]
        self._coreutility.left_click(tab_button_obj)
    
    def get_data_panel_node_object(self, node_name):
        """
        Descriptions : This method will get data tree node name.
        :Usage get_data_panel_node_object('fact')
        """
        try:
            self._utillity.synchronize_with_visble_text(Designer_chart_workbook_data_tab.NODE_BROWSER_CSS, node_name, 1)
        except ValueError:
            raise LookupError("'{0}' item not found under data.".format(node_name))
        browser_node_top_middle_location = int(self._coreutility.get_web_element_coordinate(self._utillity.validate_and_get_webdriver_object(Designer_chart_workbook_data_tab.NODE_BROWSER_CSS, 'Data panel'), element_location='top_middle')['y'])
        browser_node_bottom_middle_location = int(self._coreutility.get_web_element_coordinate(self._utillity.validate_and_get_webdriver_object(Designer_chart_workbook_data_tab.NODE_BROWSER_CSS, 'Data panel'), element_location='bottom_middle')['y'])
        self._coreutility.python_move_to_element(self._utillity.validate_and_get_webdriver_object(Designer_chart_workbook_data_tab.NODE_BROWSER_CSS, 'Data panel'))
        parent_obj = self._utillity.validate_and_get_webdriver_object(Designer_chart_workbook_data_tab.NODE_BROWSER_CSS, 'Data panel')
        node_object = parent_obj.find_element_by_xpath("//div[normalize-space() = '{0}']".format(node_name))
        node_top_middle_location = int(self._coreutility.get_web_element_coordinate(node_object, element_location='top_middle')['y'])
        direction = 'up' if node_top_middle_location < browser_node_top_middle_location else 'down'
        status_ = True
        while status_:
            node_top_middle_location = int(self._coreutility.get_web_element_coordinate(node_object, element_location='top_middle')['y'])
            node_bottom_middle_location = int(self._coreutility.get_web_element_coordinate(node_object, element_location='bottom_middle')['y'])
            if direction == 'up' and node_top_middle_location <= browser_node_top_middle_location:
                self._utillity.mouse_scroll(direction, 1, option='uiautomation',  pause=0.5)
            elif direction == 'down' and node_bottom_middle_location >= browser_node_bottom_middle_location:
                self._utillity.mouse_scroll(direction, 1, option='uiautomation',  pause=0.5)
            else:
                status_ = False
        return (node_object)
    
    def click_data_node_folder(self, node_path):
        """
        Descriptions : This method will click on node path based on hierarchy.
        example usage1 : click_data_node_folder('baseapp/facts/wf_retail_sales')
        """
        try:
            self._coreutility.switch_to_frame( frame_css=Designer_chart_workbook_data_tab.DATA_PAGE_FRAME_CSS)
            node_list = node_path.split('->') 
            for node_name in node_list:
                node_object = self.get_data_panel_node_object(node_name)
                self._coreutility.left_click(node_object, xoffset=-99)
                self._coreutility.switch_to_default_content()
        except Exception as e:
            self._coreutility.switch_to_default_content()
            raise Exception(e)
    
    def get_join_object(self, join_node_name):
        """
        Descriptions : This method will get join node on canvas.
        :Usage get_join_object('Join 1')
        """
        join_canvas_obj = self._utillity.validate_and_get_webdriver_object(Designer_chart_workbook_data_tab.JOIN_CANVAS_CSS, 'Join canvas')
        try:
            join_node_object = join_canvas_obj.find_element_by_xpath("//div[@qa='{0}']".format(join_node_name))
        except NoSuchElementException:
            raise LookupError("'{0}' join field not found on canvas.".format(join_node_name))
        return (join_node_object)
    
    def click_join_object_node_on_canvas(self, join_node_name, click_type='left', click_options=None):
        """
        Descriptions : This method will click['left', 'right'] join node on canvas.
        :Usage click_join_object_node_on_canvas('Join 1')
        """
        try:
            self._coreutility.switch_to_frame(frame_css=Designer_chart_workbook_data_tab.DATA_PAGE_FRAME_CSS)
            join_node_object = self.get_join_object(join_node_name)
            if click_type == 'left':
                self._coreutility.left_click(join_node_object)
            elif click_type == 'right':
                self._coreutility.right_click(join_node_object)
                self._wf_main_page_obj.select_context_menu_item(click_options, expire_time=self.home_page_medium_timesleep)
            else:
                raise LookupError("'{click_type}' click is not implemented, 'click_join_object_node_on_canvas' method needs modification for '{click_type}' click option.".format(click_type=click_type))
            self._coreutility.switch_to_default_content()
        except Exception as e:
            self._coreutility.switch_to_default_content()
            raise Exception(e)
                    
    def drag_and_drop_node_folder_to_join_obj(self, node_path, join_node_name):
        """
        Descriptions : This method will click on node path based on hierarchy then get data tree node and drop on join node canvas.
        :Usage drag_and_drop_node_folder_to_join_obj('baseapp/facts/wf_retail_sales', 'join 1')
        """
        node_list= node_path.split('->')
        if len(node_list) > 1:
            self.click_data_node_folder('->'.join(node_list[:-1]))
        try:
            self._coreutility.switch_to_frame(frame_css=Designer_chart_workbook_data_tab.DATA_PAGE_FRAME_CSS)
            node_object = self.get_data_panel_node_object(node_list[-1])
            join_node_object = self.get_join_object(join_node_name)
            source_cord=self._coreutility.get_web_element_coordinate(node_object)
            target_cord=self._coreutility.get_web_element_coordinate(join_node_object)
            if sys.platform == 'linux':
                mouse_.press(int(source_cord['x'])-99, int(source_cord['y']))
                time.sleep(3)
                pyautogui.moveTo(int(target_cord['x']), int(target_cord['y']), duration=3)
                time.sleep(3)
                mouse_.release(int(target_cord['x']), int(target_cord['y']))
            else:
                self._coreutility.drag_and_drop(source_cord['x']-99, source_cord['y'], target_cord['x'], target_cord['y'])
            self._coreutility.switch_to_default_content()
        except Exception as e:
            self._coreutility.switch_to_default_content()
            raise Exception(e)
    
    def _get_query_bucket_object(self, bucket_name):
        """
        Description : Return the query bucket object 
        :Usage : _get_query_bucket_object('Size')
        """
        bucket_objects = self._utillity.validate_and_get_webdriver_objects(dc_locators.QUERY_BUCKET_CSS, bucket_name)
        for bucket in bucket_objects :
            bucket_label = self._utillity.validate_and_get_webdriver_object(".wfc-bucket-header-label", bucket_name, bucket).text.strip()
            if bucket_label == bucket_name :
                return bucket
        else :
            raise_msg = "'{0}' bucket not exists".format(bucket_name)
            raise KeyError(raise_msg)
    
    def _get_query_bucket_field_object(self, bucket_name, field_name, field_position=1):
        """
        Description : Return the field object from specified query bucket
        :Usage : _get_query_bucket_field_object('Vertical', 'SALES')
        """
        bucket_object = self._get_query_bucket_object(bucket_name)
        field_objects = self._utillity.validate_and_get_webdriver_objects("div[title]", bucket_name + " fields", bucket_object)
        field_index_list = self._j_script.find_all_index_of_element_by_text(field_objects, field_name)
        if field_index_list != [] and len(field_index_list) >= field_position:
            field_index = field_index_list[field_position -1]
            field_object = field_objects[field_index]
            return field_object
        else :
            raise_msg = "'{0}' field not exists in {1}' bucket".format(field_name, bucket_name)
            raise KeyError(raise_msg)
        
    def drag_data_field_to_canvas(self, data_field_type, data_field_path,location='middle'):
        """
        Description : Drag the data field to canvas
        :Usage - drag_data_field_to_query_bucket('measure', 'SALES', 'Vertical')
        """
        field_css = dc_locators.DIMENSIONS_TAB_CSS if data_field_type == 'dimension' else dc_locators.MEASURES_TAB_CSS
        data_field_list = data_field_path.split('->')
        data_field_object = self.get_recursive_tree_object(data_field_list, field_css, 0)
        data_field_location = self._coreutility.get_web_element_coordinate(data_field_object)
        canvas_object = self._utillity.validate_and_get_webdriver_object('div[id*="chartpreview"]' ,"Canvas_object")
        canvas_location = self._coreutility.get_web_element_coordinate(canvas_object, location)
        if sys.platform == 'linux':
            mouse_.press(int(data_field_location['x']), int(data_field_location['y']))
            time.sleep(1)
            pyautogui.moveTo(int(canvas_location['x']), int(canvas_location['y']), duration=3)
            time.sleep(1)
            mouse_.release(int(canvas_location['x']), int(canvas_location['y']))
        else:
            self._coreutility.drag_and_drop(data_field_location['x'], data_field_location['y'], canvas_location['x'], canvas_location['y'])
        
    def drag_data_field_to_query_bucket(self, data_field_type, data_field_path, bucket_name, bucket_location='bottom_middle'):
        """
        Description : Drag the data field to query bucket container
        :Usage - drag_data_field_to_query_bucket('measure', 'SALES', 'Vertical')
        """
        field_css = dc_locators.DIMENSIONS_TAB_CSS if data_field_type == 'dimension' else dc_locators.MEASURES_TAB_CSS
        data_field_list = data_field_path.split('->')
        data_field_object = self.get_recursive_tree_object(data_field_list, field_css, 0)
        data_field_location = self._coreutility.get_web_element_coordinate(data_field_object)
        bucket_object = self._get_query_bucket_object(bucket_name)
        bucket_y_offset = -8 if bucket_location == 'bottom_middle' else 0
        bucket_location = self._coreutility.get_web_element_coordinate(bucket_object, bucket_location, yoffset=bucket_y_offset)
        if sys.platform == 'linux':
            mouse_.press(int(data_field_location['x']), int(data_field_location['y']))
            time.sleep(1)
            pyautogui.moveTo(int(bucket_location['x']), int(bucket_location['y']), duration=3)
            time.sleep(1)
            mouse_.release(int(bucket_location['x']), int(bucket_location['y']))
        else:
            self._coreutility.drag_and_drop(data_field_location['x'], data_field_location['y'], bucket_location['x'], bucket_location['y'])
    
    def drag_data_field_to_query_bucket_field(self, data_field_type, data_field_path, bucket_name, bucket_field, bucket_field_pos=1, bucket_field_loc='bottom_middle'):
        """
        Description : Drag the data field to query bucket field
        :Usage - drag_data_field_to_query_bucket_field('measure', 'SALES', 'Vertical', 'SEATS')
        """
        field_css = dc_locators.DIMENSIONS_TAB_CSS if data_field_type == 'dimension' else dc_locators.MEASURES_TAB_CSS
        data_field_list = data_field_path.split('->')
        data_field_object = self.get_recursive_tree_object(data_field_list, field_css, 0)
        data_field_location = self._coreutility.get_web_element_coordinate(data_field_object)
        bucket_field_object = self._get_query_bucket_field_object(bucket_name, bucket_field, bucket_field_pos)
        bucket_field_loc = self._coreutility.get_web_element_coordinate(bucket_field_object, bucket_field_loc)
        if sys.platform == 'linux':
            mouse_.press(int(data_field_location['x']), int(data_field_location['y']))
            time.sleep(1)
            pyautogui.moveTo(int(bucket_field_loc['x']), int(bucket_field_loc['y']), duration=3)
            time.sleep(1)
            mouse_.release(int(bucket_field_loc['x']), int(bucket_field_loc['y']))
        else:
            self._coreutility.drag_and_drop(data_field_location['x'], data_field_location['y'], bucket_field_loc['x'], bucket_field_loc['y'])
    
    def drag_query_bucket_field_to_another_query_bucket(self, source_bucket, source_field, target_bucket, source_field_pos=1, target_bucket_loc='bottom_middle'):
        """
        Description : Drag the query bucket field to another query bucket
        :Usage - drag_query_bucket_field_to_another_query_bucket('Vertical', 'SALES', 'Color')
        """
        source_bucket_field_obj = self._get_query_bucket_field_object(source_bucket, source_field, source_field_pos)
        target_bucket_obj = self._get_query_bucket_object(target_bucket)
        target_bucket_y_offset = -8 if target_bucket_loc == 'bottom_middle' else 0
        source_bucket_field_loc = self._coreutility.get_web_element_coordinate(source_bucket_field_obj)
        target_bucket_loc = self._coreutility.get_web_element_coordinate(target_bucket_obj, target_bucket_loc, yoffset=target_bucket_y_offset)
        if sys.platform == 'linux':
            mouse_.press(int(source_bucket_field_loc['x']), int(source_bucket_field_loc['y']))
            time.sleep(1)
            pyautogui.moveTo(int(target_bucket_loc['x']), int(target_bucket_loc['y']), duration=3)
            time.sleep(1)
            mouse_.release(int(target_bucket_loc['x']), int(target_bucket_loc['y']))
        else:
            self._coreutility.drag_and_drop(source_bucket_field_loc['x'], source_bucket_field_loc['y'], target_bucket_loc['x'], target_bucket_loc['y'])
        
    def drag_query_bucket_field_to_another_query_bucket_field(self, source_bucket, source_field, target_bucket, target_field, source_field_pos=1, target_field_pos=1, target_field_loc='bottom_middle'):
        """
        Description : Drag the query bucket field to another query bucket field
        :Usage - drag_query_bucket_field_to_another_query_bucket_field('Vertical', 'SALES', 'Horizontal', 'CAR')
        """
        source_bucket_field_obj = self._get_query_bucket_field_object(source_bucket, source_field, source_field_pos)
        target_bucket_field_obj = self._get_query_bucket_field_object(target_bucket, target_field, target_field_pos)
        source_bucket_field_loc = self._coreutility.get_web_element_coordinate(source_bucket_field_obj)
        target_bucket_field_loc = self._coreutility.get_web_element_coordinate(target_bucket_field_obj, target_field_loc)
        if sys.platform == 'linux':
            mouse_.press(int(source_bucket_field_loc['x']), int(source_bucket_field_loc['y']))
            time.sleep(1)
            pyautogui.moveTo(int(target_bucket_field_loc['x']), int(target_bucket_field_loc['y']), duration=3)
            time.sleep(1)
            mouse_.release(int(target_bucket_field_loc['x']), int(target_bucket_field_loc['y']))
        else:
            self._coreutility.drag_and_drop(source_bucket_field_loc['x'], source_bucket_field_obj['y'], target_bucket_field_loc['x'], target_bucket_field_loc['y'])
    
    def right_click_and_select_option_from_query_pane(self, bucket_name,field_name, context_menu_item_path, field_position=1, click_type='left',element_location='middle',xoffset=0):
        """
        Descriptions: This function is used to choose context menu option from query bucket
        :Usage : right_click_and_select_option_from_query_pane('Vertical', 'SALARY', 'Shape->Line')
        """
        query_pane_obj = self._get_query_bucket_field_object(bucket_name, field_name, field_position=field_position)
        self._coreutility.python_right_click(query_pane_obj)
        self._utillity.synchronize_until_element_is_visible('.pop-top',self._wf_main_page_obj.chart_medium_timesleep)
        self._wf_main_page_obj.select_context_menu_item(context_menu_item_path, pop_up_css="div[class*='pop-top'][data-ibx-type*='Menu']", click_type=click_type,element_location=element_location,xoffset=xoffset)
    
    def verify_right_click_and_option_from_query_pane(self, expected_context_menu_item_list, msg, comparision_type='asequal', context_menu_item_path=None):
        """
        Descriptions : This function will right click on query bucket and verify options
        :Usage verify_right_click_and_option_from_query_pane(['Line'], 'Step 09.00: Verify option', comparision_type='asin', context_menu_item_path='Shape')
        """
        if context_menu_item_path:
            self._wf_main_page_obj.select_context_menu_item(context_menu_item_path, pop_up_css="div[class*='pop-top'][data-ibx-type*='Menu']", click_type=None)
        self._wf_main_page_obj.verify_context_menu_item(expected_context_menu_item_list, msg, comparision_type=comparision_type)
    
    def verify_right_click_and_option_is_checked_unchecked_from_query_pane(self, expected_context_menu_item_list, status_type, msg, comparision_type='asequal', context_menu_item_path=None):
        """
        Descriptions : This function will right click on query bucket and verify option is checked/Unchecked.
        :Usage verify_right_click_and_option_is_checked_unchecked_from_query_pane(['Line'], 'check', 'Step 09.00: Verify option', comparision_type='asin', context_menu_item_path='Shape')
        """
        if context_menu_item_path:
            self._wf_main_page_obj.select_context_menu_item(context_menu_item_path, pop_up_css="div[class*='pop-top'][data-ibx-type*='Menu']", click_type=None)
        self._wf_main_page_obj.verify_context_menu_item_checked_unchecked(expected_context_menu_item_list, status_type, msg, comparision_type=comparision_type)
           
class Designer_Insight_Metadata(BasePage):
    """ Inherit attributes of the parent class = Baseclass """
    
    def __init__(self, driver):
        super(Designer_Insight_Metadata, self).__init__(driver)
        self._utillity = UtillityMethods(self.driver)
        self._coreutility = CoreUtillityMethods(self.driver)
        self._page_design_misc = PageDesignerMiscelaneous(self.driver)
        self._j_script = JavaScript(self.driver)
        self._wf_main_page_obj = Wf_Mainpage(self.driver)
        self._popup_dialog = poptop_dialog.Poptop_Dialog(self.driver)
        self.MEASURE_PARENT_CSS = None
        self.DIMENSIONS_PARENT_CSS = None
        self.QUERYBOX_PARENT_CSS = ".wfc-bucket-acc-page"
        self.QUERYTOOL_PARENT_CSS = self.QUERYBOX_PARENT_CSS + " [data-ibx-name='bucketToolbar']"
        self.QUERYBUCKET_PARENT_CSS = self.QUERYBOX_PARENT_CSS + " [data-ibx-name='bucketGroup']"
        self.CHART_PICKER_PARENT_CSS = '.chart-picker-box'
        self.FILTER_BAR_PARENT_CSS = '.filter-bar'
    
    def switch_to_insight_frame(self):
        self._coreutility.switch_to_frame(insight_locators.INSIGHT_PREVIEW_FRAME)
    
    def select_more_options(self, item_path):
        """
        Description : This method select the more option from the tool bar
        :Usage select_more_options('Export Data')
        """
        layout_button = self._utillity.validate_and_get_webdriver_object(insight_locators.MORE_OPTIONS_BUTTON, 'more-options-button')
        self._coreutility.left_click(layout_button)
        self._wf_main_page_obj.select_context_menu_item(item_path)
        
    def verify_insight_querybox(self, expected_query_list, msg, compare_type='asequal'):
        """
        Description: This function is used to verify the query buckets text
        :Usage : verify_insight_querybox(['Size', 'Color'], "Step 4.1: ")
        """
        query_buckets = self._utillity.validate_and_get_webdriver_objects(insight_locators.QUERY_BOX + " .bucket-label", 'query buckets')
        query_buckets_text =  [element.text for element in query_buckets]
        self._utillity.verify_list_values(expected_query_list, query_buckets_text, msg, assert_type=compare_type)
        
    def verify_insight_optionsbox(self, expected_options_list, msg, compare_type='asequal'):
        """
        Description: This function is used to verify the options buckets text
        :Usage : verify_insight_optionsbox(['More Options'], "Step 4.1: ")
        """
        option_buckets = self._utillity.validate_and_get_webdriver_objects(insight_locators.OPTIONS_BOX + " .er-menu-button:not([style*='none'])", 'option buckets')
        option_buckets_text = [self._utillity.get_element_attribute(element, 'title').strip() for element in option_buckets]
        self._utillity.verify_list_values(expected_options_list, list(option_buckets_text), msg, assert_type=compare_type)
        
    def select_chart_from_grid(self, chart_option=None):
        """
        Description: This function is used to select the chart
        :Usage : select_chart_from_grid(chart_option='Horizontal Bar')
        :Params : Use the following parameters are arguments
        Horizontal Bar
        Vertical Bar
        Vertical Stacked Bar
        Ring Pie
        Vertical Line
        Area
        Scatter
        Circle Plot
        Treemap
        Histogram
        Table
        Matrix
        Point Map
        Choropleth Map
        """
        chart_picker_button = self._utillity.validate_and_get_webdriver_object(insight_locators.CHART_PICKER_BUTTON, 'chart-picker-button')
        self._coreutility.left_click(chart_picker_button)
        self._utillity.synchronize_with_number_of_element(".chart-picker-grid", 1, 30)
        chart_objects = self._utillity.validate_and_get_webdriver_objects('.chart-button', 'chart-buttons')
        found = False
        for element in chart_objects:
            if self._utillity.get_element_attribute(element, 'title').strip() == chart_option:
                found = True
                self._coreutility.left_click(element)
        if not found:
            raise IndexError("{0} is not found in the chart list".format(chart_option))
    
    def verify_query_bucket_sort(self, bucket_name, sorted_value, msg, sorted_by='ascending'):
        """
        Description: This function is verify the query bucket
        :Usage : verify_query_bucket_sort('Horizontal Axis', 'Product Category', 'Step 4', sorted_by='ascending')
        :Params : sorted_by = 'ascending' or 'descending'
        """
        query_buckets = self._utillity.validate_and_get_webdriver_objects(".query-box .bucket-container", 'bucket-container')
        found = False
        for element in query_buckets:
            if self._utillity.validate_and_get_webdriver_object('.bucket-label', 'label', element).text.strip() == bucket_name:
                found = True
                sort_element_text = self._utillity.get_element_attribute(self._utillity.validate_and_get_webdriver_object('.er-fb-how-sort', 'sort', element),'aria-label').lower()
                self._utillity.asin(sorted_by, sort_element_text, msg + " {0} is sorted by {1}".format(bucket_name, sorted_by))
                element_sorted = self._utillity.validate_and_get_webdriver_object('.er-fb-field-label', 'sorted-element', element).text.strip()
                self._utillity.asequal(sorted_value, element_sorted, msg)
        if not found:
            raise Exception("{0} bucket is not found".format(bucket_name))
    
    def select_insight_optionsbox(self, select_option):
        """
        Description: This function is used to select the options
        :Usage : select_insight_optionsbox('Save')
        """
        option_buckets = self._utillity.validate_and_get_webdriver_objects(insight_locators.OPTIONS_BOX + " .er-menu-button:not([style*='none'])", 'option buckets')
        for element in option_buckets:
            if self._utillity.get_element_attribute(element, 'title').strip() == select_option:
                self._coreutility.left_click(element)
                break
        
class Designer_FontStyle(object):
    """Common class to veriy all the font styles"""
    def __init__(self, driver, parent_css=None, msg=None):
        self.parent_css = parent_css
        self.msg = msg
        self._utillity = UtillityMethods(driver)
        self._coreutility = CoreUtillityMethods(driver)
        
    def expand_title(self):
        tile_to_expand = self._utillity.validate_and_get_webdriver_object(self.parent_css, 'title')
        self._coreutility.left_click(tile_to_expand)
         
    def verify_show_title(self, present, verify_text=None):
        """
        Description: This function is used to verify the show title
        :Usage : verify_show_title(False)
        """
        title_css = self.parent_css + ' .wfc-fp-checkbox'
        try:
            self._utillity.validate_and_get_webdriver_object(title_css, 'show-title')
            found = True
        except AttributeError:
            found = False
        self._utillity.asequal(present, found, self.msg)
        if verify_text:
            observed_text = self._utillity.validate_and_get_webdriver_object(title_css, 'text').text.strip()
            self._utillity.asequal(verify_text, observed_text, self.msg)
            
