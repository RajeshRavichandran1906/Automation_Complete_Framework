import os
from openpyxl import Workbook
from .basepage import BasePage
from openpyxl import load_workbook
from common.locators import reports as Locators
from selenium.common.exceptions import ElementNotVisibleException

class Common(BasePage):
    
    def __init__(self, table_locator):
        
        super().__init__()
        self._table_locator = table_locator
    
    def wait_for_text(self, text, time_out=60, pause_time=0, case_sensitive=False, raise_error=True, javascript=False):
        
        self._webelement.wait_for_element_text(self._table_locator, text, time_out, pause_time, case_sensitive, raise_error, javascript)
        
    def convert_to_excel(self, file_name, start_row=None, start_col=None, end_row=None, end_col=None):
        """
        This function read all data from report table and convert as xlsx file using BeautifulSoup4 module
        :arg = table_css ='##ITableData0' or table['summary'] ( table_css is parent css value of table )
        :arg = file_name = 'C12345_TableData_Step_03'
        """
        save_file_path = self._get_xl_file_path_to_save(file_name)
        if os.path.exists(save_file_path):
            error_msg = '[{}] file already exists. Delete it if you want to create new excel file or do not call this function'.format(save_file_path)
            raise FileExistsError(error_msg)
        workbook = Workbook()
        sheet = workbook['Sheet']
        table_rows = self._get_beautifulsoup_table_object()
        for r, row in enumerate(table_rows[start_row:end_row]) :
            row_columns=row.select(' > td ')
            for c,column in enumerate(row_columns[start_col:end_col]) :
                sheet.cell(row=r + 1, column=c + 1).value = str(column.get_text(strip=True)).strip()
        workbook.save(save_file_path) #save the xlsx file
        
    def verify_data(self, saved_file_name, step_num, start_row=None, start_col=None, end_row=None, end_col=None):
        """
        This function compare two report table data
        :arg = table_css ='#ITableData0' or table['summary'] ( table_css is parent css value of table )
        :arg = saved_file_name = 'C12345_TableData_Step_03'
        """
        compare_status = self.__compare_report_table_data_using_bs4(saved_file_name, start_row, start_col, end_row, end_col)
        actual_result = len(compare_status)
        msg = 'Step {} : Verify table data {}'.format(step_num, compare_status)
        self._utils.asequal(1, actual_result, msg)
    
    def __compare_report_table_data_using_bs4(self, file_name, start_row=None, start_col=None, end_row=None, end_col=None):
        """
        This function read all the data from report table using BeautifulSoup4 module and 
        compare with expected report table xlsx file which created by convert_report_table_data_to_xl_using_bs4()
        It will return status=[0] if two xlsx files are same
        """
        table_rows=self._get_beautifulsoup_table_object()# get all rows of table
        sheet=self._get_workboot_sheet_object(file_name)
        status=[]
        for r, row in enumerate(table_rows[start_row:end_row]) :
            row_columns=row.select(' > td ')
            for c,column in enumerate(row_columns[start_col:end_col]) :
                expected_value = sheet.cell(row=r + 1, column=c + 1).value
                if expected_value != None : #check excel cell value is not null
                    actual_value = str(column.get_text(strip=True)).strip()
                    if actual_value == expected_value:
                        status=[0]
                        continue
                    else :
                        status = ['Actual:Table[{}][{}] = {}'.format(r+1, c+1, actual_value), 'Expected:Table[{}][{}] = {}'.format(r+1, c+1, expected_value)]
                        return status
        return status
        
    def _get_xl_file_path_to_save(self, file_name):
        """
        This method used to create xlsx format file and return xl saved file path.
        It will validate whether "data" folder available in test suite folder
        """
        save_file_name = file_name + '.xlsx'
        script_data_path = os.path.join(os.getcwd(), 'data')
        save_file_path = os.path.join(script_data_path, save_file_name)
        return save_file_path
    
    def _get_workboot_sheet_object(self, file_name):
        """
        This method used to get saved report xl file and ceate wookbook sheet object 
        """
        saved_file_name = file_name + '.xlsx'
        saved_file_path=os.path.join(os.getcwd(), 'data', saved_file_name)
        workbook=load_workbook(saved_file_path) 
        sheet = workbook['Sheet']
        return sheet
        
    def _get_beautifulsoup_table_object(self):
        """
        This function return the beautifulsoup table object
        """
        table = self._driver.find_element(*self._table_locator)
        if table.is_displayed() :
            from bs4 import BeautifulSoup
            script='return arguments[0].innerHTML'
            page_source = self._driver.execute_script(script, table)
            beautifulsoup = BeautifulSoup(page_source, 'html.parser')
            table_rows = beautifulsoup.select(' > tbody > tr')
            if len(table_rows) > 0 :
                return table_rows
            else :
                raise ValueError("Table does not have any rows")
        else :
            raise ElementNotVisibleException("Table not visible")
    
    def _get_cell_object(self, row, col):
        """
        Description: Return the given row and columns cell object
        """
        table = self._utils.validate_and_get_webdriver_object_using_locator(self._table_locator, "Report table")
        cell_css = "tbody>tr:nth-child({})>td:nth-child({})".format(row, col)
        cell_object = self._utils.validate_and_get_webdriver_object(cell_css, "Report table cell", parent_object=table)
        self._javascript.scrollIntoView(cell_object, wait_time=0.5)
        return cell_object

class HTML5(Common):
    
    Locator = Locators.HTML5

    def __init__(self, table_locator = Locator.run_table):
        
        super().__init__(table_locator)
        
    def click_on_hyperlink_cell(self, row, col):
        """
        Description: Left click on hyper link cell 
        :Usage click_on_hyperlink_cell(5, 2)
        """
        cell_object = self._get_cell_object(row, col)
        hyperlink = self._utils.validate_and_get_webdriver_object("a[href]", 'Report table hyper link cell', parent_object=cell_object)
        hyperlink.click()
        #self._core_utils.left_click(hyperlink)