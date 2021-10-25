from common.lib import utillity, javascript
from common.lib.base import BasePage
from common.pages import vfour_miscelaneous, vfour_portal_canvas
from selenium.common.exceptions import NoSuchElementException
import time, os
from openpyxl import Workbook
from openpyxl import load_workbook


class Vfour_Portal_Run(BasePage):
    """ Inherit attributes of the parent class = Baseclass """

    def __init__(self, driver):
        super(Vfour_Portal_Run, self).__init__(driver)
    
    def create_table_data_set(self,table_css,file_name, **kwargs):
        """    
        @param: table_css= ".arPivot tr:nth-child(1) table" or "table[summary='Summary']"   =  Need to provide the full parent path till table
        @param: file_name: "test1.xlsx" 
        Usage: create_table_data_set("table[summary='Summary']", "test1.xlsx" )
        """
        rows_css = table_css + " > tbody > tr:not([style*='none'])"
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        column_ = table_rows[0].find_elements_by_css_selector(rows_css + ":nth-child(1) > td")
        no_of_col = len(column_)
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        for r in range(no_of_rows):
            col_css=rows_css + " > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(no_of_col):
                value = columns[c].text
                s.cell(row=r + 1, column=c + 1).value = str(value)
#                 print(r,c,str(value))
        wb.save(os.getcwd() + "\data\\" + file_name) 
    
    def verify_table_data_set(self,table_css,file_name,msg, **kwargs):
        """
        Usage: utillobj.verify_table_data_set("table[summary='Summary']", "test1.xlsx","Step 10: fail data set")
        """
        x= self.compare_table_data_set(table_css,file_name, **kwargs)
        utillity.UtillityMethods.asequal(self,len(x),1,msg+ ' Row,Column -'+ str(x))
        
    def compare_table_data_set(self, table_css, file_name, **kwargs):
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        rows_css = table_css + " > tbody > tr:not([style*='none'])"
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        column_ = table_rows[0].find_elements_by_css_selector(rows_css + ":nth-child(1) > td")
        no_of_col = len(column_)
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        for r in range(no_of_rows):
            col_css=rows_css + " > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(no_of_col):
                value = columns[c].text
                if s1.cell(row=r + 1, column=c + 1).value != None and str(value) != '':
                    if s1.cell(row=r + 1, column=c + 1).value == str(value):
                        status=[0]
                        continue
                    else:
                        status=[r+1,c+1, 'actual value : ',str(value), 'expected value : ', s1.cell(row=r + 1, column=c + 1).value]
                        return (status)
        return (status)
         
    def select_or_verify_portal_menu_bar_item(self, **kwargs):
        """
        :param kwargs[select]='Close'
        :param kwargs[verify]=['autodevuser48', 'Tools', 'Resources', 'Help', 'Close', 'Sign Out', 'Portals', 'Theme', 'Dashboards', 'Hidden Content']
        :Param msg="Step 9: Verify portal menu bar"
        :Usage1: select_or_verify_menu_bar_item(select='Close',verify=[user_id, 'Tools', 'Resources', 'Help', 'Close', 'Sign Out', 'Portals', 'Theme', 'Dashboards', 'Hidden Content'],msg='Step 04:')
        :Usage2: select_or_verify_menu_bar_item(select='Close')
        """
        menu_css="div[class*='menu-bar-item'] span"
        menus=self.driver.find_elements_by_css_selector(menu_css)
        menu_list=[el.text.strip() for el in menus]
        if 'select' in kwargs:
            menus[menu_list.index(kwargs['select'])].click()
            if 'option' in kwargs:
                utillity.UtillityMethods.select_or_verify_bipop_menu(self, kwargs['option'])
        if 'verify' in kwargs:
            utillity.UtillityMethods.as_List_equal(self, kwargs['verify'], menu_list, kwargs['msg'])
        time.sleep(1)
        
    def verify_portal_panel_label(self, expected_label_list, msg):
        """
        :param expected_label=['Panel 1','Category Sales','Panel 3','Panel 4'] #should be in list
        :Param msg="Step 9: Verify portal panel label"
        :Usage1: verify_portal_panel_label(['Panel 1','Category Sales','Panel 3','Panel 4'],'Step 04:')
        """
        label_css="[id^='BipInnerBox'] div[id^='BiLabel'][class*='bi-label bip-title-bar pd-internal-id']"
        label=self.driver.find_elements_by_css_selector(label_css)
        actual_label_list=[el.text.strip() for el in label if el.text.strip() != '']
        utillity.UtillityMethods.as_List_equal(self, expected_label_list, actual_label_list, msg + " Verify portal panel labels.")
        time.sleep(1)
        
    def verify_number_of_portal_panel(self, expected_number, msg):
        """
        expected_number: 3 or 4 ..(Any integer value. This is the number of panels in portal)
        syntax verify_number_of_portal_panel(4, 'Step 10:')
        """
        active_page = vfour_portal_canvas.Vfour_Portal_Canvas.get_current_page(self)
        total_panel=len(active_page.find_elements_by_css_selector("[class*='bip-container'] [id^='BiDockPanel']"))
        actual_number=int(total_panel)
        utillity.UtillityMethods.asequal(self,expected_number, actual_number, msg + " Verify number of panels in portal")
    
    def expand_portal_resource_tree(self, folder_path, **kwargs):
        """
        :Param : folder_path = 'S7068->AR-RP-193'
        Syntax: expand_resource_tree('S7068->AR-RP-193')
        """
        tree_rows="div[id^='BipContentArea'] #treeView table>tbody>tr"
        folder_list=folder_path.split('->')
        for item in folder_list:
            repository_items = self.driver.find_elements_by_css_selector(tree_rows)
            try:
                for i in range(len(repository_items)):
                    if repository_items[i].text.strip() == item:
                        try:
                            folder_img = repository_items[i].find_element_by_css_selector("td>img[src*='triangle_collapsed']")
                            self.driver.execute_script("arguments[0].scrollIntoView(true);", folder_img)
                            folder_img.click()
                            time.sleep(3)
                            if item == folder_list[-1]:
                                return
                        except:
                            print("Resource " + item + " already expanded.")
                        break
            except NoSuchElementException as e:
                            print(e,item + " not found in Tree view.")
        
    def right_click_portal_resource_tree(self, item_name, click_option=1, **kwargs):                                                            
        """
        :Param : item_name = 'AR-RP-193'
        Syntax: right_click_resource_tree('AR-RP-193')
        """
        BIPtree_rows="div[id^='BipContentArea'] #treeView table>tbody>tr"
        rows = self.driver.find_elements_by_css_selector(BIPtree_rows)
        for i in range(len(rows)):
            td_item = BIPtree_rows + ":nth-child(" + str(i+1) + ")>td"
            get_td_text=self.driver.find_element_by_css_selector(td_item).text
            if get_td_text == item_name:
                td_img = self.driver.find_element_by_css_selector(td_item + ">img.icon")
                self.driver.execute_script("arguments[0].scrollIntoView(true);", td_img)
                td_img.click()
                utillity.UtillityMethods.click_on_screen(self, td_img, 'middle', click_type=click_option, **kwargs)
                
    def select_portal_resource_menu(self, folder_path, menu_item, option_click=1, **kwargs):
        """
        :Param : folder_path = 'S7068->AR-RP-193'
        :Param : menu_item = 'Edit' or 'Run'
        Syntax: select_portal_resource_menu('S7068->AR-RP-193', 'Run')
        """
        item_name=folder_path.split('->')[-1]
        if 'expand_resource_tree' in kwargs:
            Vfour_Portal_Run.expand_portal_resource_tree(self, folder_path, **kwargs)
        Vfour_Portal_Run.right_click_portal_resource_tree(self,item_name, click_option=option_click, **kwargs)
        if option_click==1:
            menu_list = menu_item.split('->')
            for j in range(len(menu_list)):
                popups = self.driver.find_elements_by_css_selector("div[id^='BiPopup'][style*='inherit']")
                menu_items = popups[-1].find_elements_by_css_selector("table>tbody>tr>[class='text']")
                for i in range(len(menu_items)):
                    if menu_items[i].text.strip() == menu_list[j]:
                        menu_items[i].click()
                        break
        utillity.UtillityMethods.switch_to_default_content(self, pause=2)  
    
    def verify_portal_resource_item(self, folder_path, item_name, msg, item_exit=True, **kwargs):
        """
        folder_path = 'P116->S7068'
        item_name= 'S7067'
        msg='9'            #Step number
        item_exit=Ture or False
        Syntax: verify_resource_item('P116', 'S7067', '9', item_exit=True)
        """
        Vfour_Portal_Run.expand_portal_resource_tree(self, folder_path, **kwargs)
        BIPtree_rows="div[id^='BipContentArea'] #treeView table>tbody>tr"
        repository_items = self.driver.find_elements_by_css_selector(BIPtree_rows)
        statx = item_name in [i.text.strip() for i in repository_items]
        if item_exit == True:
            utillity.UtillityMethods.asequal(self,item_exit, statx, "Step "+str(msg)+": Verify " + item_name + " File or Folder is Avialable.")
        else:
            utillity.UtillityMethods.asequal(self,item_exit, statx, "Step "+str(msg)+": Verify " + item_name + " File or Folder is Not Avialable.")      
        
    def drag_portal_resource_tree_item(self, folder_path, target_elem, **kwargs):
        """
        :Param : folder_path = 'P292->S10117->BIP_V4_Portal->Accordion_report' #inorder to drag 'Accordion_report' from resource tree
        :Param : target_elem = driver.find_element_by_css_selector("#BIPortalPanel") #must be the target locator
        Syntax: Usage1: vfour_runobj.drag_portal_resource_tree_item('P292->S10117->BIP_V4_Portal->Accordion_report',target_elem) #to drag any item from resource tree
        Syntax: Usage2: vfour_runobj.drag_portal_resource_tree_item('P292->S10117->BIP_V4_Portal->Accordion_report', target_elem, panel_css=True) #to drag any item from panel tree
        """
        if 'panel_css' in kwargs:
            BIPtree_rows="div[id^='BipContentArea'] #treeView table>tbody>tr"
            Vfour_Portal_Run.expand_portal_resource_tree(self, folder_path, **kwargs)
        else:
            BIPtree_rows="#bipResourcesPanel #treeView table>tbody>tr"
            vfour_miscelaneous.Vfour_Miscelaneous.expand_resource_tree(self, folder_path, **kwargs)
        item_name=folder_path.split('->')[-1]
        rows = self.driver.find_elements_by_css_selector(BIPtree_rows)
        for i in range(len(rows)):
            td_item = BIPtree_rows + ":nth-child(" + str(i+1) + ")>td"
            get_td_text=self.driver.find_element_by_css_selector(td_item).text
            if get_td_text == item_name:
                td_img = self.driver.find_element_by_css_selector(td_item + ">img.icon")
                self.driver.execute_script("arguments[0].scrollIntoView(true);", td_img)
                break
        if td_img != '':
            utillity.UtillityMethods.click_on_screen(self, td_img, 'middle', click_type=0)
            time.sleep(2)
            vfour_miscelaneous.Vfour_Miscelaneous.drag_drop_in_bip(self, td_img, target_elem, **kwargs)
        else:
            print(item_name+" not found in Web Resource tree.")
                
    def get_current_page(self):
        elems = self.driver.find_elements_by_css_selector("[class*='bip-canvas'] [class*='bip-page']")
        for el in elems:
            if el.value_of_css_property('visibility') == 'visible':
                return (el)
            
    def get_panel_obj_in_responsive(self, panel_name):
        """
        This function Retrun panel element.
        :param: panel_name='Panel 1'
        :Usage:  get_panel_obj('Panel 1')
        """
        current_page = Vfour_Portal_Run.get_current_page(self)
        panel_elems = current_page.find_elements_by_css_selector("[class*='bip-container'] [id^='BiDockPanel'] [class*='flex-panel-item']")
        return(panel_elems[[panel_name in el1 for el1 in [el.text.strip() for el in panel_elems]].index(True)])
        
    def select_panel_in_responsive(self, panel_name):
        """
        This function is select column in canvas.
        :Param :panel_name='Panel 1'
        :Usage : select_panel('Panel 1')
        """
        elem=Vfour_Portal_Run.get_panel_obj_in_responsive(self, panel_name)
        panel_titles = elem.find_elements_by_css_selector("[id^='BiLabel'][class*='bip-title-bar']")
        panel_title=panel_titles[[elem.text.strip() for elem in panel_titles].index(panel_name)]
        utillity.UtillityMethods.click_on_screen(self, panel_title, 'middle', click_type=0)
    
    def verify_panel_caption_in_responsive(self, panel_name, msg):
        '''
        Verify panel caption is exits column.
        :param : panel_name='Panel 1' or 'any panel name'
        :param : msg = 'Step X : Verify first panel title'
        :Usage : verify_panel_caption('Panel 1','Step 02 : Verify first panel title') 
        '''
        panel_elem=Vfour_Portal_Run.get_panel_obj_in_responsive(self, panel_name)
        actual_title=panel_elem.find_element_by_css_selector("[id^='BiLabel'][class*='bip-title-bar']").text.strip()
        utillity.UtillityMethods.asequal(self, panel_name, actual_title, msg)
    
    def manage_panel_title_menubar_button(self, panel_name, verify=True, **kwargs):
        panel_elem = Vfour_Portal_Run.get_panel_obj_in_responsive(self, panel_name)
        panel_elem.find_element_by_css_selector("[id^='BipTitleBarMenuButton'][class*='button']").click()
        if verify == True:
            utillity.UtillityMethods.select_or_verify_bipop_menu(self, verify=True, expected_popup_list=kwargs['expected_opt'], msg=kwargs['msg'])
        else:
            select_menu_option = kwargs['select_menu_opt']
            utillity.UtillityMethods.select_or_verify_bipop_menu(self, select_menu_option)
    
    def verify_column_panel_title_menubar_button(self, column_number, panel_name, verify=True, **kwargs):
        '''
        This function verify panel title menu option in column panels.
        '''
        column_elem = Vfour_Portal_Run.get_column_obj(self, column_number)
        panel_elems = column_elem.find_elements_by_css_selector("[class*='bip-container'] [id^='BiDockPanel']")
        panel_elem = panel_elems[[panel_name in el1 for el1 in [el.text.strip() for el in panel_elems]].index(True)]
        panel_titles = panel_elem.find_element_by_css_selector("[id^='BipTitleBarMenuButton'][class*='button']")
        utillity.UtillityMethods.default_click(self, panel_titles)
        if verify == True:
            utillity.UtillityMethods.select_or_verify_bipop_menu(self, verify=True, expected_popup_list=kwargs['expected_opt'], msg=kwargs['msg'])
        else:
            select_menu_option = kwargs['select_menu_opt']
            utillity.UtillityMethods.select_or_verify_bipop_menu(self, select_menu_option)
    
    def get_column_obj(self, column_no):
        '''
        This function Retrun column element.
        :Param : column_no = 1 or 2,3,4 ( Column number start from 1 )
        :Usage : get_column_obj(1)
         or 
        :Usage : get_column_obj(2)
        '''
        current_page = Vfour_Portal_Run.get_current_page(self)
        colums=current_page.find_elements_by_css_selector("[class*='bip-column']")
        return(colums[column_no-1])
    
    def select_column(self, column_no):
        """
        This function is selecet column in canvas.
        :Param : column_no = 1 or 2,3,4 ( Column number start from 1 )
        :Usage : select_column(1)
        """
        elem=Vfour_Portal_Run.get_column_obj(self, column_no)
        utillity.UtillityMethods.click_on_screen(self, elem, 'top_middle', click_type=0)
    
    def get_panel_obj(self, panel_name):
        """
        This function Retrun panel element.
        :param: panel_name='Panel 1'
        :Usage:  get_panel_obj('Panel 1')
        """
        current_page = Vfour_Portal_Run.get_current_page(self)
        panel_elems = current_page.find_elements_by_css_selector("[class*='bip-container'] [id^='BiDockPanel']")
        try:
            return(panel_elems[[panel_name in el1 for el1 in [el.text.strip() for el in panel_elems]].index(True)])
        except:
            print(panel_name+" panel name not in list ", [el.text.strip() for el in panel_elems])
        
    def select_panel(self, panel_name):
        """
        This function is select column in canvas.
        :Param :panel_name='Panel 1'
        :Usage : select_panel('Panel 1')
        """
        elem=Vfour_Portal_Run.get_panel_obj(self, panel_name)
        panel_titles = elem.find_elements_by_css_selector("[id^='BiLabel'][class*='bip-title-bar']")
        panel_title=panel_titles[[elem.text.strip() for elem in panel_titles].index(panel_name)]
        utillity.UtillityMethods.click_on_screen(self, panel_title, 'middle', click_type=0)
    
    def select_panel_in_column(self, column_no, panel_name):
        '''
        This function select panel in specific column.
        '''
        column_obj = Vfour_Portal_Run.get_column_obj(self, column_no)
        panel_elems = column_obj.find_elements_by_css_selector("[class*='bip-container'] [id^='BiDockPanel']")
        try:
            elem = panel_elems[[panel_name in el1 for el1 in [el.text.strip() for el in panel_elems]].index(True)]
        except:
            raise KeyError(print(panel_name+" panel name not in list ", [el.text.strip() for el in panel_elems]))
        panel_titles = elem.find_elements_by_css_selector("[id^='BiLabel'][class*='bip-title-bar']")
        panel_title=panel_titles[[elem.text.strip() for elem in panel_titles].index(panel_name)]
        utillity.UtillityMethods.click_on_screen(self, panel_title, 'middle', click_type=0)
    
    def select_panel_in_column_with_index(self, column_no, panel_index_number, panel_name):
        '''
        This function select panel in specific column.
        '''
        column_obj = Vfour_Portal_Run.get_column_obj(self, column_no)
        panel_elems = column_obj.find_elements_by_css_selector("[class*='bip-container'] [id^='BiDockPanel']")
        try:
            elem = panel_elems[panel_index_number]
        except:
            raise KeyError(print(panel_name+" panel name not in list ", [el.text.strip() for el in panel_elems]))
        panel_titles = elem.find_elements_by_css_selector("[id^='BiLabel'][class*='bip-title-bar']")
        panel_title=panel_titles[[elem.text.strip() for elem in panel_titles].index(panel_name)]
        utillity.UtillityMethods.click_on_screen(self, panel_title, 'middle', click_type=0)
    
    def verify_column_panel_title_menu_with_index(self, column_no, panel_index_number, panel_name, expected_opt, msg):
        '''
        This function select panel in specific column.
        '''
        column_obj = Vfour_Portal_Run.get_column_obj(self, column_no)
        panel_elems = column_obj.find_elements_by_css_selector("[class*='bip-container'] [id^='BiDockPanel']")
        try:
            elem = panel_elems[panel_index_number]
        except:
            raise KeyError(print(panel_name+" panel name not in list ", [el.text.strip() for el in panel_elems]))
        panel_titles = elem.find_element_by_css_selector("[id^='BipTitleBarMenuButton'][class*='button']")
        utillity.UtillityMethods.default_click(self, panel_titles)
        utillity.UtillityMethods.select_or_verify_bipop_menu(self, verify=True, expected_popup_list=expected_opt, msg=msg)
    
    def expand_tree(self, folder_path, tree_css=None, scroll_elem=None):
        """
        :Param : folder_path = 'S7068->AR-RP-193'
        :Usage expand_tree('S7068->AR-RP-193')
        @author: AAkhan
        """
        if scroll_elem!=None:
            scroll_obj=scroll_elem
        else:
            scroll_obj=self.driver.find_element_by_css_selector("div[id^='BipContentArea'] #treeView table>tbody>tr")
        utillity.UtillityMethods.click_on_screen(self, scroll_obj, 'middle')
        time.sleep(2)
        utillity.UtillityMethods.click_on_screen(self, scroll_obj, 'middle', y_offset=-9)
        if tree_css != None:
            tree_rows=tree_css
        else:
            tree_rows="div[id^='BipContentArea'] #treeView table>tbody>tr"
        folder_list=folder_path.split('->')
        for item in folder_list:
            repository_items = self.driver.find_elements_by_css_selector(tree_rows)
            resource_tree_list = javascript.JavaScript.get_elements_text(self, repository_items)
            if item not in resource_tree_list:
                raise KeyError(item+ " Not Exist in setup "+ self.driver.current_url)
            try:
                for i in range(len(repository_items)):
                    if repository_items[i].text.strip() == item:
                        try:
                            folder_img = repository_items[i].find_element_by_css_selector("td>img[src*='triangle_collapsed']")
                            vfour_miscelaneous.Vfour_Miscelaneous.scroll_within_resource_tree(self, scroll_obj, folder_img)
                            utillity.UtillityMethods.default_click(self, folder_img)
                            time.sleep(3)
                            if item == folder_list[-1]:
                                return (i)
                        except:
#                             print("Resource " + item + " already expanded.")
                            if item == folder_list[-1]:
                                return (i)
                        break
            except NoSuchElementException as e:
                print(e,item + " not found in Repository Tree view. It might be a Bug also.")

    def right_click_on_tree_elem(self, elem, tree_css=None):
        """
        :Param : elem = elem                        '''Need to pass object of the element.'''
        :Usage right_click_on_tree_elem(elem)
        @author: AAkhan
        """ 
        if tree_css != None:
            tree_rows=tree_css
        else:
            tree_rows="div[id^='BipContentArea'] #treeView table>tbody>tr"
        repository_items = self.driver.find_elements_by_css_selector(tree_rows)
        td_img = repository_items[elem].find_element_by_css_selector("img.icon")
        self.driver.execute_script("arguments[0].scrollIntoView(true);", td_img)
        utillity.UtillityMethods.default_click(self, td_img, click_option=1)
        
    def select_menu(self, folder_path, menu_item=None, tree_elem_css=None, scrollable_elem=None, **kwargs):
        """
        :Param : folder_path = 'S7068->AR-RP-193'
        :Param : menu_item = 'Edit' or 'Run'
        :kwargs['item_exit']= True or False
        :kwargs['expected_menu_list']= ['TextEditor', 'URL'] 
        :kwargs['msg']= "Step 11"
        :Syntax: select_menu('S7068->AR-RP-193', 'Run')
        @author: AAkhan
        """
        expand = Vfour_Portal_Run.expand_tree(self, folder_path, tree_css=tree_elem_css, scroll_elem=scrollable_elem)
        Vfour_Portal_Run.right_click_on_tree_elem(self, expand, tree_css=tree_elem_css)
        popup_css="div[id^='BiPopup'][style*='inherit']"
        popup_menu_css="table>tbody>tr>[class='text']"
        utillity.UtillityMethods.synchronize_with_number_of_element(self, popup_css, 1, 19)
        if menu_item == None:
            pass
        else:
            menu_list = menu_item.split('->')
            for j in range(len(menu_list)):
                popups = self.driver.find_elements_by_css_selector(popup_css)
                if self.browser.lower() == 'firefox':
                    utillity.UtillityMethods.click_on_screen(self, popups[-1], 'top_middle', y_offset=3)
                menu_items = popups[-1].find_elements_by_css_selector(popup_menu_css)
                for i in range(len(menu_items)):
                    if menu_items[i].text.strip() == menu_list[j]:
                        utillity.UtillityMethods.default_click(self, menu_items[i])
                        break
        if 'expected_menu_list' in kwargs:
            popup_menu_css="table>tbody>tr"
            popups = self.driver.find_elements_by_css_selector(popup_css)
            menu_items = popups[-1].find_elements_by_css_selector(popup_menu_css)
            actual_elem=[el.text.strip().replace(' ','') for el in menu_items]
            actual_elem = [elem for elem in actual_elem if elem != '']
            expected_elem = [exp.replace(' ', '') for exp in kwargs['expected_menu_list']]
            count_= 1
            for elem in expected_elem:
                status_=True if elem in actual_elem else False
                verification_msg = "Visible" if kwargs['item_exit'] == True else "Not Visible"
                utillity.UtillityMethods.asequal(self, kwargs['item_exit'], status_, str(kwargs['msg']) + "." + str(count_) + " : Verify "+ elem + " Option "+verification_msg+" in Repository menu")
                count_ += 1
        utillity.UtillityMethods.switch_to_default_content(self, pause=2)   
    