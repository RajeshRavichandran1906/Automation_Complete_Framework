from common.lib import mobile_utillity
from common.locators.mobile_web_locators import MobileLoginPageLocators
from common.lib.mobile_base import BasePage
from selenium.webdriver import ActionChains
#from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from appium.webdriver.common.mobileby import By
from selenium.webdriver.support.color import Color
from openpyxl import Workbook
from openpyxl import load_workbook
import platform,time,os,re

class Mobile_View_Miscellaneous(BasePage):
    """ Inherit attributes of the parent class = Baseclass """

    def __init__(self, driver):
        super(Mobile_View_Miscellaneous, self).__init__(driver)
        self.test=mobile_utillity.UtillityMethods.parseinitfile(self,'test')
    
    def invoke_WF(self):
        '''
        :Usage : invoke_mobile_view()
        '''
        node_id=mobile_utillity.UtillityMethods.parseinitfile(self,'nodeid')
        port_num=mobile_utillity.UtillityMethods.parseinitfile(self,'httpport')
        context=mobile_utillity.UtillityMethods.parseinitfile(self,'wfcontext')
        url='http://'+node_id+':'+port_num+context+'/signin'
        self.driver.get(url)
        login_button=MobileLoginPageLocators.mv_signin
        mobile_utillity.UtillityMethods.wait_for_element(self, login_button, wait_method=0)
       
    def invoke_mobile_view(self):
        '''
        :Usage : invoke_mobile_view()
        '''
        node_id=mobile_utillity.UtillityMethods.parseinitfile(self,'nodeid')
        port_num=mobile_utillity.UtillityMethods.parseinitfile(self,'httpport')
        context=mobile_utillity.UtillityMethods.parseinitfile(self,'wfcontext')
        mv_folder_path=mobile_utillity.UtillityMethods.parseinitfile(self,'mv_folder_path')
        url='http://'+node_id+':'+port_num+context+'/mv/'+mv_folder_path
        self.driver.get(url)
    
    def login_mobile_view(self,mrid,mrpass):
        '''
        :Param : mrid='mrid' or 'mrid01' or 'mrid02' 
        :Param : mrpass='mrpass' or 'mrpass01', or 'mrpass02'
        :Usage : login_mobile_view('mrid','mrpass')
        '''
        username=mobile_utillity.UtillityMethods.parseinitfile(self,mrid)
        password=mobile_utillity.UtillityMethods.parseinitfile(self,mrpass)
        self.invoke_WF()
        mobile_utillity.UtillityMethods.browser_height=mobile_utillity.UtillityMethods.get_browser_height(self)
        self.driver.find_element(*MobileLoginPageLocators.mv_username).send_keys(username)
        self.driver.find_element(*MobileLoginPageLocators.mv_password).send_keys(password)
        if self.test=='android_web' :
            current_context=self.driver.current_context
            self.driver.switch_to.context('NATIVE_APP')
            self.driver.press_keycode(66)
            self.driver.switch_to.context(current_context)
        else :
            self.driver.find_element(*MobileLoginPageLocators.mv_signin).click()
        self.invoke_mobile_view()
        ribbon=(By.CSS_SELECTOR,"div[data-ibx-type='ibxButtonSimple'][class*='hamburger-button'] div[class*='material-icons']")
        mobile_utillity.UtillityMethods.wait_for_element(self,ribbon, wait_method=0)
    
    def select_fex(self,fex_name, click_slider_menu=True):
        '''
        :Param : fex_name = 'S10141_ACT:AR_RP_001'
        :Usage : select_fex('S10141_ACT')
        '''
        if click_slider_menu == True :
            slider_menu = self.driver.find_element_by_css_selector("div[data-ibx-type='ibxButtonSimple'][class*='hamburger-button']")
            ActionChains(self.driver).move_to_element(slider_menu).perform()
            time.sleep(1)
            slider_menu.click()
        fex_elements=self.driver.find_elements_by_css_selector("div[class*='sidebar-button'] div[class*='ibx-label-text']")
        for fex in fex_elements :
            if fex.text.strip()==fex_name :
                ActionChains(self.driver).move_to_element(fex).perform()
                fex.click()
                time.sleep(2)
                selected_fex=self.driver.find_element_by_css_selector("div[class*='title-box']").text.strip()
                mobile_utillity.UtillityMethods.asequal(self,selected_fex,fex_name,'Step X : Verify '+fex_name+' is selected')
                break
            
    def mobileview_switch_to_frame(self,fex_name,**kwargs):
        '''
        :Param : fex_name='S10141_ACT:AR_RP_001'
        :Usage : mobileview_switch_to_frame('S10141_ACT:AR_RP_001')
        '''
        iframe_height=self.driver.find_element_by_css_selector("div[class^='top-bar']").size['height']
        if 'ignore_frame_height' not in kwargs :
            mobile_utillity.UtillityMethods.browser_height+=iframe_height
        frame_css="div[class*='render-content'] iframe[src]"
        WebDriverWait(self.driver, 120).until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR, frame_css)))
        
    def run_fex(self,fex_name,mrid,mrpass):
        '''
        :Param : fex_name= 'S10141_ACT:AR_RP_001'
        :Param : mrid='mrid' or 'mrid01' or 'mrid02' 
        :Param : mrpass='mrpass' or 'mrpass01', or 'mrpass02'
        :Usage : run_fex('S10141_ACT:AR_RP_001','mrid','mrpass')
        '''
        self.login_mobile_view(mrid, mrpass)
        self.select_fex(fex_name)
        self.mobileview_switch_to_frame(fex_name)
        
    def select_menu_item(self,table_id,column_no,menu_nagivation_patten,**kwargs):
        '''
        :Param : table_id='ITableData0'
        :Param : column_no= 0 or 1,2... (column number start from 0 which is table columns name position)
        :Param : menu_nagivation_patten = 'Chart->Pie->State' or 'Rollup->State' or 'Grid Tool'
        :Usage : select_menu_item("ITableData0",2,"Chart->Pie->State")
        :Usage : select_menu_item("ITableData0",5,"Rollup->State")
        :Usage : select_menu_item("ITableData0",0,"Grid Tool")
        '''
        table_index=list(table_id)[-1]
        main_id=str(table_index)+"_"+str(column_no)
        main_id_css="dt"+main_id+"_0"
        menu_icon_css="#popid"+main_id+" div[onclick*='ibiMenu.Showmenu']"
        column_css="#TCOL_"+str(table_index)+"_C_"+str(column_no)
        self.driver.find_element_by_css_selector(column_css).click()
        time.sleep(1)
        self.driver.find_element_by_css_selector(menu_icon_css).click()
        time.sleep(2)
        menu_items=menu_nagivation_patten.split('->')
        sub_items_css="[style*='block']>table>tbody>tr[id^='t"+main_id+"']>td>span[id^='set"+main_id+"']"
        for item in menu_items :
            main_items_css="div[id='"+main_id_css+"']"
            full_css=main_items_css+sub_items_css
            item_elements=self.driver.find_elements_by_css_selector(full_css)
            item_text=[ele.text.strip() for ele in item_elements]
            sub_item_index=item_text.index(item)
            if 'verify' in kwargs and item==menu_items[-1] :
                actual_menu=item_text
                if '' in actual_menu :
                    actual_menu.remove('')
                #print(actual_menu)
                mobile_utillity.UtillityMethods.asequal(self,actual_menu,kwargs['verify'],kwargs['msg'])
            item_elements[sub_item_index].click()
            time.sleep(2)
            main_id_css=main_id_css+"_"+str(sub_item_index)
    
    def verify_menu_items(self,table_id,column_no,menu_nagivation_patten,expected_menu_list,msg):
        '''
        :Param : table_id='ITableData0'
        :Param : column_no= 0 or 1,2... (column number start from 0 which is table columns name position)
        :Param : menu_nagivation_patten = 'Chart->Pie->State' or 'Rollup->State' or 'Grid Tool'
        :Usage : select_menu_item("ITableData0",2,"Chart->Pie->State")
        :Usage : select_menu_item("ITableData0",5,"Rollup->State")
        :Usage : select_menu_item("ITableData0",0,"Grid Tool")
        '''
        table_index=list(table_id)[-1]
        main_id=str(table_index)+"_"+str(column_no)
        main_id_css="dt"+main_id+"_0"
        menu_icon_css="#popid"+main_id+" div[onclick*='ibiMenu.Showmenu']"
        column_css="#TCOL_"+str(table_index)+"_C_"+str(column_no)
        self.driver.find_element_by_css_selector(column_css).click()
        time.sleep(1)
        self.driver.find_element_by_css_selector(menu_icon_css).click()
        time.sleep(2)
        sub_items_css="[style*='block']>table>tbody>tr[id^='t"+main_id+"']>td>span[id^='set"+main_id+"']"
        actual_menu_list=[]
        if menu_nagivation_patten==None :
            main_items_css="div[id='"+main_id_css+"']"
            full_css=main_items_css+sub_items_css
            item_elements=self.driver.find_elements_by_css_selector(full_css)
            actual_menu_list=[ele.text.strip() for ele in item_elements if ele.text.strip()!='']
        else :
            menu_items=menu_nagivation_patten.split('->')
            for item in menu_items :
                main_items_css="div[id='"+main_id_css+"']"
                full_css=main_items_css+sub_items_css
                item_elements=self.driver.find_elements_by_css_selector(full_css)
                item_text=[ele.text.strip() for ele in item_elements]
                sub_item_index=item_text.index(item)
                item_elements[sub_item_index].click()
                time.sleep(2)
                main_id_css=main_id_css+"_"+str(sub_item_index)
                if item==menu_items[-1] :
                    main_items_css="div[id='"+main_id_css+"']"
                    full_css=main_items_css+sub_items_css
                    item_elements=self.driver.find_elements_by_css_selector(full_css)
                    actual_menu_list=[ele.text.strip() for ele in item_elements if ele.text.strip()!='']
        #print(actual_menu_list)
        mobile_utillity.UtillityMethods.asequal(self,actual_menu_list,expected_menu_list,msg)
    
    def verify_popup_title(self,popup_id,expected_title,msg) :
        '''
        :Param : popup_id ='wall1'
        :Param : expected_title ='Product ID BY Product ID'
        :Param : msg ='Step 01.1 : Verify popup window title'
        :Usage : verify_popup_title('wall1','Product ID BY Product ID','Step 01.3 : Verify popup window title')
        '''
        popup_css="#"+popup_id+" .arWindowBarTitle"
        actual_title=self.driver.find_element_by_css_selector(popup_css).text.strip()
        mobile_utillity.UtillityMethods.asequal(self,actual_title,expected_title,msg)
    
    def close_popup_window(self,popup_index):
        '''
        :Param : popup_index=1 or 2 
        :Usage : close_popup_window(1)
        '''
        popup_css="#wall"+str(popup_index)+" .arWindowBar div[id^='WCS'] div[onclick^='closewin']"
        self.driver.find_element_by_css_selector(popup_css).click()
        time.sleep(2)
        
    def minimize_popup_window(self,popup_index,**kwargs):
        '''
        :Param : popup_index=1 or 2 
        :Usage : close_popup_window(1)
        '''
        popup_css="#wall"+str(popup_index)+" .arWindowBar div[id^='WCS'] div[onclick^='minwin']"
        self.driver.find_element_by_css_selector(popup_css).click()
        time.sleep(2)
        if 'verify' in kwargs :
            self.verify_minimized_popup_windows(kwargs['verify'], kwargs['msg'])
            
    def verify_report_column_heading(self, table_id, expected_list, msg):

        """
        :Param : table_id='ITableData0'
        :Param : column=['Category', 'Product ID', 'Product', 'State', 'Unit Sales', 'Dollar Sales']
        :Usage : verify_column_heading(table_id, expected_list,'Step 02: Verify column heading')
        Author: Prabhakaran
        """
        column_heading_css = "#" + table_id + " .arGridColumnHeading > td [id^='TCOL'] span"
        columns=self.driver.find_elements_by_css_selector(column_heading_css)
        actual_list=[el1.text.strip() for el1 in columns]
        #print(actual_list)
        mobile_utillity.UtillityMethods.asequal(self, expected_list, actual_list, msg)
    
    def verify_report_page_summary(self, parent_id, expected_title, msg):
        '''
        :Param  : parent_id='MAINTABLE_wbody0 '
        :Syntax : verify_report_summary('IWindowBody0', '107of107records,Page1of2', 'Step 04: Verify Pagination shows: SUB/TOT')
        :Author : Prabhakaran
        '''
        parent_css="#"+parent_id+" .arGridBar table td[align*='LEFT']>table>tbody>tr>td:nth-child(2)"
        actual_title=self.driver.find_element_by_css_selector(parent_css).text.strip()
        #print(actual_title)
        mobile_utillity.UtillityMethods.asequal(self,expected_title, actual_title, msg)
    
    def move_to_invisible_element(self,element):
        '''
        :Param : element = webelement
        :Usage : move_to_invisible_element(element)
        '''
        if self.test=='android_web' or self.test=='android_app' :
            ActionChains(self.driver).move_to_element(element).perform()
        else :
            self.driver.execute_script("return arguments[0].scrollIntoView();",element)
        time.sleep(2)
    
    def create_data_set(self,table_id,tr_id,file_name, color='default'):
        """
        Usage: utillobj.create_data_set('ITableData0','I0r','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','rgb','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','notrgb','C2140681.xlsx')
        Usage: utillobj.create_data_set('ITableData0','bgcolor','C2140681.xlsx', color='green')
        """
        if tr_id == 'rgb':
            css="[id='{0}'] tr[style*='{1}']"
        elif tr_id == 'notrgb':
            css="[id='{0}'] tr:not([style*='rgb'])[id*='I0']"
        elif tr_id=='bgcolor':
            bgcolor=self.color_picker(color, 'rgb')
            css="[id='{0}'] tr[style*='background-color: " + bgcolor + "']"
        else:
            css="[id='{0}'] tr[id ^='{1}']"
        #from openpyxl import Workbook
        wb = Workbook()
        #sheet = wb.get_active_sheet
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(css.format(table_id, tr_id))
        for r in range(0, len(table_rows)):
            columns = table_rows[r].find_elements_by_css_selector("td")
            for c in range(0, len(columns)):
                s.cell(row=r + 1, column=c + 1).value = str(columns[c].text.strip())
        os_platform=platform.system()
        if os_platform=='Windows' :
            wb.save(os.getcwd() + "\data\\" + file_name)
        else :
            wb.save(os.getcwd() + "/data/" + file_name)
    
    def compare_data_set(self, table_id, tr_id, file_name, color='default'):
        '''
        Usage: utillobj.verify_data_set('ITableData0','I0r','C2140681.xlsx')
        '''
        if tr_id == 'rgb':
            css="[id='{0}'] tr[style*='{1}']"
        elif tr_id == 'notrgb':
            css="[id='{0}'] tr:not([style*='rgb'])[id*='I0']"
        elif tr_id=='bgcolor':
            bgcolor=self.color_picker(color, 'rgb')
            css="[id='{0}'] tr[style*='background-color: " + bgcolor + "']"
        else:
            css="[id='{0}'] tr[id ^='{1}']"
        os_platform=platform.system()
        if os_platform=='Windows' :
            wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        else :
            wb1 = load_workbook(os.getcwd() + "/data/" + file_name)
        status=[]
        #sheet = wb.get_active_sheet
        #s = wb.get_sheet_by_name('Sheet')
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(css.format(table_id, tr_id))
        for r in range(0, len(table_rows)):
            columns = table_rows[r].find_elements_by_css_selector("td")
            for c in range(0, len(columns)):
                if s1.cell(row=r + 1, column=c + 1).value != None:
                    if s1.cell(row=r + 1, column=c + 1).value.strip() == str(columns[c].text.strip()):
                        status=[0]
                        continue
                    else:
                        status=[r+1,c]
                        return (status)
        return (status)
    
    def verify_data_set(self,table_id, tr_id, file_name,msg, color='default'):
        """
        Usage: utillobj.verify_data_set('ITableData0','I0r','C2140681.xlsx',"Step 10: fail data set")
        Usage: utillobj.verify_data_set('ITableData0','rgb','C2140681.xlsx',"Step 10: fail data set")
        Usage: utillobj.verify_data_set('ITableData0','notrgb','C2140681.xlsx',"Step 10: fail data set")
        Usage: utillobj.verify_data_set('ITableData0','bgcolor','C2140681.xlsx',"Step 10: fail data set", color='green')
        """
        x= self.compare_data_set(table_id,tr_id,file_name, color)
        mobile_utillity.UtillityMethods.asequal(self,len(x),1,msg+ ' Row,Column -'+ str(x))
    
    def click_refresh_and_select_fex(self,fex_name):
        '''
        :Param : fex_name='S10141_ACT:AR_RP_001'
        :Param : msg ='Expect back to ACT:AR_RP_001'
        :Usage : click_refresh_and_select_fex('S10141_ACT:AR_RP_001','Step 03.1 : Expect back to ACT:AR_RP_001')
        '''
        self.driver.switch_to_default_content()
        slider_menu = self.driver.find_element_by_css_selector("div[data-ibx-type='ibxButtonSimple'][class*='hamburger-button']")
        ActionChains(self.driver).move_to_element(slider_menu).perform()
        time.sleep(1)
        slider_menu.click()
        time.sleep(2)
        self.driver.find_element_by_css_selector("div[data-ibx-type='ibxButtonSimple'][class*='refresh-button']").click()
        if self.test=='ios_web' or self.test=='ios_app' :
            time.sleep(6)
        self.select_fex(fex_name, False)
        if self.test=='ios_web' or self.test=='ios_app' :
            time.sleep(5)
        self.mobileview_switch_to_frame(fex_name)
    
    def get_chart_riser_css_path(self,chart_type):
        
        riser_css={
            'bar':"svg g[class='risers']>g>rect[class^='riser'][class*='mbar']",
            'pie':"svg g[class='chartPanel'] g>path[class^='riser'][class*='edge']",
            'line':"svg g[class='risers']>g>path[class^='riser'][class*='line']",
            'scatter':"svg g[class='markers']>g>circle[class^='riser'][class*='marker']"
            }
        return riser_css[chart_type]
    
    def verify_number_of_chart_risers(self,parent_id,chart_type,expected_risers,msg):
        '''
        :Param : parent_id ='wbody1_f'
        :Param : chart_type ='bar' or 'pie' or 'line' or 'scatter'
        :Param : expected_total_risers =10
        :Param : msg ='Step 02 : Verify number of pie chart risers'
        :Usage : verify_number_of_chart_risers('wbody1_f','pie',10,'Step 02 : Verify number of pie chart risers')
        '''
        riser_css=self.get_chart_riser_css_path(chart_type)
        css="#"+parent_id+" "+riser_css
        total_risers=self.driver.find_elements_by_css_selector(css)
        actual_risers=0
        for riser in total_risers :
            if riser.is_displayed()==True :
                actual_risers+=1
        #print(actual_risers)
        mobile_utillity.UtillityMethods.asequal(self,actual_risers,expected_risers,msg)
    
    def verify_number_of_legend_risers(self,parent_id,expected_legend_risers,msg):
        '''
        :Param : parent_id ='wbody1_f'
        :Param : expected_legend_risers =10
        :Param : msg ='Step 02 : Verify number of legend risers'
        :Usage : verify_number_of_legend_risers('wbody1_f',0,'Step 04.2 : Verify nummber of legend risers')
        '''
        legend_risers_css="#"+parent_id+" g[class='legend'] g>path[class^='legend-markers']"
        total_risers=self.driver.find_elements_by_css_selector(legend_risers_css)
        actual_legend_risers=0
        for riser in total_risers :
            if riser.is_displayed()==True :
                actual_legend_risers+=1
        #print(actual_legend_risers)
        mobile_utillity.UtillityMethods.asequal(self,actual_legend_risers,expected_legend_risers,msg)
    
    def verify_legend_riser_color(self,parent_id,riser_css,color,msg,**kwargs):
        '''
        :Param : parent_id ='wbody1_f'
        :Param : riser_class ='legend-markers!s2!'
        :Param : color ='bar_blue'
        :Param : kwargs['attribute_type'] ='stroke' or 'color' or 'backgroud-color'
        :Param : kwargs['custom_css'] ='g[class='legend'] g>rect[class='legend-markers!s2!']'
        :Usage1: verify_legend_riser_color('parent_id','legend-markers!s2!','bar_blue','Step 1: Verify legend color')
        :Usage1: verify_legend_riser_color('parent_id',None,'bar_blue','Step 1: Verify legend color',custom_css='g[class='legend'] g>rect[class='legend-markers!s2!']',attribute_type='stroke')
        '''
        attribute_type=kwargs['attribute_type'] if 'attribute_type' in kwargs else "fill"           
        custom_css=kwargs['custom_css'] if 'custom_css' in kwargs else "g[class='legend'] g>path[class='"+riser_css+"']"
        legend_riser_css="#"+ parent_id + " " + custom_css
        actual_color = Color.from_string(self.driver.find_element_by_css_selector(legend_riser_css).value_of_css_property(attribute_type)).rgba
        expected_color=mobile_utillity.UtillityMethods.color_picker(self,color, 'rgba')
        #print(actual_color)
        mobile_utillity.UtillityMethods.asequal(self,actual_color, expected_color, msg)  
    
    def verify_pagination_arrows(self,parent_id,msg_step,**kwargs):
        '''
        :Param : parent_id ='MAINTABLE_wbody0'
        :Param : msg_step ='Step 02'
        :Param : kwargs['move_first_arrow']=True
        :Param : kwargs['move_previous_arrow']=True
        :Param : kwargs['move_next_arrow']=True
        :Param : kwargs['move_last_arrow']=True
        :Usage : verify_pagination_arrows('MAINTABLE_wbody0','Step 02',move_first_arrow=True,move_previous_arrow=True,move_next_arrow=True,move_last_arrow=True)
        '''
        if self.test=='ios_web' or self.test=='ios_app' :
            expected_move_first_arrow='◄◄'
            expected_move_previous_arrow='|◄'
            expected_move_next_arrow='►|' 
            expected_move_last_arrow='►►'
        else :
            expected_move_first_arrow='<<'
            expected_move_previous_arrow='|<'
            expected_move_next_arrow='>|' 
            expected_move_last_arrow='>>'
        icon_css="#"+parent_id+" .arGridBar span[title*='Move'] img"
        icon=self.driver.find_element_by_css_selector(icon_css).is_displayed()
        mobile_utillity.UtillityMethods.asequal(self,icon,True,msg_step+' : Verify pagination icon')
        if 'move_first_arrow' in kwargs :
            move_first_css="#"+parent_id+" .arGridBar span[title='First Page']>span"
            actual_move_first_arrows=self.driver.find_element_by_css_selector(move_first_css).text.strip()
            mobile_utillity.UtillityMethods.asequal(self,actual_move_first_arrows,expected_move_first_arrow,msg_step+' : Verify move first page arrow')
        else :
            move_first_css="#"+parent_id+" .arGridBar span[title='First Page']>span"
            move_first=len(self.driver.find_elements_by_css_selector(move_first_css))
            mobile_utillity.UtillityMethods.asequal(self,move_first,0,msg_step+' : Verify move first page arrow not appear')
            
        if 'move_previous_arrow' in kwargs :
            move_previous_css="#"+parent_id+" .arGridBar span[title='Previous Page']>span"
            actual_move_previous_arrows=self.driver.find_element_by_css_selector(move_previous_css).text.strip()
            mobile_utillity.UtillityMethods.asequal(self,actual_move_previous_arrows,expected_move_previous_arrow,msg_step+' : Verify move previous page arrow')
        else :
            move_previous_css="#"+parent_id+" .arGridBar span[title='Previous Page']>span"
            move_previous=len(self.driver.find_elements_by_css_selector(move_previous_css))
            mobile_utillity.UtillityMethods.asequal(self,move_previous,0,msg_step+' : Verify move previous page arrow not appear')    
            
        if 'move_next_arrow' in kwargs :
            move_next_css="#"+parent_id+" .arGridBar span[title='Next Page']>span"
            actual_move_next_arrows=self.driver.find_element_by_css_selector(move_next_css).text.strip()
            mobile_utillity.UtillityMethods.asequal(self,actual_move_next_arrows,expected_move_next_arrow,msg_step+' : Verify move next page arrow')
        else :
            move_next_css="#"+parent_id+" .arGridBar span[title='Next Page']>span"
            move_next=len(self.driver.find_elements_by_css_selector(move_next_css))
            mobile_utillity.UtillityMethods.asequal(self,move_next,0,msg_step+' : Verify move next page arrow not appear')
            
        if 'move_last_arrow' in kwargs :
            move_last_css="#"+parent_id+" .arGridBar span[title='Last Page']>span"
            actual_move_last_arrows=self.driver.find_element_by_css_selector(move_last_css).text.strip()
            mobile_utillity.UtillityMethods.asequal(self,actual_move_last_arrows,expected_move_last_arrow,msg_step+' : Verify move last page arrow')
        else :
            move_last_css="#"+parent_id+" .arGridBar span[title='Last Page']>span"
            move_last=len(self.driver.find_elements_by_css_selector(move_last_css))
            mobile_utillity.UtillityMethods.asequal(self,move_last,0,msg_step+' : Verify move last page arrow not appear')
    
    def verify_minimized_popup_windows(self,expected_windows,msg):
        '''
        :Param : expected_windows=['Filter Selection'] or expected_windows=None
        :Param : msg='Step 03.5 : Verify minimized popup windows'
        :Usage : verify_minimized_popup_windows(['Filter Selection'],'Step 03.5 : Verify minimized popup windows')
        '''
        window_elements=self.driver.find_elements_by_css_selector("#wall0 table[style*='border: 1px solid black'] td[class^='tabPagingText']")
        if expected_windows!=None :
            actual_windows=[window.text.strip() for window in window_elements]
            #print(actual_windows)
            mobile_utillity.UtillityMethods.asequal(self,actual_windows,expected_windows,msg)
        else :
            mobile_utillity.UtillityMethods.asequal(self,0,len(window_elements),msg)
            
    def verify_chart_tooltip(self, parent_id, raiser_class, expected_tooltip_list, msg,**kwargs):
        """
        :Param: parent_id='iosTabs0_f'
        :Param: raiser_class='riser!s0!g5!mbar!'
        :Param: expected=['DEALER_COST, ENGLAND: 37,853']
        :Param: touch_tab=True ( This keyword arguments using for ios to tab on riser because some times action chain not working for ios to verify tooltip)
        :Usage1: verify_active_chart_tooltip('iosTabs0_f','riser!s0!g5!mbar!',['DEALER_COST, ENGLAND: 37,853'],'Step 02 : verify tooltip value') 
        :Usage2: verify_chart_tooltip('MAINTABLE_wbody0','riser!s0!g8!mbar!',expected_tooplip,'Step 03.6 : Verify chart tooltip value',touch_tab=True)
        """
        action1 = ActionChains(self.driver)
        move1 = self.driver.find_element_by_css_selector("#"+ parent_id)
        action1.move_to_element_with_offset(move1,1,1).perform()
        time.sleep(1)
        del action1
        action = ActionChains(self.driver)        
        raiser_css="#"+ parent_id + " [class*='" + raiser_class + "']"
        tooltip_css="#"+ parent_id + " span[id*='tdgchart-tooltip'] > div"
        riser=self.driver.find_element_by_css_selector(raiser_css)
        if 'touch_tab' in kwargs and self.test=='ios_web' :
            utliobj=mobile_utillity.UtillityMethods(self.driver)
            param=utliobj.get_element_screen_coordinate(riser,'middle',browser_height=True)
            self.driver.execute_script('mobile: tap',param)
        else :
            action.move_to_element(riser).perform()
            del action
            time.sleep(1)
        tooltips=self.driver.find_element_by_css_selector(tooltip_css).text.split("\n")
        actual_tooltip_list=[el.strip() for el in tooltips]
        #print(actual_tooltip_list)
        mobile_utillity.UtillityMethods.asequal(self,expected_tooltip_list,actual_tooltip_list,msg)
        time.sleep(1)   
    
    def select_minimized_popup_windows(self,index_num,**kwargs):
        '''
        :Param : expected_windows=['Filter Selection']
        :Param : msg='Step 03.5 : Verify minimized popup windows'
        :Usage : veridy_minimized_popup_windows(['Filter Selection'],'Step 03.5 : Verify minimized popup windows')
        '''
        window_elements=self.driver.find_elements_by_css_selector("#wall0 table[style*='border: 1px solid black'] td[class^='tabPagingText']")
        window_elements[index_num-1].click()
        time.sleep(1)
        if 'verify' in kwargs :
            self.verify_minimized_popup_windows(kwargs['verify'], kwargs['msg'])
    
    def verify_rollup_report_title(self,parent_id,expected_title,msg):
        '''
        :Param : Parent_id = 'ITableData1'
        :Param : expected_title='Unit Sales BY Product'
        :Usage : verify_rollup_report_title('ITableData1','Unit Sales BY Product','Step 07.2 : Verify rollup report title')
        '''
        parent_css="#"+parent_id+" tt>div"
        actual_title=self.driver.find_element_by_css_selector(parent_css).text.strip()
        #print(actual_title)
        mobile_utillity.UtillityMethods.asequal(self,actual_title,expected_title,msg)
    
    def compare_pivot_data_set(self, table_id, file_name):
        """
        :Param : table_id ='piv1'
        :Param : file_name= 'C2140681_Ds01.xlsx'
        :Usage : compare_pivot_data_set('piv1', 'C2140681_Ds01.xlsx')
        """
        os_platform=platform.system()
        if os_platform=='Windows' :
            wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
        else :
            wb1 = load_workbook(os.getcwd() + "/data/" + file_name)
        rows_css = "#" + table_id + " > tbody > tr:nth-child(2) > td >table > tbody > tr"
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        for r in range(len(table_rows)):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                if (s1.cell(row=r + 1, column=c + 1).value):
                    value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text.strip()
                    if s1.cell(row=r + 1, column=c + 1).value.strip() == str(value):
                        status=[0]
                        continue
                    else:
                        status=[r+1,c]
                        return (status)
        return (status)

    def create_pivot_data_set(self, table_id, file_name):
        """
        :Param : table_id ='piv1'
        :Param : file_name= 'C2140681_Ds01.xlsx'
        :Usage : create_pivot_data_set('piv1', 'C2140681_Ds01.xlsx')
        """
        rows_css = "#" + table_id + " > tbody > tr:nth-child(2) > td >table > tbody > tr"
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        for r in range(len(table_rows)):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text.strip()
                s.cell(row=r + 1, column=c + 1).value = str(value)
        os_platform=platform.system()
        if os_platform=='Windows' :
            wb.save(os.getcwd() + "\data\\" + file_name)
        else :
            wb.save(os.getcwd() + "/data/" + file_name)

    def verify_pivot_data_set(self, table_id, file_name, msg):
        """
        :Param : table_id ='piv1'
        :Param : file_name= 'C2140681_Ds01.xlsx'
        :Usage : verify_pivot_data_set('piv1', 'C2140681_Ds01.xlsx','Step 02 : Verify pivot table data')
        """
        x= self.compare_pivot_data_set(table_id, file_name)
        mobile_utillity.UtillityMethods.asequal(self,len(x),1,msg+ ' Row,Column -'+ str(x))
    
    def navigate_page(self, navigate_type,*args):
        '''
        Syntax: navigate_page('goto_page', 2)
        Syntax: navigate_page('next_page')
        Syntax: navigate_page('last_page')
        Syntax: navigate_page('previous_page')
        Syntax: navigate_page('first_page')
        '''
        btn_type={'goto_page':'Goto Page','next_page':'Next Page','Last Page':'Last Page','previous_page':'Previous Page','first_page':'First Page'}
        btn_css="table .arGridBar span[title='"+btn_type[navigate_type]+"']"
        self.driver.find_element_by_css_selector(btn_css).click()
        time.sleep(1)
        if navigate_type == 'goto_page':
            page_no=self.driver.find_element_by_id("formgoto")
            page_no.clear()
            page_no.send_keys(args[0])
            page_no.send_keys(Keys.ENTER)
            time.sleep(1)
    
    def verify_visualization(self, table_id, tr_id, colnum, color_code, msg,**kwargs):
        '''
        Syntax: verify_visualization('ITableData0', 'I0r', 5, 'green', 'Step 05: Verify visualization added')
        Syntax: verify_visualization('ITableData0', 'I0r', 5, 'green', 'Step 05: Verify visualization removed',visualization_removed=True)
        param : kwargs=visualization_removed=True (if you want to verify visualization bar removed from report)
        column number starts with 0
        '''
        expected_color_code=mobile_utillity.UtillityMethods.color_picker(self,color_code, 'rgba')
        occurence = 0
        total_rows=self.driver.find_elements_by_css_selector("[id='{0}'] tr[id ^='{1}']".format(table_id, tr_id))
        if 'visualization_removed' in kwargs :
            for i in range(0, len(total_rows)):
                no_of_visualization=self.driver.find_elements_by_css_selector("[id=" + table_id + "] tr[id ^= '" + tr_id + str(i) +"'] > td:nth-child(" + str(colnum+2) + ") > table table td")
                if len(no_of_visualization)==0:
                    occurence=0
                else :
                    occurence = occurence + 1
                    break 
            mobile_utillity.UtillityMethods.asequal(self,0,occurence, msg)
        else :
            for i in range(0, len(total_rows)):
                actual_color_code=self.driver.find_element_by_css_selector("[id=" + table_id + "] tr[id ^= '" + tr_id + str(i) +"'] > td:nth-child(" + str(colnum+2) + ") > table table td").value_of_css_property('background-color')
                reobj=re.match('.*\((\d+,\s\d+,\s\d+).*', actual_color_code)
                bg_color=reobj.group(1)
                if bg_color in expected_color_code:
                    occurence = occurence + 1 
                else:
                    break
            mobile_utillity.UtillityMethods.asequal(self, len(total_rows), occurence, msg)
    
    def create_lasso(self,parent_id,source_riser,target_riser,**kwargs):
        '''
        :param : parent_id='MAINTABLE_wbody0_f'
        :param : source_riser='riser!s0!g2!mbar!'
        :param : target_riser='riser!s0!g7!mbar!'
        :param : kwargs['verify']=['1 points', 'Filter Chart', 'Exclude from Chart']
        :param : kwargs['msg']='Step X : Verify lasso filters'
        :param : kwargs['select']='Filter Chart'
        :usage1: create_lasso('MAINTABLE_wbody0_f','riser!s0!g0!mbar!','riser!s0!g0!mbar!',verify=['1 points', 'Filter Chart', 'Exclude from Chart'],msg='Step 01 : ',select='Filter Chart')
        :usage2: create_lasso('MAINTABLE_wbody0_f','riser!s0!g0!mbar!','riser!s0!g0!mbar!',select='Filter Chart') 
        '''
        parent_css="#"+parent_id
        source_riser_css=parent_css+" [class*='" + source_riser + "']"
        target_riser_css=parent_css+" [class*='" + target_riser + "']"
        source=self.driver.find_element_by_css_selector(source_riser_css)
        target=self.driver.find_element_by_css_selector(target_riser_css)
        if self.test=='android_web' or self.test=='android_app' :
            action=ActionChains(self.driver)
            action.drag_and_drop(source, target).perform()
        else :
            source_offset=mobile_utillity.UtillityMethods.get_element_screen_coordinate(self,source,'middle',browser_height=True)
            target_offset=mobile_utillity.UtillityMethods.get_element_screen_coordinate(self,target,'bottom_middle',y_offset=-8,browser_height=True)
            drag_drop_params={'duration':2.0,'fromX':source_offset['x'],'fromY':source_offset['y'],'toX':target_offset['x'],'toY':target_offset['y']}
            self.driver.execute_script('mobile: dragFromToForDuration',drag_drop_params)
        time.sleep(2)
        lasso_filter_menus=self.driver.find_elements_by_css_selector("div[id^='ibi'][class*=tdgchart-tooltip] .tdgchart-tooltip-pad")
        menu_list=[menu.text.strip() for menu in lasso_filter_menus if menu.text.strip()!='']
        #print(menu_list)
        if 'verify' in kwargs:
            mobile_utillity.UtillityMethods.asequal(self,menu_list,kwargs['verify'],kwargs['msg'])
        if 'select' in kwargs:
            lasso_filter_menus[menu_list.index(kwargs['select'])].click()
        time.sleep(3)
    
    def select_drilldown_tooltip_menu(self,parent_id,raiser_class,menu,**kwargs):
        '''
        :param: parent_id=MAINTABLE_wbody0_f'
        :param: raiser_class='riser!s0!g3!mbar!'
        :param: menu='Filter'
        :usage: select_drilldown_tooltip_menu('iosTabs0_f','riser!s0!g3!mbar!','IBI')
        '''
        parent_css="#"+parent_id
        raiser_css=parent_css+" [class*='" + raiser_class + "']"
        tooltip_css=parent_css+" span[id*='tdgchart-tooltip']>div>ul>li[class*='tdgchart-tooltip-pad'] [onclick]"
        parent_element=self.driver.find_element_by_css_selector(parent_css)
        ActionChains(self.driver).move_to_element(parent_element).perform()
        raiser=self.driver.find_element_by_css_selector(raiser_css)
        if 'touch_tab' in kwargs and self.test=='ios_web' :
            utliobj=mobile_utillity.UtillityMethods(self.driver)
            utliobj.webview_touch_actions(raiser,'tap')
        else :
            ActionChains(self.driver).move_to_element(raiser).perform()
        tooltip=self.driver.find_elements_by_css_selector(tooltip_css)
        tooltip_menus=[item.text.strip() for item in tooltip]
        tooltip_element=tooltip[tooltip_menus.index(menu)]
        if 'touch_tab' in kwargs and self.test=='ios_web' :
            utliobj.webview_touch_actions(tooltip_element,'tap')
        else :
            tooltip_element.click()
        time.sleep(2)
    
    def verify_drilldown_navigation_menu(self,parent_id,expected_menu_list,msg):
        '''
        :Param : parent_id='MAINTABLE_wbody0_f'
        :Param : expected_menu_list=['Home', '->', 'Stereo Systems']
        :Usage : verify_drilldown_navigation_menu('MAINTABLE_wbody0_f',['Home', '->', 'Stereo Systems'],'Step 03.1 : Verify drilldown menu option')
        '''
        navigation_css="#"+parent_id+" foreignObject div>span"
        actual_menu_list=[]
        menu_elements=self.driver.find_elements_by_css_selector(navigation_css)
        for menu in menu_elements :
            if '\u2192' in menu.text.strip() :
                actual_menu_list.append('->')
            else :
                actual_menu_list.append(menu.text.strip())
        mobile_utillity.UtillityMethods.asequal(self,actual_menu_list,expected_menu_list,msg)
        
    def create_table_data_set(self,table_css,file_name, **kwargs):
        """    
        @param: table_css= ".arPivot tr:nth-child(1) table" or "table[summary='Summary']"   =  Need to provide the full parent path till table
        @param: file_name: "test1.xlsx" 
        Usage: create_table_data_set("table[summary='Summary']", "test1.xlsx" )
        """
        rows_css = table_css + " > tbody > tr"
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        for r in range(no_of_rows):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text
                s.cell(row=r + 1, column=c + 1).value = str(value)
        if platform.system()=='Windows' :
            wb.save(os.getcwd() + "\data\\" + file_name) 
        else :
            wb.save(os.getcwd() + "/data/" + file_name) 
    
    def verify_table_data_set(self,table_css,file_name,msg, **kwargs):
        """
        Usage: utillobj.verify_table_data_set("table[summary='Summary']", "test1.xlsx","Step 10: fail data set")
        """
        x= self.compare_table_data_set(table_css,file_name, **kwargs)
        mobile_utillity.UtillityMethods.asequal(self,len(x),1,msg+ ' Row,Column -'+ str(x))
        
    def compare_table_data_set(self, table_css, file_name, **kwargs):
        
        wb1 = load_workbook(os.getcwd() + "\data\\" + file_name) if platform.system()=='Windows' else load_workbook(os.getcwd() + "/data/" + file_name)
        status=[]
        s1 = wb1.get_sheet_by_name('Sheet')
        rows_css = table_css + " > tbody > tr"
        table_rows = self.driver.find_elements_by_css_selector(rows_css)
        if 'desired_no_of_rows' in kwargs:
            no_of_rows=kwargs['desired_no_of_rows']
        else:
            no_of_rows=len(table_rows)
        if 'starting_rownum' in kwargs:
            start_rownum=kwargs['starting_rownum']
        else:
            start_rownum=0
        for r in range(start_rownum, no_of_rows):
            col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
            columns = table_rows[r].find_elements_by_css_selector(col_css)
            for c in range(0, len(columns)):
                value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text
                if s1.cell(row=r + 1, column=c + 1).value != None and str(value) != '':
                    if s1.cell(row=r + 1, column=c + 1).value == str(value):
                        status=[0]
                        continue
                    else:
                        status=[r+1,c]
                        return (status)
        return (status)
    
    def verify_drilldown_autolink_property(self,table_css,row_index,column_index,msg,**kwargs):
        '''
        :params : table_css='#IWindowBodyFTB_0_0'
        :params : row_index=1 (index start from 1 )
        :params : column_index=1 (index start from 1 )
        :params : kwargs['total_link']=7,  kwargs['text_color']='cerulean_blue_2', kwargs['under_line']=True
        :usage : verify_drilldown_autolink_property('#IWindowBodyFTB_0_0',1,1,'Step 01 :',total_link=7,text_color='cerulean_blue_2',under_line=True)
        '''
        autolink_css=table_css+">tbody>tr>td:nth-child("+str(column_index)+") a[href]"
        table_index_css=table_css+">tbody>tr:nth-child("+str(row_index)+")>td:nth-child("+str(column_index)+") a[href]"
        table_index=self.driver.find_element_by_css_selector(table_index_css)
        if 'total_link' in kwargs :
            total_autolink=len(self.driver.find_elements_by_css_selector(autolink_css))
            mobile_utillity.UtillityMethods.asequal(self,total_autolink,kwargs['total_link'],msg+' Verify number of drilldown hyper links displayed')
        if 'text_color' in kwargs :
            expected_color=mobile_utillity.UtillityMethods.color_picker(self,kwargs['text_color'],'rgba')
            actual_text_color=Color.from_string(table_index.value_of_css_property("color")).rgba
            mobile_utillity.UtillityMethods.asequal(self,expected_color,actual_text_color,msg+' Verify drilldown hyper link color')
        if 'under_line' in kwargs :
            style=table_index.value_of_css_property("text-decoration")
            mobile_utillity.UtillityMethods.asequal(self,True,('underline' in style),msg+' Verify drilldown hyper links displayed with underline')
    
    def verify_report_drilldown_navigation_menu(self,parent_css,expected_menu_list,msg):
        '''
        :Param : parent_id='MAINTABLE_wbody0_f'
        :Param : expected_menu_list=['Home', '->', 'Stereo Systems']
        :Usage : drilldown_navigation_menu('MAINTABLE_wbody0_f',['Home', '->', 'Stereo Systems'],'Step 03.1 : Verify drilldown menu option')
        '''
        navigation_css=parent_css+" span"
        actual_menu_list=[]
        menu_elements=self.driver.find_elements_by_css_selector(navigation_css)
        for menu in menu_elements :
            if '\u2192' in menu.text.strip() :
                actual_menu_list.append('->')
            else :
                actual_menu_list.append(menu.text.strip())
        mobile_utillity.UtillityMethods.asequal(self,actual_menu_list,expected_menu_list,msg)
    
    def select_report_drilldown_menu(self,table_css,row_index,column_index,menu,**kwargs):
        '''
        :param : table_css='table[summary]'
        :param : row_index=3
        :param : column_index=1
        :Usage : select_report_drilldown_menu('table[summary]',4,1,'Drill down to Product Subcategory',verify=['Drill down to Product Subcategory'],msg='Step 09.1 : Verify drilldwon option')
        '''
        table_index_css=table_css+">tbody>tr:nth-child("+str(row_index)+")>td:nth-child("+str(column_index)+") a[href]"
        self.driver.find_element_by_css_selector(table_index_css).click()
        time.sleep(2)
        drilldown_css="#div_menu  span[id='my_tooltip_id']>div>ul[class*='tdgchart-tooltip-list']>li[class*='tdgchart-tooltip-pad']"
        drilldown_elements=self.driver.find_elements_by_css_selector(drilldown_css)
        drilldown_menus=[item.text.strip() for item in drilldown_elements]
        if 'verify' in kwargs :
            mobile_utillity.UtillityMethods.asequal(self,drilldown_menus,kwargs['verify'],kwargs['msg'])
        drilldown_elements=drilldown_elements[drilldown_menus.index(menu)].click()
        time.sleep(2)
    
    def select_table_field_menu_items(self,table_id,row_index,column_index,menu,**kwargs):
        '''
        :param : table_id='ITableData0'
        :param : row_index=0 (row_index start from 0)
        :param : column_index=1 (column_index start from 0)
        :param : menu=''Highlight Value
        :param : kwargs['verify']=['Comments','Highlight Value','Highlight Row','Unhighlight All', 'Filter Cell']
        :usage : select_table_field_menu_items('ITableData0',1,0,'Comments')
        '''
        table_cell_css="#"+table_id+" tr[id*='r" + str(row_index) + ".'] td[id$='C" + str(column_index) + "']"
        self.driver.find_element_by_css_selector(table_cell_css).click()
        time.sleep(2)
        table_index=table_id[-1]
        parent_css="div[id='dt"+str(table_index)+"_I0r"+str(row_index)+".0C"+str(column_index)+"_x__0'][style*='display: block']"
        table_css=parent_css+">table>tbody>tr"
        table_rows=self.driver.find_elements_by_css_selector(table_css)
        rows_text=[row.text.strip() for row in table_rows]
        table_row_element=table_rows[rows_text.index(menu)]
        if 'verify' in kwargs :
            actual_list=[value for value in rows_text if value!='']
            mobile_utillity.UtillityMethods.asequal(self,actual_list,kwargs['verify'],kwargs['msg'])
        if 'touch_tab' in kwargs and self.test=='ios_web' :
            utliobj=mobile_utillity.UtillityMethods(self.driver)
            utliobj.webview_touch_actions(table_row_element,'tap')
        else :
            table_row_element.click()
        time.sleep(2)