from selenium.common.exceptions import NoSuchElementException
from common.lib.javascript import JavaScript
from common.lib.utillity import UtillityMethods as utils
from common.lib.base import BasePage
import time
from common.pages import vfour_miscelaneous
from common.lib.core_utility import CoreUtillityMethods as core_utils
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver import ActionChains
from openpyxl import Workbook
from openpyxl import load_workbook
import os

class Security_Center(BasePage):
    """ Inherit attributes of the parent class = Baseclass """
 
    def __init__(self, driver):
        super(Security_Center, self).__init__(driver)
     
    def create_user_(self, uname, **kwargs):
        """
        Desc: This function will allow you to create user in Security Center. 
        uname:- 'basic'
        desc:- 'New basic user'
        email:- 'automation@ibi.com'
        pwd1:- 'something'
        pwd2:- 'something'
        group:- 'P292/developers'
        status:- 'active'
        btn:- 'create' OR 'OK OR 'cancel'
        @Author: Niranjan 
        """
        add_newuser_icon=self.driver.find_element_by_css_selector("#dlgSecurityManager #SecurityManagerDialog_btnNewUser")
        if self.browser.lower() == 'firefox':
            action1 = ActionChains(self.driver)
            action1.move_to_element(add_newuser_icon).click(add_newuser_icon).perform()
            del action1
        else:
            add_newuser_icon.click()
        utils.synchronize_with_visble_text(self, "#SecurityUserEditDialog_btnOK", "OK", expire_time=3)
        elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtID")
        utils.set_text_to_textbox_using_keybord(self, text_string=uname, text_box_elem=elem)
        if 'desc' in kwargs:
            elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtDesc")
            utils.set_text_to_textbox_using_keybord(self, text_string=kwargs['desc'], text_box_elem=elem)
        if 'email' in kwargs:
            elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtEmail")
            utils.set_text_to_textbox_using_keybord(self, text_string=kwargs['email'], text_box_elem=elem)
        if 'pwd' in kwargs:
            elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtP1")
            utils.set_text_to_textbox_using_keybord(self, text_string=kwargs['pwd'], text_box_elem=elem)
            elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtP2")
            utils.set_text_to_textbox_using_keybord(self, text_string=kwargs['pwd'], text_box_elem=elem)
        if 'group' in kwargs:
            combobox_btn_css="#dlgSecurityUserEdit #SecurityUserEditDialog_cbDefGroup div[id^='BiButton']"
            utils.scroll_and_select_combobox_item_(self, combobox_btn_css, kwargs['group'])
        if 'status' in kwargs:
            combo_btn_elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_cbDefGroup div[id^='BiButton']")
            utils.select_combo_box_item(self, kwargs['status'], combobox_dropdown_elem=combo_btn_elem)
        if 'btn' in kwargs:
            btn_id = 'Apply' if kwargs['btn'] == 'create' else 'OK' if kwargs['btn'] == 'ok' else 'Cancel'
            btn_css="#dlgSecurityUserEdit #SecurityUserEditDialog_btn" + btn_id
            ok_btn_obj=self.driver.find_element_by_css_selector(btn_css)
            core_utils.left_click(self, ok_btn_obj)
        try:
            warning_dialog_text_css = "#dlgError div[class*='active'] #ErrorDialog_msgText"
            warning_text = self.driver.find_element_by_css_selector(warning_dialog_text_css).text.strip()
            raise FileExistsError(warning_text)
        except NoSuchElementException:
            '''
            No message required
            '''
    
    def close_New_User_dialog(self):  
        """
        Desc: This function will close New USer dialog in Security Centre after user creation
        """      
        btn_css="#dlgSecurityUserEdit [class*='active'] [class*='window-close-button']"
        self.driver.find_element_by_css_selector(btn_css).click()
        vfour_miscelaneous.Vfour_Miscelaneous.synchronize_until_element_disappear(self, btn_css, 0, 90)
        time.sleep(2)
    
    def close_security_center_window(self):   
        """
        Desc: This function will close Security centre window, this is applicable for new Home Page Security Center which opens as window
        """     
        btn_css="#dlgSecurityManager [class*='active'][class*='window'] #SecurityManagerDialog_btnExit"
        self.driver.find_element_by_css_selector(btn_css).click()
        time.sleep(2)
        
    def select_group_role_(self, group_path, select_type='left'):
        """
        Desc: This method used to expand and select specific group in group list window
        example usage : select_group_role_("S10141->GroupAdmins")
        """
        group_table_css="#SecurityManagerDialog_treeGroups table[class='bi-tree-view-table']>tbody>tr>td:nth-child(1)"
        Security_Center.expand_group_section_(self, group_path)
        group_rows_obj=self.driver.find_elements_by_css_selector(group_table_css)
        group_role=group_path.split('->')[-1]
        group_item_obj=JavaScript.find_elements_by_text(self, group_rows_obj, group_role)
        if select_type == 'left':
            core_utils.python_left_click(self, group_item_obj[0])
        elif select_type == 'right':
            core_utils.right_click(self, group_item_obj[0])
            
    def get_group_table_row_elements(self):
        '''
        This function is used to get the list groups in group name
        '''
        group_table_row_css="#SecurityManagerDialog_treeGroups table[class='bi-tree-view-table']>tbody>tr"
        group_table_row_description='group table rows in security center window'
        group_table_row_elements=utils.validate_and_get_webdriver_objects(self, group_table_row_css, group_table_row_description)
        return group_table_row_elements

    def get_group_table_row_element(self, group_name):
        '''
        This function is used to verify the group which is available in the groups section
        '''
        group_table_row_elements=Security_Center.get_group_table_row_elements(self) 
        for group_table_row_element in group_table_row_elements:
            if group_name in group_table_row_element.text:
                return group_name
        error_msg='The requested Group name "' + group_name + '" is not available under Groups section in security center window.'
        raise ValueError(error_msg)

    def select_group_context_menu(self, context_menu_path):
        """
        Desc : This method used to select group context menu
        """
        select_item_list=context_menu_path.split('->')
        for select_item in select_item_list:
            utils.select_or_verify_bipop_menu(self, select_item)

    def verify_group_context_menu(self, context_menu_path, expected_popup_list, msg):
        """
        This function is used to verify group context menu.
        context_menu_path='Security->Rules'
        """ 
        utils.select_or_verify_bipop_menu(self, verify='true', expected_popup_list=expected_popup_list, msg=msg)
        
    def verify_user_in_group_(self, group_path, username_list, msg, **kwargs):
        """
        Desc: This will verify user in group
        usage: verify_user_in_group_('Retail_Samples->BasicUsers', 'Basic_user', "Step04: Verify")
        """
        Security_Center.select_group_role_(self, group_path, **kwargs)
        time.sleep(2)
        Security_Center.verify_user_(self, username_list, msg,**kwargs)
        time.sleep(1)
        parent_section=group_path.split('->')[0]
        Security_Center.collapse_group_section_(self, parent_section)
        if 'close_btn' in kwargs:
            Security_Center.close_security_center_window(self)
            
    def verify_user_(self, username_list, msg, **kwargs):
        """
        username:- 'autodevuser100'
        """
        added_user_scrollable_css="#SecurityManagerDialog_treeUsersInGroup div[class='bi-tree-view-body']"
        JavaScript.scroll_element(self, added_user_scrollable_css, 0, wait_time=2)
        USER_TABLE_ROWS_CSS="#dlgSecurityManager #SecurityManagerDialog_treeUsersInGroup .bi-tree-view-body-content table td:nth-child(1)"
        first_row=self.driver.find_elements_by_css_selector(USER_TABLE_ROWS_CSS)[0]
        core_utils.python_left_click(self, first_row)
#         utils.click_on_screen(self, first_row, 'middle', click_type=0)
        for username in username_list :
            utils.scroll_down_and_find_item_using_mouse(self, USER_TABLE_ROWS_CSS, username)
            user_rows_obj=self.driver.find_elements_by_css_selector(USER_TABLE_ROWS_CSS)
            actual_user_list=JavaScript.get_elements_text(self, user_rows_obj)
            utils.asin(self, username, actual_user_list, str(msg)+" Verify " + username + " is added.")
    
    def expand_group_section_(self, group_path):
        """
        example usage : select_group_role("P116->GroupAdmins")
        """
        Security_Center.__expand_or_collapse_group_(self, group_path, 'EXPAND')
        
    def collapse_group_section_(self, group_path):
        """
        example usage : collapse_group_section_("P116")
        """
        Security_Center.__expand_or_collapse_group_(self, group_path, 'COLLAPSE') 
    
    def __expand_or_collapse_group_(self, group_path, target):
        """
        This method used to expand and select specific group in group list window
        example usage : __expand_or_collapse_group_("P116->GroupAdmins")
        """
        if target.upper() == 'EXPAND' :
            expand_icon="img[src*='view-plus']"
        if target.upper() == 'COLLAPSE' :
            expand_icon="img[src*='view-minus']"
        group_table_css="#SecurityManagerDialog_treeGroups table[class='bi-tree-view-table']>tbody>tr>td:nth-child(1)"
        group_scrollable_css="#SecurityManagerDialog_treeGroups div[class='bi-tree-view-body']"
        JavaScript.scroll_element(self, group_scrollable_css, 0, wait_time=2)
        group_list=group_path.split('->')
        first_row=self.driver.find_elements_by_css_selector(group_table_css)[1]
        core_utils.python_left_click(self, first_row)
        time.sleep(2)
        for group in group_list :
            utils.scroll_down_and_find_item_using_mouse(self, group_table_css, group, pause_time=1)
            group_rows_obj=self.driver.find_elements_by_css_selector(group_table_css)
            group_item_index=JavaScript.find_element_index_by_text(self, group_rows_obj, group)
            if group_item_index == None :
                ERROR_MSG="[{0}] Group not found in group list".format(group)
                raise NoSuchElementException(ERROR_MSG)
            else :
                expand_icon_obj=group_rows_obj[group_item_index].find_elements_by_css_selector(expand_icon)
                if len(expand_icon_obj)>0 :
                    core_utils.python_move_to_element(self, expand_icon_obj[0])
                    time.sleep(2)
                    core_utils.python_left_click(self, expand_icon_obj[0])
    
    
        
    '''**********************************Old Function*************************************'''
     
    def create_user(self, uname, **kwargs):
        """
        uname:- 'something'
        desc:- 'something'
        email:- 'something'
        pwd1:- 'something'
        pwd2:- 'something'
        group:- 'P292/developers'
        status:- 'active'
        btn:- 'create' OR 'OK OR 'cancel'
        @Author: Niranjan 
        """
        add_newuser_icon=self.driver.find_element_by_css_selector("#dlgSecurityManager #SecurityManagerDialog_btnNewUser")
        utils.default_click(self, add_newuser_icon)
        utils.synchronize_with_visble_text(self, "#SecurityUserEditDialog_btnOK", "OK", expire_time=3)
        elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtID")
        utils.set_text_field_using_actionchains(self, elem, uname)
        if 'desc' in kwargs:
            self.driver.execute_script()
            elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtDesc")
            utils.set_text_field_using_actionchains(self, elem, kwargs['desc'])
        if 'email' in kwargs:
            elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtEmail")
            utils.set_text_field_using_actionchains(self, elem, kwargs['email'])
        if 'pwd' in kwargs:
            elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtP1")
            utils.set_text_field_using_actionchains(self, elem, kwargs['pwd'])
            elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_txtP2")
            utils.set_text_field_using_actionchains(self, elem, kwargs['pwd'])
        if 'group' in kwargs:
            combobox_btn_css="#dlgSecurityUserEdit #SecurityUserEditDialog_cbDefGroup div[id^='BiButton']"
            utils.scroll_and_select_combobox_item(self, combobox_btn_css, kwargs['group'])
        if 'status' in kwargs:
            combo_btn_elem=self.driver.find_element_by_css_selector("#dlgSecurityUserEdit #SecurityUserEditDialog_cbDefGroup div[id^='BiButton']")
            utils.select_any_combobox_item(self, combo_btn_elem, kwargs['status'])
        if 'btn' in kwargs:
            btn_id = 'Apply' if kwargs['btn'] == 'create' else 'OK' if kwargs['btn'] == 'ok' else 'Cancel'
            btn_css="#dlgSecurityUserEdit #SecurityUserEditDialog_btn" + btn_id
            ok_btn_obj=self.driver.find_element_by_css_selector(btn_css)
            utils.default_click(self, ok_btn_obj)
    
    def delete_user(self, uname_list, option='Yes', Step_num_msg='X'):
        """
        :param uname_list=['basicuser', 'advanceduser', 'developeruser', 'groupadminuser']
        :usage delete_user(['basicuser'])
        @author: Aftab(AA14564)
        """
        dialog_css="#dlgSecurityManager [class*='active']"
        delete_user_icon_css=dialog_css+" #SecurityManagerDialog_btnDeleteUser img"
        selected_user_row_css=dialog_css+" #SecurityManagerDialog_treeUsers table[class='bi-tree-view-table']>tbody>tr.selected"
        delete_confirm_dialog_css="[id*='BiDialog'] [class*='active']"
        count=1
        try:
            for uname in uname_list:
                Security_Center.select_user(self, uname)
                utils.synchronize_with_number_of_element(self, selected_user_row_css, 1, 90)
                delete_user_icon_elem = self.driver.find_element_by_css_selector(delete_user_icon_css)
                utils.default_click(self, delete_user_icon_elem)
                utils.synchronize_with_number_of_element(self, delete_confirm_dialog_css, 1, 90)
                utils.click_dialog_button(self, delete_confirm_dialog_css, option)
                vfour_miscelaneous.Vfour_Miscelaneous.synchronize_until_element_disappear(self, delete_confirm_dialog_css, 0, 90)
                utils.asequal(self, True, True, 'Step '+str(Step_num_msg)+'.'+str(count)+': User name ' +str(uname)+ ' removed.')
                count+=1
        except:
            utils.asequal(self, uname, False, 'Step '+str(Step_num_msg)+':Error to list of user name')
            
    def verify_user(self, username_list, msg, **kwargs):
        """
        username:- 'autodevuser100'
        """
        added_user_scrollable_css="#SecurityManagerDialog_treeUsersInGroup div[class='bi-tree-view-body']"
        JavaScript.scroll_element(self, added_user_scrollable_css, 0, wait_time=2)
        USER_TABLE_ROWS_CSS="#dlgSecurityManager #SecurityManagerDialog_treeUsersInGroup .bi-tree-view-body-content table td:nth-child(1)"
        first_row=self.driver.find_elements_by_css_selector(USER_TABLE_ROWS_CSS)[0]
        utils.click_on_screen(self, first_row, 'middle', click_type=0)
        for username in username_list :
            utils.scroll_down_and_find_item_using_mouse(self, USER_TABLE_ROWS_CSS, username, number_of_times=3)
            user_rows_obj=self.driver.find_elements_by_css_selector(USER_TABLE_ROWS_CSS)
            actual_user_list=JavaScript.get_elements_text(self, user_rows_obj)
            utils.asin(self, username, actual_user_list, str(msg)+" Verify " + username + " is added.")
        
    def verify_user_in_group(self, group_path, username_list, msg, **kwargs):
        """
        """
        Security_Center.select_group_role(self, group_path, **kwargs)
        time.sleep(2)
        Security_Center.verify_user(self, username_list, msg,**kwargs)
        time.sleep(1)
        parent_section=group_path.split('->')[0]
        Security_Center.collapse_group_section(self, parent_section)
        if 'close_btn' in kwargs:
            Security_Center.close_security_center_dialog(self)
        
    def add_user_to_group(self, username, group_path, username_list, msg, collapse_group=True, **kwargs):
        """
        This method used to assign user to group and verify user is added 
        usage : add_user_to_group('public', 'S10141->GroupAdmins', ['wfpenbas1'], 'Step 01')
        """
        Security_Center.select_user(self, username)
        Security_Center.select_group_role(self, group_path)
        add_user_obj=self.driver.find_element_by_css_selector("#dlgSecurityManager #SecurityManagerDialog_btnAddUser")
        utils.default_click(self, add_user_obj)
        Security_Center.verify_user(self, username_list, msg, **kwargs)
        parent_section=group_path.split('->')[0]
        collapse_group and Security_Center.collapse_group_section(self, parent_section)
        if 'close_btn' in kwargs:
            Security_Center.close_security_center_dialog(self)
        
    def close_security_center_dialog(self):        
        btn_css="#dlgSecurityManager [class*='active'][class*='window'] #SecurityManagerDialog_btnExit"
        self.driver.find_element_by_css_selector(btn_css).click()
        vfour_miscelaneous.Vfour_Miscelaneous.synchronize_until_element_disappear(self, btn_css, 0, 90)
        time.sleep(2)
        
    def select_security_rules_dialog_buttons(self, btn_name): 
        '''
        btn_name='OK' OR 'Apply' OR 'Cancel'
        '''       
        btn_css="#dlgSecurityResource [class*='active'][class*='window'] #SecurityResourceDialog_btn"+btn_name
        self.driver.find_element_by_css_selector(btn_css).click()
        time.sleep(2)
    
    def select_user(self, username):
        """
        This method used to select username in user list window
        example usage : select_user('admin')
        """
        user_scrollable_css="#SecurityManagerDialog_treeUsers div[class='bi-tree-view-body']"
        JavaScript.scroll_element(self, user_scrollable_css, 0, wait_time=2)
        USER_TABLE_ROWS_CSS="#SecurityManagerDialog_treeUsers table[class='bi-tree-view-table']>tbody>tr>td:nth-child(1)"
        ERROR_MSG="[{0}] user not found in user list".format(username)
        first_row=self.driver.find_elements_by_css_selector(USER_TABLE_ROWS_CSS)[1]
        utils.click_on_screen(self, first_row, 'middle', click_type=0)
        utils.scroll_down_and_find_item_using_mouse(self, USER_TABLE_ROWS_CSS, username, number_of_times=10)
        user_table_rows_obj=self.driver.find_elements_by_css_selector(USER_TABLE_ROWS_CSS)
        user_name_obj=JavaScript.find_elements_by_text(self, user_table_rows_obj, username)
        if len(user_name_obj)==0 :
            raise NoSuchElementException(ERROR_MSG)
        else :
            utils.default_click(self, user_name_obj[0])
            time.sleep(1)
            
    def __expand_or_collapse_group(self, group_path, target):
        """
        This method used to expand and select specific group in group list window
        example usage : select_group_role("P116->GroupAdmins")
        """
        if target.upper() == 'EXPAND' :
            expand_icon="img[src*='view-plus']"
        if target.upper() == 'COLLAPSE' :
            expand_icon="img[src*='view-minus']"
        group_table_css="#SecurityManagerDialog_treeGroups table[class='bi-tree-view-table']>tbody>tr>td:nth-child(1)"
        group_scrollable_css="#SecurityManagerDialog_treeGroups div[class='bi-tree-view-body']"
        JavaScript.scroll_element(self, group_scrollable_css, 0, wait_time=2)
        group_list=group_path.split('->')
        first_row=self.driver.find_elements_by_css_selector(group_table_css)[1]
        utils.click_on_screen(self, first_row, 'middle', click_type=0)
        for group in group_list :
            utils.scroll_down_and_find_item_using_mouse(self, group_table_css, group, pause_time=1, number_of_times=5)
            group_rows_obj=self.driver.find_elements_by_css_selector(group_table_css)
            group_item_obj=JavaScript.find_elements_by_text(self, group_rows_obj, group)
            if len(group_item_obj)==0 :
                ERROR_MSG="[{0}] Group not found in group list".format(group)
                raise NoSuchElementException(ERROR_MSG)
            else :
                expand_icon_obj=group_item_obj[0].find_elements_by_css_selector(expand_icon)
                if len(expand_icon_obj)>0 :
                    utils.default_click(self, expand_icon_obj[0])
    
    def expand_group_section(self, group_path):
        """
        Example usage : select_group_role("P116->GroupAdmins")
        """
        Security_Center.__expand_or_collapse_group(self, group_path, 'EXPAND')
        
    def collapse_group_section(self, group_path):
        """
        example usage : select_group_role("P116")
        """
        Security_Center.__expand_or_collapse_group(self, group_path, 'COLLAPSE') 
    
    def select_group_role(self, group_path):
        """
        This method used to expand and select specific group in group list window
        example usage : select_group_role("S10141->GroupAdmins")
        """
        group_table_css="#SecurityManagerDialog_treeGroups table[class='bi-tree-view-table']>tbody>tr>td:nth-child(1)"
        Security_Center.expand_group_section(self, group_path)
        group_rows_obj=self.driver.find_elements_by_css_selector(group_table_css)
        group_role=group_path.split('->')[-1]
        group_item_obj=JavaScript.find_elements_by_text(self, group_rows_obj, group_role)
        utils.default_click(self, group_item_obj[0])
    
    def delete_group(self, group_name, option='Yes', Step_num_msg='X'):
        '''
        :param group_name='P292_S10117_G169813'
        :usage delete_user('P292_S10117_G169813')
        @author: Aftab(AA14564)
        '''
        dialog_css="#dlgSecurityManager [class*='active']"
        delete_group_icon_css=dialog_css+" #SecurityManagerDialog_btnDeleteGroup img"
        selected_group_row_css=dialog_css+" #SecurityManagerDialog_treeGroups table[class='bi-tree-view-table']>tbody>tr.selected"
        delete_confirm_dialog_css="[id*='BiDialog'] [class*='active']"
        try:
            Security_Center.select_group_role(self, group_name)
            utils.synchronize_with_number_of_element(self, selected_group_row_css, 1, 90)
            delete_group_icon_elem = self.driver.find_element_by_css_selector(delete_group_icon_css)
            utils.default_click(self, delete_group_icon_elem)
            utils.synchronize_with_number_of_element(self, delete_confirm_dialog_css, 1, 90)
            utils.click_dialog_button(self, delete_confirm_dialog_css, option)
            vfour_miscelaneous.Vfour_Miscelaneous.synchronize_until_element_disappear(self, delete_confirm_dialog_css, 0, 90)
            utils.asequal(self, True, True, 'Step '+str(Step_num_msg)+': User name ' +str(group_name)+ ' removed.')
        except:
            utils.asequal(self, str(group_name), False, 'Step '+str(Step_num_msg)+':Error to get '+str(group_name)+' group_name name.')
        
        
class Security_Rules_Dialog(BasePage):
    
    def __init__(self, driver):
        super(Security_Rules_Dialog, self).__init__(driver)
        
    def close_dialog(self):
        pass

class Rules_For_This_Group_Dialog(BasePage):
    
    dialog_css='#resourceRulesDlg'
    dialog_name = 'Rules for this Group'
    
    def __init__(self, driver):
        super(Security_Rules_Dialog, self).__init__(driver)
    
    def verify_caption(self, expected_title, msg):
        """
        This function is used to verify caption text is present in the rules dialog
        verify_caption_in_rules_for_this_group_dialog('Rules for this Group- GroupAdmins)
        """
        caption_css = Rules_For_This_Group_Dialog.dialog_css + ' .window-caption'
        caption_description=Rules_For_This_Group_Dialog.dialog_name + ' Caption'
        caption_title_obj=utils.validate_and_get_webdriver_object(self, caption_css, caption_description)
        caption_title = caption_title_obj.text.strip()
        utils.asin(self, expected_title, caption_title, msg)
    
    def create_grid_data(self, file_name):
        """
        This function is used to verify rules for groups dialog grid data
        verify_grid_data("Test1.xlsx")
        """
        wb = Workbook()
        s = wb.get_sheet_by_name('Sheet')
        grid_body_css = Rules_For_This_Group_Dialog.dialog_css + ' .bi-grid-body'
        grid_row_css = grid_body_css + ' .grid-row'
        grid_cell_css = grid_body_css + ' .grid-cell'
        total_rows=self.driver.find_elements_by_css_selector(grid_row_css)
        total_cols=self.driver.find_elements_by_css_selector(grid_cell_css)
        for r in range(0, len(grid_row_css)[-1]):
            col_val=grid_row_css+":nth-child(" +str(r+1) + ") > .grid-cell"
            print(col_val)
            columns = total_rows[r].find_elements_by_css_selector(".grid-cell")
            print(columns)
            for c in range(len(total_cols)):
                value=self.driver.find_element_by_css_selector(grid_cell_css + ":nth-child(" + str(c+1) + ")").text
                print(value)
                s.cell(row=r + 1, column=c + 1).value = str(value).replace(' ','').strip()
        wb.save(os.getcwd() + "\data\\" + file_name)
         
    def select_buttons(self, button_name):
        """
        This function is used to select any button preesent in the rules for this group dialog.
        select_button('close')
        """
        button_name_css_dict={'help' : ' #rorHelpBtn',
                     'create_report' : ' #btnReport',
                     'close' : ' #btnOK'}
        button_css = Rules_For_This_Group_Dialog.dialog_css + button_name_css_dict[button_name.lower()]
        button_description=button_name + ' button in ' + Rules_For_This_Group_Dialog.dialog_name + ' dialog'
        caption_title_obj=utils.validate_and_get_webdriver_object(self, button_css, button_description)
        core_utils.left_click(self, caption_title_obj)
        