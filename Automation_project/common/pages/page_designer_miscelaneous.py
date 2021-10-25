from common.locators.loginpage_locators import LoginPageLocators
from common.locators import page_designer_design as DesignLocators
from common.locators.page_designer_locators import PageDesigner as PD_LOCATORS
from common.lib.core_utility import CoreUtillityMethods as coreutils
from common.lib.utillity import UtillityMethods as utils
from common.lib.javascript import JavaScript as javascript
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementNotVisibleException
from selenium.webdriver.support.color import Color
from selenium.webdriver.common.by import By
from common.lib.base import BasePage
from openpyxl import Workbook
from openpyxl import load_workbook
import time, os
from common.lib.global_variables import Global_variables
from common.pages.wf_mainpage import Wf_Mainpage

WORKBOOK_SHEET_NAME='Sheet'
EXCEL_FILE_EXTENSION='.xlsx'
SCRIPT_DATA_FOLDER_NAME='data'

class PageDesignerMiscelaneous(BasePage):
    
    """ Inherit attributes of the parent class = Baseclass """

    def __init__(self, driver):
        super(PageDesignerMiscelaneous, self).__init__(driver)
     
    def invoke_webfocus(self):
        """
        Descriptions : This method used to invoke given setup wf url 
        example url : http://unxrh7:25164/ibi_apps
        """
        setup_url=utils.get_setup_url(self)
        self.driver.get(setup_url)
    
    def login_webfocus(self, mrid=None, mrpass=None):
        """
        Descriptions : This method used to enter the valid webfocus user id and password to login webfocus
        """
        loginid = coreutils.parseinitfile(self, 'mrid') if mrid==None else mrid
        loginpwd = coreutils.parseinitfile(self, 'mrpass') if mrpass==None else mrpass
        uname1=(By.CSS_SELECTOR, 'input[id=SignonUserName]')
        self.longwait.until(EC.visibility_of_element_located(uname1))
        usename= self.driver.find_element(*LoginPageLocators.uname)
        usename.click()
        time.sleep(2)
        usename.send_keys(loginid)
        time.sleep(1)
        if loginpwd!=None:
            passwd =self.driver.find_element(*LoginPageLocators.pword)
            passwd.send_keys(loginpwd)
            time.sleep(1)
        coreutils.update_current_working_area_browser_specification(self)
        sign_in =self.driver.find_element(*LoginPageLocators.submit)
        sign_in.click()
        time.sleep(4)
    
    def get_page_designer_folder_path(self):
        """
        Descriptions : This method used to read project_id, suite_id,group_id and create folder path format to navigate pd folder
        It will return as 'P292_S10660->G433312 '
        """
        project_id=coreutils.parseinitfile(self, 'project_id')
        suite_id=coreutils.parseinitfile(self, 'suite_id')
        group_id=coreutils.parseinitfile(self, 'group_id')
        folder_path=project_id + '_' + suite_id + '->' + group_id
        return folder_path
    
    def get_page_designer_reference_folder_path(self):
        """
        Descriptions : This method used to read project_id, suite_id,group_id, reference folder and create folder path format to navigate 
        It will return as 'P292_S10660->G433312->Reference Items'
        """
        pd_foder=PageDesignerMiscelaneous.get_page_designer_folder_path(self)
        reference_folder=coreutils.parseinitfile(self, 'reference_folder')
        reference_folder_path=pd_foder + '->' + reference_folder if reference_folder!='' else pd_foder
        return reference_folder_path
    
    def invoke_and_login_webfocus(self, mrid=None, mrpass=None):
        """
        Descriptions : This method used invoke given webfocus setup and enter the valid user id and password to login
        """
        PageDesignerMiscelaneous.invoke_webfocus(self)
        PageDesignerMiscelaneous.login_webfocus(self, mrid, mrpass)
        try :
            user_icon_css=(By.CSS_SELECTOR, "div[class*='menu-btn']>div>div[class*='ibx-glyph-user']")
            self.longwait.until(EC.visibility_of_element_located(user_icon_css))
        except :
            pass
    
    def expand_pd_content_folder(self, folder_path, left_click=True):
        """
        Descriptions : This method used to expand the content folder by given folder path order
        example usage1 : expand_pd_content_folder_and_select_item('Retail Samples->Reports->Auto Link Targets')
        """
        folder_list = folder_path.split('->') 
        for folder in folder_list :
            folder_obj = PageDesignerMiscelaneous.find_pd_content_item_and_scroll_into_view(self, folder)
            expand_icon_obj = folder_obj.find_element_by_css_selector(DesignLocators.ContentTab.CONTENT_EXPAND_ICON_CSS)
            if left_click == True:
                coreutils.left_click(self, expand_icon_obj)
            else:
                coreutils.python_doubble_click(self, expand_icon_obj)
        
    def find_pd_content_item_and_scroll_into_view(self, item_name):
        """
        Descriptions :  This method used to find content item and scrollinto visible area.
        example usage : find_pd_content_item_and_scroll_into_view('P292')
        """
        expanded_content_obj =  self.driver.find_element_by_css_selector(DesignLocators.ContentTab.EXPANDED_CONTENT_CSS)
        expanded_content_label_obj = self.driver.find_element_by_css_selector(DesignLocators.ContentTab.EXPANDED_CONTENT_LABEL_CSS)
        expanded_content_bottom_y = int(coreutils.get_web_element_coordinate(self, expanded_content_obj, 'bottom_middle')['y'])
        expanded_content_top_y = int(coreutils.get_web_element_coordinate(self, expanded_content_label_obj, 'bottom_middle')['y'])
        item_elements = utils.validate_and_get_webdriver_objects(self, DesignLocators.ContentTab.EXPANDED_CONTENT_ITEMS_CSS, 'DesignLocators content items')
        item_index = javascript.find_element_index_by_text(self, item_elements, item_name)
        max_scroll = len(item_elements)
        scroll_count = 0
        if item_index != None :
            coreutils.python_move_to_element(self, expanded_content_obj)
            item_obj = item_elements[item_index]    
            while True :
                item_bottom_y = int(coreutils.get_web_element_coordinate(self, item_obj, 'bottom_middle')['y'])
                item_top_y = int(coreutils.get_web_element_coordinate(self, item_obj, 'top_middle')['y'])
                if expanded_content_bottom_y < item_bottom_y and scroll_count < max_scroll:
                    utils.mouse_scroll(self, 'down', '1', option='uiautomation')
                    scroll_count+=1
                elif item_top_y < expanded_content_top_y and scroll_count < max_scroll:
                    utils.mouse_scroll(self, 'up', '1', option='uiautomation')
                    scroll_count+=1
                else : 
                    break
            return item_obj
        else :
            error_msg = "{0} item not found in page designer content".format(item_name)
            raise FileNotFoundError(error_msg)
        
    def expand_domain_folders(self, folder_path):
        """
        Descriptions : This method used to expand home page domains folder 
        example usage : expand_domain_folders('Retail Samples->Reports->Auto Link Targets')
        """
        Wf_Mainpage.expand_repository_folders(self, folder_path)
    
    def click_on_page_designer_icon(self, from_designer_group=False):
        """
        Descriptions : This method used to click on Page icon in home page to create new page designer 
        """
        is_expanded=self.driver.find_element_by_css_selector("div[data-ibxp-glyph='keyboard_arrow_up']").is_displayed()
        if is_expanded == False :
            exapnd=self.driver.find_element_by_css_selector("div[data-ibxp-glyph='keyboard_arrow_down']")
            coreutils.left_click(self, exapnd)
        if from_designer_group :
            Wf_Mainpage.select_action_bar_tab(self, 'Designer')
            Wf_Mainpage.select_action_bar_tab_option(self, 'Page')
        else :
            icon_css="div[class*='create-new-item'][data-ibxp-text='Page']"
            utils.synchronize_with_visble_text(self, icon_css, visble_element_text='Page', expire_time=10)
            icon_obj=self.driver.find_element_by_css_selector(icon_css)
            coreutils.left_click(self, icon_obj)
    
    def select_homepage_context_menu_item(self, context_menu_path):
        """
        Descriptions : This method used to select home page context menu 
        """
        Wf_Mainpage.select_context_menu_item(self, context_menu_path)
    
    def __get_domain_folder_item(self, folder_path, item_name):
        """
        Descriptions : This method used to find domain folder items and get it object.
        example usage :  __get_domain_folder_item('Retail Samples->Reports->Auto Link Targets', 'Act_fex')
        """
        PageDesignerMiscelaneous.expand_domain_folders(self, folder_path)
        time.sleep(2)
        item_obj = Wf_Mainpage.get_domain_folder_item(self, item_name)
        return item_obj
        
    def select_domain_folder_item_context_menu(self, folder_path, item_name, context_menu_path):
        """
        Descriptions : This method select domain folder item context menu
        example usage : select_domain_folder_item_context_menu('Retail Samples->Reports->Auto Link Targets', 'Act_fex', 'Delete')
        """
        item_obj=PageDesignerMiscelaneous.__get_domain_folder_item(self, folder_path, item_name)
        coreutils.python_right_click(self, item_obj)
        PageDesignerMiscelaneous.select_homepage_context_menu_item(self, context_menu_path)
        
    def run_page_designer_by_double_click(self, folder_path, item_name):
        """
        Descriptions : This method used to run saved page designer file by double click
        example usage : run_page_designer_by_double_click('Retail Samples->Reports->Auto Link Targets', 'Act_fex')
        """
        folder_path=PageDesignerMiscelaneous.get_page_designer_folder_path(self) if folder_path==None else folder_path
        item_obj=PageDesignerMiscelaneous.__get_domain_folder_item(self, folder_path, item_name)
        coreutils.double_click(self, item_obj)
        
    def delete_saved_page_designer(self, folder_path, saved_page_name):
        """
        Descriptions : This method used to delete the saved page designer from domains folder 
        example usage : delete_saved_page_designer('Retail Samples->Reports->Auto Link Targets', 'Act_fex')
        """
        delete_button_css="div[class*='ibx-dialog-ok-button']"
        PageDesignerMiscelaneous.select_domain_folder_item_context_menu(self, folder_path, saved_page_name, 'Delete')
        utils.synchronize_with_visble_text(self, delete_button_css, visble_element_text='OK', expire_time=5)
        delete_button=self.driver.find_element_by_css_selector(delete_button_css)
        coreutils.left_click(self, delete_button)
    
    def invoke_page_designer(self, folder_path, mrid=None, mrpass=None, from_designer_group=False):
        """
        Descriptions : This method used to invoke page designer page from given folder path
        example usage : invoke_page_designer('P116->S7074')
        example usage1 : invoke_page_designer('P116->S7074', 'admin', 'admin')
        """
        PageDesignerMiscelaneous.invoke_and_login_webfocus(self, mrid, mrpass)
        Wf_Mainpage.select_left_panel(self, 'content')
        PageDesignerMiscelaneous.expand_domain_folders(self, folder_path)
        PageDesignerMiscelaneous.click_on_page_designer_icon(self, from_designer_group)
        coreutils.switch_to_new_window(self)
        coreutils.update_current_working_area_browser_specification(self)
        
    def select_page_designer_template(self, template_name):
        """
        Descriptions : This method used to select the page designer template
        template names : 'Blank', 'Grid 2-1', 'Grid 4-2-1', 'Grid 3-3-3'
        example usage : select_page_designer_template('Blank')
        """
        error_message="[{0}] template not list out to select".format(template_name)
        utils.synchronize_with_visble_text(self, DesignLocators.NewPagePopModal.TEMPLATES_CSS, visble_element_text='Blank', expire_time=10)
        template_obj_list=self.driver.find_elements_by_css_selector(DesignLocators.NewPagePopModal.TEMPLATES_CSS)
        template_index=javascript.find_element_index_by_text(self, template_obj_list, template_name)
        if template_index == None :
            raise KeyError(error_message)
        else :
            coreutils.left_click(self, template_obj_list[template_index])
        utils.synchronize_with_visble_text(self, "div[class*='pd-page-header'] div[class*='ibx-label-text']", visble_element_text='Page Heading', expire_time=20)
    
    def select_page_designer_application_menu(self, menu_name):
        """
        Descriptions : This method used to select page designer application menu such as Save, Save as, close, open
        Application menus : 'Open...' ,'New', 'Save', 'Save as...', 'Close'
        example usage : select_page_designer_application_menu('Save as...')
        """
        error_msg="[{0}] option not list out in page designer application menu".format(menu_name)
        application_icon_ob=self.driver.find_element_by_css_selector(DesignLocators.ToolBar.APPLICATION_BUTTON_CSS)
        coreutils.left_click(self, application_icon_ob)
        utils.synchronize_with_number_of_element(self, DesignLocators.ApplicationMenu.PARENT_CSS, expected_number=1, expire_time=8)
        application_menu_obj_list=self.driver.find_elements_by_css_selector(DesignLocators.ApplicationMenu.MENU_ITEMS_CSS)
        menu_index=javascript.find_element_index_by_text(self, application_menu_obj_list, menu_name)
        if menu_index==None :
            raise KeyError(error_msg)
        else :
            coreutils.left_click(self, application_menu_obj_list[menu_index])
    
    def page_designer_open_dialog_resources(self, folder_path=None, title=None, name=None, ok_button=False, cancel_button=False):
        """
        Descriptions : This method used to perform different operation such as save page, save as page, open page, etc
        example usage1 : page_designer_open_dialog_resources(folder_path='P116->S7074', title='Test')
        example usage2 : page_designer_open_dialog_resources(ok_button=True)
        """
        parent_css="div[class*='open-dialog-resources']"
        cancel_btn_css=parent_css + " div[data-ibx-name='btnCancel']"
        utils.synchronize_with_visble_text(self, cancel_btn_css, visble_element_text='Cancel', expire_time=15)
        if folder_path !=None :
            pass # Need to implement
        if title !=None :
            title_textbox_css = "div[data-ibx-name='sdtxtFileTitle'] input"
            title_textbox_obj = utils.validate_and_get_webdriver_object(self, title_textbox_css, 'Save dialog title text box')
            coreutils.left_click(self, title_textbox_obj)
            title_textbox_obj.clear()
            time.sleep(1)
            title_textbox_obj.send_keys(title)
            time.sleep(1)
        if name !=None :
            pass # Need to implement
        if ok_button == True :
            ok_button_css=parent_css + " div[data-ibx-name='btnOK']"
            ok_button_obj=self.driver.find_element_by_css_selector(ok_button_css)
            coreutils.left_click(self, ok_button_obj)
        if cancel_button == True :
            cancel_button_obj=self.driver.find_element_by_css_selector(cancel_btn_css)
            coreutils.left_click(self, cancel_button_obj)
        time.sleep(2)
        try : #Checking whether already file exits dialog box appear
            dialog_box_obs=self.driver.find_elements_by_css_selector("div[data-ibx-type='ibxDialog'][class*='pop-top']")
            if len(dialog_box_obs)>0 :
                PageDesignerMiscelaneous.dialog_box(self,  button_name_to_click='OK')
        except :
            pass
        
    def drag_and_drop_from_content_to_container_obj(self, content_folder_path, target_file_to_drag, container_obj_to_drop):
        """
        Descriptions : This method used to drag content items to page designer container
        :arg = container_obj_to_drop : container element object to drop target file
        example usage : drag_and_drop_from_content_to_container_obj('P116->S7074' ,'ACT-123', element_object)
        """
        (content_folder_path !=None) and PageDesignerMiscelaneous.expand_pd_content_folder(self, content_folder_path)
        target_file_obj = PageDesignerMiscelaneous.find_pd_content_item_and_scroll_into_view(self, target_file_to_drag)
        source_cord=coreutils.get_web_element_coordinate(self, target_file_obj)
        target_cord=coreutils.get_web_element_coordinate(self, container_obj_to_drop)
        coreutils.drag_and_drop(self, source_cord['x'], source_cord['y'], target_cord['x'], target_cord['y'])

    def verify_quick_filter_properties(self, property_name, expected_value, msg_step):
        """
        Descriptions : This method used to quick filter properties such as filter count value, backgroud_color, font size and etc..
        current property name = 'text' ,'background_color', 'font_size', 'position', 'text_align'
        example usage1 : verify_quick_filter_properties('background_color', 'mandy', 'Step 01.1 : Verify quick filter display')
        example usage2 : verify_quick_filter_properties('font_size', '2px', 'Step 01.1 : Verify quick filter display')
        """
        quick_filter=self.driver.find_element_by_css_selector(DesignLocators.ToolBar.QUICK_FILTER_BUTTON_CSS)
        if quick_filter.is_displayed()==True :
            if property_name=='text' :
                actaul_value=javascript.get_element_before_style_properties(self, quick_filter, 'content').replace('"', '')
                utils.asequal(self, expected_value, actaul_value, msg_step + ' : Verify quick filter display with ' + expected_value)
            if property_name=='background_color' :
                actaul_color=Color.from_string(javascript.get_element_before_style_properties(self, quick_filter, 'background-color')).rgb
                expected_color=utils.color_picker(self, expected_value, 'rgb')
                utils.asequal(self, expected_color, actaul_color, msg_step + ' : Verify quick filter display with ' + expected_color + ' color')
            if property_name=='font_size' :
                actaul_value=javascript.get_element_before_style_properties(self, quick_filter, 'font-size')
                utils.asequal(self, expected_value, actaul_value, msg_step + ' : Verify quick filter font size')
            if property_name=='position' :
                actaul_value=javascript.get_element_before_style_properties(self, quick_filter, 'position')
                utils.asequal(self, expected_value, actaul_value, msg_step + ' : Verify quick filter position')
            if property_name=='text_align' :
                actaul_value=javascript.get_element_before_style_properties(self, quick_filter, 'text-align')
                utils.asequal(self, expected_value, actaul_value, msg_step + ' : Verify quick filter text with '  + expected_value)
        else :
            raise ElementNotVisibleException('Quick Filter button not display')
    
    def get_pd_container_object(self, container_title, container_position=1):
        """
        Descriptions : This method used to pd container object based on container_title, container_position
        :arg- container_title = title of container
        :arg- container_position = if two containers have same title, that time we can pass container_position to get which container object
        example usage : get_pd_container_object('Panel 9', 1)
        """
        container_obj_list=self.driver.find_elements_by_css_selector(PD_LOCATORS.PD_CONTAINER_CSS)
        container_title_css=PD_LOCATORS.PD_CONTAINER_CSS + " div[data-ibx-name='_title']"
        container_title_obj=self.driver.find_elements_by_css_selector(container_title_css)
        container_title_list=javascript.get_elements_text(self, container_title_obj)
        container_index_list=[index for index, title in enumerate(container_title_list) if title==container_title]
        if len(container_index_list)>0 :
            container_object=container_obj_list[container_index_list[container_position-1]]
            return container_object
        else :
            error_msg="['{0}'] container not display in page designer canvas".format(container_title)
            raise KeyError(error_msg)
    
    def get_pd_filter_control_object(self, filter_control_name, filter_control_position=1, grid_container_title=None, model_window=False):
        """
        Descriptions : This method used to pd filter control object based on filter control name and container_position
        :arg- filter_control_name = filter control label
        :arg- filter_control_position = if two filter control have same title, that time we can pass filter_control_position to get which control object
        example usage : get_pd_filter_control_object('Business Region:', 1)
        """
        filter_grid = PageDesignerMiscelaneous.get_filter_grid_object(self, grid_container_title, model_window)
        filter_panel_parent_css="div[data-ibx-type='pdFilterPanel']"
        filter_panel_objs=filter_grid.find_elements_by_css_selector(filter_panel_parent_css)
        filter_condition_labels_css=filter_panel_parent_css + " div[class*='pd-amper-label'] div[class='ibx-label-text']"
        filter_condition_objs=filter_grid.find_elements_by_css_selector(filter_condition_labels_css)
        filter_control_name_list=javascript.get_elements_text(self, filter_condition_objs)
        container_index_list=[index for index, name in enumerate(filter_control_name_list) if name==filter_control_name]
        if len(container_index_list)>0 :
            container_object=filter_panel_objs[container_index_list[filter_control_position-1]]
            return container_object
        else :
            error_msg="['{0}'] filter condition not display in filter panel".format(filter_control_name)
            raise KeyError(error_msg)
    
    def get_filter_grid_object(self, grid_container_title=None, model_window=False):
        """
        Descriptions : This method will return filter grid object
        :arg - grid_container_title = 'Panel1'. If you want to get filter grid object from inside the grid container then pass grid container title. 
        :arg - model_window = True. If you want to get filter grid object from filter model window then pass model_window=True
        :Usage - get_filter_grid_object()
        """
        parent_object = None
        if model_window == True :
            filter_grid_css = PD_LOCATORS.FILTER_MODEL_WINDOW_GRID_CSS
            object_name = "Filter model window grid"
        elif grid_container_title != None :
            parent_object =  PageDesignerMiscelaneous.get_pd_container_object(self, grid_container_title)
            filter_grid_css = PD_LOCATORS.FILTER_GRID_PARENT_CSS
            object_name = "[{0}] container filter grid".format(grid_container_title)
        else :
            filter_grid_css = PD_LOCATORS.FILTER_BAR_GRID_CSS
            object_name = "Filter bar grid"
        filter_grid_object = utils.validate_and_get_webdriver_object(self, filter_grid_css, object_name, parent_object=parent_object)
        if filter_grid_object.is_displayed():
            return filter_grid_object
        else :
            error_msg = "[{0}] not visible".format(object_name)
            raise ElementNotVisibleException(error_msg)
        
    def get_pd_container_title_bar_visible_button_name(self, container_title, container_position=1):
        """
        Descriptions : This method used to get visible container title bar button names
        current visible icon names=['Maximize', 'Options'] and ['Restore', 'Options']
        """
        visible_button_name=[]
        current_button_title={'Maximize' : 'expand', 'Options':'ellipsis-v', 'Restore' : 'compress'}
        button_title_css="div[class*='pd-container-title-button']"
        icon_css=button_title_css + " div[class*='ibx-icons ibx-glyph-{0}']"
        container_obj=PageDesignerMiscelaneous.get_pd_container_object(self, container_title, container_position)
        button_name_obj=container_obj.find_elements_by_css_selector(button_title_css)
        actual_button_title=[icon.get_attribute('title') for icon in button_name_obj]
        for title in actual_button_title :
            icon_obj=container_obj.find_elements_by_css_selector(icon_css.format(current_button_title[title]))
            if len(icon_obj)==1 :
                if icon_obj[0].is_displayed()==True :
                    visible_button_name.append(title)
        return visible_button_name
    
    def get_page_heading_visible_button_name(self):
        """
        Descriptions : This method used to get visible container title bar button names
        current visible icon names=['Refresh', 'Show filters', 'Share', 'Delete']
        """
        visible_button_name=[]
        current_page_heading_buttons={'Refresh' : 'refresh', 'Show filters':'filter', 'Share' : 'share', 'Delete' : 'trash'}
        button_css=" div[class*='pd-header-button'][title]:not([style*='none'])"
        icon_css="div[class*='ibx-label-icon'][class*='{0}']"
        page_header_obj=self.driver.find_element(*PD_LOCATORS.PAGE_HEADER)
        button_obj_list=page_header_obj.find_elements_by_css_selector(button_css)
        for button_obj in button_obj_list :
            button_title=button_obj.get_attribute('title').strip()
            icon_obj=button_obj.find_elements_by_css_selector(icon_css.format(current_page_heading_buttons[button_title]))
            if len(icon_obj)==1 :
                if icon_obj[0].is_displayed()==True :
                    visible_button_name.append(button_title)
        return visible_button_name
    
    def verify_page_designer_component_is_selected(self, component_parent_obj, border_width, border_rgb_value, border_style, position, msg, selection_class='pd-selection'):
        """
        Descriptions : This method used to verify any page designer component is selected such as page section , filter box, page canvas
        all page component selected by border
        example usage : verify_page_designer_component_is_selected(parent_obj, border_width='1px', border_rgb_value='rgb(212, 73, 73)', border_style='dashed', position='absolute', msg='Step 01.1 : Verify page container selected')
        """
        directions=['north', 'south', 'east', 'west']
        selection_xpath="./div[@class='{0} {1}']"
        for direction in directions :
            selection_obj=component_parent_obj.find_element_by_xpath(selection_xpath.format(selection_class, direction))
            selection_visible=selection_obj.is_displayed()
            selection_border_width=utils.get_element_border_property_value(self, selection_obj, 'width')
            selection_border_rgb=Color.from_string(utils.get_element_border_property_value(self, selection_obj, 'color')).rgb
            selection_border_style=utils.get_element_border_property_value(self, selection_obj, 'style')
            selection_position=selection_obj.value_of_css_property('position')
            if selection_visible==True and selection_border_width==border_width and selection_border_rgb==border_rgb_value and selection_border_style==border_style and selection_position==position :
                status=True
            else :
                print(selection_visible, selection_border_width, selection_border_rgb, selection_border_style, selection_position)
                status=False
                break
        utils.asequal(self, True, status, msg)
    
    def switch_to_previous_window(self, driver_close=True):
        '''
        Descriptions:- This function will switch the control back to previous window by closing the current window.
        :arg - driver_close = if already current window closed then pass driver_close=False
        '''
        if driver_close==True :
            self.driver.close()
        time.sleep(Global_variables.longwait)
        coreutils.update_window_handles_list(self, update='remove')
        self.driver.switch_to.window(Global_variables.windows_handles[-1])
        coreutils.update_current_working_area_browser_specification(self)
        time.sleep(Global_variables.shortwait)
    
    def create_report_data_set_using_bs4(self, file_name_to_save, table_css):
        """
        This function read all data from report table and convert as xlsx file using BeautifulSoup4 module
        :arg = table_css ='##ITableData0' or table['summary'] ( table_css is parent css value of table )
        :arg = file_name = 'C12345_TableData_Step_03'
        example usage = create_report_data_set_using_bs4('C12345', table['summary'])
        """
        save_file_name=file_name_to_save + EXCEL_FILE_EXTENSION
        save_file_path=os.path.join(os.getcwd(), SCRIPT_DATA_FOLDER_NAME, save_file_name)
        if os.path.exists(save_file_path) :
            error_msg = "'{0}' file already exists in data folder. If you want to create new data set file then delete '{0}' file from data folder".format(save_file_name)
            raise FileExistsError(error_msg)
        else :
            workbook = Workbook()
            sheet = workbook.get_sheet_by_name(WORKBOOK_SHEET_NAME)
            table_rows=PageDesignerMiscelaneous.get_beautifulsoup_table_object(self, table_css)
            for r, row in enumerate(table_rows) :
                row_columns=row.select(' > td ')
                for c,column in enumerate(row_columns) :
                    sheet.cell(row=r + 1, column=c + 1).value = str(column.get_text(strip=True)).strip()
            workbook.save(save_file_path) #save the xlsx file
    
    def verify_report_table_data_using_bs4(self, saved_file_name, message, table_css):
        """
        This function compare two report table data
        :arg = table_css ='#ITableData0' or table['summary'] ( table_css is parent css value of table )
        :arg = saved_file_name = 'C12345_TableData_Step_03'
        example usage = create_report_data_set_using_bs4('C12345', table['summary'])
        """
        compare_status=PageDesignerMiscelaneous.__compare_report_table_data_using_bs4(self, table_css, saved_file_name)
        expected_result=1
        actual_result=len(compare_status)
        utils.asequal(self, expected_result, actual_result, message + ' : ' + str(compare_status))
    
    def __compare_report_table_data_using_bs4(self, table_css, saved_file_name):
        """
        This function read all the data from report table using BeautifulSoup4 module and 
        compare with expected report table xlsx file which created by convert_report_table_data_to_xl_using_bs4()
        It will return status=[0] if two xlsx files are same
        """
        table_rows=PageDesignerMiscelaneous.get_beautifulsoup_table_object(self, table_css)# get all rows of table
        sheet=PageDesignerMiscelaneous.get_workboot_sheet_object(self, saved_file_name)
        status=[]
        for r, row in enumerate(table_rows) :
            row_columns=row.select(' > td ')
            for c,column in enumerate(row_columns) :
                cell_value=sheet.cell(row=r + 1, column=c + 1).value
                cell_value='' if cell_value==None else cell_value
                if cell_value== str(column.get_text(strip=True)).strip():
                    status=[0]
                    continue
                else :
                    status=['Row = ' + str(r+1), 'Column = ' + str(c+1), 'Expected = ' + cell_value, 'Actual = ' + str(column.get_text(strip=True)).strip()]
                    return status
        return status
    
    def get_workboot_sheet_object(self, file_name):
        """
        This method used to get saved report xl file and ceate wookbook sheet object 
        """
        saved_file_name=file_name + EXCEL_FILE_EXTENSION
        saved_file_path=os.path.join(os.getcwd(), SCRIPT_DATA_FOLDER_NAME , saved_file_name)
        if os.path.exists(saved_file_path) :
            workbook=load_workbook(saved_file_path) 
            sheet = workbook.get_sheet_by_name(WORKBOOK_SHEET_NAME)
            return sheet
        else :
            error_message=saved_file_name +' data set file not found in data folder'
            raise FileNotFoundError (error_message)
        
    def get_beautifulsoup_table_object(self, table_css):
        """
        This function return the beautifulsoup table object
        """
        NO_TABLE_ROWS_ERROR_MSG="Unable not find any rows in {0} table css".format(table_css)
        TABLE_NOT_VISIBLE_ERROR="The given table ['{0}'] css element not visible in screen".format(table_css)
        table_visible=self.driver.find_element_by_css_selector(table_css).is_displayed()
        if table_visible == True :
            from bs4 import BeautifulSoup
            script='return document.querySelector("{0}").innerHTML'.format(table_css)
            page_source=self.driver.execute_script(script)
            beautifulsoup = BeautifulSoup(page_source, 'html.parser')
            table_rows=beautifulsoup.select(' > tbody > tr')
            if len(table_rows)>0 :
                return table_rows
            else :
                raise ValueError(NO_TABLE_ROWS_ERROR_MSG)
        else :
            raise ElementNotVisibleException(TABLE_NOT_VISIBLE_ERROR)
    
    def dialog_box(self, button_name_to_click=False, close_dialog=False, title_text=False, dialog_msg=False):
        """
        Descriptions : This method used to handle dialog box such as click on Ok, Yes, No and etc
        example usage : dialog_box(button_name_to_click = 'Yes')
        """
        dialogbox_parent_css = "div[data-ibx-type='ibxDialog'][class*='pop-top']"
        dialog_buttons_css = dialogbox_parent_css + " div[data-ibx-type='ibxButton']:not([style*='none'])>div[class='ibx-label-text']"
        utils.synchronize_with_number_of_element(self, dialogbox_parent_css, expected_number=1, expire_time=10)
        dialogbox_parent_obj=self.driver.find_element_by_css_selector(dialogbox_parent_css)
        if title_text==True :
            pass
        if dialog_msg==True :
            pass
        if button_name_to_click != None :
            dialog_buttons_obj = self.driver.find_elements_by_css_selector(dialog_buttons_css)
            dialog_button_index = javascript.find_element_index_by_text(self, dialog_buttons_obj, button_name_to_click)
            if dialog_button_index == None :
                error = '{0} button not displayed in dialog box'
                raise KeyError(error)
            else :
                dialog_button_obj = dialog_buttons_obj[dialog_button_index]
                coreutils.left_click(self, dialog_button_obj)
        if close_dialog == True :
            close_btn_obj=dialogbox_parent_obj.find_element_by_css_selector("div[data-ibx-name='titleClose']")
            coreutils.left_click(self, close_btn_obj)
        time.sleep(2)
    
    def select_page_designer_context_menu(self, menu_name_to_select):
        """
        Descriptions : This method used to select page designer context menu
        example usage : select_page_designer_context_menu('Settings') 
        """
        popup_css="div[data-ibx-type='ibxMenu'][class*='pop-top']"
        menu_item_css=" div[data-ibx-type='ibxMenuItem']:not([style*='none']) div[class='ibx-label-text']"
        utils.synchronize_with_number_of_element(self, popup_css, 1, expire_time=8)
        context_menus_css=popup_css + menu_item_css
        context_menu_obj_list=self.driver.find_elements_by_css_selector(context_menus_css)
        context_menu_index=javascript.find_element_index_by_text(self, context_menu_obj_list, menu_name_to_select)
        if context_menu_index==None :
            error = '[{0}] menu not displayed in context menu'.format(menu_name_to_select)
            raise KeyError(error)
        else :
            menu_obj=context_menu_obj_list[context_menu_index]
            coreutils.left_click(self, menu_obj)
    
    def run_page_using_api(self, page_name, mrid=None, mrpass=None, folder=None, login=True):
        """
        Description : Run the saved page using API url
        :Usage - run_page_using_api("C1234567")
        """
        node = coreutils.parseinitfile(self, 'nodeid')
        port = coreutils.parseinitfile(self, 'httpport')
        context = coreutils.parseinitfile(self, 'wfcontext')
        if folder == None :
            project =  coreutils.parseinitfile(self, 'project_id')
            suite =  coreutils.parseinitfile(self, 'suite_id')
            group = coreutils.parseinitfile(self, 'group_id')
            folder = project + "_" + suite + "/" + group
        api_url = "http://{0}:{1}{2}/run.bip?BIP_REQUEST_TYPE=BIP_LAUNCH&BIP_folder=IBFS:/WFC/Repository/{3}/&BIP_item={4}".format(node, port, context, folder, page_name.lower())
        self.driver.get(api_url)
        if login :
            mrid = 'mrid' if mrid==None else mrid
            mrpass ='mrpass' if mrpass==None else mrpass
            utils.login_wf(self, mrid, mrpass)
        utils.synchronize_until_element_is_visible(self, ".runner", 240)