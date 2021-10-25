'''
Created on Mar 22, 2017

@author: Robert

Test Case = http://lnxtestrail.ibi.com/testrail/index.php?/cases/view/2229123
Test case Name =  Verify Export data
'''
import unittest
from common.pages import visualization_resultarea, visualization_ribbon, visualization_metadata,wf_map
from common.lib import utillity
from common.lib import core_utility
import time, os
from common.lib.basetestcase import BaseTestCase
from openpyxl import load_workbook

class C2229123_TestClass(BaseTestCase):

    def test_C2229123(self):
        utillobj = utillity.UtillityMethods(self.driver)
        resultobj = visualization_resultarea.Visualization_Resultarea(self.driver)
        ribbonobj = visualization_ribbon.Visualization_Ribbon(self.driver)
        metaobj=visualization_metadata.Visualization_Metadata(self.driver)
        core_utility_obj=core_utility.CoreUtillityMethods(self.driver)
        wfmapobj=wf_map.Wf_Map(self.driver)
        base_file="C2229123_Ds01.xlsx"
        actual_file="C2229123_Ds01_actual.xlsx"
        
        def compare_using_openpyxl(rows, columns):
            workbook1 = load_workbook(os.getcwd() + "\\data\\" + base_file)
            sheet1 = workbook1.get_sheet_by_name('Sheet1')
            workbook2 = load_workbook(os.getcwd() + "\\data\\" + actual_file)
            sheet2 = workbook2.get_sheet_by_name('Sheet1')
            status={'status':'pass'}
            for r in range(1, rows+1):
                for c in range(1, columns+1):
                    if type(sheet1.cell(row=r, column=c).value) == float:
                        base=round(sheet1.cell(row=r, column=c).value, 2)
                        act=round(sheet2.cell(row=r, column=c).value, 2)
                    else:
                        base=sheet1.cell(row=r, column=c).value
                        act=sheet2.cell(row=r, column=c).value
                    if base != act:   
                        status['status']= 'fail at row number [' + str(r) + '] and column number [' + str(c) + '].' + ' Expected is: [' + str(base) + '].Actual is: ['+str(act)+ '].'
                        return(status)
            return(status)
        
        
        '''    TESTCASE ID Variable    '''
        Test_Case_ID = "C2229123"
        
        '''    1. Launch the IA API with Chart in edit mode (edit the domain, port and alias of the URL - do not use the URL as is):    '''
        
        utillobj.infoassist_api_login('idis','baseapp/wf_retail_lite','P292/S10032', 'mrid', 'mrpass')
        parent_css='#TableChart_1'
        utillobj.synchronize_with_number_of_element(parent_css, 1, ribbonobj.chart_long_timesleep)
         
        '''    2. Click "Change" dropdown > "Map"    '''
        '''    3. Select "Leaflet Bubble" > OK (Keep default Territory "United States" selected)    '''
        
        ribbonobj.change_chart_type('map')
        utillobj.synchronize_until_element_is_visible("div[id^='QbDialog']", ribbonobj.home_page_long_timesleep)
        ribbonobj.select_map('bubble', teritory='United States of America', btn_click='ok')
        utillobj.synchronize_until_element_disappear("div[id^='QbDialog']", ribbonobj.home_page_long_timesleep)
        
        '''    4. drag "Store,State,Province" into Geolocation bucket    '''
        metaobj.drag_drop_data_tree_items_to_query_tree('Store,State,Province', 1, 'Layer', 0)
        
        '''    5. Select "state_name" from Geographic Role dropdown > OK   '''
        utillobj.synchronize_until_element_is_visible("div[id^='QbDialog'] div[id$='geoRoleComboBox'] div[id^='BiButton']", ribbonobj.chart_medium_timesleep)
        wfmapobj.set_location_geo_role(role_name='state_name (Alabama, Alaska, Arizona)', btn_click='Ok')
        utillobj.synchronize_until_element_disappear("div[id^='QbDialog'] div[id$='geoRoleComboBox'] div[id^='BiButton']", ribbonobj.chart_medium_timesleep)
        
        '''    6. Right click "Revenue" > "Add To Query" > "Size"    '''
        
        metaobj.datatree_field_click('Revenue', 1, 1, 'Add To Query', 'Size')
        
        '''    7. Verify the map is updated    '''
        
        parentcss="TableChart_1"
        utillobj.synchronize_with_number_of_element("#TableChart_1 path[class^='riser!s0!g25!mstate!']", 1, ribbonobj.chart_long_timesleep)
        
        expected_label_list=['Revenue']
        msg="Step 7.1. Verify Map Legends"
        resultobj.verify_riser_pie_labels_and_legends(parentcss, expected_label_list, msg, custom_css="text[class^='legend-labels']", same_group=True)
        expected_title_list=['Revenue']
        resultobj.verify_riser_pie_labels_and_legends(parentcss, expected_title_list, 'Step 7.2. Verify Legend title', custom_css="text[class^='sizeLegend-title']", same_group=True)
          
        utillobj.verify_chart_color(parentcss, "riser!s0!g25!mstate!", 'bar_blue', 'Step 7.3a Verify map color')
        utillobj.verify_chart_color(parentcss, "riser!s0!g3!mstate!", 'bar_blue', 'Step 7.3b Verify map color')
        
        '''    '8. Click the Show Data menu dropdown    '''
        '''    '9. Select "Export Data" > "Summary"    '''
        
        resultobj.select_panel_caption_btn(0, select_type='menu')
        utillobj.synchronize_until_element_is_visible("div[id^='BiPopup'][style*='inherit']", resultobj.home_page_medium_timesleep)
        utillobj.select_bipopup_list_item('Export Data',element_location='middle')
        utillobj.synchronize_until_element_is_visible("div[id^='BiPopup'][style*='inherit']", resultobj.home_page_short_timesleep)
        utillobj.select_or_verify_bipop_menu('Summary')
        utillobj.synchronize_until_element_disappear("div[id^='BiPopup'][style*='inherit']", resultobj.home_page_short_timesleep)
        
        '''    10. Download and Open the Excel file    '''
        '''    11. Verify the excel spreadsheet contains the data from the fields    '''
        '''    12. Dismiss Excel    '''
        
        core_utility_obj.switch_to_new_window()
        self.driver.maximize_window()
        utillobj.save_file_from_browser('C2229123_Ds01_actual',custom_cr_re=".*busy.*Chrome", custom_ie_re='Visual1.*')
        time.sleep(10)
        val=compare_using_openpyxl(15, 2)
        utillobj.asequal(val["status"],'pass', 'Step 11. Verify excel sheet after export')
        
        '''    13. Click "Save" icon    '''
        '''    14. Enter Title "C2229123"    '''
        '''    15. Click "Save" and dismiss IA    '''
        
        filter_box_css="#qbFilterBox"
        filter_box_elem=utillobj.validate_and_get_webdriver_object(filter_box_css,"Filter_pane_css")
        utillobj.click_on_screen(filter_box_elem,"bottom_right",0)
        ribbonobj.select_top_toolbar_item('toolbar_save')
        utillobj.synchronize_until_element_is_visible("#IbfsOpenFileDialog7_cbFileName input", resultobj.home_page_short_timesleep)
        utillobj.ibfs_save_as(Test_Case_ID)
        utillobj.synchronize_until_element_disappear("#IbfsOpenFileDialog7_cbFileName input", resultobj.home_page_short_timesleep)
        
if __name__ == '__main__':
    unittest.main()