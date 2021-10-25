'''
Created on Sep 13, 2016

@author: Sindhuja

Test Suite = http://lnxtestrail.ibi.com/testrail/index.php?/suites/view/7074
Test Case = http://lnxtestrail.ibi.com/testrail/index.php?/cases/view/2055507
'''
from common.lib.basetestcase import BaseTestCase
from common.lib import utillity
import unittest
from openpyxl import Workbook
from openpyxl import load_workbook
import os

class C2055507_TestClass(BaseTestCase):

    def test_C2055507(self):
        
        """
            Step 01: Execute C2055507.fex form repro
        """
        driver = self.driver #Driver reference object created'
#         driver.implicitly_wait(35) #Intializing common implicit wait for throughout the test
        utillobj = utillity.UtillityMethods(self.driver)
        utillobj.active_run_fex_api_login("C2055507.fex", "S7215", 'mrid', 'mrpass')
        
        """
        Step 02: Check all the 20 lines records are displayed properly.
        """
            
        def compare_pivot_data_set( table_id, file_name):
            wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
            rows_css = "#"+table_id+" > tbody > tr:nth-child(3) > td > table > tbody > tr"
            status=[]
            s1 = wb1.get_sheet_by_name('Sheet')
            table_rows = self.driver.find_elements_by_css_selector(rows_css)
            for r in range(len(table_rows)):
                col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
                columns = table_rows[r].find_elements_by_css_selector(col_css)
                for c in range(len(columns)):
                    if (s1.cell(row=r + 1, column=c + 1).value):
                        value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text.strip()
                        if s1.cell(row=r + 1, column=c + 1).value.strip() == str(value):
                            status=[0]
                            continue
                        else:
                            status=[r+1,c]
                            return (status)
            return (status)
        
        def verify_pivot_data_set(table_id, file_name, msg):
            x= compare_pivot_data_set(table_id, file_name)
            utillobj.asequal(len(x),1,msg+ ' Row,Column -'+ str(x))
        
        verify_pivot_data_set('piv-999', 'C2055507.xlsx', 'Step 02: Check all the 20 lines records are displayed properly')


if __name__ == '__main__':
    unittest.main()     
    
    
    
    
    
    
    
    
    
    
    
    