'''
Created on Aug 17, 2016

@author: Sindhuja

Test Suite = http://lnxtestrail.ibi.com/testrail/index.php?/suites/view/7070&group_by=cases:section_id&group_order=asc
Test Case = http://lnxtestrail.ibi.com/testrail/index.php?/cases/view/2050448
'''
import unittest
from common.lib.basetestcase import BaseTestCase
from common.lib import utillity
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import os
from openpyxl import Workbook



class C2050448_TestClass(BaseTestCase):

    def test_C2050448(self):
        driver = self.driver
        utillobj = utillity.UtillityMethods(self.driver)
        """
        1. Execute the 89280.fex.
        """
        utillobj.active_run_fex_api_login("89280.fex", "S7070", 'mrid', 'mrpass')
        
        
        def verify_pivot_data_set(self, table_id, file_name, msg):
            x= compare_pivot_data_set(self,table_id, file_name)
            utillobj.asequal(len(x),1,msg+ ' Row,Column -'+ str(x))
            
        def compare_pivot_data_set(self, table_id, file_name):
            wb1 = load_workbook(os.getcwd() + "\data\\" + file_name)
            rows_css = "#piv-999 > tbody > tr:nth-child(3) > td >table > tbody > tr"
            status=[]
            s1 = wb1.get_sheet_by_name('Sheet')
            table_rows = self.driver.find_elements_by_css_selector(rows_css)
            for r in range(len(table_rows)):
                col_css=rows_css + ":nth-child(" + str(r+1) + ") > td"
                columns = table_rows[r].find_elements_by_css_selector(col_css)
                for c in range(len(columns)):
                    if (s1.cell(row=r + 1, column=c + 1).value):
                        value=self.driver.find_element_by_css_selector(col_css + ":nth-child(" + str(c+1) + ")").text
                        if s1.cell(row=r + 1, column=c + 1).value == str(value):
                            status=[0]
                            continue
                        else:
                            status=[r+1,c]
                            return (status)
            return (status)

        verify_pivot_data_set(self,'#piv-999', 'C2050448_Ds01.xlsx', 'Step 01: Verify data set')
        
        """
        Step 02: From the Pivot report, click the curved arrow above the CAR column
        """
        self.driver.find_element(By.CSS_SELECTOR, '[onclick="ibiChart.makeChartAddYcol(-999,1,false,2,1,0,-1)"]').click()
    
        verify_pivot_data_set(self,'#piv-999', 'C2050448_Ds02.xlsx', 'Step 02: Expect to see the CAR column become the Across sort, with COUNTRY remaining as the BY column')
        
       
if __name__ == '__main__':
    unittest.main()
    
        
        
        
        
        